#!/usr/bin/env python3

"""
mAIpper v0.7 - Pentest Tool Analysis & Obsidian Export Tool

Changes from v0.6:
  - Nessus parser: ingests .nessus XML files from scans/nessus/; adds
    ## Nessus Findings sections to host notes (severity-sorted); creates
    Scans/<name> - Nessus.md; AI analysis focused on CVEs and exploitation
    priority; nessus_max_severity stored in frontmatter for canvas coloring
  - Burp Suite parser: ingests Burp Pro Scanner XML from scans/burp/; adds
    ## Burp Suite Findings sections to host notes; creates
    Scans/<name> - Burp.md; AI analysis focused on web attack surface
  - Excel export (--excel): generates <vault>/Export.xlsx with Summary,
    Nmap, Nessus, and Burp sheets; severity color-coding, filters, frozen
    headers, auto-fit columns
  - --no-nessus / --no-burp flags to skip those parsers
  - Canvas Campaign Overview and Next Steps updated with Nessus/Burp data
  - Nmap re-runs preserve existing Nessus/Burp sections in host notes
  - Canvas "Priority Targets" text node: AI-ranked list of highest-priority
    hosts (by CVSS, CVEs, service exposure, Burp findings); positioned between
    Campaign Overview and subnet groups; edges Overview→PriorityTargets and
    PriorityTargets→each subnet group; falls back to static severity sort when
    Ollama is skipped; stable node ID "priority-targets"

Requirements:
  pip install requests
  pip install openpyxl   # only needed for --excel

Author: Zachary Levine
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import logging
import math
import re
from pathlib import Path
import xml.etree.ElementTree as ET

import requests

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# Constants
# ============================================================

IPV4_RE = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$")

STATUS_COLORS: dict[str, str | None] = {
    "not-started": None,
    "in-progress":  "3",   # yellow
    "done":         "4",   # green
    "exploited":    "1",   # red
    "blocked":      "6",   # purple
}

# Canvas colors for Nessus severity (applied when status == not-started)
NESSUS_SEVERITY_CANVAS_COLORS: dict[int, str] = {
    4: "1",   # Critical → red
    3: "5",   # High     → orange
    2: "3",   # Medium   → yellow
}

OPERATOR_NOTES_SENTINEL = "## Operator Notes"
OPERATOR_NOTES_HINT = "_Add your own findings, observations, and next steps below._"

# Canonical section order in host notes
BODY_SECTION_ORDER = [
    "## Open Ports",
    "## Nessus Findings",
    "## Burp Suite Findings",
    "## Scan References",
    OPERATOR_NOTES_SENTINEL,
]

CARD_W          = 360
CARD_H          = 220
CARD_GAP_X      = 60
CARD_GAP_Y      = 60
GROUP_PAD       = 60
GROUP_LABEL_H   = 40
SCAN_CARD_W     = 360
SCAN_CARD_H     = 160
OVERVIEW_W      = 720
OVERVIEW_H      = 220
NEXT_STEPS_W    = 640
NEXT_STEPS_H    = 420
GROUP_GAP       = 120
ROW_GAP         = 120


# ============================================================
# Basic helpers
# ============================================================

def safe_filename(s: str) -> str:
    for ch in '<>:"/\\|?*\n\r\t':
        s = s.replace(ch, "_")
    return s.strip().strip(".")


def stable_id(key: str) -> str:
    return hashlib.md5(key.encode()).hexdigest()[:12]


def is_ipv4(value: str) -> bool:
    return bool(value and IPV4_RE.match(value.strip()))


def is_probable_fqdn(value: str) -> bool:
    v = (value or "").strip().rstrip(".")
    return bool(v and "." in v and not is_ipv4(v))


def ensure_md_suffix(name: str) -> str:
    return name if name.lower().endswith(".md") else f"{name}.md"


def get_subnet_label(ip: str) -> str:
    if not is_ipv4(ip):
        return "unknown"
    parts = ip.split(".")
    return f"{parts[0]}.{parts[1]}.{parts[2]}.0/24"


def severity_int_to_str(n: int) -> str:
    return {0: "Info", 1: "Low", 2: "Medium", 3: "High", 4: "Critical"}.get(n, "Unknown")


def severity_str_to_int(s: str) -> int:
    return {"information": 0, "info": 0, "low": 1, "medium": 2, "high": 3, "critical": 4}.get(
        (s or "").lower(), 0
    )


# ============================================================
# Frontmatter helpers
# ============================================================

def _fm_encode(v: object) -> str:
    if isinstance(v, list):
        return json.dumps(v, ensure_ascii=False)
    if v is None:
        return "null"
    return str(v)


def write_frontmatter(fm: dict) -> str:
    lines = ["---"]
    for k, v in fm.items():
        lines.append(f"{k}: {_fm_encode(v)}")
    lines.append("---")
    return "\n".join(lines) + "\n"


def read_frontmatter(text: str) -> tuple[dict, str]:
    if not text.startswith("---\n"):
        return {}, text
    end = text.find("\n---\n", 4)
    if end == -1:
        return {}, text
    fm_text = text[4:end]
    body = text[end + 5:]
    fm: dict = {}
    for line in fm_text.splitlines():
        if ": " not in line and not line.endswith(":"):
            continue
        key, _, raw = line.partition(": ")
        key = key.strip()
        raw = raw.strip()
        if not key:
            continue
        try:
            fm[key] = json.loads(raw)
        except (json.JSONDecodeError, ValueError):
            fm[key] = raw if raw else None
    return fm, body


def extract_operator_notes(body: str) -> str:
    idx = body.find(OPERATOR_NOTES_SENTINEL)
    if idx == -1:
        return ""
    after = body[idx + len(OPERATOR_NOTES_SENTINEL):]
    lines = after.splitlines()
    content_lines: list[str] = []
    past_hint = False
    for line in lines:
        if not past_hint and line.strip() in ("", OPERATOR_NOTES_HINT):
            continue
        past_hint = True
        content_lines.append(line)
    while content_lines and not content_lines[-1].strip():
        content_lines.pop()
    return "\n".join(content_lines)


def extract_body_section(body: str, section_header: str) -> str:
    """Extract content of a named ## section (without its header line)."""
    pattern = rf"(?:^|\n){re.escape(section_header)}\n(.*?)(?=\n## |\Z)"
    m = re.search(pattern, body, re.DOTALL)
    return m.group(1).strip() if m else ""


# ============================================================
# Port hints & tags
# ============================================================

def get_tags_from_ports(open_ports: list) -> list[str]:
    tags: set[str] = set()
    port_nums = {p.get("port", 0) for p in open_ports}
    svc_names = {(p.get("service", {}).get("name", "") or "").lower() for p in open_ports}

    if port_nums & {80, 81, 443, 8000, 8080, 8443} or svc_names & {"http", "https", "http-proxy"}:
        tags.add("web")
    if 445 in port_nums or svc_names & {"microsoft-ds", "smb"}:
        tags.add("smb")
    if port_nums & {88, 464} or any("kerberos" in s for s in svc_names):
        tags.add("kerberos")
    if port_nums & {389, 636, 3268, 3269} or svc_names & {"ldap", "ldaps", "globalcatldap", "globalcatldapssl"}:
        tags.add("ldap")
    if 135 in port_nums or "msrpc" in svc_names:
        tags.add("rpc")
    if port_nums & {1433, 1434} or any("mssql" in s for s in svc_names):
        tags.add("mssql")
    if 3306 in port_nums or "mysql" in svc_names:
        tags.add("mysql")
    if 5432 in port_nums or any("postgres" in s for s in svc_names):
        tags.add("postgres")
    if 22 in port_nums or "ssh" in svc_names:
        tags.add("ssh")
    if 3389 in port_nums or "ms-wbt-server" in svc_names:
        tags.add("rdp")
    if port_nums & {5985, 5986} or any("winrm" in s for s in svc_names):
        tags.add("winrm")
    if 21 in port_nums or "ftp" in svc_names:
        tags.add("ftp")
    if 53 in port_nums or "domain" in svc_names:
        tags.add("dns")
    if port_nums & {161, 162} or "snmp" in svc_names:
        tags.add("snmp")
    if port_nums & {25, 465, 587} or svc_names & {"smtp", "smtps", "submission"}:
        tags.add("smtp")
    if 2049 in port_nums or "nfs" in svc_names:
        tags.add("nfs")
    if 6379 in port_nums or any("redis" in s for s in svc_names):
        tags.add("redis")
    if {"kerberos", "ldap", "smb"} <= tags:
        tags.add("domain-controller")
    return sorted(tags)


def get_port_hints(port: int, service_name: str, product: str = "", extrainfo: str = "") -> list[str]:
    svc   = (service_name or "").lower()
    prod  = (product or "").lower()
    extra = (extrainfo or "").lower()
    hints: list[str] = []

    if port in (80, 81, 443, 8000, 8080, 8443) or svc in ("http", "https", "http-proxy"):
        hints.append("Web service detected; consider whatweb, nikto, gobuster/feroxbuster/ffuf, curl, and manual browsing.")
        hints.append("Check for default credentials, admin panels, exposed APIs, virtual hosts, and interesting headers.")
    if port == 445 or svc in ("microsoft-ds", "smb", "netbios-ssn"):
        hints.append("SMB detected; consider netexec smb, smbclient, smbmap, and enum4linux-ng.")
        hints.append("Check share access, null sessions, SMB signing, local admin reuse, and domain membership clues.")
    if port in (88, 464) or "kerberos" in svc or "kerberos" in prod:
        hints.append("Kerberos-related service detected; consider kerbrute, netexec, GetUserSPNs, and AS-REP roast checks.")
        hints.append("Look for domain naming clues, valid usernames, SPNs, and clock skew issues.")
    if port in (389, 636, 3268, 3269) or svc in ("ldap", "ldaps", "globalcatldap", "globalcatldapssl"):
        hints.append("LDAP detected; consider ldapsearch, netexec ldap, and directory enumeration.")
        hints.append("Look for anonymous bind, naming contexts, domain structure, users, groups, and password policy info.")
    if port == 135 or svc == "msrpc":
        hints.append("MSRPC detected; consider rpcclient, netexec, and enumerate Windows service exposure and domain clues.")
    if port == 139 or svc == "netbios-ssn":
        hints.append("NetBIOS detected; enumerate shares, names, and Windows host information.")
    if port in (5985, 5986) or "winrm" in svc:
        hints.append("WinRM detected; consider netexec winrm and evil-winrm if credentials are obtained.")
    if port in (1433, 1434) or "mssql" in svc or "sql server" in prod:
        hints.append("MSSQL detected; consider impacket-mssqlclient, netexec mssql, and SQL login enumeration.")
    if port == 3306 or "mysql" in svc:
        hints.append("MySQL detected; test authentication paths and enumerate version-specific exposure carefully.")
    if port == 5432 or "postgres" in svc:
        hints.append("PostgreSQL detected; test for weak/default credentials and accessible databases.")
    if port == 27017 or "mongodb" in svc:
        hints.append("MongoDB detected; verify whether authentication is required and whether remote access is overly exposed.")
    if port == 22 or svc == "ssh":
        hints.append("SSH detected; consider ssh-audit, banner review, key-based auth checks, and cautious credential testing.")
    if port == 21 or svc == "ftp":
        hints.append("FTP detected; check for anonymous access, writable locations, and plaintext credential exposure.")
    if port in (25, 465, 587) or svc in ("smtp", "smtps", "submission"):
        hints.append("SMTP detected; consider smtp-user-enum or swaks and look for open relay or user enumeration behavior.")
    if port == 53 or svc == "domain":
        hints.append("DNS detected; consider dig, nslookup, dnsrecon, and zone transfer testing where appropriate.")
    if port == 69 or svc == "tftp":
        hints.append("TFTP detected; check for unauthenticated file retrieval or upload opportunities.")
    if port == 111 or svc == "rpcbind":
        hints.append("RPCbind detected; consider rpcinfo and follow-on enumeration for NFS and related services.")
    if port == 2049 or svc == "nfs":
        hints.append("NFS detected; use showmount and test for accessible exports and weak export permissions.")
    if port in (161, 162) or svc == "snmp":
        hints.append("SNMP detected; try snmpwalk/onesixtyone and test common community strings carefully.")
    if port == 3389 or svc == "ms-wbt-server":
        hints.append("RDP detected; consider netexec rdp, xfreerdp, and review NLA / domain clues.")
    if port in (5900, 5901, 5902) or svc == "vnc":
        hints.append("VNC detected; verify authentication requirements and assess screenshot or password attack viability.")
    if port == 6379 or "redis" in svc:
        hints.append("Redis detected; verify bind/auth configuration and whether dangerous commands are exposed remotely.")
    if port == 11211 or "memcached" in svc:
        hints.append("Memcached detected; check whether it is exposed beyond localhost and assess information disclosure risk.")
    if port in (9200, 9300) or "elasticsearch" in svc:
        hints.append("Elasticsearch detected; check for unauthenticated cluster or index access.")
    if port in (2375, 2376) or "docker" in svc:
        hints.append("Docker API exposure may be high risk; verify remote daemon access and authentication/TLS posture.")
    if port == 6443 or "kubernetes" in svc:
        hints.append("Kubernetes-related service detected; look for exposed API endpoints and auth configuration issues.")
    if "ssl" in extra or "tls" in extra:
        hints.append("TLS-related context detected; review certificate details, protocol support, and any weak crypto indicators.")
    return hints


# ============================================================
# Nmap XML parsing
# ============================================================

def parse_nmap_xml(xml_path: Path) -> dict:
    logging.info(f"Parsing Nmap XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    scan_info = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "nmap_args":   root.get("args", ""),
        "nmap_version":root.get("version", ""),
        "hosts":       [],
    }

    for host in root.findall("host"):
        state_el = host.find("status")
        state = state_el.get("state") if state_el is not None else "unknown"

        addresses = [
            {"addr": a.get("addr", ""), "addrtype": a.get("addrtype", "")}
            for a in host.findall("address")
        ]

        hostnames = []
        hn_parent = host.find("hostnames")
        if hn_parent is not None:
            for hn in hn_parent.findall("hostname"):
                hostnames.append({"name": hn.get("name", ""), "type": hn.get("type", "")})

        ports = []
        ports_parent = host.find("ports")
        if ports_parent is not None:
            for port_el in ports_parent.findall("port"):
                state_el2 = port_el.find("state")
                if state_el2 is None or state_el2.get("state") != "open":
                    continue
                service_el = port_el.find("service")
                service = {}
                if service_el is not None:
                    service = {
                        "name":      service_el.get("name", ""),
                        "product":   service_el.get("product", ""),
                        "version":   service_el.get("version", ""),
                        "extrainfo": service_el.get("extrainfo", ""),
                        "tunnel":    service_el.get("tunnel", ""),
                    }
                scripts = [
                    {"id": s.get("id", ""), "output": s.get("output", "")}
                    for s in port_el.findall("script")
                ]
                ports.append({
                    "protocol": port_el.get("protocol", ""),
                    "port":     int(port_el.get("portid", "0")),
                    "service":  service,
                    "scripts":  scripts,
                })

        host_record = {
            "state":      state,
            "addresses":  addresses,
            "hostnames":  hostnames,
            "open_ports": sorted(ports, key=lambda p: (p["protocol"], p["port"])),
        }
        scan_info["hosts"].append(host_record)
        logging.debug(
            f"Nmap host: {choose_host_display_name(host_record)} "
            f"state={state} ports={len(host_record['open_ports'])}"
        )

    logging.info(f"Nmap done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# Nessus XML parsing
# ============================================================

def parse_nessus_xml(xml_path: Path) -> dict:
    """Parse a .nessus (NessusClientData_v2) file."""
    logging.info(f"Parsing Nessus XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Report name from the Report element
    report_el = root.find("Report")
    report_name = report_el.get("name", xml_path.stem) if report_el is not None else xml_path.stem

    scan_info: dict = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "report_name": report_name,
        "hosts":       [],
    }

    report_hosts = root.findall(".//ReportHost")
    for rh in report_hosts:
        # Host properties
        props: dict[str, str] = {}
        host_props_el = rh.find("HostProperties")
        if host_props_el is not None:
            for tag_el in host_props_el.findall("tag"):
                props[tag_el.get("name", "")] = (tag_el.text or "").strip()

        ip       = props.get("host-ip", rh.get("name", ""))
        hostname = props.get("hostname", props.get("host-fqdn", ""))

        findings: list[dict] = []
        for item in rh.findall("ReportItem"):
            plugin_id   = item.get("pluginID", "")
            plugin_name = item.get("pluginName", "")
            severity    = int(item.get("severity", "0"))
            port        = int(item.get("port", "0"))
            protocol    = item.get("protocol", "tcp")

            def _txt(tag: str) -> str:
                el = item.find(tag)
                return (el.text or "").strip() if el is not None else ""

            cves = [el.text.strip() for el in item.findall("cve") if el.text]

            findings.append({
                "plugin_id":    plugin_id,
                "plugin_name":  plugin_name,
                "severity_int": severity,
                "port":         port,
                "protocol":     protocol,
                "description":  _txt("description"),
                "solution":     _txt("solution"),
                "cves":         cves,
                "cvss_base":    _txt("cvss_base_score"),
                "cvss3_base":   _txt("cvss3_base_score"),
                "plugin_output":_txt("plugin_output"),
            })

        # Sort findings: highest severity first, then plugin name
        findings.sort(key=lambda f: (-f["severity_int"], f["plugin_name"]))

        host_record = {
            "ip":       ip,
            "hostname": hostname,
            "findings": findings,
        }
        scan_info["hosts"].append(host_record)
        logging.debug(
            f"Nessus host: {ip} ({hostname}) findings={len(findings)}"
        )

    logging.info(f"Nessus done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# Burp Suite XML parsing
# ============================================================

def parse_burp_xml(xml_path: Path) -> dict:
    """Parse a Burp Suite Pro Scanner Issues XML export."""
    logging.info(f"Parsing Burp XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    scan_info: dict = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "hosts":       [],
    }

    # Group issues by host IP
    host_map: dict[str, dict] = {}  # ip → {ip, url, issues}

    for issue_el in root.findall("issue"):
        host_el = issue_el.find("host")
        if host_el is None:
            continue

        ip  = host_el.get("ip", "")
        url = (host_el.text or "").strip()

        # Derive a canonical host key (prefer IP, fall back to URL base)
        host_key = ip if ip else url

        def _txt(tag: str) -> str:
            el = issue_el.find(tag)
            return (el.text or "").strip() if el is not None else ""

        issue = {
            "name":                _txt("name"),
            "path":                _txt("path"),
            "location":            _txt("location"),
            "severity":            _txt("severity"),
            "confidence":          _txt("confidence"),
            "issue_detail":        _txt("issueDetail"),
            "issue_background":    _txt("issueBackground"),
            "remediation_detail":  _txt("remediationDetail"),
            "remediation_background": _txt("remediationBackground"),
        }

        if host_key not in host_map:
            host_map[host_key] = {"ip": ip, "url": url, "issues": []}
        host_map[host_key]["issues"].append(issue)

    sev_order = {"high": 0, "medium": 1, "low": 2, "information": 3}
    for host_rec in host_map.values():
        host_rec["issues"].sort(
            key=lambda i: (sev_order.get(i["severity"].lower(), 4), i["name"])
        )
        scan_info["hosts"].append(host_rec)
        logging.debug(
            f"Burp host: {host_rec['ip']} ({host_rec['url']}) issues={len(host_rec['issues'])}"
        )

    logging.info(f"Burp done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# Host naming & port rendering
# ============================================================

def choose_host_display_name(host: dict) -> str:
    hostnames = [h["name"].strip() for h in host.get("hostnames", []) if h.get("name")]
    for hn in hostnames:
        if is_probable_fqdn(hn):
            return hn.rstrip(".")
    for hn in hostnames:
        if hn:
            return hn.rstrip(".")
    for addr in host.get("addresses", []):
        if addr.get("addrtype") == "ipv4" and addr.get("addr"):
            return addr["addr"]
    for addr in host.get("addresses", []):
        if addr.get("addr"):
            return addr["addr"]
    return "unknown-host"


def get_primary_ipv4(host: dict) -> str | None:
    for addr in host.get("addresses", []):
        if addr.get("addrtype") == "ipv4" and addr.get("addr"):
            return addr["addr"]
    return None


def render_service_string(port_info: dict) -> str:
    svc = port_info.get("service", {})
    parts = [svc.get("name", "")]
    if svc.get("product"):   parts.append(svc["product"])
    if svc.get("version"):   parts.append(svc["version"])
    if svc.get("extrainfo"): parts.append(f"({svc['extrainfo']})")
    if svc.get("tunnel"):    parts.append(f"[tunnel: {svc['tunnel']}]")
    return " ".join(p for p in parts if p).strip() or "—"


def summarize_open_ports(host: dict) -> list[str]:
    return [
        f"- **{p['protocol']}/{p['port']}** — {render_service_string(p)}"
        for p in host.get("open_ports", [])
    ]


# ============================================================
# Ollama — Nmap
# ============================================================

def build_ollama_prompt(scan_data: dict) -> str:
    hosts = scan_data.get("hosts", [])

    instructions = """
You are an experienced penetration tester analyzing Nmap scan results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.

Priorities:
- Identify notable exposed services and likely attack surface
- Suggest useful follow-up enumeration tools with a ready-to-run example command
  for each, substituting the actual target IP or hostname from the scan data
  (short, focused commands preferred)
- Highlight likely misconfigurations, risky exposures, or common weaknesses
- Infer likely technology stacks when reasonable, clearly tagging [INFERRED]
- Prefer real-world offensive security workflow recommendations over generic advice

Active Directory detection:
- If you observe Kerberos (88), LDAP (389/636/3268), DNS (53), and SMB (445)
  across one or more hosts, explicitly identify this as an Active Directory
  environment in the Environment Assessment section and tailor all attack path
  and enumeration suggestions to AD-specific methodology:
  AS-REP roasting, Kerberoasting, BloodHound, Pass-the-Hash, Pass-the-Ticket,
  DCSync, LDAP anonymous bind enumeration, GPO abuse, etc.

Cross-host correlation:
- Identify patterns across multiple hosts. If a finding applies network-wide
  (e.g., SMB signing disabled everywhere, same service version on all hosts),
  report it once as a network-level finding rather than repeating it per host.

When suggesting tools, favor practical enumeration tools such as:
- SMB: netexec, smbclient, smbmap, enum4linux-ng
- LDAP: ldapsearch, netexec, bloodyAD
- Kerberos: kerbrute, impacket tools (GetUserSPNs, GetNPUsers), netexec
- MSSQL: impacket-mssqlclient, netexec
- WinRM: evil-winrm, netexec
- RDP: netexec, xfreerdp
- SSH: ssh-audit, hydra
- Web: whatweb, nikto, gobuster, feroxbuster, ffuf, curl
- SNMP: snmpwalk, onesixtyone
- DNS: dig, nslookup, dnsrecon
- NFS: showmount, mount
- SMTP: swaks, smtp-user-enum
- FTP: ftp, Nmap NSE scripts
- RPC: rpcclient, netexec
- VNC: vncsnapshot, hydra
""".strip()

    lines = [
        f"Nmap args: {scan_data.get('nmap_args', '')}",
        f"Nmap version: {scan_data.get('nmap_version', '')}",
        f"Hosts count: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        name  = choose_host_display_name(host)
        state = host.get("state", "unknown")
        ip    = get_primary_ipv4(host) or "unknown"
        lines += [f"- Host: {name}", f"  State: {state}", f"  IP: {ip}"]
        open_ports = host.get("open_ports", [])
        if not open_ports:
            lines += ["  Open ports: none", ""]
            continue
        lines.append("  Open ports:")
        for p in open_ports:
            svc      = p.get("service", {})
            svc_name = svc.get("name", "") or "unknown"
            product  = svc.get("product", "")
            version  = svc.get("version", "")
            extrainfo= svc.get("extrainfo", "")
            tunnel   = svc.get("tunnel", "")
            parts = [svc_name]
            if product:   parts.append(product)
            if version:   parts.append(version)
            if extrainfo: parts.append(f"({extrainfo})")
            if tunnel:    parts.append(f"[tunnel: {tunnel}]")
            lines.append(f"    - {p['protocol']}/{p['port']}: {' '.join(parts).strip()}")
            for script in p.get("scripts", [])[:3]:
                out = " ".join((script.get("output", "") or "").split())
                if out:
                    if len(out) > 220:
                        out = out[:217] + "..."
                    lines.append(f"      script:{script.get('id', '')} -> {out}")
            for hint in get_port_hints(p["port"], svc_name, product, extrainfo):
                lines.append(f"      hint: {hint}")
        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections in this order:

## Environment Assessment
- State the inferred environment type (Active Directory domain, standalone Linux
  hosts, mixed Windows/Linux, web application stack, OT/ICS, etc.)
- Provide a confidence level: HIGH / MEDIUM / LOW
- List the specific evidence from the scan data that supports the assessment
- If AD is detected, state the inferred domain name if visible in hostnames or
  DNS responses

## Key Observations
- Brief bullets of the most important findings
- Each bullet must include a [CONFIRMED], [INFERRED], or [ASSUMED] tag

## Enumeration Suggestions
- Group by service or host
- For each suggestion, include a ready-to-run example command with the actual
  target IP or hostname substituted in
- Tag each with [CONFIRMED], [INFERRED], or [ASSUMED]

## Potential Attack Paths
- Realistic next steps or attack paths with supporting scan evidence for each step
- Tag each path as [CONFIRMED], [INFERRED], or [ASSUMED]
- Do not include a path if the scan data does not support it

## Notable Risks or Misconfigurations
- Network-wide findings first, then per-host
- Tag each with [CONFIRMED], [INFERRED], or [ASSUMED]
""".strip()

    prompt = f"{instructions}\n\nNmap Scan Summary\n=================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Nmap Ollama prompt ({len(prompt)} chars)")
    return prompt


# ============================================================
# Ollama — Nessus
# ============================================================

def build_nessus_ollama_prompt(nessus_data: dict) -> str:
    hosts = nessus_data.get("hosts", [])

    instructions = """
You are an experienced penetration tester analyzing Nessus vulnerability scan results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.
- Nessus plugin findings can produce false positives, particularly for
  version-based checks. If a finding relies solely on version detection without
  plugin output confirming exploitation, note it as
  [VERIFY — possible false positive].

Priorities:
- Identify the highest-impact vulnerabilities by CVSS score and exploitability
- For each CVE, tag its weaponization status:
    [MSF]         — known Metasploit module exists
    [POC]         — public PoC exists but may require adaptation
    [THEORETICAL] — no known public exploit at time of analysis
- Flag any findings suggesting credential exposure, default credentials,
  cleartext protocols, or password policy weaknesses — these often have
  outsized impact even when CVSS scores are moderate
- Version-based detection findings (where plugin_output is absent or says only
  "version X detected") must be flagged [VERIFY — version-based only]
- Cross-host correlation: if the same critical/high finding appears on multiple
  hosts, group them and call out the network-wide scope rather than listing each
  host separately
- Group findings by exploitation priority: critical/weaponized first, then
  high, then medium
""".strip()

    lines = [
        f"Report: {nessus_data.get('report_name', '')}",
        f"Hosts: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        ip       = host.get("ip", "unknown")
        hostname = host.get("hostname", "")
        findings = host.get("findings", [])

        label = f"{hostname} ({ip})" if hostname and hostname != ip else ip
        lines.append(f"- Host: {label}")

        # Skip info (severity 0) in AI prompt to reduce noise
        notable = [f for f in findings if f["severity_int"] >= 1]
        if not notable:
            lines.append("  Findings: none above informational")
            lines.append("")
            continue

        # Count by severity
        counts = {4: 0, 3: 0, 2: 0, 1: 0}
        for f in notable:
            counts[f["severity_int"]] = counts.get(f["severity_int"], 0) + 1
        lines.append(
            f"  Critical:{counts[4]} High:{counts[3]} Medium:{counts[2]} Low:{counts[1]}"
        )
        lines.append("  Top findings:")

        for f in notable[:15]:
            sev      = severity_int_to_str(f["severity_int"])
            cvss     = f.get("cvss3_base") or f.get("cvss_base") or "N/A"
            cves     = ", ".join(f["cves"]) if f["cves"] else "no CVE"
            has_out  = "plugin_output:yes" if f.get("plugin_output") else "plugin_output:none"
            lines.append(
                f"    [{sev}] {f['plugin_name']} | CVSS:{cvss} | {cves} | port {f['protocol']}/{f['port']} | {has_out}"
            )
            if f.get("solution"):
                sol = f["solution"][:200].replace("\n", " ")
                lines.append(f"      solution: {sol}")

        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections:

## Key Observations
- Most impactful vulnerabilities and what they mean for the engagement
- Each bullet must include a [CONFIRMED], [INFERRED], or [ASSUMED] tag
- Group any finding that appears on multiple hosts as a single network-wide bullet

## Exploitation Priority
- Rank the top exploitable findings with reasoning
- For each CVE include its [MSF], [POC], or [THEORETICAL] tag
- Mark version-based-only detections with [VERIFY — version-based only]
- Flag credential/cleartext findings prominently

## Attack Chains
- How findings could be combined for privilege escalation or lateral movement
- Each step must cite specific scan evidence; tag with [CONFIRMED]/[INFERRED]/[ASSUMED]

## Remediation Focus
- Highest-priority patches or configuration fixes
- Network-wide issues first, then per-host
""".strip()

    prompt = f"{instructions}\n\nNessus Scan Summary\n===================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Nessus Ollama prompt ({len(prompt)} chars)")
    return prompt


# ============================================================
# Ollama — Burp Suite
# ============================================================

def build_burp_ollama_prompt(burp_data: dict) -> str:
    hosts = burp_data.get("hosts", [])

    instructions = """
You are an experienced web application penetration tester analyzing Burp Suite scanner results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.
- Burp "Tentative" findings have a high false positive rate. Always flag these
  as [TENTATIVE — VERIFY MANUALLY] and do not build attack chains primarily on
  tentative findings.

Priorities:
- Identify the most exploitable web vulnerabilities (SQLi, XSS, SSRF, IDOR,
  auth bypass, etc.) and classify each by its OWASP Top 10 2021 category
  (e.g., A01:Broken Access Control, A03:Injection, A07:Auth Failures, etc.)
- For each Certain or Firm finding, suggest a specific manual verification step
  or payload to confirm exploitability beyond the scanner result
- Suggest how findings could be chained for greater impact
- Assess the web attack surface: authentication, input handling, business logic
- Highlight systemic issues (framework versions, CSP absence, insecure headers)
- Be explicit about what the scanner confirmed vs. what requires manual testing
""".strip()

    lines = [
        f"Hosts scanned: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        ip     = host.get("ip", "")
        url    = host.get("url", "")
        issues = host.get("issues", [])
        label  = url if url else ip
        if ip and url and ip not in url:
            label = f"{url} ({ip})"

        lines.append(f"- Host: {label}")

        notable = [i for i in issues if i["severity"].lower() not in ("information", "info")]
        if not notable:
            lines.append("  Issues: informational only")
            lines.append("")
            continue

        counts: dict[str, int] = {}
        for i in issues:
            counts[i["severity"]] = counts.get(i["severity"], 0) + 1
        lines.append(
            "  " + " ".join(f"{k}:{v}" for k, v in sorted(counts.items()))
        )
        lines.append("  Issues:")
        for i in notable[:15]:
            conf = i.get("confidence", "")
            loc  = i.get("path") or i.get("location") or ""
            lines.append(f"    [{i['severity']}|{conf}] {i['name']} @ {loc}")
            if i.get("issue_detail"):
                detail = i["issue_detail"][:200].replace("\n", " ")
                lines.append(f"      detail: {detail}")
        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections:

## Key Observations
- Most significant web vulnerabilities and their potential impact
- Each bullet must include its OWASP Top 10 2021 category and a
  [CONFIRMED], [INFERRED], or [ASSUMED] tag
- Tentative findings must be tagged [TENTATIVE — VERIFY MANUALLY]

## Exploitation Priority
- Which issues to pursue first and why
- For each Certain/Firm finding, include a specific manual verification step
  or payload (e.g., exact parameter to test, HTTP method, expected response)
- Do not elevate Tentative findings above Certain/Firm without explanation

## Attack Chains
- How findings could be combined for greater impact
- Each step must cite specific scanner evidence; tag [CONFIRMED]/[INFERRED]/[ASSUMED]
- Do not build chains primarily on Tentative findings

## False Positive Risk Assessment
- List findings that are likely false positives, with reasoning
- Consider: finding type, confidence level (Tentative = higher FP risk), context,
  and whether the issue_detail provides concrete evidence or just a generic description

## Remediation Focus
- Top development/configuration fixes to eliminate the highest-impact issues
""".strip()

    prompt = f"{instructions}\n\nBurp Suite Scan Summary\n=======================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Burp Ollama prompt ({len(prompt)} chars)")
    return prompt


def ollama_chat(base_url: str, model: str, prompt: str) -> str:
    url = base_url.rstrip("/") + "/api/generate"
    payload = {"model": model, "prompt": prompt, "stream": False}
    logging.info(f"Querying Ollama model: {model}")
    r = requests.post(url, json=payload, timeout=180)
    r.raise_for_status()
    response_text = r.json().get("response", "")
    logging.debug(f"Ollama response: {len(response_text)} chars")
    return response_text


# ============================================================
# Host note section rendering
# ============================================================

def _render_nessus_section(findings: list[dict]) -> str:
    """Render Nessus findings list as markdown body (no section header)."""
    notable = [f for f in findings if f["severity_int"] >= 1]
    info    = [f for f in findings if f["severity_int"] == 0]

    if not findings:
        return "_No Nessus findings._"

    lines: list[str] = []

    if notable:
        for f in notable:
            sev_str  = severity_int_to_str(f["severity_int"])
            cvss_val = f.get("cvss3_base") or f.get("cvss_base") or "N/A"
            cves_str = ", ".join(f["cves"]) if f["cves"] else "N/A"
            port_str = f"{f['protocol']}/{f['port']}" if f["port"] else "N/A"

            lines.append(f"#### [{sev_str}] {f['plugin_name']} (Plugin {f['plugin_id']})")
            lines.append(
                f"**Port:** {port_str}  ·  **CVSS v3:** {cvss_val}  ·  **CVE(s):** {cves_str}"
            )

            if f.get("description"):
                desc = f["description"].strip()
                if len(desc) > 500:
                    desc = desc[:497] + "..."
                lines.append(f"\n{desc}")

            if f.get("solution"):
                sol = f["solution"].strip()
                if len(sol) > 300:
                    sol = sol[:297] + "..."
                lines.append(f"\n**Solution:** {sol}")

            if f.get("plugin_output"):
                out = f["plugin_output"].strip()
                if len(out) > 300:
                    out = out[:297] + "..."
                lines.append(f"\n**Plugin Output:**\n```\n{out}\n```")

            lines.append("")

    if info:
        lines.append(f"_Plus {len(info)} informational finding(s) — omitted for brevity._")

    return "\n".join(lines).strip()


def _render_burp_section(issues: list[dict]) -> str:
    """Render Burp Suite issues list as markdown body (no section header)."""
    if not issues:
        return "_No Burp Suite findings._"

    sev_order = {"high": 0, "medium": 1, "low": 2, "information": 3}
    sorted_issues = sorted(
        issues, key=lambda i: (sev_order.get(i["severity"].lower(), 4), i["name"])
    )

    lines: list[str] = []
    for issue in sorted_issues:
        sev  = issue["severity"]
        conf = issue.get("confidence", "")
        conf_str = f" ({conf})" if conf else ""

        lines.append(f"#### [{sev}{conf_str}] {issue['name']}")

        if issue.get("path"):
            lines.append(f"**Path:** `{issue['path']}`")
        if issue.get("location"):
            lines.append(f"**Location:** {issue['location']}")

        if issue.get("issue_detail"):
            detail = issue["issue_detail"].strip()
            if len(detail) > 500:
                detail = detail[:497] + "..."
            lines.append(f"\n{detail}")

        if issue.get("remediation_detail"):
            rem = issue["remediation_detail"].strip()
            if len(rem) > 300:
                rem = rem[:297] + "..."
            lines.append(f"\n**Remediation:** {rem}")
        elif issue.get("remediation_background"):
            rem = issue["remediation_background"].strip()
            if len(rem) > 300:
                rem = rem[:297] + "..."
            lines.append(f"\n**Remediation:** {rem}")

        lines.append("")

    return "\n".join(lines).strip()


# ============================================================
# Vault writing — Nmap
# ============================================================

def _find_host_note_by_ip(hosts_dir: Path, ip: str) -> Path | None:
    """Find an existing host note whose frontmatter ip matches."""
    if not hosts_dir.exists() or not ip:
        return None
    for hp in hosts_dir.glob("*.md"):
        try:
            fm, _ = read_frontmatter(hp.read_text(encoding="utf-8"))
            if fm.get("ip") == ip:
                return hp
        except Exception:
            pass
    return None


def _write_host_note(
    hosts_dir: Path,
    host: dict,
    scan_stem: str,
    scan_display: str,
    tool_name: str,
) -> tuple[str, str]:
    """
    Write (or merge) a host note for an Nmap host.
    Preserves existing Nessus Findings and Burp Suite Findings sections.
    Returns (display_name, host_stem).
    """
    display    = choose_host_display_name(host)
    host_stem  = safe_filename(display)
    host_path  = hosts_dir / ensure_md_suffix(host_stem)
    primary_ip = get_primary_ipv4(host)
    open_ports = host.get("open_ports", [])

    existing_fm: dict  = {}
    existing_op_notes  = ""
    existing_nessus    = ""
    existing_burp      = ""

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes = extract_operator_notes(old_body)
        existing_nessus   = extract_body_section(old_body, "## Nessus Findings")
        existing_burp     = extract_body_section(old_body, "## Burp Suite Findings")
        logging.debug(f"Merging Nmap host note: {host_path.name}")
    else:
        logging.debug(f"Creating Nmap host note: {host_path.name}")

    # Merge frontmatter
    existing_hostnames: list = existing_fm.get("hostnames", [])
    new_hostnames = [h["name"].strip() for h in host.get("hostnames", []) if h.get("name")]
    merged_hostnames = existing_hostnames + [h for h in new_hostnames if h not in existing_hostnames]

    scan_source = f"{scan_display} - {tool_name}"
    existing_sources: list = existing_fm.get("sources", [])
    merged_sources = existing_sources if scan_source in existing_sources else existing_sources + [scan_source]

    new_tags = get_tags_from_ports(open_ports)
    merged_tags = sorted(set(existing_fm.get("tags", [])) | set(new_tags))

    fm: dict = {
        "ip":        primary_ip or existing_fm.get("ip"),
        "hostnames": merged_hostnames,
        "status":    existing_fm.get("status", "not-started"),
        "tags":      merged_tags,
        "sources":   merged_sources,
    }
    # Preserve nessus_max_severity if present
    if "nessus_max_severity" in existing_fm:
        fm["nessus_max_severity"] = existing_fm["nessus_max_severity"]

    # Build body
    lines: list[str] = [f"**State:** {host['state']}"]
    if primary_ip:
        lines.append(f"**IP:** {primary_ip}")
    lines.append(f"**Open Ports:** {len(open_ports)}")

    lines += ["", "## Open Ports"]
    if open_ports:
        lines.extend(summarize_open_ports(host))
    else:
        lines.append("_No open ports detected._")

    if existing_nessus:
        lines += ["", "## Nessus Findings", existing_nessus]

    if existing_burp:
        lines += ["", "## Burp Suite Findings", existing_burp]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_stem


def create_obsidian_vault(
    vault_dir: Path,
    scan_name: str,
    scan_data: dict,
    tool_name: str,
    analysis_text: str | None,
    model_name: str | None,
) -> dict:
    """Write all host notes and the scan note for one Nmap scan file."""
    logging.info(f"Writing Nmap vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = scan_name
    if len(scan_data["hosts"]) == 1:
        scan_display = choose_host_display_name(scan_data["hosts"][0])

    scan_stem = safe_filename(f"{scan_display} - {tool_name}")
    scan_path = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in scan_data["hosts"]:
        display, host_stem = _write_host_note(
            hosts_dir, host, scan_stem, scan_display, tool_name
        )
        host_entries.append((display, host_stem))

    scan_lines = [
        f"# {scan_display} - {tool_name}",
        "",
        f"- **Source:** {scan_data['source_file']}",
        f"- **Parsed:** {scan_data['parsed_at']}",
        f"- **Nmap Args:** `{scan_data.get('nmap_args', '')}`",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    scan_lines += ["", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Nmap scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — Nessus
# ============================================================

def _update_host_note_nessus(
    hosts_dir: Path,
    ip: str,
    hostname: str,
    findings: list[dict],
    scan_source: str,
) -> tuple[str, str]:
    """
    Write or update the ## Nessus Findings section in a host note.
    Returns (display_name, host_stem).
    """
    # Try to find an existing note by IP first
    existing_path = _find_host_note_by_ip(hosts_dir, ip)

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        # Determine display name
        if hostname and is_probable_fqdn(hostname):
            display = hostname
        else:
            display = ip or hostname or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports_section = ""
    existing_burp_section = ""
    existing_preamble_lines: list[str] = []
    existing_scan_refs: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes           = extract_operator_notes(old_body)
        existing_open_ports_section = extract_body_section(old_body, "## Open Ports")
        existing_burp_section       = extract_body_section(old_body, "## Burp Suite Findings")
        # Capture preamble (lines before first ## section)
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating Nessus section in: {host_path.name}")
    else:
        # Minimal preamble for Nessus-only host
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if hostname and hostname != ip:
            existing_preamble_lines.append(f"**Hostname:** {hostname}")
        logging.debug(f"Creating Nessus-only host note: {host_path.name}")

    # Update frontmatter
    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname and hostname not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname]

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    max_sev = max((f["severity_int"] for f in findings), default=0)
    existing_max_sev = existing_fm.get("nessus_max_severity", 0)
    try:
        existing_max_sev = int(existing_max_sev)
    except (TypeError, ValueError):
        existing_max_sev = 0
    merged_max_sev = max(max_sev, existing_max_sev)

    fm: dict = {
        "ip":                  ip or existing_fm.get("ip"),
        "hostnames":           existing_hostnames,
        "status":              existing_fm.get("status", "not-started"),
        "tags":                existing_fm.get("tags", []),
        "sources":             existing_sources,
        "nessus_max_severity": merged_max_sev,
    }

    # Build body
    preamble = "\n".join(existing_preamble_lines).rstrip()
    lines: list[str] = []

    if preamble:
        lines.append(preamble)

    if existing_open_ports_section:
        lines += ["", "## Open Ports", existing_open_ports_section]

    lines += ["", "## Nessus Findings", _render_nessus_section(findings)]

    if existing_burp_section:
        lines += ["", "## Burp Suite Findings", existing_burp_section]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_nessus_vault(
    vault_dir: Path,
    scan_name: str,
    nessus_data: dict,
    analysis_text: str | None,
    model_name: str | None,
) -> dict:
    """Write Nessus host note sections and scan note."""
    logging.info(f"Writing Nessus vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = nessus_data.get("report_name", scan_name)
    scan_stem    = safe_filename(f"{scan_display} - Nessus")
    scan_source  = f"{scan_display} - Nessus"
    scan_path    = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in nessus_data.get("hosts", []):
        display, host_stem = _update_host_note_nessus(
            hosts_dir,
            host["ip"],
            host.get("hostname", ""),
            host["findings"],
            scan_source,
        )
        host_entries.append((display, host_stem))

    # Build scan note
    total_findings = sum(len(h["findings"]) for h in nessus_data.get("hosts", []))
    critical_count = sum(
        sum(1 for f in h["findings"] if f["severity_int"] == 4)
        for h in nessus_data.get("hosts", [])
    )
    high_count = sum(
        sum(1 for f in h["findings"] if f["severity_int"] == 3)
        for h in nessus_data.get("hosts", [])
    )

    scan_lines = [
        f"# {scan_display} - Nessus",
        "",
        f"- **Source:** {nessus_data['source_file']}",
        f"- **Parsed:** {nessus_data['parsed_at']}",
        f"- **Total Findings:** {total_findings}  ·  **Critical:** {critical_count}  ·  **High:** {high_count}",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    scan_lines += ["", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Nessus scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — Burp Suite
# ============================================================

def _update_host_note_burp(
    hosts_dir: Path,
    ip: str,
    url: str,
    issues: list[dict],
    scan_source: str,
) -> tuple[str, str]:
    """
    Write or update the ## Burp Suite Findings section in a host note.
    Returns (display_name, host_stem).
    """
    existing_path = _find_host_note_by_ip(hosts_dir, ip) if ip else None

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        # Derive display name from URL host or IP
        if url:
            from urllib.parse import urlparse
            parsed = urlparse(url)
            netloc = parsed.netloc or url
            display = netloc.split(":")[0] if netloc else (ip or url)
        else:
            display = ip or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports_section = ""
    existing_nessus_section = ""
    existing_preamble_lines: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes            = extract_operator_notes(old_body)
        existing_open_ports_section  = extract_body_section(old_body, "## Open Ports")
        existing_nessus_section      = extract_body_section(old_body, "## Nessus Findings")
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating Burp section in: {host_path.name}")
    else:
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if url:
            existing_preamble_lines.append(f"**URL:** {url}")
        logging.debug(f"Creating Burp-only host note: {host_path.name}")

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    # Derive hostname from URL for frontmatter
    hostname_for_fm = ""
    if url:
        from urllib.parse import urlparse
        parsed = urlparse(url)
        hostname_for_fm = parsed.netloc.split(":")[0] if parsed.netloc else ""

    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname_for_fm and hostname_for_fm not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname_for_fm]

    fm: dict = {
        "ip":        ip or existing_fm.get("ip"),
        "hostnames": existing_hostnames,
        "status":    existing_fm.get("status", "not-started"),
        "tags":      existing_fm.get("tags", []),
        "sources":   existing_sources,
    }
    if "nessus_max_severity" in existing_fm:
        fm["nessus_max_severity"] = existing_fm["nessus_max_severity"]

    preamble = "\n".join(existing_preamble_lines).rstrip()
    lines: list[str] = []

    if preamble:
        lines.append(preamble)

    if existing_open_ports_section:
        lines += ["", "## Open Ports", existing_open_ports_section]

    if existing_nessus_section:
        lines += ["", "## Nessus Findings", existing_nessus_section]

    lines += ["", "## Burp Suite Findings", _render_burp_section(issues)]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_burp_vault(
    vault_dir: Path,
    scan_name: str,
    burp_data: dict,
    analysis_text: str | None,
    model_name: str | None,
) -> dict:
    """Write Burp Suite host note sections and scan note."""
    logging.info(f"Writing Burp vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = scan_name
    scan_stem    = safe_filename(f"{scan_display} - Burp")
    scan_source  = f"{scan_display} - Burp"
    scan_path    = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in burp_data.get("hosts", []):
        display, host_stem = _update_host_note_burp(
            hosts_dir,
            host.get("ip", ""),
            host.get("url", ""),
            host["issues"],
            scan_source,
        )
        host_entries.append((display, host_stem))

    total_issues = sum(len(h["issues"]) for h in burp_data.get("hosts", []))
    high_count   = sum(
        sum(1 for i in h["issues"] if i["severity"].lower() == "high")
        for h in burp_data.get("hosts", [])
    )

    scan_lines = [
        f"# {scan_display} - Burp",
        "",
        f"- **Source:** {burp_data['source_file']}",
        f"- **Parsed:** {burp_data['parsed_at']}",
        f"- **Total Issues:** {total_issues}  ·  **High:** {high_count}",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    scan_lines += ["", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Burp scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Canvas node / edge constructors
# ============================================================

def _text_node(nid: str, text: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "text", "text": text, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _file_node(nid: str, file: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "file", "file": file, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _group_node(nid: str, label: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "group", "label": label, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _edge(eid: str, from_id: str, to_id: str, label: str = "") -> dict:
    e: dict = {
        "id":       eid,
        "fromNode": from_id,
        "fromSide": "bottom",
        "toNode":   to_id,
        "toSide":   "top",
    }
    if label:
        e["label"] = label
    return e


# ============================================================
# Canvas content helpers
# ============================================================

def _read_host_frontmatter(host_path: Path) -> dict:
    try:
        fm, _ = read_frontmatter(host_path.read_text(encoding="utf-8"))
        return fm
    except Exception:
        return {}


def _build_campaign_overview(
    vault_dir: Path,
    scan_host_map: dict,
    all_analyses: list[tuple[str, str]],
) -> str:
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    total_hosts = len(host_notes)
    total_scans = len(scan_host_map)
    status_counts: dict[str, int] = {}
    all_tags: set[str] = set()
    total_ports = 0
    nessus_critical = 0
    nessus_high     = 0
    nessus_medium   = 0
    burp_high       = 0
    burp_total      = 0

    for hp in host_notes:
        text = ""
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue

        fm, body = read_frontmatter(text)
        status = fm.get("status", "not-started")
        status_counts[status] = status_counts.get(status, 0) + 1
        all_tags.update(fm.get("tags", []))

        # Count Nmap open ports
        total_ports += sum(
            1 for line in body.splitlines()
            if line.strip().startswith("- **") and "/" in line
            and line.strip().startswith("- **tcp") or
            (line.strip().startswith("- **") and "/" in line and not line.strip().startswith("- **tcp"))
        )

        # Count Nessus severity from frontmatter
        max_sev = fm.get("nessus_max_severity", 0)
        try:
            max_sev = int(max_sev)
        except (TypeError, ValueError):
            max_sev = 0
        if max_sev >= 4:
            nessus_critical += 1
        elif max_sev >= 3:
            nessus_high += 1
        elif max_sev >= 2:
            nessus_medium += 1

        # Count Burp issues from section
        burp_section = extract_body_section(body, "## Burp Suite Findings")
        if burp_section and "_No Burp" not in burp_section:
            burp_total += burp_section.count("#### [")
            burp_high  += burp_section.count("#### [High")

    # Re-count ports more simply
    total_ports = 0
    for hp in host_notes:
        try:
            _, body = read_frontmatter(hp.read_text(encoding="utf-8"))
            ports_section = extract_body_section(body, "## Open Ports")
            total_ports += sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))
        except Exception:
            pass

    lines = [
        "# Campaign Overview",
        "",
        f"**Hosts:** {total_hosts}  ·  **Scans:** {total_scans}  ·  **Open Ports:** {total_ports}",
        "",
    ]

    if status_counts:
        parts = [f"{k}: {v}" for k, v in sorted(status_counts.items())]
        lines += [f"**Progress:** {' · '.join(parts)}", ""]

    display_tags = sorted(all_tags - {"rpc"})
    if display_tags:
        lines += [f"**Services seen:** {', '.join(display_tags)}", ""]

    nessus_parts = []
    if nessus_critical:
        nessus_parts.append(f"critical hosts: {nessus_critical}")
    if nessus_high:
        nessus_parts.append(f"high hosts: {nessus_high}")
    if nessus_medium:
        nessus_parts.append(f"medium hosts: {nessus_medium}")
    if nessus_parts:
        lines += [f"**Nessus:** {' · '.join(nessus_parts)}", ""]

    if burp_total:
        lines += [f"**Burp:** {burp_total} issues ({burp_high} high)", ""]

    # Key Observations bullets from the most recent AI analysis
    if all_analyses:
        text = all_analyses[-1][1]
        match = re.search(
            r"##\s+Key Observations(.*?)(?=\n##\s|\Z)", text, re.DOTALL | re.IGNORECASE
        )
        if match:
            bullets = [
                l.strip() for l in match.group(1).splitlines()
                if l.strip().startswith(("-", "*", "•"))
            ][:4]
            if bullets:
                lines += ["**Key Findings:**"] + bullets

    return "\n".join(lines)


def build_priority_targets_prompt(vault_dir: Path, all_analyses: list[tuple[str, str]]) -> str:
    """
    Build an Ollama prompt asking it to rank all discovered hosts by exploitation
    priority, drawing on Nmap service tags, Nessus CVEs/CVSS, and Burp findings.
    """
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    instructions = """
You are an experienced penetration tester. Given scan data about multiple hosts,
rank them from highest to lowest exploitation priority.
This ranking will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

GROUNDING RULES — follow these strictly:
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- Only rank a host as a top target if there is concrete scan evidence supporting
  exploitability. Do not rank based on port count alone.
- If evidence is insufficient to justify a high ranking, say so rather than
  padding the list.
- Tag each ranked entry as one of:
    [CONFIRMED]  — directly in the scan data (CVE with plugin output, confirmed vuln)
    [INFERRED]   — reasonable conclusion from the scan data (service fingerprint, role)
    [ASSUMED]    — requires verification; explain why

Output ONLY a numbered list — no preamble, no section headers, no trailing text.
Each entry must be a single line in this exact format:
  N. <hostname or IP> [TAG] — <primary evidence: CVE ID / finding name / service>; <brief impact>

Base your ranking on:
- CVEs with confirmed plugin output and known weaponized exploits ([MSF] first)
- CVSS scores (critical ≥ 9.0, high ≥ 7.0) — but only when plugin_output:yes
- Exposed management interfaces with evidence of weak auth (RDP, WinRM, SSH)
- Domain role evidence in hostnames, ports, or service banners (DC, CA, DB)
- Chaining potential supported by scan data (pivot hosts, credential stores)
- Web findings from Burp (Certain/Firm auth bypass, SQLi — not Tentative)
- Unencrypted or unauthenticated services confirmed in scan data

Do not rank a host highly based solely on port count or service tags without
supporting vulnerability or misconfiguration evidence.
Maximum 10 entries.
""".strip()

    host_lines: list[str] = []
    for hp in host_notes:
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue
        fm, body = read_frontmatter(text)

        ip        = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        display   = hostnames[0] if hostnames else (ip or hp.stem)
        tags      = fm.get("tags", [])
        nessus_max = fm.get("nessus_max_severity", 0)
        try:
            nessus_max = int(nessus_max)
        except (TypeError, ValueError):
            nessus_max = 0

        ports_section  = extract_body_section(body, "## Open Ports")
        port_count     = sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))

        nessus_section = extract_body_section(body, "## Nessus Findings")
        crit_count     = nessus_section.count("[Critical]") if nessus_section else 0
        high_count_n   = nessus_section.count("[High]")     if nessus_section else 0
        cves           = re.findall(r"CVE-\d{4}-\d+", nessus_section)[:4] if nessus_section else []

        burp_section   = extract_body_section(body, "## Burp Suite Findings")
        burp_high      = burp_section.count("[High") if burp_section else 0
        burp_med       = burp_section.count("[Medium") if burp_section else 0

        parts = [f"{display} (IP: {ip})"]
        if tags:
            parts.append(f"services: {', '.join(tags)}")
        parts.append(f"open-ports: {port_count}")
        if nessus_max:
            parts.append(f"nessus-max: {severity_int_to_str(nessus_max)}")
        if crit_count or high_count_n:
            parts.append(f"critical/high findings: {crit_count}/{high_count_n}")
        if cves:
            parts.append(f"CVEs: {', '.join(cves)}")
        if burp_high or burp_med:
            parts.append(f"burp high/medium: {burp_high}/{burp_med}")

        host_lines.append("- " + " | ".join(parts))

    # Include Key Observations snippets from up to 3 most recent analyses for context
    analysis_ctx: list[str] = []
    for scan_display, text in all_analyses[-3:]:
        m = re.search(r"##\s+Key Observations(.*?)(?=\n##\s|\Z)", text, re.DOTALL | re.IGNORECASE)
        if m:
            bullets = [
                l.strip() for l in m.group(1).splitlines()
                if l.strip().startswith(("-", "*", "•"))
            ][:3]
            if bullets:
                analysis_ctx.append(f"[{scan_display}]")
                analysis_ctx.extend(bullets)

    prompt = f"{instructions}\n\nHosts to rank:\n" + "\n".join(host_lines)
    if analysis_ctx:
        prompt += "\n\nKey scan findings for context:\n" + "\n".join(analysis_ctx)
    logging.debug(f"Built priority targets prompt ({len(prompt)} chars)")
    return prompt


def _build_priority_targets_fallback(vault_dir: Path) -> str:
    """
    Build a static Priority Targets list (no AI) sorted by Nessus severity,
    Burp high count, and open port count.  Used when Ollama is skipped or
    no analyses have been generated yet.
    """
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    entries: list[dict] = []
    for hp in host_notes:
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue
        fm, body = read_frontmatter(text)

        ip        = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        display   = hostnames[0] if hostnames else (ip or hp.stem)
        tags      = fm.get("tags", [])
        nessus_max = fm.get("nessus_max_severity", 0)
        try:
            nessus_max = int(nessus_max)
        except (TypeError, ValueError):
            nessus_max = 0

        ports_section  = extract_body_section(body, "## Open Ports")
        port_count     = sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))

        nessus_section = extract_body_section(body, "## Nessus Findings")
        cves           = re.findall(r"CVE-\d{4}-\d+", nessus_section)[:2] if nessus_section else []

        burp_section   = extract_body_section(body, "## Burp Suite Findings")
        burp_high      = burp_section.count("[High") if burp_section else 0

        reasons: list[str] = []
        if cves:
            reasons.append(", ".join(cves))
        if "domain-controller" in tags:
            reasons.append("domain controller")
        if nessus_max >= 4:
            reasons.append("critical Nessus findings")
        elif nessus_max >= 3:
            reasons.append("high-severity Nessus findings")
        elif nessus_max >= 2:
            reasons.append("medium-severity Nessus findings")
        if burp_high:
            reasons.append(f"{burp_high} high web issue(s)")
        exposed = [t for t in tags if t in ("rdp", "winrm", "smb", "kerberos", "ldap", "mssql", "ftp", "snmp")]
        if exposed:
            reasons.append(f"exposed: {', '.join(exposed)}")
        if not reasons and port_count:
            service_str = ", ".join(tags[:4]) if tags else "unknown services"
            reasons.append(f"{port_count} open port(s) — {service_str}")

        entries.append({
            "display":  display,
            "sort_key": (-nessus_max, -burp_high, -port_count),
            "reason":   "; ".join(reasons) if reasons else "no notable findings",
        })

    entries.sort(key=lambda e: e["sort_key"])

    if not entries:
        return "## Priority Targets\n\n_No hosts found. Run scans to populate._"

    lines = ["## Priority Targets", ""]
    for i, entry in enumerate(entries[:10], 1):
        lines.append(f"{i}. {entry['display']} — {entry['reason']}")
    return "\n".join(lines)


def _build_next_steps(all_analyses: list[tuple[str, str]], max_items: int = 14) -> str:
    items: list[str] = []

    for _scan_display, text in all_analyses:
        for section in (
            "Potential Attack Paths",
            "Enumeration Suggestions",
            "Exploitation Priority",
            "Attack Chains",
        ):
            match = re.search(
                rf"##\s+{re.escape(section)}(.*?)(?=\n##\s|\Z)",
                text, re.DOTALL | re.IGNORECASE,
            )
            if not match:
                continue
            for line in match.group(1).splitlines():
                stripped = line.strip()
                if not stripped or stripped.startswith("#"):
                    continue
                cleaned = re.sub(r"^[-*•\d]+[.)]\s*", "", stripped).strip()
                if cleaned and len(cleaned) > 12 and cleaned not in items:
                    items.append(cleaned)
                if len(items) >= max_items:
                    break
            if len(items) >= max_items:
                break

    if not items:
        return "# Next Steps\n\n_Run scans with AI analysis enabled to populate next steps._"

    lines = ["# Next Steps", ""]
    for item in items[:max_items]:
        lines.append(f"- {item}")
    return "\n".join(lines)


# ============================================================
# Canvas layout engine
# ============================================================

def _group_dimensions(n_hosts: int, cols: int) -> tuple[int, int]:
    if n_hosts == 0:
        return 300, 200
    actual_cols = min(n_hosts, cols)
    rows = math.ceil(n_hosts / cols)
    inner_w = actual_cols * CARD_W + (actual_cols - 1) * CARD_GAP_X
    inner_h = rows * CARD_H + (rows - 1) * CARD_GAP_Y
    w = GROUP_PAD * 2 + inner_w
    h = GROUP_PAD + GROUP_LABEL_H + inner_h + GROUP_PAD
    return w, h


def build_canvas(
    vault_dir: Path,
    canvas_name: str,
    scan_host_map: dict,
    all_analyses: list[tuple[str, str]],
    canvas_cols: int = 2,
    max_groups_per_row: int = 3,
    priority_targets_text: str | None = None,
) -> Path:
    hosts_dir   = vault_dir / "Hosts"
    canvas_path = vault_dir / canvas_name

    existing_canvas: dict = {"nodes": [], "edges": []}
    if canvas_path.exists():
        try:
            existing_canvas = json.loads(canvas_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    managed_ids: set[str] = set()
    our_nodes:   list[dict] = []
    our_edges:   list[dict] = []

    # 1. Campaign Overview
    overview_id = stable_id("campaign_overview")
    managed_ids.add(overview_id)
    overview_x = -(OVERVIEW_W // 2)
    overview_y = -900
    our_nodes.append(_text_node(
        overview_id,
        _build_campaign_overview(vault_dir, scan_host_map, all_analyses),
        overview_x, overview_y, OVERVIEW_W, OVERVIEW_H,
    ))

    # 2. Scan note cards
    scan_stems = list(scan_host_map.keys())
    total_scan_row_w = (
        len(scan_stems) * SCAN_CARD_W + max(0, len(scan_stems) - 1) * CARD_GAP_X
    )
    scan_row_x0 = -(total_scan_row_w // 2)
    scan_row_y  = overview_y + OVERVIEW_H + 100

    scan_node_ids: dict[str, str] = {}

    for i, scan_stem in enumerate(scan_stems):
        sid = stable_id(f"scan_{scan_stem}")
        managed_ids.add(sid)
        scan_node_ids[scan_stem] = sid
        sx  = scan_row_x0 + i * (SCAN_CARD_W + CARD_GAP_X)
        rel = f"Scans/{ensure_md_suffix(scan_stem)}"
        our_nodes.append(_file_node(sid, rel, sx, scan_row_y, SCAN_CARD_W, SCAN_CARD_H))
        eid = stable_id(f"edge_overview_{sid}")
        managed_ids.add(eid)
        our_edges.append(_edge(eid, overview_id, sid))

    # 2b. Priority Targets node (centered, between scan cards and subnet groups)
    pt_id = stable_id("priority-targets")
    managed_ids.add(pt_id)
    pt_text = priority_targets_text or _build_priority_targets_fallback(vault_dir)
    # Ensure the markdown header is present
    if not pt_text.lstrip().startswith("## Priority Targets"):
        pt_text = "## Priority Targets\n\n" + pt_text.lstrip()
    pt_x = -(NEXT_STEPS_W // 2)
    pt_y = scan_row_y + SCAN_CARD_H + ROW_GAP
    our_nodes.append(_text_node(pt_id, pt_text, pt_x, pt_y, NEXT_STEPS_W, NEXT_STEPS_H, color="2"))

    # Overview → Priority Targets
    eid_ov_pt = stable_id("edge_overview_pt")
    managed_ids.add(eid_ov_pt)
    our_edges.append(_edge(eid_ov_pt, overview_id, pt_id))

    # 3. Subnet groups with host cards
    subnet_hosts: dict[str, list[tuple[str, Path]]] = {}
    if hosts_dir.exists():
        for hp in sorted(hosts_dir.glob("*.md")):
            fm     = _read_host_frontmatter(hp)
            ip     = fm.get("ip")
            subnet = get_subnet_label(ip) if ip and is_ipv4(str(ip)) else "unknown"
            subnet_hosts.setdefault(subnet, []).append((hp.stem, hp))

    subnets  = sorted(subnet_hosts.keys())
    n_groups = len(subnets)
    gdims    = [_group_dimensions(len(subnet_hosts[s]), canvas_cols) for s in subnets]

    # Groups start below the Priority Targets node
    cur_y = pt_y + NEXT_STEPS_H + ROW_GAP
    group_positions: list[tuple[int, int]] = []

    for row_i in range(math.ceil(n_groups / max_groups_per_row)):
        start = row_i * max_groups_per_row
        end   = min(start + max_groups_per_row, n_groups)
        row_indices = range(start, end)
        row_total_w = sum(gdims[j][0] for j in row_indices) + (end - start - 1) * GROUP_GAP
        row_height  = max((gdims[j][1] for j in row_indices), default=200)
        x = -(row_total_w // 2)
        for j in row_indices:
            group_positions.append((x, cur_y))
            x += gdims[j][0] + GROUP_GAP
        cur_y += row_height + ROW_GAP

    host_node_ids: dict[str, str] = {}

    for i, subnet in enumerate(subnets):
        if i >= len(group_positions):
            break
        gx, gy = group_positions[i]
        gw, gh = gdims[i]

        gid = stable_id(f"subnet_{subnet}")
        managed_ids.add(gid)
        our_nodes.append(_group_node(gid, subnet, gx, gy, gw, gh))

        # Priority Targets → subnet group edge
        eid_pt_g = stable_id(f"edge_pt_{gid}")
        managed_ids.add(eid_pt_g)
        our_edges.append(_edge(eid_pt_g, pt_id, gid))

        for j, (host_stem, hp) in enumerate(subnet_hosts[subnet]):
            col = j % canvas_cols
            row = j // canvas_cols
            hx  = gx + GROUP_PAD + col * (CARD_W + CARD_GAP_X)
            hy  = gy + GROUP_PAD + GROUP_LABEL_H + row * (CARD_H + CARD_GAP_Y)

            hid = stable_id(f"host_{host_stem}")
            managed_ids.add(hid)
            host_node_ids[host_stem] = hid

            fm     = _read_host_frontmatter(hp)
            status = fm.get("status", "not-started")
            color  = STATUS_COLORS.get(status)

            # If not-started, use Nessus severity for color hint
            if color is None and status == "not-started":
                nessus_max = fm.get("nessus_max_severity", 0)
                try:
                    nessus_max = int(nessus_max)
                except (TypeError, ValueError):
                    nessus_max = 0
                color = NESSUS_SEVERITY_CANVAS_COLORS.get(nessus_max)

            rel = f"Hosts/{ensure_md_suffix(host_stem)}"
            our_nodes.append(_file_node(hid, rel, hx, hy, CARD_W, CARD_H, color))

    # Scan → host edges
    for scan_stem, host_stems in scan_host_map.items():
        sid = scan_node_ids.get(scan_stem)
        if not sid:
            continue
        for host_stem in host_stems:
            hid = host_node_ids.get(host_stem)
            if not hid:
                continue
            eid = stable_id(f"edge_{sid}_{hid}")
            managed_ids.add(eid)
            our_edges.append(_edge(eid, sid, hid))

    # 4. Next Steps
    next_id = stable_id("next_steps")
    managed_ids.add(next_id)
    next_x = -(NEXT_STEPS_W // 2)
    next_y = cur_y + 40
    our_nodes.append(_text_node(
        next_id,
        _build_next_steps(all_analyses),
        next_x, next_y, NEXT_STEPS_W, NEXT_STEPS_H,
        color="5",
    ))

    preserved   = [n for n in existing_canvas.get("nodes", []) if n.get("id") not in managed_ids]
    final_nodes = preserved + our_nodes
    final_edges = our_edges

    canvas_path.write_text(
        json.dumps({"nodes": final_nodes, "edges": final_edges}, indent=2),
        encoding="utf-8",
    )
    logging.info(
        f"Canvas written: {canvas_path.name} "
        f"({len(final_nodes)} nodes, {len(final_edges)} edges, "
        f"{len(preserved)} preserved)"
    )
    return canvas_path


# ============================================================
# Excel export
# ============================================================

def _xl_header_row(ws, headers: list[str]) -> None:
    """Write bold header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=False)


def _xl_severity_fill(severity_str: str) -> "PatternFill | None":
    fills = {
        "critical": PatternFill("solid", fgColor="FF4444"),
        "high":     PatternFill("solid", fgColor="FF8C00"),
        "medium":   PatternFill("solid", fgColor="FFD700"),
        "low":      PatternFill("solid", fgColor="ADD8E6"),
    }
    return fills.get((severity_str or "").lower())


def _xl_autofit(ws, max_width: int = 60) -> None:
    """Approximate column auto-fit based on max content length."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=0,
        )
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _trunc(s: str | None, n: int) -> str:
    if not s:
        return ""
    s = s.strip()
    return s[:n - 3] + "..." if len(s) > n else s


def export_excel(
    vault_dir: Path,
    all_nmap_scans: list[tuple[str, str, dict]],   # (scan_stem, scan_display, scan_data)
    all_nessus_scans: list[tuple[str, str, dict]], # (scan_stem, scan_display, nessus_data)
    all_burp_scans: list[tuple[str, str, dict]],   # (scan_stem, scan_display, burp_data)
) -> Path | None:
    if not HAS_OPENPYXL:
        logging.error(
            "openpyxl is not installed. Run: pip install openpyxl"
        )
        return None

    wb = openpyxl.Workbook()

    # ---- Summary sheet (first tab) ----
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = [
        "IP", "Hostname", "Status", "Open Ports",
        "Highest Nessus Severity", "Critical", "High", "Medium", "Low",
        "Burp Issue Count", "Sources",
    ]
    _xl_header_row(ws_summary, summary_headers)

    hosts_dir = vault_dir / "Hosts"
    host_rows: list[list] = []

    if hosts_dir.exists():
        for hp in sorted(hosts_dir.glob("*.md")):
            try:
                text = hp.read_text(encoding="utf-8")
            except Exception:
                continue
            fm, body = read_frontmatter(text)

            ip       = fm.get("ip", "")
            hostname = ", ".join(fm.get("hostnames", [])) if fm.get("hostnames") else ""
            status   = fm.get("status", "not-started")
            sources  = ", ".join(fm.get("sources", []))

            # Open ports from ## Open Ports section
            ports_section = extract_body_section(body, "## Open Ports")
            open_port_strs = [
                l.strip().lstrip("- ").split(" — ")[0].strip("**")
                for l in ports_section.splitlines()
                if l.strip().startswith("- **")
            ]
            open_ports_str = ", ".join(open_port_strs)

            # Nessus severity counts from ## Nessus Findings section
            nessus_section = extract_body_section(body, "## Nessus Findings")
            sev_counts = {4: 0, 3: 0, 2: 0, 1: 0}
            if nessus_section:
                for line in nessus_section.splitlines():
                    m = re.match(r"####\s+\[(Critical|High|Medium|Low)\]", line)
                    if m:
                        sev_counts[severity_str_to_int(m.group(1))] += 1

            max_sev_int = fm.get("nessus_max_severity", 0)
            try:
                max_sev_int = int(max_sev_int)
            except (TypeError, ValueError):
                max_sev_int = 0
            max_sev_str = severity_int_to_str(max_sev_int) if max_sev_int > 0 else ""

            # Burp issue count
            burp_section = extract_body_section(body, "## Burp Suite Findings")
            burp_count   = burp_section.count("#### [") if burp_section else 0

            host_rows.append([
                ip, hostname, status, open_ports_str,
                max_sev_str,
                sev_counts[4], sev_counts[3], sev_counts[2], sev_counts[1],
                burp_count, sources,
            ])

    for row_data in host_rows:
        ws_summary.append(row_data)
        # Color severity cell (col 5 = Highest Nessus Severity)
        row_idx = ws_summary.max_row
        sev_cell = ws_summary.cell(row=row_idx, column=5)
        fill = _xl_severity_fill(str(sev_cell.value or ""))
        if fill:
            sev_cell.fill = fill

    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions
    _xl_autofit(ws_summary)

    # ---- Nmap sheet ----
    ws_nmap = wb.create_sheet("Nmap")
    nmap_headers = [
        "IP", "Hostname", "Port", "Protocol", "Service", "Version",
        "State", "Scripts/Banner", "Source Scan", "Status",
    ]
    _xl_header_row(ws_nmap, nmap_headers)

    # Build IP→frontmatter map for status lookup
    host_fm_by_ip: dict[str, dict] = {}
    if hosts_dir.exists():
        for hp in hosts_dir.glob("*.md"):
            try:
                fm, _ = read_frontmatter(hp.read_text(encoding="utf-8"))
                if fm.get("ip"):
                    host_fm_by_ip[fm["ip"]] = fm
            except Exception:
                pass

    for scan_stem, scan_display, scan_data in all_nmap_scans:
        for host in scan_data.get("hosts", []):
            ip       = get_primary_ipv4(host) or ""
            hostname = choose_host_display_name(host)
            h_fm     = host_fm_by_ip.get(ip, {})
            status   = h_fm.get("status", "not-started")

            open_ports = host.get("open_ports", [])
            if not open_ports:
                ws_nmap.append([ip, hostname, "", "", "", "", host.get("state", ""), "", scan_display, status])
                continue

            for p in open_ports:
                svc     = p.get("service", {})
                scripts = "; ".join(
                    f"{s['id']}: {' '.join(s['output'].split())[:80]}"
                    for s in p.get("scripts", [])[:3] if s.get("output")
                )
                version = " ".join(filter(None, [
                    svc.get("product", ""), svc.get("version", ""), svc.get("extrainfo", "")
                ])).strip()
                ws_nmap.append([
                    ip, hostname,
                    p["port"], p["protocol"],
                    svc.get("name", ""), version,
                    "open", scripts,
                    scan_display, status,
                ])

    ws_nmap.freeze_panes = "A2"
    ws_nmap.auto_filter.ref = ws_nmap.dimensions
    _xl_autofit(ws_nmap)

    # ---- Nessus sheet ----
    ws_nessus = wb.create_sheet("Nessus")
    nessus_headers = [
        "IP", "Hostname", "Port", "Protocol",
        "Plugin ID", "Plugin Name", "Severity",
        "CVSS v3", "CVSS v2", "CVE(s)",
        "Description", "Solution", "Plugin Output",
        "Source Scan",
    ]
    _xl_header_row(ws_nessus, nessus_headers)

    for scan_stem, scan_display, nessus_data in all_nessus_scans:
        for host in nessus_data.get("hosts", []):
            ip       = host.get("ip", "")
            hostname = host.get("hostname", "")
            for f in host.get("findings", []):
                sev_str  = severity_int_to_str(f["severity_int"])
                cves_str = ", ".join(f["cves"]) if f["cves"] else ""
                row = [
                    ip, hostname,
                    f["port"], f["protocol"],
                    f["plugin_id"], f["plugin_name"], sev_str,
                    f.get("cvss3_base", ""), f.get("cvss_base", ""),
                    cves_str,
                    _trunc(f.get("description"), 500),
                    _trunc(f.get("solution"), 300),
                    _trunc(f.get("plugin_output"), 300),
                    scan_display,
                ]
                ws_nessus.append(row)
                row_idx  = ws_nessus.max_row
                sev_cell = ws_nessus.cell(row=row_idx, column=7)
                fill     = _xl_severity_fill(sev_str)
                if fill:
                    sev_cell.fill = fill

    ws_nessus.freeze_panes = "A2"
    ws_nessus.auto_filter.ref = ws_nessus.dimensions
    _xl_autofit(ws_nessus)

    # ---- Burp sheet ----
    ws_burp = wb.create_sheet("Burp")
    burp_headers = [
        "Host", "Path", "Issue Name", "Severity", "Confidence",
        "Location", "Issue Detail", "Remediation", "Source Scan",
    ]
    _xl_header_row(ws_burp, burp_headers)

    for scan_stem, scan_display, burp_data in all_burp_scans:
        for host in burp_data.get("hosts", []):
            host_label = host.get("url") or host.get("ip") or ""
            for issue in host.get("issues", []):
                remediation = issue.get("remediation_detail") or issue.get("remediation_background") or ""
                row = [
                    host_label,
                    issue.get("path", ""),
                    issue.get("name", ""),
                    issue.get("severity", ""),
                    issue.get("confidence", ""),
                    issue.get("location", ""),
                    _trunc(issue.get("issue_detail"), 500),
                    _trunc(remediation, 300),
                    scan_display,
                ]
                ws_burp.append(row)
                row_idx  = ws_burp.max_row
                sev_cell = ws_burp.cell(row=row_idx, column=4)
                fill     = _xl_severity_fill(issue.get("severity", ""))
                if fill:
                    sev_cell.fill = fill

    ws_burp.freeze_panes = "A2"
    ws_burp.auto_filter.ref = ws_burp.dimensions
    _xl_autofit(ws_burp)

    out_path = vault_dir / "Export.xlsx"
    wb.save(str(out_path))
    logging.info(f"Excel export written: {out_path}")
    return out_path


# ============================================================
# Main
# ============================================================

def main() -> None:
    ap = argparse.ArgumentParser(
        description="mAIpper v0.7 — Pentest scan analysis and Obsidian vault generator"
    )
    ap.add_argument("--xml",        default=None,  help="Process a single Nmap XML file")
    ap.add_argument("--scans-dir",  default="scans", help="Base scans directory (default: ./scans)")
    ap.add_argument("--vault",      default="Obsidian", help="Obsidian vault output directory")
    ap.add_argument("--tool-name",  default="Nmap")
    ap.add_argument("--model",      default="qwen2.5:14b-instruct-q5_K_M")
    ap.add_argument("--ollama-url", default="http://localhost:11434")
    ap.add_argument("--no-ollama",  action="store_true", help="Skip AI analysis")
    ap.add_argument("--no-canvas",  action="store_true", help="Skip canvas generation")
    ap.add_argument("--no-nessus",  action="store_true", help="Skip Nessus scan processing")
    ap.add_argument("--no-burp",    action="store_true", help="Skip Burp Suite scan processing")
    ap.add_argument("--excel",      action="store_true", help="Generate Export.xlsx (requires openpyxl)")
    ap.add_argument("--canvas-name",           default="Assessment Canvas.canvas")
    ap.add_argument("--canvas-cols",           type=int, default=2,
                    help="Host cards per row within each subnet group (default: 2)")
    ap.add_argument("--canvas-groups-per-row", type=int, default=3,
                    help="Subnet groups per canvas row (default: 3)")
    ap.add_argument("-v", "--verbose", action="count", default=1,
                    help="Increase verbosity: -v = INFO, -vv = DEBUG")

    args = ap.parse_args()

    level = logging.WARNING
    if args.verbose >= 1: level = logging.INFO
    if args.verbose >= 2: level = logging.DEBUG
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")
    logging.debug(f"Args: {vars(args)}")

    vault_dir = Path(args.vault).resolve()
    base      = Path(args.scans_dir).resolve()

    scan_host_map: dict[str, list[str]]   = {}
    all_analyses:  list[tuple[str, str]]  = []

    # ------------------------------------------------------------------ #
    # Nmap                                                                 #
    # ------------------------------------------------------------------ #
    all_nmap_scans: list[tuple[str, str, dict]] = []

    xml_files: list[Path] = []
    if args.xml:
        xml_path = Path(args.xml).resolve()
        if not xml_path.exists():
            logging.error(f"File not found: {xml_path}")
            return
        xml_files = [xml_path]
    else:
        nmap_dir = base / "nmap"
        if not nmap_dir.exists():
            nmap_dir = base / "Nmap"
        if nmap_dir.exists():
            xml_files = sorted(nmap_dir.glob("*.xml"))
        else:
            logging.info("No scans/nmap/ directory found; skipping Nmap processing")

    if xml_files:
        logging.info(f"Processing {len(xml_files)} Nmap file(s)")
    for xml_path in xml_files:
        logging.info(f"Nmap → {xml_path.name}")
        scan_data = parse_nmap_xml(xml_path)

        analysis: str | None = None
        if not args.no_ollama:
            try:
                analysis = ollama_chat(
                    args.ollama_url, args.model, build_ollama_prompt(scan_data)
                )
            except Exception as exc:
                logging.warning(f"Ollama failed ({xml_path.name}): {exc}")
                analysis = f"_Ollama failed: {exc}_"
        else:
            logging.info("Skipping Ollama (--no-ollama)")

        result = create_obsidian_vault(
            vault_dir,
            xml_path.stem,
            scan_data,
            args.tool_name,
            analysis,
            None if args.no_ollama else args.model,
        )

        scan_host_map[result["scan_stem"]] = result["host_stems"]
        if analysis and not args.no_ollama:
            all_analyses.append((result["scan_display"], analysis))
        all_nmap_scans.append((result["scan_stem"], result["scan_display"], scan_data))

    # ------------------------------------------------------------------ #
    # Nessus                                                               #
    # ------------------------------------------------------------------ #
    all_nessus_scans: list[tuple[str, str, dict]] = []

    if not args.no_nessus:
        nessus_dir = base / "nessus"
        if not nessus_dir.exists():
            nessus_dir = base / "Nessus"
        nessus_files: list[Path] = sorted(nessus_dir.glob("*.nessus")) if nessus_dir.exists() else []

        if nessus_files:
            logging.info(f"Processing {len(nessus_files)} Nessus file(s)")
        for nessus_path in nessus_files:
            logging.info(f"Nessus → {nessus_path.name}")
            nessus_data = parse_nessus_xml(nessus_path)

            analysis = None
            if not args.no_ollama:
                try:
                    analysis = ollama_chat(
                        args.ollama_url, args.model, build_nessus_ollama_prompt(nessus_data)
                    )
                except Exception as exc:
                    logging.warning(f"Ollama failed ({nessus_path.name}): {exc}")
                    analysis = f"_Ollama failed: {exc}_"

            result = create_nessus_vault(
                vault_dir,
                nessus_path.stem,
                nessus_data,
                analysis,
                None if args.no_ollama else args.model,
            )

            scan_host_map[result["scan_stem"]] = result["host_stems"]
            if analysis and not args.no_ollama:
                all_analyses.append((result["scan_display"], analysis))
            all_nessus_scans.append((result["scan_stem"], result["scan_display"], nessus_data))
    else:
        logging.info("Skipping Nessus (--no-nessus)")

    # ------------------------------------------------------------------ #
    # Burp Suite                                                           #
    # ------------------------------------------------------------------ #
    all_burp_scans: list[tuple[str, str, dict]] = []

    if not args.no_burp:
        burp_dir = base / "burp"
        if not burp_dir.exists():
            burp_dir = base / "Burp"
        burp_files: list[Path] = sorted(burp_dir.glob("*.xml")) if burp_dir.exists() else []

        if burp_files:
            logging.info(f"Processing {len(burp_files)} Burp file(s)")
        for burp_path in burp_files:
            logging.info(f"Burp → {burp_path.name}")
            # Distinguish Burp XML from Nmap XML by checking root tag
            try:
                root_tag = ET.parse(burp_path).getroot().tag
            except Exception as exc:
                logging.warning(f"Failed to parse {burp_path.name}: {exc}")
                continue
            if root_tag != "issues":
                logging.debug(f"Skipping non-Burp XML: {burp_path.name} (root tag: {root_tag})")
                continue

            burp_data = parse_burp_xml(burp_path)

            analysis = None
            if not args.no_ollama:
                try:
                    analysis = ollama_chat(
                        args.ollama_url, args.model, build_burp_ollama_prompt(burp_data)
                    )
                except Exception as exc:
                    logging.warning(f"Ollama failed ({burp_path.name}): {exc}")
                    analysis = f"_Ollama failed: {exc}_"

            result = create_burp_vault(
                vault_dir,
                burp_path.stem,
                burp_data,
                analysis,
                None if args.no_ollama else args.model,
            )

            scan_host_map[result["scan_stem"]] = result["host_stems"]
            if analysis and not args.no_ollama:
                all_analyses.append((result["scan_display"], analysis))
            all_burp_scans.append((result["scan_stem"], result["scan_display"], burp_data))
    else:
        logging.info("Skipping Burp Suite (--no-burp)")

    # ------------------------------------------------------------------ #
    # Canvas                                                               #
    # ------------------------------------------------------------------ #
    if not args.no_canvas:
        # Build Priority Targets text: ask Ollama if we have any analyses,
        # otherwise fall back to a static severity-sorted list.
        priority_targets_text: str | None = None
        if not args.no_ollama and all_analyses:
            try:
                pt_prompt = build_priority_targets_prompt(vault_dir, all_analyses)
                raw_pt    = ollama_chat(args.ollama_url, args.model, pt_prompt)
                # Prepend the standard header so the canvas node is self-labelled
                priority_targets_text = "## Priority Targets\n\n" + raw_pt.strip()
            except Exception as exc:
                logging.warning(f"Ollama priority targets failed: {exc}")
                priority_targets_text = _build_priority_targets_fallback(vault_dir)
        else:
            priority_targets_text = _build_priority_targets_fallback(vault_dir)

        build_canvas(
            vault_dir,
            args.canvas_name,
            scan_host_map,
            all_analyses,
            canvas_cols=args.canvas_cols,
            max_groups_per_row=args.canvas_groups_per_row,
            priority_targets_text=priority_targets_text,
        )
    else:
        logging.info("Skipping canvas (--no-canvas)")

    # ------------------------------------------------------------------ #
    # Excel export                                                         #
    # ------------------------------------------------------------------ #
    if args.excel:
        out = export_excel(vault_dir, all_nmap_scans, all_nessus_scans, all_burp_scans)
        if out:
            logging.info(f"Excel export: {out}")
        else:
            logging.error("Excel export failed — is openpyxl installed?")

    logging.info("Done.")


if __name__ == "__main__":
    main()
