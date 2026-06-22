#!/usr/bin/env python3

"""
mAIpper v0.8 - Pentest Tool Analysis & Obsidian Export Tool

Changes from v0.7:
  - AutoRecon integration: ingests AutoRecon results directories
    (per-target subdirs with scans/xml/, scans/tcp*/, scans/udp*/)
  - Layer 1: feeds AutoRecon's nmap XMLs through existing Nmap pipeline
  - Layer 2: Python extractors parse structured facts from tool outputs
    (gobuster, nikto, enum4linux, smbmap, smbclient, whatweb, snmpwalk,
    onesixtyone, sslscan, dnsrecon, curl) before LLM analysis
  - Two-pass LLM analysis for AutoRecon (fact extraction → analysis),
    matching the Nessus approach for accuracy
  - ## AutoRecon Enumeration section in host notes with structured tables
  - Scans/<target> - AutoRecon.md scan notes with collapsible manual
    commands and commands log
  - --autorecon DIR / --no-autorecon flags; auto-discovers scans/autorecon/
  - AutoRecon sheet in Excel export
  - Canvas Campaign Overview updated with AutoRecon stats

Changes from v0.6:
  - Nessus parser: ingests .nessus XML files from scans/nessus/; adds
    ## Nessus Findings sections to host notes (severity-sorted); creates
    Scans/<name> - Nessus.md; AI analysis focused on CVEs and exploitation
    priority; nessus_max_severity stored in frontmatter for canvas coloring
  - Burp Suite parser: ingests Burp Pro Scanner XML from scans/burp/; adds
    ## Burp Suite Findings sections to host notes; creates
    Scans/<name> - Burp.md; AI analysis focused on web attack surface
  - Excel export (--excel): generates <vault>/Export.xlsx with Summary,
    Nmap, Nessus, and Burp sheets; severity color-coding, filters, frozen
    headers, auto-fit columns
  - --no-nessus / --no-burp flags to skip those parsers
  - Canvas Campaign Overview and Next Steps updated with Nessus/Burp data
  - Nmap re-runs preserve existing Nessus/Burp sections in host notes
  - Canvas "Priority Targets" text node: AI-ranked list of highest-priority
    hosts (by CVSS, CVEs, service exposure, Burp findings); positioned between
    Campaign Overview and subnet groups; edges Overview→PriorityTargets and
    PriorityTargets→each subnet group; falls back to static severity sort when
    Ollama is skipped; stable node ID "priority-targets"

Requirements:
  pip install requests
  pip install openpyxl   # only needed for --excel

Author: Zachary Levine
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import logging
import math
import re
from pathlib import Path
import xml.etree.ElementTree as ET

import requests

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# Constants
# ============================================================

IPV4_RE          = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$")
CVE_RE           = re.compile(r"\bCVE-\d{4}-\d+\b", re.IGNORECASE)
PORT_IN_TEXT_RE  = re.compile(r"(?:port\s+(\d{1,5})|(\d{1,5})/(?:tcp|udp))", re.IGNORECASE)
IP_IN_TEXT_RE    = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
FQDN_IN_TEXT_RE  = re.compile(
    r"\b(?:[a-zA-Z0-9](?:[a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,6}\b"
)
AUTORECON_FILENAME_RE = re.compile(
    r"^(tcp|udp)_(\d+)_(\w+)_(.+)\.(txt|html|xml|json|log)$"
)
AUTORECON_PORT_DIR_RE = re.compile(r"^(tcp|udp)_?(\d+)$")

STATUS_COLORS: dict[str, str | None] = {
    "not-started": None,
    "in-progress":  "3",   # yellow
    "done":         "4",   # green
    "exploited":    "1",   # red
    "blocked":      "6",   # purple
}

# Canvas colors for Nessus severity (applied when status == not-started)
NESSUS_SEVERITY_CANVAS_COLORS: dict[int, str] = {
    4: "1",   # Critical → red
    3: "5",   # High     → orange
    2: "3",   # Medium   → yellow
}

OPERATOR_NOTES_SENTINEL = "## Operator Notes"
OPERATOR_NOTES_HINT = "_Add your own findings, observations, and next steps below._"

# Canonical section order in host notes
BODY_SECTION_ORDER = [
    "## Open Ports",
    "## Nessus Findings",
    "## Burp Suite Findings",
    "## AutoRecon Enumeration",
    "## Scan References",
    OPERATOR_NOTES_SENTINEL,
]

CARD_W          = 360
CARD_H          = 220
CARD_GAP_X      = 60
CARD_GAP_Y      = 60
GROUP_PAD       = 60
GROUP_LABEL_H   = 40
SCAN_CARD_W     = 360
SCAN_CARD_H     = 160
OVERVIEW_W      = 720
OVERVIEW_H      = 220
NEXT_STEPS_W    = 640
NEXT_STEPS_H    = 420
GROUP_GAP       = 120
ROW_GAP         = 120


# ============================================================
# Basic helpers
# ============================================================

def safe_filename(s: str) -> str:
    for ch in '<>:"/\\|?*\n\r\t':
        s = s.replace(ch, "_")
    return s.strip().strip(".")


def stable_id(key: str) -> str:
    return hashlib.md5(key.encode()).hexdigest()[:12]


def is_ipv4(value: str) -> bool:
    return bool(value and IPV4_RE.match(value.strip()))


def is_probable_fqdn(value: str) -> bool:
    v = (value or "").strip().rstrip(".")
    return bool(v and "." in v and not is_ipv4(v))


def ensure_md_suffix(name: str) -> str:
    return name if name.lower().endswith(".md") else f"{name}.md"


def get_subnet_label(ip: str) -> str:
    if not is_ipv4(ip):
        return "unknown"
    parts = ip.split(".")
    return f"{parts[0]}.{parts[1]}.{parts[2]}.0/24"


def severity_int_to_str(n: int) -> str:
    return {0: "Info", 1: "Low", 2: "Medium", 3: "High", 4: "Critical"}.get(n, "Unknown")


def severity_str_to_int(s: str) -> int:
    return {"information": 0, "info": 0, "low": 1, "medium": 2, "high": 3, "critical": 4}.get(
        (s or "").lower(), 0
    )


# ============================================================
# Frontmatter helpers
# ============================================================

def _fm_encode(v: object) -> str:
    if isinstance(v, list):
        return json.dumps(v, ensure_ascii=False)
    if v is None:
        return "null"
    return str(v)


def write_frontmatter(fm: dict) -> str:
    lines = ["---"]
    for k, v in fm.items():
        lines.append(f"{k}: {_fm_encode(v)}")
    lines.append("---")
    return "\n".join(lines) + "\n"


def read_frontmatter(text: str) -> tuple[dict, str]:
    if not text.startswith("---\n"):
        return {}, text
    end = text.find("\n---\n", 4)
    if end == -1:
        return {}, text
    fm_text = text[4:end]
    body = text[end + 5:]
    fm: dict = {}
    for line in fm_text.splitlines():
        if ": " not in line and not line.endswith(":"):
            continue
        key, _, raw = line.partition(": ")
        key = key.strip()
        raw = raw.strip()
        if not key:
            continue
        try:
            fm[key] = json.loads(raw)
        except (json.JSONDecodeError, ValueError):
            fm[key] = raw if raw else None
    return fm, body


def extract_operator_notes(body: str) -> str:
    idx = body.find(OPERATOR_NOTES_SENTINEL)
    if idx == -1:
        return ""
    after = body[idx + len(OPERATOR_NOTES_SENTINEL):]
    lines = after.splitlines()
    content_lines: list[str] = []
    past_hint = False
    for line in lines:
        if not past_hint and line.strip() in ("", OPERATOR_NOTES_HINT):
            continue
        past_hint = True
        content_lines.append(line)
    while content_lines and not content_lines[-1].strip():
        content_lines.pop()
    return "\n".join(content_lines)


def extract_body_section(body: str, section_header: str) -> str:
    """Extract content of a named ## section (without its header line)."""
    pattern = rf"(?:^|\n){re.escape(section_header)}\n(.*?)(?=\n## |\Z)"
    m = re.search(pattern, body, re.DOTALL)
    return m.group(1).strip() if m else ""


# ============================================================
# Port hints & tags
# ============================================================

def get_tags_from_ports(open_ports: list) -> list[str]:
    tags: set[str] = set()
    port_nums = {p.get("port", 0) for p in open_ports}
    svc_names = {(p.get("service", {}).get("name", "") or "").lower() for p in open_ports}

    if port_nums & {80, 81, 443, 8000, 8080, 8443} or svc_names & {"http", "https", "http-proxy"}:
        tags.add("web")
    if 445 in port_nums or svc_names & {"microsoft-ds", "smb"}:
        tags.add("smb")
    if port_nums & {88, 464} or any("kerberos" in s for s in svc_names):
        tags.add("kerberos")
    if port_nums & {389, 636, 3268, 3269} or svc_names & {"ldap", "ldaps", "globalcatldap", "globalcatldapssl"}:
        tags.add("ldap")
    if 135 in port_nums or "msrpc" in svc_names:
        tags.add("rpc")
    if port_nums & {1433, 1434} or any("mssql" in s for s in svc_names):
        tags.add("mssql")
    if 3306 in port_nums or "mysql" in svc_names:
        tags.add("mysql")
    if 5432 in port_nums or any("postgres" in s for s in svc_names):
        tags.add("postgres")
    if 22 in port_nums or "ssh" in svc_names:
        tags.add("ssh")
    if 3389 in port_nums or "ms-wbt-server" in svc_names:
        tags.add("rdp")
    if port_nums & {5985, 5986} or any("winrm" in s for s in svc_names):
        tags.add("winrm")
    if 21 in port_nums or "ftp" in svc_names:
        tags.add("ftp")
    if 53 in port_nums or "domain" in svc_names:
        tags.add("dns")
    if port_nums & {161, 162} or "snmp" in svc_names:
        tags.add("snmp")
    if port_nums & {25, 465, 587} or svc_names & {"smtp", "smtps", "submission"}:
        tags.add("smtp")
    if 2049 in port_nums or "nfs" in svc_names:
        tags.add("nfs")
    if 6379 in port_nums or any("redis" in s for s in svc_names):
        tags.add("redis")
    if {"kerberos", "ldap", "smb"} <= tags:
        tags.add("domain-controller")
    return sorted(tags)


def get_port_hints(port: int, service_name: str, product: str = "", extrainfo: str = "") -> list[str]:
    svc   = (service_name or "").lower()
    prod  = (product or "").lower()
    extra = (extrainfo or "").lower()
    hints: list[str] = []

    if port in (80, 81, 443, 8000, 8080, 8443) or svc in ("http", "https", "http-proxy"):
        hints.append("Web service detected; consider whatweb, nikto, gobuster/feroxbuster/ffuf, curl, and manual browsing.")
        hints.append("Check for default credentials, admin panels, exposed APIs, virtual hosts, and interesting headers.")
    if port == 445 or svc in ("microsoft-ds", "smb", "netbios-ssn"):
        hints.append("SMB detected; consider netexec smb, smbclient, smbmap, and enum4linux-ng.")
        hints.append("Check share access, null sessions, SMB signing, local admin reuse, and domain membership clues.")
    if port in (88, 464) or "kerberos" in svc or "kerberos" in prod:
        hints.append("Kerberos-related service detected; consider kerbrute, netexec, GetUserSPNs, and AS-REP roast checks.")
        hints.append("Look for domain naming clues, valid usernames, SPNs, and clock skew issues.")
    if port in (389, 636, 3268, 3269) or svc in ("ldap", "ldaps", "globalcatldap", "globalcatldapssl"):
        hints.append("LDAP detected; consider ldapsearch, netexec ldap, and directory enumeration.")
        hints.append("Look for anonymous bind, naming contexts, domain structure, users, groups, and password policy info.")
    if port == 135 or svc == "msrpc":
        hints.append("MSRPC detected; consider rpcclient, netexec, and enumerate Windows service exposure and domain clues.")
    if port == 139 or svc == "netbios-ssn":
        hints.append("NetBIOS detected; enumerate shares, names, and Windows host information.")
    if port in (5985, 5986) or "winrm" in svc:
        hints.append("WinRM detected; consider netexec winrm and evil-winrm if credentials are obtained.")
    if port in (1433, 1434) or "mssql" in svc or "sql server" in prod:
        hints.append("MSSQL detected; consider impacket-mssqlclient, netexec mssql, and SQL login enumeration.")
    if port == 3306 or "mysql" in svc:
        hints.append("MySQL detected; test authentication paths and enumerate version-specific exposure carefully.")
    if port == 5432 or "postgres" in svc:
        hints.append("PostgreSQL detected; test for weak/default credentials and accessible databases.")
    if port == 27017 or "mongodb" in svc:
        hints.append("MongoDB detected; verify whether authentication is required and whether remote access is overly exposed.")
    if port == 22 or svc == "ssh":
        hints.append("SSH detected; consider ssh-audit, banner review, key-based auth checks, and cautious credential testing.")
    if port == 21 or svc == "ftp":
        hints.append("FTP detected; check for anonymous access, writable locations, and plaintext credential exposure.")
    if port in (25, 465, 587) or svc in ("smtp", "smtps", "submission"):
        hints.append("SMTP detected; consider smtp-user-enum or swaks and look for open relay or user enumeration behavior.")
    if port == 53 or svc == "domain":
        hints.append("DNS detected; consider dig, nslookup, dnsrecon, and zone transfer testing where appropriate.")
    if port == 69 or svc == "tftp":
        hints.append("TFTP detected; check for unauthenticated file retrieval or upload opportunities.")
    if port == 111 or svc == "rpcbind":
        hints.append("RPCbind detected; consider rpcinfo and follow-on enumeration for NFS and related services.")
    if port == 2049 or svc == "nfs":
        hints.append("NFS detected; use showmount and test for accessible exports and weak export permissions.")
    if port in (161, 162) or svc == "snmp":
        hints.append("SNMP detected; try snmpwalk/onesixtyone and test common community strings carefully.")
    if port == 3389 or svc == "ms-wbt-server":
        hints.append("RDP detected; consider netexec rdp, xfreerdp, and review NLA / domain clues.")
    if port in (5900, 5901, 5902) or svc == "vnc":
        hints.append("VNC detected; verify authentication requirements and assess screenshot or password attack viability.")
    if port == 6379 or "redis" in svc:
        hints.append("Redis detected; verify bind/auth configuration and whether dangerous commands are exposed remotely.")
    if port == 11211 or "memcached" in svc:
        hints.append("Memcached detected; check whether it is exposed beyond localhost and assess information disclosure risk.")
    if port in (9200, 9300) or "elasticsearch" in svc:
        hints.append("Elasticsearch detected; check for unauthenticated cluster or index access.")
    if port in (2375, 2376) or "docker" in svc:
        hints.append("Docker API exposure may be high risk; verify remote daemon access and authentication/TLS posture.")
    if port == 6443 or "kubernetes" in svc:
        hints.append("Kubernetes-related service detected; look for exposed API endpoints and auth configuration issues.")
    if "ssl" in extra or "tls" in extra:
        hints.append("TLS-related context detected; review certificate details, protocol support, and any weak crypto indicators.")
    return hints


# ============================================================
# Nmap XML parsing
# ============================================================

def parse_nmap_xml(xml_path: Path) -> dict:
    logging.info(f"Parsing Nmap XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    scan_info = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "nmap_args":   root.get("args", ""),
        "nmap_version":root.get("version", ""),
        "hosts":       [],
    }

    for host in root.findall("host"):
        state_el = host.find("status")
        state = state_el.get("state") if state_el is not None else "unknown"

        addresses = [
            {"addr": a.get("addr", ""), "addrtype": a.get("addrtype", "")}
            for a in host.findall("address")
        ]

        hostnames = []
        hn_parent = host.find("hostnames")
        if hn_parent is not None:
            for hn in hn_parent.findall("hostname"):
                hostnames.append({"name": hn.get("name", ""), "type": hn.get("type", "")})

        ports = []
        ports_parent = host.find("ports")
        if ports_parent is not None:
            for port_el in ports_parent.findall("port"):
                state_el2 = port_el.find("state")
                if state_el2 is None or state_el2.get("state") != "open":
                    continue
                service_el = port_el.find("service")
                service = {}
                if service_el is not None:
                    service = {
                        "name":      service_el.get("name", ""),
                        "product":   service_el.get("product", ""),
                        "version":   service_el.get("version", ""),
                        "extrainfo": service_el.get("extrainfo", ""),
                        "tunnel":    service_el.get("tunnel", ""),
                    }
                scripts = [
                    {"id": s.get("id", ""), "output": s.get("output", "")}
                    for s in port_el.findall("script")
                ]
                ports.append({
                    "protocol": port_el.get("protocol", ""),
                    "port":     int(port_el.get("portid", "0")),
                    "service":  service,
                    "scripts":  scripts,
                })

        host_record = {
            "state":      state,
            "addresses":  addresses,
            "hostnames":  hostnames,
            "open_ports": sorted(ports, key=lambda p: (p["protocol"], p["port"])),
        }
        scan_info["hosts"].append(host_record)
        logging.debug(
            f"Nmap host: {choose_host_display_name(host_record)} "
            f"state={state} ports={len(host_record['open_ports'])}"
        )

    logging.info(f"Nmap done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# Nessus XML parsing
# ============================================================

def parse_nessus_xml(xml_path: Path) -> dict:
    """Parse a .nessus (NessusClientData_v2) file."""
    logging.info(f"Parsing Nessus XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Report name from the Report element
    report_el = root.find("Report")
    report_name = report_el.get("name", xml_path.stem) if report_el is not None else xml_path.stem

    scan_info: dict = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "report_name": report_name,
        "hosts":       [],
    }

    report_hosts = root.findall(".//ReportHost")
    for rh in report_hosts:
        # Host properties
        props: dict[str, str] = {}
        host_props_el = rh.find("HostProperties")
        if host_props_el is not None:
            for tag_el in host_props_el.findall("tag"):
                props[tag_el.get("name", "")] = (tag_el.text or "").strip()

        ip       = props.get("host-ip", rh.get("name", ""))
        hostname = props.get("hostname", props.get("host-fqdn", ""))

        findings: list[dict] = []
        for item in rh.findall("ReportItem"):
            plugin_id   = item.get("pluginID", "")
            plugin_name = item.get("pluginName", "")
            severity    = int(item.get("severity", "0"))
            port        = int(item.get("port", "0"))
            protocol    = item.get("protocol", "tcp")

            def _txt(tag: str) -> str:
                el = item.find(tag)
                return (el.text or "").strip() if el is not None else ""

            cves = [el.text.strip() for el in item.findall("cve") if el.text]

            findings.append({
                "plugin_id":    plugin_id,
                "plugin_name":  plugin_name,
                "severity_int": severity,
                "port":         port,
                "protocol":     protocol,
                "description":  _txt("description"),
                "solution":     _txt("solution"),
                "cves":         cves,
                "cvss_base":    _txt("cvss_base_score"),
                "cvss3_base":   _txt("cvss3_base_score"),
                "plugin_output":_txt("plugin_output"),
            })

        # Sort findings: highest severity first, then plugin name
        findings.sort(key=lambda f: (-f["severity_int"], f["plugin_name"]))

        host_record = {
            "ip":       ip,
            "hostname": hostname,
            "findings": findings,
        }
        scan_info["hosts"].append(host_record)
        logging.debug(
            f"Nessus host: {ip} ({hostname}) findings={len(findings)}"
        )

    logging.info(f"Nessus done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# Burp Suite XML parsing
# ============================================================

def parse_burp_xml(xml_path: Path) -> dict:
    """Parse a Burp Suite Pro Scanner Issues XML export."""
    logging.info(f"Parsing Burp XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    scan_info: dict = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "hosts":       [],
    }

    # Group issues by host IP
    host_map: dict[str, dict] = {}  # ip → {ip, url, issues}

    for issue_el in root.findall("issue"):
        host_el = issue_el.find("host")
        if host_el is None:
            continue

        ip  = host_el.get("ip", "")
        url = (host_el.text or "").strip()

        # Derive a canonical host key (prefer IP, fall back to URL base)
        host_key = ip if ip else url

        def _txt(tag: str) -> str:
            el = issue_el.find(tag)
            return (el.text or "").strip() if el is not None else ""

        issue = {
            "name":                _txt("name"),
            "path":                _txt("path"),
            "location":            _txt("location"),
            "severity":            _txt("severity"),
            "confidence":          _txt("confidence"),
            "issue_detail":        _txt("issueDetail"),
            "issue_background":    _txt("issueBackground"),
            "remediation_detail":  _txt("remediationDetail"),
            "remediation_background": _txt("remediationBackground"),
        }

        if host_key not in host_map:
            host_map[host_key] = {"ip": ip, "url": url, "issues": []}
        host_map[host_key]["issues"].append(issue)

    sev_order = {"high": 0, "medium": 1, "low": 2, "information": 3}
    for host_rec in host_map.values():
        host_rec["issues"].sort(
            key=lambda i: (sev_order.get(i["severity"].lower(), 4), i["name"])
        )
        scan_info["hosts"].append(host_rec)
        logging.debug(
            f"Burp host: {host_rec['ip']} ({host_rec['url']}) issues={len(host_rec['issues'])}"
        )

    logging.info(f"Burp done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# AutoRecon tool-output extractors
# ============================================================

def _extract_dirbusting(text: str) -> dict:
    """Extract discovered paths from gobuster / feroxbuster / dirsearch output."""
    paths: list[dict] = []
    # gobuster: /path (Status: 200) [Size: 1234]
    for m in re.finditer(
        r"^(/\S+)\s+\(Status:\s*(\d+)\)\s+\[Size:\s*(\d+)\](?:\s+\[--> (\S+)\])?",
        text, re.MULTILINE,
    ):
        paths.append({"path": m.group(1), "status": int(m.group(2)),
                       "size": int(m.group(3)), "redirect": m.group(4)})
    # feroxbuster: 200  GET  10l  20w  1234c http://target/path
    if not paths:
        for m in re.finditer(
            r"^(\d{3})\s+\w+\s+\d+l\s+\d+w\s+(\d+)c\s+https?://[^/]+(/.*)$",
            text, re.MULTILINE,
        ):
            paths.append({"path": m.group(3), "status": int(m.group(1)),
                           "size": int(m.group(2)), "redirect": None})
    # dirsearch: 200  1234B  /path
    if not paths:
        for m in re.finditer(
            r"^\s*(\d{3})\s+[\d.]+[KMB]*\s+(/.+)$", text, re.MULTILINE,
        ):
            paths.append({"path": m.group(2).strip(), "status": int(m.group(1)),
                           "size": 0, "redirect": None})

    paths = [p for p in paths if p["status"] != 404]
    interest_order = {200: 0, 301: 1, 302: 1, 401: 2, 403: 2}
    paths.sort(key=lambda p: (interest_order.get(p["status"], 9), p["path"]))
    interesting = [p for p in paths if p["status"] in interest_order]
    return {
        "tool": "dirbusting",
        "paths": paths[:100],
        "interesting": interesting[:50],
        "total_found": len(paths),
    }


def _extract_nikto(text: str) -> dict:
    """Extract findings from nikto output."""
    server = ""
    headers: dict[str, str] = {}
    findings: list[dict] = []

    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("+"):
            continue
        body = line[1:].strip()

        sm = re.match(r"Server:\s+(.+)", body)
        if sm:
            server = sm.group(1).strip()
            continue

        hm = re.match(r"Retrieved\s+(\S+)\s+header:\s+(.+)", body, re.IGNORECASE)
        if hm:
            headers[hm.group(1).lower()] = hm.group(2).strip()
            continue

        fm = re.match(r"(?:OSVDB-(\d+):\s+)?(/\S*):\s+(.+)", body)
        if fm:
            findings.append({
                "osvdb": fm.group(1) or "",
                "path": fm.group(2),
                "description": fm.group(3).strip(),
            })

    return {
        "tool": "nikto",
        "server": server,
        "headers": headers,
        "findings": findings,
        "total_findings": len(findings),
    }


def _extract_enum4linux(text: str) -> dict:
    """Extract SMB/NetBIOS enumeration data from enum4linux output."""
    result: dict = {
        "tool": "enum4linux",
        "os_info": "",
        "domain": "",
        "workgroup": "",
        "null_session": False,
        "users": [],
        "shares": [],
        "groups": [],
        "password_policy": {},
    }

    os_m = re.search(r"OS information on \S+.*?:\s*(.+)", text)
    if not os_m:
        os_m = re.search(r"OS:\s+(.+)", text)
    if os_m:
        result["os_info"] = os_m.group(1).strip()

    dm = re.search(r"Domain Name:\s+(\S+)", text)
    if dm:
        result["domain"] = dm.group(1)
    wm = re.search(r"Workgroup:\s+(\S+)", text)
    if wm:
        result["workgroup"] = wm.group(1)

    if re.search(r"(?:Null session|anonymous.*(?:allowed|access|success))", text, re.IGNORECASE):
        result["null_session"] = True

    # RID cycling users: 500: DOMAIN\Administrator (Local User)
    for m in re.finditer(r"\d+:\s+\S+\\(\S+)\s+\(", text):
        u = m.group(1)
        if u not in result["users"]:
            result["users"].append(u)

    # Shares: //host/SHARE  Mapping: OK  Listing: OK
    for m in re.finditer(
        r"//\S+/(\S+)\s+Mapping:\s+(\w+)\s+Listing:\s+(\w+)", text
    ):
        access = "read" if m.group(2) == "OK" else "denied"
        if m.group(3) == "OK":
            access = "read/write" if access == "read" else "read"
        result["shares"].append({"name": m.group(1), "access": access})
    # Fallback: Disk share lines
    if not result["shares"]:
        for m in re.finditer(r"^\s*(\S+)\s+(Disk|IPC|Printer)", text, re.MULTILINE):
            result["shares"].append({"name": m.group(1), "type": m.group(2), "access": ""})

    for m in re.finditer(r"group:\[([^\]]+)\]", text):
        g = m.group(1).strip()
        if g and g not in result["groups"]:
            result["groups"].append(g)

    pp = result["password_policy"]
    ml = re.search(r"Minimum password length:\s*(\d+)", text)
    if ml:
        pp["min_length"] = int(ml.group(1))
    lt = re.search(r"Account Lockout Threshold:\s*(\d+)", text)
    if lt:
        pp["lockout_threshold"] = int(lt.group(1))
    if re.search(r"Password Complexity.*Enabled", text, re.IGNORECASE):
        pp["complexity"] = True

    return result


def _extract_smbmap(text: str) -> dict:
    """Extract share permissions from smbmap output."""
    shares: list[dict] = []
    for m in re.finditer(
        r"^\s+(\S+)\s+(READ ONLY|READ,\s*WRITE|READ/WRITE|NO ACCESS)\s*(.*?)$",
        text, re.MULTILINE,
    ):
        perm = m.group(2).strip().replace(",", "/").replace("  ", " ")
        shares.append({
            "name": m.group(1),
            "permissions": perm,
            "comment": m.group(3).strip(),
        })
    writable = [s["name"] for s in shares if "WRITE" in s["permissions"]]
    readable = [s["name"] for s in shares
                if "READ" in s["permissions"] and s["name"] not in writable]
    return {
        "tool": "smbmap",
        "shares": shares,
        "writable_shares": writable,
        "readable_shares": readable,
    }


def _extract_smbclient(text: str) -> dict:
    """Extract share listing from smbclient output."""
    shares: list[dict] = []
    for m in re.finditer(r"^\s+(\S+)\s+(Disk|IPC|Printer)\s*(.*?)$", text, re.MULTILINE):
        shares.append({
            "name": m.group(1),
            "type": m.group(2),
            "comment": m.group(3).strip(),
        })
    return {"tool": "smbclient", "shares": shares}


def _extract_whatweb(text: str) -> dict:
    """Extract technology fingerprints from whatweb output."""
    technologies: list[dict] = []
    title = ""
    status_code = 0

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # WhatWeb format: http://target [200 OK] Apache[2.4.41], PHP[7.4.3], ...
        sc_m = re.search(r"\[(\d{3})\s", line)
        if sc_m:
            status_code = int(sc_m.group(1))
        for seg in re.split(r",\s*", line):
            seg = seg.strip()
            tm = re.match(r"([\w][\w\s./-]*?)(?:\[([^\]]*)\])?$", seg)
            if tm:
                name = tm.group(1).strip()
                version = (tm.group(2) or "").strip()
                if name.lower() == "title":
                    title = version
                elif name and not re.match(r"^https?://", name) and not re.match(r"^\[\d{3}", name):
                    technologies.append({"name": name, "version": version})

    seen: set[str] = set()
    deduped: list[dict] = []
    for t in technologies:
        key = t["name"].lower()
        if key not in seen and len(key) > 1:
            seen.add(key)
            deduped.append(t)

    return {
        "tool": "whatweb",
        "technologies": deduped,
        "title": title,
        "status_code": status_code,
    }


def _extract_snmpwalk(text: str) -> dict:
    """Extract key SNMP data from snmpwalk output."""
    result: dict = {
        "tool": "snmpwalk",
        "sys_descr": "",
        "sys_name": "",
        "sys_contact": "",
        "sys_location": "",
        "interfaces": [],
        "ip_addresses": [],
        "running_processes": [],
        "raw_oid_count": 0,
    }
    result["raw_oid_count"] = sum(1 for line in text.splitlines() if "=" in line)

    oid_values: dict[str, str] = {}
    for m in re.finditer(r"^(\S+)\s+=\s+\S+:\s+(.+)$", text, re.MULTILINE):
        oid_values[m.group(1).lower()] = m.group(2).strip().strip('"')

    for oid_key, val in oid_values.items():
        if "sysdescr" in oid_key or "1.3.6.1.2.1.1.1.0" in oid_key:
            result["sys_descr"] = val
        elif "sysname" in oid_key or "1.3.6.1.2.1.1.5.0" in oid_key:
            result["sys_name"] = val
        elif "syscontact" in oid_key or "1.3.6.1.2.1.1.4.0" in oid_key:
            result["sys_contact"] = val
        elif "syslocation" in oid_key or "1.3.6.1.2.1.1.6.0" in oid_key:
            result["sys_location"] = val
        elif "ifdescr" in oid_key:
            if val and val not in result["interfaces"]:
                result["interfaces"].append(val)
        elif "ipadentaddr" in oid_key or "ipaddressaddr" in oid_key:
            if IPV4_RE.match(val) and val not in result["ip_addresses"]:
                result["ip_addresses"].append(val)
        elif "hrswrunname" in oid_key:
            if val and val not in result["running_processes"]:
                result["running_processes"].append(val)

    return result


def _extract_onesixtyone(text: str) -> dict:
    """Extract SNMP community strings from onesixtyone output."""
    community_strings: list[dict] = []
    for m in re.finditer(r"^(\S+)\s+\[(\S+)\]\s+(.+)$", text, re.MULTILINE):
        community_strings.append({
            "ip": m.group(1),
            "community": m.group(2),
            "sys_descr": m.group(3).strip(),
        })
    return {"tool": "onesixtyone", "community_strings": community_strings}


def _extract_sslscan(text: str) -> dict:
    """Extract TLS/SSL scan results from sslscan output."""
    protocols: dict[str, bool] = {}
    weak_ciphers: list[str] = []
    cert_subject = cert_issuer = ""
    cert_not_after = ""
    cert_expired = False
    key_size = 0
    heartbleed_vulnerable = False

    for m in re.finditer(
        r"(SSLv[23]|TLSv1\.[0-3])\s+(enabled|disabled)", text, re.IGNORECASE,
    ):
        protocols[m.group(1)] = m.group(2).lower() == "enabled"

    for m in re.finditer(
        r"Accepted\s+\S+\s+\d+\s+bits\s+(\S+)", text,
    ):
        cipher = m.group(1)
        if re.search(r"DES|RC4|NULL|EXPORT|MD5", cipher, re.IGNORECASE):
            if cipher not in weak_ciphers:
                weak_ciphers.append(cipher)

    sm = re.search(r"Subject:\s*(.+)", text)
    if sm:
        cert_subject = sm.group(1).strip()
    im = re.search(r"Issuer:\s*(.+)", text)
    if im:
        cert_issuer = im.group(1).strip()
    na = re.search(r"Not valid after:\s*(.+)", text)
    if na:
        cert_not_after = na.group(1).strip()
    if re.search(r"expired", text, re.IGNORECASE):
        cert_expired = True
    ks = re.search(r"RSA Key Strength:\s*(\d+)", text)
    if ks:
        key_size = int(ks.group(1))
    if re.search(r"vulnerable to heartbleed", text, re.IGNORECASE):
        heartbleed_vulnerable = True

    weak_protos = [p for p, enabled in protocols.items()
                   if enabled and p in ("SSLv2", "SSLv3", "TLSv1.0", "TLSv1.1")]

    return {
        "tool": "sslscan",
        "protocols": protocols,
        "weak_protocols": weak_protos,
        "weak_ciphers": weak_ciphers,
        "cert_subject": cert_subject,
        "cert_issuer": cert_issuer,
        "cert_not_after": cert_not_after,
        "cert_expired": cert_expired,
        "key_size": key_size,
        "heartbleed_vulnerable": heartbleed_vulnerable,
    }


def _extract_dnsrecon(text: str) -> dict:
    """Extract DNS records from dnsrecon output."""
    records: list[dict] = []
    zone_transfer = False

    for m in re.finditer(r"\[\*\]\s+(\w+)\s+(\S+)\s+(\S+)(?:\s+(.+))?$", text, re.MULTILINE):
        records.append({
            "type": m.group(1),
            "name": m.group(2),
            "value": m.group(3),
            "extra": (m.group(4) or "").strip(),
        })
    if re.search(r"Zone Transfer.*success|AXFR.*records", text, re.IGNORECASE):
        zone_transfer = True

    return {
        "tool": "dnsrecon",
        "records": records,
        "zone_transfer_successful": zone_transfer,
    }


def _extract_curl_headers(text: str) -> dict:
    """Extract useful info from curl output (HTML or header dump)."""
    title = ""
    headers: dict[str, str] = {}

    tm = re.search(r"<title[^>]*>([^<]+)</title>", text, re.IGNORECASE)
    if tm:
        title = tm.group(1).strip()

    for m in re.finditer(r"^([\w-]+):\s+(.+)$", text, re.MULTILINE):
        hname = m.group(1).lower()
        if hname in ("server", "x-powered-by", "x-aspnet-version",
                      "x-generator", "x-redirect-by", "www-authenticate",
                      "content-type", "set-cookie"):
            headers[hname] = m.group(2).strip()

    interesting = [h for h in headers if h != "content-type"]
    return {
        "tool": "curl",
        "title": title,
        "headers": headers,
        "interesting_headers": interesting,
    }


def _extract_generic(filename: str, text: str) -> dict:
    """Fallback extractor for unrecognized tool output."""
    return {
        "tool": "unknown",
        "filename": filename,
        "line_count": len(text.splitlines()),
        "first_lines": text[:500],
        "non_empty": bool(text.strip()),
    }


AUTORECON_EXTRACTORS: dict[str, any] = {
    r"gobuster|feroxbuster|dirsearch": _extract_dirbusting,
    r"nikto":        _extract_nikto,
    r"enum4linux":   _extract_enum4linux,
    r"smbmap":       _extract_smbmap,
    r"smbclient":    _extract_smbclient,
    r"whatweb":      _extract_whatweb,
    r"snmpwalk":     _extract_snmpwalk,
    r"onesixtyone":  _extract_onesixtyone,
    r"sslscan":      _extract_sslscan,
    r"dnsrecon":     _extract_dnsrecon,
    r"curl":         _extract_curl_headers,
}


def _get_extractor(tool_name: str):
    """Look up the extractor function for a given tool name."""
    for pattern, func in AUTORECON_EXTRACTORS.items():
        if re.search(pattern, tool_name, re.IGNORECASE):
            return func
    return None


# ============================================================
# AutoRecon directory parsing
# ============================================================

def _build_autorecon_target_summary(tool_results: dict) -> dict:
    """Compute a structured summary from parsed tool results (no LLM)."""
    summary: dict = {
        "total_tools_run": 0,
        "tools_with_findings": 0,
        "technologies": [],
        "writable_shares": [],
        "null_session": False,
        "users_found": [],
        "weak_tls": [],
        "community_strings": [],
    }
    seen_techs: set[str] = set()
    seen_users: set[str] = set()

    for port_key, results in tool_results.items():
        for entry in results:
            summary["total_tools_run"] += 1
            data = entry.get("data", {})
            tool = data.get("tool", "unknown")
            has_findings = False

            if tool == "dirbusting" and data.get("paths"):
                has_findings = True
            elif tool == "nikto" and data.get("findings"):
                has_findings = True
            elif tool in ("enum4linux", "smbmap", "smbclient") and data.get("shares"):
                has_findings = True
            elif tool == "whatweb" and data.get("technologies"):
                has_findings = True
                for t in data["technologies"]:
                    key = t["name"]
                    if t.get("version"):
                        key += f"/{t['version']}"
                    if key not in seen_techs:
                        seen_techs.add(key)
                        summary["technologies"].append(key)
            elif tool == "snmpwalk" and data.get("sys_descr"):
                has_findings = True
            elif tool == "onesixtyone" and data.get("community_strings"):
                has_findings = True
                for cs in data["community_strings"]:
                    c = cs.get("community", "")
                    if c and c not in summary["community_strings"]:
                        summary["community_strings"].append(c)
            elif tool == "sslscan" and (data.get("weak_protocols") or data.get("weak_ciphers")):
                has_findings = True
                summary["weak_tls"].extend(data.get("weak_protocols", []))
            elif tool == "dnsrecon" and data.get("records"):
                has_findings = True
            elif tool == "curl" and data.get("interesting_headers"):
                has_findings = True
            elif tool == "unknown" and data.get("non_empty"):
                has_findings = True

            if has_findings:
                summary["tools_with_findings"] += 1

            if tool == "enum4linux":
                if data.get("null_session"):
                    summary["null_session"] = True
                for u in data.get("users", []):
                    if u not in seen_users:
                        seen_users.add(u)
                        summary["users_found"].append(u)
                for s in data.get("shares", []):
                    if s.get("access") in ("read/write",):
                        if s["name"] not in summary["writable_shares"]:
                            summary["writable_shares"].append(s["name"])

            if tool == "smbmap":
                for s_name in data.get("writable_shares", []):
                    if s_name not in summary["writable_shares"]:
                        summary["writable_shares"].append(s_name)

    return summary


def parse_autorecon_results(results_dir: Path) -> dict:
    """Walk an AutoRecon results directory and extract structured data per target."""
    logging.info(f"Parsing AutoRecon results: {results_dir}")

    ar_data: dict = {
        "source_file": str(results_dir),
        "parsed_at": dt.datetime.now().isoformat(timespec="seconds"),
        "report_name": results_dir.name,
        "targets": [],
        "hosts": [],
    }

    for target_dir in sorted(results_dir.iterdir()):
        if not target_dir.is_dir():
            continue
        scans_dir = target_dir / "scans"
        if not scans_dir.exists():
            logging.debug(f"Skipping {target_dir.name}: no scans/ directory")
            continue

        target_name = target_dir.name
        ip = target_name if IPV4_RE.match(target_name) else ""
        hostname = "" if ip else target_name

        # Layer 1: collect nmap XML paths
        xml_dir = scans_dir / "xml"
        nmap_xml_files: list[Path] = []
        if xml_dir.exists():
            nmap_xml_files = sorted(xml_dir.glob("*.xml"))
            # Try to resolve IP/hostname from nmap XML
            if nmap_xml_files and (not ip):
                try:
                    nmap_data = parse_nmap_xml(nmap_xml_files[0])
                    for h in nmap_data.get("hosts", []):
                        for addr in h.get("addresses", []):
                            if addr.get("addrtype") == "ipv4" and addr.get("addr"):
                                ip = addr["addr"]
                                break
                        if not hostname:
                            for hn in h.get("hostnames", []):
                                if hn.get("name"):
                                    hostname = hn["name"]
                                    break
                        if ip:
                            break
                except Exception as exc:
                    logging.debug(f"Failed to pre-parse nmap XML for IP resolution: {exc}")

        # Layer 2: parse tool output files per port directory
        tool_results: dict[str, list[dict]] = {}

        for port_dir in sorted(scans_dir.iterdir()):
            if not port_dir.is_dir():
                continue
            dir_match = AUTORECON_PORT_DIR_RE.match(port_dir.name)
            if not dir_match:
                continue
            protocol = dir_match.group(1)
            port = int(dir_match.group(2))
            port_key = f"{protocol}/{port}"

            port_results: list[dict] = []
            for filepath in sorted(port_dir.iterdir()):
                if not filepath.is_file():
                    continue
                fname = filepath.name
                fn_match = AUTORECON_FILENAME_RE.match(fname)
                tool_name = fn_match.group(4) if fn_match else filepath.stem

                # Skip nmap text files (XML handled in Layer 1)
                if tool_name.lower().endswith("nmap") or tool_name.lower().startswith("nmap"):
                    continue

                try:
                    raw_text = filepath.read_text(encoding="utf-8", errors="replace")
                except Exception as exc:
                    logging.debug(f"Cannot read {filepath}: {exc}")
                    continue
                if not raw_text.strip():
                    logging.debug(f"Empty file: {filepath}")
                    continue

                extractor = _get_extractor(tool_name)
                if extractor:
                    data = extractor(raw_text)
                else:
                    data = _extract_generic(fname, raw_text)

                port_results.append({
                    "tool": data.get("tool", tool_name),
                    "filename": fname,
                    "data": data,
                })

            if port_results:
                tool_results[port_key] = port_results

        # Read _commands.log and _manual_commands.txt
        commands_log = ""
        manual_commands = ""
        cl_path = scans_dir / "_commands.log"
        if cl_path.exists():
            try:
                commands_log = cl_path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                pass
        mc_path = scans_dir / "_manual_commands.txt"
        if mc_path.exists():
            try:
                manual_commands = mc_path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                pass

        summary = _build_autorecon_target_summary(tool_results)

        target_record = {
            "target": target_name,
            "ip": ip,
            "hostname": hostname,
            "nmap_xml_files": nmap_xml_files,
            "commands_log": commands_log,
            "manual_commands": manual_commands,
            "tool_results": tool_results,
            "summary": summary,
        }
        ar_data["targets"].append(target_record)
        ar_data["hosts"].append(target_record)
        logging.info(
            f"AutoRecon target: {target_name} | ports={len(tool_results)} "
            f"tools={summary['total_tools_run']} with_findings={summary['tools_with_findings']}"
        )

    logging.info(f"AutoRecon done: {results_dir.name} | targets={len(ar_data['targets'])}")
    return ar_data


# ============================================================
# Host naming & port rendering
# ============================================================

def choose_host_display_name(host: dict) -> str:
    hostnames = [h["name"].strip() for h in host.get("hostnames", []) if h.get("name")]
    for hn in hostnames:
        if is_probable_fqdn(hn):
            return hn.rstrip(".")
    for hn in hostnames:
        if hn:
            return hn.rstrip(".")
    for addr in host.get("addresses", []):
        if addr.get("addrtype") == "ipv4" and addr.get("addr"):
            return addr["addr"]
    for addr in host.get("addresses", []):
        if addr.get("addr"):
            return addr["addr"]
    return "unknown-host"


def get_primary_ipv4(host: dict) -> str | None:
    for addr in host.get("addresses", []):
        if addr.get("addrtype") == "ipv4" and addr.get("addr"):
            return addr["addr"]
    return None


def render_service_string(port_info: dict) -> str:
    svc = port_info.get("service", {})
    parts = [svc.get("name", "")]
    if svc.get("product"):   parts.append(svc["product"])
    if svc.get("version"):   parts.append(svc["version"])
    if svc.get("extrainfo"): parts.append(f"({svc['extrainfo']})")
    if svc.get("tunnel"):    parts.append(f"[tunnel: {svc['tunnel']}]")
    return " ".join(p for p in parts if p).strip() or "—"


def summarize_open_ports(host: dict) -> list[str]:
    return [
        f"- **{p['protocol']}/{p['port']}** — {render_service_string(p)}"
        for p in host.get("open_ports", [])
    ]


# ============================================================
# Ollama — Nmap
# ============================================================

def build_ollama_prompt(scan_data: dict) -> str:
    hosts = scan_data.get("hosts", [])

    instructions = """
You are an experienced penetration tester analyzing Nmap scan results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.

Priorities:
- Identify notable exposed services and likely attack surface
- Suggest useful follow-up enumeration tools with a ready-to-run example command
  for each, substituting the actual target IP or hostname from the scan data
  (short, focused commands preferred)
- Highlight likely misconfigurations, risky exposures, or common weaknesses
- Infer likely technology stacks when reasonable, clearly tagging [INFERRED]
- Prefer real-world offensive security workflow recommendations over generic advice

Active Directory detection:
- If you observe Kerberos (88), LDAP (389/636/3268), DNS (53), and SMB (445)
  across one or more hosts, explicitly identify this as an Active Directory
  environment in the Environment Assessment section and tailor all attack path
  and enumeration suggestions to AD-specific methodology:
  AS-REP roasting, Kerberoasting, BloodHound, Pass-the-Hash, Pass-the-Ticket,
  DCSync, LDAP anonymous bind enumeration, GPO abuse, etc.

Cross-host correlation:
- Identify patterns across multiple hosts. If a finding applies network-wide
  (e.g., SMB signing disabled everywhere, same service version on all hosts),
  report it once as a network-level finding rather than repeating it per host.

When suggesting tools, favor practical enumeration tools such as:
- SMB: netexec, smbclient, smbmap, enum4linux-ng
- LDAP: ldapsearch, netexec, bloodyAD
- Kerberos: kerbrute, impacket tools (GetUserSPNs, GetNPUsers), netexec
- MSSQL: impacket-mssqlclient, netexec
- WinRM: evil-winrm, netexec
- RDP: netexec, xfreerdp
- SSH: ssh-audit, hydra
- Web: whatweb, nikto, gobuster, feroxbuster, ffuf, curl
- SNMP: snmpwalk, onesixtyone
- DNS: dig, nslookup, dnsrecon
- NFS: showmount, mount
- SMTP: swaks, smtp-user-enum
- FTP: ftp, Nmap NSE scripts
- RPC: rpcclient, netexec
- VNC: vncsnapshot, hydra
""".strip()

    lines = [
        f"Nmap args: {scan_data.get('nmap_args', '')}",
        f"Nmap version: {scan_data.get('nmap_version', '')}",
        f"Hosts count: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        name  = choose_host_display_name(host)
        state = host.get("state", "unknown")
        ip    = get_primary_ipv4(host) or "unknown"
        lines += [f"- Host: {name}", f"  State: {state}", f"  IP: {ip}"]
        open_ports = host.get("open_ports", [])
        if not open_ports:
            lines += ["  Open ports: none", ""]
            continue
        lines.append("  Open ports:")
        for p in open_ports:
            svc      = p.get("service", {})
            svc_name = svc.get("name", "") or "unknown"
            product  = svc.get("product", "")
            version  = svc.get("version", "")
            extrainfo= svc.get("extrainfo", "")
            tunnel   = svc.get("tunnel", "")
            parts = [svc_name]
            if product:   parts.append(product)
            if version:   parts.append(version)
            if extrainfo: parts.append(f"({extrainfo})")
            if tunnel:    parts.append(f"[tunnel: {tunnel}]")
            lines.append(f"    - {p['protocol']}/{p['port']}: {' '.join(parts).strip()}")
            for script in p.get("scripts", [])[:3]:
                out = " ".join((script.get("output", "") or "").split())
                if out:
                    if len(out) > 220:
                        out = out[:217] + "..."
                    lines.append(f"      script:{script.get('id', '')} -> {out}")
            for hint in get_port_hints(p["port"], svc_name, product, extrainfo):
                lines.append(f"      hint: {hint}")
        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections in this order:

## Environment Assessment
- State the inferred environment type (Active Directory domain, standalone Linux
  hosts, mixed Windows/Linux, web application stack, OT/ICS, etc.)
- Provide a confidence level: HIGH / MEDIUM / LOW
- List the specific evidence from the scan data that supports the assessment
- If AD is detected, state the inferred domain name if visible in hostnames or
  DNS responses

## Key Observations
- Brief bullets of the most important findings
- Each bullet must include a [CONFIRMED], [INFERRED], or [ASSUMED] tag

## Enumeration Suggestions
- Group by service or host
- For each suggestion, include a ready-to-run example command with the actual
  target IP or hostname substituted in
- Tag each with [CONFIRMED], [INFERRED], or [ASSUMED]

## Potential Attack Paths
- Realistic next steps or attack paths with supporting scan evidence for each step
- Tag each path as [CONFIRMED], [INFERRED], or [ASSUMED]
- Do not include a path if the scan data does not support it

## Notable Risks or Misconfigurations
- Network-wide findings first, then per-host
- Tag each with [CONFIRMED], [INFERRED], or [ASSUMED]
""".strip()

    prompt = f"{instructions}\n\nNmap Scan Summary\n=================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Nmap Ollama prompt ({len(prompt)} chars)")
    return prompt


# ============================================================
# Ollama — Nessus
# ============================================================

def build_nessus_ollama_prompt(nessus_data: dict) -> str:
    hosts = nessus_data.get("hosts", [])

    instructions = """
You are an experienced penetration tester analyzing Nessus vulnerability scan results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.
- Nessus plugin findings can produce false positives, particularly for
  version-based checks. If a finding relies solely on version detection without
  plugin output confirming exploitation, note it as
  [VERIFY — possible false positive].

Priorities:
- Identify the highest-impact vulnerabilities by CVSS score and exploitability
- For each CVE, tag its weaponization status:
    [MSF]         — known Metasploit module exists
    [POC]         — public PoC exists but may require adaptation
    [THEORETICAL] — no known public exploit at time of analysis
- Flag any findings suggesting credential exposure, default credentials,
  cleartext protocols, or password policy weaknesses — these often have
  outsized impact even when CVSS scores are moderate
- Version-based detection findings (where plugin_output is absent or says only
  "version X detected") must be flagged [VERIFY — version-based only]
- Cross-host correlation: if the same critical/high finding appears on multiple
  hosts, group them and call out the network-wide scope rather than listing each
  host separately
- Group findings by exploitation priority: critical/weaponized first, then
  high, then medium
""".strip()

    lines = [
        f"Report: {nessus_data.get('report_name', '')}",
        f"Hosts: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        ip       = host.get("ip", "unknown")
        hostname = host.get("hostname", "")
        findings = host.get("findings", [])

        label = f"{hostname} ({ip})" if hostname and hostname != ip else ip
        lines.append(f"- Host: {label}")

        # Skip info (severity 0) in AI prompt to reduce noise
        notable = [f for f in findings if f["severity_int"] >= 1]
        if not notable:
            lines.append("  Findings: none above informational")
            lines.append("")
            continue

        # Count by severity
        counts = {4: 0, 3: 0, 2: 0, 1: 0}
        for f in notable:
            counts[f["severity_int"]] = counts.get(f["severity_int"], 0) + 1
        lines.append(
            f"  Critical:{counts[4]} High:{counts[3]} Medium:{counts[2]} Low:{counts[1]}"
        )
        lines.append("  Top findings:")

        for f in notable[:15]:
            sev      = severity_int_to_str(f["severity_int"])
            cvss     = f.get("cvss3_base") or f.get("cvss_base") or "N/A"
            cves     = ", ".join(f["cves"]) if f["cves"] else "no CVE"
            has_out  = "plugin_output:yes" if f.get("plugin_output") else "plugin_output:none"
            lines.append(
                f"    [{sev}] {f['plugin_name']} | CVSS:{cvss} | {cves} | port {f['protocol']}/{f['port']} | {has_out}"
            )
            if f.get("solution"):
                sol = f["solution"][:200].replace("\n", " ")
                lines.append(f"      solution: {sol}")

        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections:

## Key Observations
- Most impactful vulnerabilities and what they mean for the engagement
- Each bullet must include a [CONFIRMED], [INFERRED], or [ASSUMED] tag
- Group any finding that appears on multiple hosts as a single network-wide bullet

## Exploitation Priority
- Rank the top exploitable findings with reasoning
- For each CVE include its [MSF], [POC], or [THEORETICAL] tag
- Mark version-based-only detections with [VERIFY — version-based only]
- Flag credential/cleartext findings prominently

## Attack Chains
- How findings could be combined for privilege escalation or lateral movement
- Each step must cite specific scan evidence; tag with [CONFIRMED]/[INFERRED]/[ASSUMED]

## Remediation Focus
- Highest-priority patches or configuration fixes
- Network-wide issues first, then per-host
""".strip()

    prompt = f"{instructions}\n\nNessus Scan Summary\n===================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Nessus Ollama prompt ({len(prompt)} chars)")
    return prompt


# ============================================================
# Ollama — Nessus two-pass helpers
# ============================================================

def _build_nessus_fact_extraction_prompt(nessus_data: dict) -> str:
    """
    Pass 1 prompt: instruct the model to extract ONLY facts explicitly present
    in the Nessus data — no inference, no recommendations.
    """
    hosts = nessus_data.get("hosts", [])
    lines = [
        "You are a data extraction assistant. From the following Nessus scan data, "
        "extract ONLY facts that are explicitly present. Do not infer, recommend, or analyze. "
        "Output a structured list: for each host, list confirmed CVEs with their CVSS scores, "
        "confirmed severity levels, and confirmed affected ports/services. "
        "Mark everything as [CONFIRMED]. Do not add any information not present in the data.",
        "",
        "Nessus data:",
    ]
    for host in hosts:
        ip       = host.get("ip", "unknown")
        hostname = host.get("hostname", "")
        label    = f"{hostname} ({ip})" if hostname and hostname != ip else ip
        lines.append(f"\nHost: {label}")
        notable = [f for f in host.get("findings", []) if f["severity_int"] >= 1]
        if not notable:
            lines.append("  No notable findings.")
            continue
        for f in notable[:20]:
            sev  = severity_int_to_str(f["severity_int"])
            cvss = f.get("cvss3_base") or f.get("cvss_base") or "N/A"
            cves = ", ".join(f["cves"]) if f["cves"] else "none"
            lines.append(
                f"  [{sev}] {f['plugin_name']} | CVSS:{cvss} | CVE:{cves} "
                f"| port:{f['protocol']}/{f['port']}"
            )
    return "\n".join(lines)


# ============================================================
# Ollama — Burp Suite
# ============================================================

def build_burp_ollama_prompt(burp_data: dict) -> str:
    hosts = burp_data.get("hosts", [])

    instructions = """
You are an experienced web application penetration tester analyzing Burp Suite scanner results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.
- Burp "Tentative" findings have a high false positive rate. Always flag these
  as [TENTATIVE — VERIFY MANUALLY] and do not build attack chains primarily on
  tentative findings.

Priorities:
- Identify the most exploitable web vulnerabilities (SQLi, XSS, SSRF, IDOR,
  auth bypass, etc.) and classify each by its OWASP Top 10 2021 category
  (e.g., A01:Broken Access Control, A03:Injection, A07:Auth Failures, etc.)
- For each Certain or Firm finding, suggest a specific manual verification step
  or payload to confirm exploitability beyond the scanner result
- Suggest how findings could be chained for greater impact
- Assess the web attack surface: authentication, input handling, business logic
- Highlight systemic issues (framework versions, CSP absence, insecure headers)
- Be explicit about what the scanner confirmed vs. what requires manual testing
""".strip()

    lines = [
        f"Hosts scanned: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        ip     = host.get("ip", "")
        url    = host.get("url", "")
        issues = host.get("issues", [])
        label  = url if url else ip
        if ip and url and ip not in url:
            label = f"{url} ({ip})"

        lines.append(f"- Host: {label}")

        notable = [i for i in issues if i["severity"].lower() not in ("information", "info")]
        if not notable:
            lines.append("  Issues: informational only")
            lines.append("")
            continue

        counts: dict[str, int] = {}
        for i in issues:
            counts[i["severity"]] = counts.get(i["severity"], 0) + 1
        lines.append(
            "  " + " ".join(f"{k}:{v}" for k, v in sorted(counts.items()))
        )
        lines.append("  Issues:")
        for i in notable[:15]:
            conf = i.get("confidence", "")
            loc  = i.get("path") or i.get("location") or ""
            lines.append(f"    [{i['severity']}|{conf}] {i['name']} @ {loc}")
            if i.get("issue_detail"):
                detail = i["issue_detail"][:200].replace("\n", " ")
                lines.append(f"      detail: {detail}")
        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections:

## Key Observations
- Most significant web vulnerabilities and their potential impact
- Each bullet must include its OWASP Top 10 2021 category and a
  [CONFIRMED], [INFERRED], or [ASSUMED] tag
- Tentative findings must be tagged [TENTATIVE — VERIFY MANUALLY]

## Exploitation Priority
- Which issues to pursue first and why
- For each Certain/Firm finding, include a specific manual verification step
  or payload (e.g., exact parameter to test, HTTP method, expected response)
- Do not elevate Tentative findings above Certain/Firm without explanation

## Attack Chains
- How findings could be combined for greater impact
- Each step must cite specific scanner evidence; tag [CONFIRMED]/[INFERRED]/[ASSUMED]
- Do not build chains primarily on Tentative findings

## False Positive Risk Assessment
- List findings that are likely false positives, with reasoning
- Consider: finding type, confidence level (Tentative = higher FP risk), context,
  and whether the issue_detail provides concrete evidence or just a generic description

## Remediation Focus
- Top development/configuration fixes to eliminate the highest-impact issues
""".strip()

    prompt = f"{instructions}\n\nBurp Suite Scan Summary\n=======================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Burp Ollama prompt ({len(prompt)} chars)")
    return prompt


# ============================================================
# Ollama — AutoRecon
# ============================================================

def _build_autorecon_fact_extraction_prompt(target_data: dict) -> str:
    """Pass 1: present Python-extracted structured data; ask LLM to organize only."""
    target = target_data.get("target", "unknown")
    ip = target_data.get("ip", target)
    hostname = target_data.get("hostname", "")
    label = f"{ip} ({hostname})" if hostname else ip

    lines: list[str] = [f"Target: {label}", ""]

    for port_key in sorted(target_data.get("tool_results", {})):
        results = target_data["tool_results"][port_key]
        lines.append(f"Port {port_key}:")
        for entry in results:
            data = entry.get("data", {})
            tool = data.get("tool", entry.get("tool", "unknown"))

            if tool == "dirbusting":
                interesting = data.get("interesting", [])[:20]
                if interesting:
                    lines.append(f"  Directory enumeration ({data.get('total_found', 0)} total):")
                    for p in interesting:
                        redir = f" -> {p['redirect']}" if p.get("redirect") else ""
                        lines.append(f"    {p['status']}  {p['path']}  ({p['size']}B){redir}")
            elif tool == "nikto":
                if data.get("server"):
                    lines.append(f"  Nikto server: {data['server']}")
                for h_name, h_val in data.get("headers", {}).items():
                    lines.append(f"  Header: {h_name}: {h_val}")
                for f in data.get("findings", [])[:15]:
                    osvdb = f"OSVDB-{f['osvdb']}: " if f.get("osvdb") else ""
                    lines.append(f"  Nikto: {osvdb}{f['path']} — {f['description'][:200]}")
            elif tool == "enum4linux":
                if data.get("os_info"):
                    lines.append(f"  OS: {data['os_info']}")
                if data.get("domain"):
                    lines.append(f"  Domain: {data['domain']}")
                lines.append(f"  Null session: {'YES' if data.get('null_session') else 'No'}")
                if data.get("users"):
                    lines.append(f"  Users ({len(data['users'])}): {', '.join(data['users'][:20])}")
                if data.get("shares"):
                    for s in data["shares"]:
                        lines.append(f"  Share: {s['name']} — {s.get('access', s.get('type', ''))}")
                if data.get("groups"):
                    lines.append(f"  Groups: {', '.join(data['groups'][:15])}")
                pp = data.get("password_policy", {})
                if pp:
                    parts = []
                    if "min_length" in pp:
                        parts.append(f"min_length={pp['min_length']}")
                    if "lockout_threshold" in pp:
                        parts.append(f"lockout={pp['lockout_threshold']}")
                    if pp.get("complexity"):
                        parts.append("complexity=on")
                    if parts:
                        lines.append(f"  Password policy: {', '.join(parts)}")
            elif tool == "smbmap":
                for s in data.get("shares", []):
                    lines.append(f"  Share: {s['name']} — {s['permissions']}")
            elif tool == "smbclient":
                for s in data.get("shares", []):
                    lines.append(f"  Share: {s['name']} ({s['type']})")
            elif tool == "whatweb":
                techs = data.get("technologies", [])
                if techs:
                    parts = [f"{t['name']}/{t['version']}" if t.get("version") else t["name"]
                             for t in techs]
                    lines.append(f"  Technologies: {', '.join(parts)}")
                if data.get("title"):
                    lines.append(f"  Title: {data['title']}")
            elif tool == "snmpwalk":
                if data.get("sys_descr"):
                    lines.append(f"  SNMP sysDescr: {data['sys_descr']}")
                if data.get("sys_name"):
                    lines.append(f"  SNMP sysName: {data['sys_name']}")
                if data.get("running_processes"):
                    lines.append(f"  Processes: {', '.join(data['running_processes'][:15])}")
                lines.append(f"  Total OIDs: {data.get('raw_oid_count', 0)}")
            elif tool == "onesixtyone":
                for cs in data.get("community_strings", []):
                    lines.append(f"  SNMP community: {cs['community']} — {cs['sys_descr'][:100]}")
            elif tool == "sslscan":
                if data.get("weak_protocols"):
                    lines.append(f"  Weak TLS: {', '.join(data['weak_protocols'])}")
                if data.get("weak_ciphers"):
                    lines.append(f"  Weak ciphers: {', '.join(data['weak_ciphers'][:5])}")
                if data.get("cert_subject"):
                    lines.append(f"  Cert: {data['cert_subject']}")
                if data.get("cert_expired"):
                    lines.append(f"  Cert expired: {data.get('cert_not_after', 'yes')}")
                if data.get("heartbleed_vulnerable"):
                    lines.append("  HEARTBLEED: VULNERABLE")
            elif tool == "dnsrecon":
                for r in data.get("records", [])[:20]:
                    lines.append(f"  DNS {r['type']}: {r['name']} -> {r['value']}")
                if data.get("zone_transfer_successful"):
                    lines.append("  ZONE TRANSFER: SUCCESSFUL")
            elif tool == "curl":
                if data.get("title"):
                    lines.append(f"  Page title: {data['title']}")
                for h_name, h_val in data.get("headers", {}).items():
                    lines.append(f"  Header: {h_name}: {h_val}")
        lines.append("")

    instructions = """You are a data extraction assistant. Your ONLY job is to organize
and deduplicate the following AutoRecon enumeration data into a clean, structured summary.

RULES:
- Do NOT analyze, infer, recommend, or draw conclusions.
- Do NOT add any information not present in the data below.
- Organize by port/service. Combine duplicate findings across tools.
- Tag every item as [CONFIRMED] — it comes directly from tool output.
- Output a structured list, not prose."""

    return f"{instructions}\n\n{chr(10).join(lines)}"


def build_autorecon_ollama_prompt(target_data: dict) -> str:
    """Pass 2: analyze structured AutoRecon findings with grounding rules."""
    target = target_data.get("target", "unknown")
    ip = target_data.get("ip", target)
    hostname = target_data.get("hostname", "")
    label = f"{ip} ({hostname})" if hostname else ip
    summary = target_data.get("summary", {})

    lines: list[str] = [
        f"Target: {label}",
        f"Tools run: {summary.get('total_tools_run', 0)}  |  "
        f"With findings: {summary.get('tools_with_findings', 0)}",
    ]
    if summary.get("technologies"):
        lines.append(f"Technologies: {', '.join(summary['technologies'][:15])}")
    if summary.get("null_session"):
        lines.append("Null session: YES")
    if summary.get("writable_shares"):
        lines.append(f"Writable shares: {', '.join(summary['writable_shares'])}")
    if summary.get("users_found"):
        lines.append(f"Users found ({len(summary['users_found'])}): "
                      f"{', '.join(summary['users_found'][:20])}")
    if summary.get("weak_tls"):
        lines.append(f"Weak TLS: {', '.join(summary['weak_tls'])}")
    if summary.get("community_strings"):
        lines.append(f"SNMP communities: {', '.join(summary['community_strings'])}")
    lines.append("")

    for port_key in sorted(target_data.get("tool_results", {})):
        results = target_data["tool_results"][port_key]
        lines.append(f"--- Port {port_key} ---")
        for entry in results:
            data = entry.get("data", {})
            tool = data.get("tool", entry.get("tool", "unknown"))

            if tool == "dirbusting":
                interesting = data.get("interesting", [])[:20]
                if interesting:
                    lines.append(f"[dirbusting] {data.get('total_found', 0)} paths found. Notable:")
                    for p in interesting[:15]:
                        lines.append(f"  {p['status']}  {p['path']}  ({p['size']}B)")
            elif tool == "nikto":
                lines.append(f"[nikto] {data.get('total_findings', 0)} findings "
                              f"(server: {data.get('server', 'unknown')})")
                for f in data.get("findings", [])[:10]:
                    osvdb = f"OSVDB-{f['osvdb']} " if f.get("osvdb") else ""
                    lines.append(f"  {osvdb}{f['path']}: {f['description'][:150]}")
            elif tool == "enum4linux":
                lines.append(f"[enum4linux] OS={data.get('os_info', '?')} "
                              f"Domain={data.get('domain', '?')} "
                              f"NullSession={'YES' if data.get('null_session') else 'no'}")
                if data.get("users"):
                    lines.append(f"  Users: {', '.join(data['users'][:20])}")
                if data.get("shares"):
                    for s in data["shares"]:
                        lines.append(f"  Share: {s['name']} [{s.get('access', s.get('type', ''))}]")
                pp = data.get("password_policy", {})
                if pp:
                    lines.append(f"  PwPolicy: {pp}")
            elif tool == "smbmap":
                for s in data.get("shares", []):
                    lines.append(f"[smbmap] {s['name']}: {s['permissions']}")
            elif tool == "smbclient":
                for s in data.get("shares", []):
                    lines.append(f"[smbclient] {s['name']} ({s['type']})")
            elif tool == "whatweb":
                techs = data.get("technologies", [])
                if techs:
                    parts = [f"{t['name']}/{t['version']}" if t.get("version") else t["name"]
                             for t in techs[:10]]
                    lines.append(f"[whatweb] {', '.join(parts)}")
            elif tool == "snmpwalk":
                if data.get("sys_descr"):
                    lines.append(f"[snmpwalk] sysDescr: {data['sys_descr']}")
                if data.get("running_processes"):
                    lines.append(f"  Processes: {', '.join(data['running_processes'][:10])}")
            elif tool == "onesixtyone":
                for cs in data.get("community_strings", []):
                    lines.append(f"[onesixtyone] community={cs['community']}")
            elif tool == "sslscan":
                parts = []
                if data.get("weak_protocols"):
                    parts.append(f"weak_proto={','.join(data['weak_protocols'])}")
                if data.get("cert_expired"):
                    parts.append("cert_expired")
                if data.get("heartbleed_vulnerable"):
                    parts.append("HEARTBLEED")
                if data.get("cert_subject"):
                    parts.append(f"cert={data['cert_subject']}")
                if parts:
                    lines.append(f"[sslscan] {' | '.join(parts)}")
            elif tool == "dnsrecon":
                lines.append(f"[dnsrecon] {len(data.get('records', []))} records")
                if data.get("zone_transfer_successful"):
                    lines.append("  ZONE TRANSFER SUCCESSFUL")
                for r in data.get("records", [])[:10]:
                    lines.append(f"  {r['type']} {r['name']} -> {r['value']}")
            elif tool == "curl":
                if data.get("title"):
                    lines.append(f"[curl] title={data['title']}")
                for h_name in data.get("interesting_headers", []):
                    lines.append(f"  {h_name}: {data['headers'].get(h_name, '')}")
        lines.append("")

    instructions = f"""You are an experienced penetration tester analyzing AutoRecon
enumeration results for target {label}.

GROUNDING RULES — follow these strictly:
- All data below was pre-extracted from tool outputs by an automated parser.
  Every finding is [CONFIRMED] unless stated otherwise.
- Only reference paths, shares, users, technologies, and findings explicitly
  present in the data. Do not invent CVEs, paths, or services.
- If evidence is insufficient, say "insufficient data" — do not guess.
- Tag each claim: [CONFIRMED] = in the data, [INFERRED] = reasonable
  conclusion from confirmed data, [ASSUMED] = requires verification.
- Nikto has a high false-positive rate. Tag unverified nikto-only findings
  as [NIKTO — VERIFY MANUALLY].
- Provide ready-to-run commands with actual target IP {ip} substituted.
- Do not suggest an attack path unless enumeration data supports each step.
- Cross-correlate findings across tools (e.g., whatweb tech + nikto vuln +
  gobuster path = confirmed attack surface).
- Prioritize: writable shares, null sessions, admin panels, weak TLS,
  exposed credentials, directory listings, version-specific vulnerabilities."""

    output_format = """
Respond using EXACTLY these section headers:

## Key Observations
- Most actionable findings; cross-tool correlation where possible
- Each bullet: [CONFIRMED]/[INFERRED]/[ASSUMED] tag

## Service-Specific Findings
- Grouped by port/service
- What was confirmed vs needs manual verification
- Discovered paths with security implications (admin panels, backups, APIs)

## Credential & Access Findings
- Null sessions, user lists, password policies, writable shares
- Default credentials to test based on confirmed technologies

## Attack Paths
- How enumeration findings chain together for exploitation
- Each step must cite specific tool evidence
- Tag [CONFIRMED]/[INFERRED]/[ASSUMED]

## Manual Follow-Up
- Specific commands to run next, with actual IPs/hostnames
- Prioritized by likely impact
""".strip()

    prompt = (f"{instructions}\n\nAutoRecon Enumeration Summary\n"
              f"{'=' * 30}\n{chr(10).join(lines)}\n\n{output_format}")
    logging.debug(f"Built AutoRecon Ollama prompt ({len(prompt)} chars)")
    return prompt


def ollama_chat(base_url: str, model: str, prompt: str, temperature: float = 0.15) -> str:
    url = base_url.rstrip("/") + "/api/generate"
    payload = {"model": model, "prompt": prompt, "stream": False, "temperature": temperature}
    logging.info(f"Querying Ollama model: {model} (temperature={temperature})")
    r = requests.post(url, json=payload, timeout=180)
    r.raise_for_status()
    response_text = r.json().get("response", "")
    logging.debug(f"Ollama response: {len(response_text)} chars")
    return response_text


# ============================================================
# AI output validation
# ============================================================

def validate_ai_output(
    raw_text: str, source_data: dict, source_type: str
) -> tuple[str, list[str]]:
    """
    Post-process AI output to flag potential hallucinations.

    Returns (annotated_text, warnings_list).
    source_type: "nmap", "nessus", "burp", or "autorecon"
    """
    warnings: list[str] = []
    hosts = source_data.get("hosts", [])

    # --- Build source truth sets ---
    source_ips:       set[str] = set()
    source_ports:     set[int] = set()
    source_cves:      set[str] = set()
    source_hostnames: set[str] = set()

    if source_type == "nmap":
        for h in hosts:
            for addr in h.get("addresses", []):
                if addr.get("addrtype") == "ipv4" and addr.get("addr"):
                    source_ips.add(addr["addr"])
            for p in h.get("open_ports", []):
                source_ports.add(p["port"])
            for hn in h.get("hostnames", []):
                if hn.get("name"):
                    source_hostnames.add(hn["name"].lower().rstrip("."))

    elif source_type == "nessus":
        for h in hosts:
            if h.get("ip"):
                source_ips.add(h["ip"])
            if h.get("hostname"):
                source_hostnames.add(h["hostname"].lower().rstrip("."))
            for f in h.get("findings", []):
                if f.get("port"):
                    source_ports.add(int(f["port"]))
                for cve in f.get("cves", []):
                    source_cves.add(cve.strip().upper())

    elif source_type == "burp":
        for h in hosts:
            if h.get("ip"):
                source_ips.add(h["ip"])
            if h.get("url"):
                from urllib.parse import urlparse
                parsed = urlparse(h["url"])
                if parsed.netloc:
                    source_hostnames.add(parsed.netloc.split(":")[0].lower())

    elif source_type == "autorecon":
        for target in hosts:
            if target.get("ip"):
                source_ips.add(target["ip"])
            if target.get("hostname"):
                source_hostnames.add(target["hostname"].lower().rstrip("."))
            for nmap_scan in target.get("nmap_scans", []):
                for h in nmap_scan.get("hosts", []):
                    for hn in h.get("hostnames", []):
                        if hn.get("name"):
                            source_hostnames.add(hn["name"].lower().rstrip("."))
                    for p in h.get("open_ports", []):
                        source_ports.add(p["port"])
            for port_key in target.get("tool_results", {}):
                try:
                    port_num = int(port_key.split("/")[1])
                    source_ports.add(port_num)
                except (IndexError, ValueError):
                    pass

    annotated = raw_text

    # --- CVE validation (nessus only) ---
    if source_type == "nessus":
        found_cves = {m.upper() for m in CVE_RE.findall(annotated)}
        for cve in sorted(found_cves):
            if cve not in source_cves:
                inline_warn = (
                    f"\n> ⚠️ HALLUCINATION WARNING: {cve} was not found in the "
                    "scan data — verify before acting."
                )
                # Insert after the first occurrence of this CVE
                pos = 0
                while pos < len(annotated):
                    m = CVE_RE.search(annotated, pos)
                    if not m:
                        break
                    if m.group(0).upper() == cve:
                        end = m.end()
                        annotated = annotated[:end] + inline_warn + annotated[end:]
                        break
                    pos = m.end()
                warnings.append(
                    f"{cve} mentioned in analysis but not present in Nessus input"
                )

    # --- IP cross-reference (all types, only when we have source IPs to compare) ---
    if source_ips:
        found_ips = set(IP_IN_TEXT_RE.findall(annotated))
        for ip in sorted(found_ips):
            if ip not in source_ips:
                warnings.append(
                    f"IP {ip} appears in analysis but was not in input scan"
                )

    # --- Port cross-reference (all types, only when we have source ports) ---
    if source_ports:
        found_ports: set[int] = set()
        for m in PORT_IN_TEXT_RE.finditer(annotated):
            port_str = m.group(1) or m.group(2)
            try:
                found_ports.add(int(port_str))
            except (ValueError, TypeError):
                pass
        for port in sorted(found_ports):
            if port not in source_ports:
                warnings.append(
                    f"Port {port} mentioned in analysis but was not open on any host in the scan"
                )

    # --- Hostname cross-reference (all types, only when we have source hostnames) ---
    if source_hostnames:
        found_fqdns: set[str] = set()
        for m in FQDN_IN_TEXT_RE.finditer(annotated):
            val = m.group(0).lower().rstrip(".")
            if not is_ipv4(val):
                found_fqdns.add(val)
        for fqdn in sorted(found_fqdns):
            if not any(
                fqdn == sh
                or fqdn.endswith("." + sh)
                or sh.endswith("." + fqdn)
                for sh in source_hostnames
            ):
                warnings.append(
                    f"Hostname {fqdn} appears in analysis but was not in input scan"
                )

    # --- Prepend summary block if warnings exist ---
    if warnings:
        n = len(warnings)
        block_lines = [
            f"> ⚠️ **Validation Warnings** "
            f"({n} issue{'s' if n != 1 else ''} found — review before acting)"
        ]
        for w in warnings:
            block_lines.append(f"> - {w}")
        annotated = "\n".join(block_lines) + "\n\n" + annotated

    return annotated, warnings


# ============================================================
# Host note section rendering
# ============================================================

def _render_nessus_section(findings: list[dict]) -> str:
    """Render Nessus findings list as markdown body (no section header)."""
    notable = [f for f in findings if f["severity_int"] >= 1]
    info    = [f for f in findings if f["severity_int"] == 0]

    if not findings:
        return "_No Nessus findings._"

    lines: list[str] = []

    if notable:
        for f in notable:
            sev_str  = severity_int_to_str(f["severity_int"])
            cvss_val = f.get("cvss3_base") or f.get("cvss_base") or "N/A"
            cves_str = ", ".join(f["cves"]) if f["cves"] else "N/A"
            port_str = f"{f['protocol']}/{f['port']}" if f["port"] else "N/A"

            lines.append(f"#### [{sev_str}] {f['plugin_name']} (Plugin {f['plugin_id']})")
            lines.append(
                f"**Port:** {port_str}  ·  **CVSS v3:** {cvss_val}  ·  **CVE(s):** {cves_str}"
            )

            if f.get("description"):
                desc = f["description"].strip()
                if len(desc) > 500:
                    desc = desc[:497] + "..."
                lines.append(f"\n{desc}")

            if f.get("solution"):
                sol = f["solution"].strip()
                if len(sol) > 300:
                    sol = sol[:297] + "..."
                lines.append(f"\n**Solution:** {sol}")

            if f.get("plugin_output"):
                out = f["plugin_output"].strip()
                if len(out) > 300:
                    out = out[:297] + "..."
                lines.append(f"\n**Plugin Output:**\n```\n{out}\n```")

            lines.append("")

    if info:
        lines.append(f"_Plus {len(info)} informational finding(s) — omitted for brevity._")

    return "\n".join(lines).strip()


def _render_burp_section(issues: list[dict]) -> str:
    """Render Burp Suite issues list as markdown body (no section header)."""
    if not issues:
        return "_No Burp Suite findings._"

    sev_order = {"high": 0, "medium": 1, "low": 2, "information": 3}
    sorted_issues = sorted(
        issues, key=lambda i: (sev_order.get(i["severity"].lower(), 4), i["name"])
    )

    lines: list[str] = []
    for issue in sorted_issues:
        sev  = issue["severity"]
        conf = issue.get("confidence", "")
        conf_str = f" ({conf})" if conf else ""

        lines.append(f"#### [{sev}{conf_str}] {issue['name']}")

        if issue.get("path"):
            lines.append(f"**Path:** `{issue['path']}`")
        if issue.get("location"):
            lines.append(f"**Location:** {issue['location']}")

        if issue.get("issue_detail"):
            detail = issue["issue_detail"].strip()
            if len(detail) > 500:
                detail = detail[:497] + "..."
            lines.append(f"\n{detail}")

        if issue.get("remediation_detail"):
            rem = issue["remediation_detail"].strip()
            if len(rem) > 300:
                rem = rem[:297] + "..."
            lines.append(f"\n**Remediation:** {rem}")
        elif issue.get("remediation_background"):
            rem = issue["remediation_background"].strip()
            if len(rem) > 300:
                rem = rem[:297] + "..."
            lines.append(f"\n**Remediation:** {rem}")

        lines.append("")

    return "\n".join(lines).strip()


# ============================================================
# Vault writing — Nmap
# ============================================================

def _find_host_note_by_ip(hosts_dir: Path, ip: str) -> Path | None:
    """Find an existing host note whose frontmatter ip matches."""
    if not hosts_dir.exists() or not ip:
        return None
    for hp in hosts_dir.glob("*.md"):
        try:
            fm, _ = read_frontmatter(hp.read_text(encoding="utf-8"))
            if fm.get("ip") == ip:
                return hp
        except Exception:
            pass
    return None


def _write_host_note(
    hosts_dir: Path,
    host: dict,
    scan_stem: str,
    scan_display: str,
    tool_name: str,
) -> tuple[str, str]:
    """
    Write (or merge) a host note for an Nmap host.
    Preserves existing Nessus Findings, Burp Suite Findings, and AutoRecon Enumeration sections.
    Returns (display_name, host_stem).
    """
    display    = choose_host_display_name(host)
    host_stem  = safe_filename(display)
    host_path  = hosts_dir / ensure_md_suffix(host_stem)
    primary_ip = get_primary_ipv4(host)
    open_ports = host.get("open_ports", [])

    existing_fm: dict  = {}
    existing_op_notes  = ""
    existing_nessus    = ""
    existing_burp      = ""
    existing_autorecon = ""

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes  = extract_operator_notes(old_body)
        existing_nessus    = extract_body_section(old_body, "## Nessus Findings")
        existing_burp      = extract_body_section(old_body, "## Burp Suite Findings")
        existing_autorecon = extract_body_section(old_body, "## AutoRecon Enumeration")
        logging.debug(f"Merging Nmap host note: {host_path.name}")
    else:
        logging.debug(f"Creating Nmap host note: {host_path.name}")

    # Merge frontmatter
    existing_hostnames: list = existing_fm.get("hostnames", [])
    new_hostnames = [h["name"].strip() for h in host.get("hostnames", []) if h.get("name")]
    merged_hostnames = existing_hostnames + [h for h in new_hostnames if h not in existing_hostnames]

    scan_source = f"{scan_display} - {tool_name}"
    existing_sources: list = existing_fm.get("sources", [])
    merged_sources = existing_sources if scan_source in existing_sources else existing_sources + [scan_source]

    new_tags = get_tags_from_ports(open_ports)
    merged_tags = sorted(set(existing_fm.get("tags", [])) | set(new_tags))

    fm: dict = {
        "ip":        primary_ip or existing_fm.get("ip"),
        "hostnames": merged_hostnames,
        "status":    existing_fm.get("status", "not-started"),
        "tags":      merged_tags,
        "sources":   merged_sources,
    }
    if "nessus_max_severity" in existing_fm:
        fm["nessus_max_severity"] = existing_fm["nessus_max_severity"]
    if "autorecon_tools_run" in existing_fm:
        fm["autorecon_tools_run"] = existing_fm["autorecon_tools_run"]

    # Build body
    lines: list[str] = [f"**State:** {host['state']}"]
    if primary_ip:
        lines.append(f"**IP:** {primary_ip}")
    lines.append(f"**Open Ports:** {len(open_ports)}")

    lines += ["", "## Open Ports"]
    if open_ports:
        lines.extend(summarize_open_ports(host))
    else:
        lines.append("_No open ports detected._")

    if existing_nessus:
        lines += ["", "## Nessus Findings", existing_nessus]

    if existing_burp:
        lines += ["", "## Burp Suite Findings", existing_burp]

    if existing_autorecon:
        lines += ["", "## AutoRecon Enumeration", existing_autorecon]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_stem


def create_obsidian_vault(
    vault_dir: Path,
    scan_name: str,
    scan_data: dict,
    tool_name: str,
    analysis_text: str | None,
    model_name: str | None,
    validation_warnings: list[str] | None = None,
) -> dict:
    """Write all host notes and the scan note for one Nmap scan file."""
    logging.info(f"Writing Nmap vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = scan_name
    if len(scan_data["hosts"]) == 1:
        scan_display = choose_host_display_name(scan_data["hosts"][0])

    scan_stem = safe_filename(f"{scan_display} - {tool_name}")
    scan_path = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in scan_data["hosts"]:
        display, host_stem = _write_host_note(
            hosts_dir, host, scan_stem, scan_display, tool_name
        )
        host_entries.append((display, host_stem))

    scan_lines = [
        f"# {scan_display} - {tool_name}",
        "",
        f"- **Source:** {scan_data['source_file']}",
        f"- **Parsed:** {scan_data['parsed_at']}",
        f"- **Nmap Args:** `{scan_data.get('nmap_args', '')}`",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    scan_lines += ["", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    if validation_warnings:
        scan_lines += ["", "## Validation Warnings", ""]
        for w in validation_warnings:
            scan_lines.append(f"- {w}")

    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Nmap scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — Nessus
# ============================================================

def _update_host_note_nessus(
    hosts_dir: Path,
    ip: str,
    hostname: str,
    findings: list[dict],
    scan_source: str,
) -> tuple[str, str]:
    """
    Write or update the ## Nessus Findings section in a host note.
    Returns (display_name, host_stem).
    """
    # Try to find an existing note by IP first
    existing_path = _find_host_note_by_ip(hosts_dir, ip)

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        # Determine display name
        if hostname and is_probable_fqdn(hostname):
            display = hostname
        else:
            display = ip or hostname or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports_section = ""
    existing_burp_section = ""
    existing_autorecon_section = ""
    existing_preamble_lines: list[str] = []
    existing_scan_refs: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes           = extract_operator_notes(old_body)
        existing_open_ports_section = extract_body_section(old_body, "## Open Ports")
        existing_burp_section       = extract_body_section(old_body, "## Burp Suite Findings")
        existing_autorecon_section  = extract_body_section(old_body, "## AutoRecon Enumeration")
        # Capture preamble (lines before first ## section)
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating Nessus section in: {host_path.name}")
    else:
        # Minimal preamble for Nessus-only host
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if hostname and hostname != ip:
            existing_preamble_lines.append(f"**Hostname:** {hostname}")
        logging.debug(f"Creating Nessus-only host note: {host_path.name}")

    # Update frontmatter
    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname and hostname not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname]

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    max_sev = max((f["severity_int"] for f in findings), default=0)
    existing_max_sev = existing_fm.get("nessus_max_severity", 0)
    try:
        existing_max_sev = int(existing_max_sev)
    except (TypeError, ValueError):
        existing_max_sev = 0
    merged_max_sev = max(max_sev, existing_max_sev)

    fm: dict = {
        "ip":                  ip or existing_fm.get("ip"),
        "hostnames":           existing_hostnames,
        "status":              existing_fm.get("status", "not-started"),
        "tags":                existing_fm.get("tags", []),
        "sources":             existing_sources,
        "nessus_max_severity": merged_max_sev,
    }

    # Build body
    preamble = "\n".join(existing_preamble_lines).rstrip()
    lines: list[str] = []

    if preamble:
        lines.append(preamble)

    if existing_open_ports_section:
        lines += ["", "## Open Ports", existing_open_ports_section]

    lines += ["", "## Nessus Findings", _render_nessus_section(findings)]

    if existing_burp_section:
        lines += ["", "## Burp Suite Findings", existing_burp_section]

    if existing_autorecon_section:
        lines += ["", "## AutoRecon Enumeration", existing_autorecon_section]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_nessus_vault(
    vault_dir: Path,
    scan_name: str,
    nessus_data: dict,
    analysis_text: str | None,
    model_name: str | None,
    validation_warnings: list[str] | None = None,
) -> dict:
    """Write Nessus host note sections and scan note."""
    logging.info(f"Writing Nessus vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = nessus_data.get("report_name", scan_name)
    scan_stem    = safe_filename(f"{scan_display} - Nessus")
    scan_source  = f"{scan_display} - Nessus"
    scan_path    = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in nessus_data.get("hosts", []):
        display, host_stem = _update_host_note_nessus(
            hosts_dir,
            host["ip"],
            host.get("hostname", ""),
            host["findings"],
            scan_source,
        )
        host_entries.append((display, host_stem))

    # Build scan note
    total_findings = sum(len(h["findings"]) for h in nessus_data.get("hosts", []))
    critical_count = sum(
        sum(1 for f in h["findings"] if f["severity_int"] == 4)
        for h in nessus_data.get("hosts", [])
    )
    high_count = sum(
        sum(1 for f in h["findings"] if f["severity_int"] == 3)
        for h in nessus_data.get("hosts", [])
    )

    scan_lines = [
        f"# {scan_display} - Nessus",
        "",
        f"- **Source:** {nessus_data['source_file']}",
        f"- **Parsed:** {nessus_data['parsed_at']}",
        f"- **Total Findings:** {total_findings}  ·  **Critical:** {critical_count}  ·  **High:** {high_count}",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    scan_lines += ["", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    if validation_warnings:
        scan_lines += ["", "## Validation Warnings", ""]
        for w in validation_warnings:
            scan_lines.append(f"- {w}")

    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Nessus scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — Burp Suite
# ============================================================

def _update_host_note_burp(
    hosts_dir: Path,
    ip: str,
    url: str,
    issues: list[dict],
    scan_source: str,
) -> tuple[str, str]:
    """
    Write or update the ## Burp Suite Findings section in a host note.
    Returns (display_name, host_stem).
    """
    existing_path = _find_host_note_by_ip(hosts_dir, ip) if ip else None

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        # Derive display name from URL host or IP
        if url:
            from urllib.parse import urlparse
            parsed = urlparse(url)
            netloc = parsed.netloc or url
            display = netloc.split(":")[0] if netloc else (ip or url)
        else:
            display = ip or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports_section = ""
    existing_nessus_section = ""
    existing_autorecon_section = ""
    existing_preamble_lines: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes            = extract_operator_notes(old_body)
        existing_open_ports_section  = extract_body_section(old_body, "## Open Ports")
        existing_nessus_section      = extract_body_section(old_body, "## Nessus Findings")
        existing_autorecon_section   = extract_body_section(old_body, "## AutoRecon Enumeration")
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating Burp section in: {host_path.name}")
    else:
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if url:
            existing_preamble_lines.append(f"**URL:** {url}")
        logging.debug(f"Creating Burp-only host note: {host_path.name}")

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    # Derive hostname from URL for frontmatter
    hostname_for_fm = ""
    if url:
        from urllib.parse import urlparse
        parsed = urlparse(url)
        hostname_for_fm = parsed.netloc.split(":")[0] if parsed.netloc else ""

    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname_for_fm and hostname_for_fm not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname_for_fm]

    fm: dict = {
        "ip":        ip or existing_fm.get("ip"),
        "hostnames": existing_hostnames,
        "status":    existing_fm.get("status", "not-started"),
        "tags":      existing_fm.get("tags", []),
        "sources":   existing_sources,
    }
    if "nessus_max_severity" in existing_fm:
        fm["nessus_max_severity"] = existing_fm["nessus_max_severity"]

    preamble = "\n".join(existing_preamble_lines).rstrip()
    lines: list[str] = []

    if preamble:
        lines.append(preamble)

    if existing_open_ports_section:
        lines += ["", "## Open Ports", existing_open_ports_section]

    if existing_nessus_section:
        lines += ["", "## Nessus Findings", existing_nessus_section]

    lines += ["", "## Burp Suite Findings", _render_burp_section(issues)]

    if existing_autorecon_section:
        lines += ["", "## AutoRecon Enumeration", existing_autorecon_section]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_burp_vault(
    vault_dir: Path,
    scan_name: str,
    burp_data: dict,
    analysis_text: str | None,
    model_name: str | None,
    validation_warnings: list[str] | None = None,
) -> dict:
    """Write Burp Suite host note sections and scan note."""
    logging.info(f"Writing Burp vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = scan_name
    scan_stem    = safe_filename(f"{scan_display} - Burp")
    scan_source  = f"{scan_display} - Burp"
    scan_path    = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in burp_data.get("hosts", []):
        display, host_stem = _update_host_note_burp(
            hosts_dir,
            host.get("ip", ""),
            host.get("url", ""),
            host["issues"],
            scan_source,
        )
        host_entries.append((display, host_stem))

    total_issues = sum(len(h["issues"]) for h in burp_data.get("hosts", []))
    high_count   = sum(
        sum(1 for i in h["issues"] if i["severity"].lower() == "high")
        for h in burp_data.get("hosts", [])
    )

    scan_lines = [
        f"# {scan_display} - Burp",
        "",
        f"- **Source:** {burp_data['source_file']}",
        f"- **Parsed:** {burp_data['parsed_at']}",
        f"- **Total Issues:** {total_issues}  ·  **High:** {high_count}",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    scan_lines += ["", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    if validation_warnings:
        scan_lines += ["", "## Validation Warnings", ""]
        for w in validation_warnings:
            scan_lines.append(f"- {w}")

    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Burp scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — AutoRecon
# ============================================================

def _render_autorecon_section(target_data: dict) -> str:
    """Render AutoRecon enumeration data as markdown body (no section header)."""
    summary = target_data.get("summary", {})
    tool_results = target_data.get("tool_results", {})

    if not tool_results:
        return "_AutoRecon ran but produced no additional findings._"

    lines: list[str] = [
        f"**AutoRecon target:** {target_data.get('target', '?')} "
        f"| **Tools run:** {summary.get('total_tools_run', 0)} "
        f"| **With findings:** {summary.get('tools_with_findings', 0)}",
        "",
    ]

    empty_ports = 0

    for port_key in sorted(tool_results):
        results = tool_results[port_key]
        proto, port = port_key.split("/", 1)
        # Derive service hint from first result filename
        service_hint = ""
        for r in results:
            fn_match = AUTORECON_FILENAME_RE.match(r.get("filename", ""))
            if fn_match:
                service_hint = fn_match.group(3)
                break

        port_has_content = False
        port_lines: list[str] = [f"### {port_key} — {service_hint or proto.upper()}"]

        for entry in results:
            data = entry.get("data", {})
            tool = data.get("tool", "unknown")

            if tool == "dirbusting" and data.get("interesting"):
                port_has_content = True
                interesting = data["interesting"][:20]
                port_lines.append(f"\n#### Discovered Paths ({data.get('total_found', 0)} total)")
                port_lines.append("| Status | Path | Size |")
                port_lines.append("|--------|------|------|")
                for p in interesting:
                    redir = f" → {p['redirect']}" if p.get("redirect") else ""
                    port_lines.append(f"| {p['status']} | `{p['path']}` | {p['size']}{redir} |")

            elif tool == "nikto" and data.get("findings"):
                port_has_content = True
                port_lines.append("\n#### Nikto Findings")
                if data.get("server"):
                    port_lines.append(f"**Server:** {data['server']}")
                for f in data["findings"][:15]:
                    osvdb = f"OSVDB-{f['osvdb']}: " if f.get("osvdb") else ""
                    port_lines.append(f"- {osvdb}`{f['path']}` — {f['description'][:200]}")

            elif tool == "whatweb" and data.get("technologies"):
                port_has_content = True
                techs = [f"{t['name']}/{t['version']}" if t.get("version") else t["name"]
                         for t in data["technologies"]]
                port_lines.append(f"\n**Technologies:** {', '.join(techs)}")
                if data.get("title"):
                    port_lines.append(f"**Title:** {data['title']}")

            elif tool == "enum4linux":
                has_data = any([data.get("os_info"), data.get("users"),
                                data.get("shares"), data.get("null_session")])
                if has_data:
                    port_has_content = True
                    if data.get("os_info"):
                        port_lines.append(f"\n**OS:** {data['os_info']}")
                    if data.get("domain"):
                        port_lines.append(f"**Domain:** {data['domain']}")
                    if data.get("null_session"):
                        port_lines.append("**Null Session:** Yes")
                    if data.get("shares"):
                        port_lines.append("\n#### Shares")
                        port_lines.append("| Share | Access |")
                        port_lines.append("|-------|--------|")
                        for s in data["shares"]:
                            access = s.get("access", s.get("type", ""))
                            bold = "**" if "write" in access.lower() else ""
                            port_lines.append(f"| {s['name']} | {bold}{access}{bold} |")
                    if data.get("users"):
                        port_lines.append("\n#### Users Found")
                        port_lines.append(", ".join(f"`{u}`" for u in data["users"][:30]))
                    pp = data.get("password_policy", {})
                    if pp:
                        parts = []
                        if "min_length" in pp:
                            parts.append(f"Min length: {pp['min_length']}")
                        if "lockout_threshold" in pp:
                            val = pp["lockout_threshold"]
                            parts.append(f"Lockout threshold: {val}" + (" (no lockout)" if val == 0 else ""))
                        if pp.get("complexity"):
                            parts.append("Complexity: on")
                        if parts:
                            port_lines.append(f"\n**Password Policy:** {' | '.join(parts)}")

            elif tool == "smbmap" and data.get("shares"):
                port_has_content = True
                port_lines.append("\n#### SMB Shares (smbmap)")
                port_lines.append("| Share | Permissions |")
                port_lines.append("|-------|------------|")
                for s in data["shares"]:
                    bold = "**" if "WRITE" in s["permissions"] else ""
                    port_lines.append(f"| {s['name']} | {bold}{s['permissions']}{bold} |")

            elif tool == "smbclient" and data.get("shares"):
                port_has_content = True
                port_lines.append("\n#### Shares (smbclient)")
                for s in data["shares"]:
                    port_lines.append(f"- {s['name']} ({s['type']})")

            elif tool == "sslscan":
                has_data = any([data.get("weak_protocols"), data.get("weak_ciphers"),
                                data.get("cert_expired"), data.get("heartbleed_vulnerable")])
                if has_data:
                    port_has_content = True
                    port_lines.append("\n#### TLS/SSL")
                    if data.get("weak_protocols"):
                        port_lines.append(f"- **Weak protocols:** {', '.join(data['weak_protocols'])}")
                    if data.get("weak_ciphers"):
                        port_lines.append(f"- **Weak ciphers:** {', '.join(data['weak_ciphers'][:5])}")
                    if data.get("cert_subject"):
                        port_lines.append(f"- **Certificate:** {data['cert_subject']}")
                    if data.get("cert_expired"):
                        port_lines.append(f"- **Expired:** {data.get('cert_not_after', 'yes')}")
                    if data.get("heartbleed_vulnerable"):
                        port_lines.append("- **HEARTBLEED: VULNERABLE**")

            elif tool == "snmpwalk" and data.get("sys_descr"):
                port_has_content = True
                port_lines.append(f"\n**SNMP sysDescr:** {data['sys_descr']}")
                if data.get("sys_name"):
                    port_lines.append(f"**sysName:** {data['sys_name']}")
                if data.get("running_processes"):
                    port_lines.append(f"**Processes:** {', '.join(data['running_processes'][:15])}")

            elif tool == "onesixtyone" and data.get("community_strings"):
                port_has_content = True
                port_lines.append("\n#### SNMP Community Strings")
                for cs in data["community_strings"]:
                    port_lines.append(f"- `{cs['community']}` — {cs['sys_descr'][:100]}")

            elif tool == "dnsrecon" and data.get("records"):
                port_has_content = True
                port_lines.append("\n#### DNS Records")
                if data.get("zone_transfer_successful"):
                    port_lines.append("**ZONE TRANSFER SUCCESSFUL**")
                for r in data["records"][:20]:
                    port_lines.append(f"- {r['type']} `{r['name']}` → {r['value']}")

            elif tool == "curl":
                if data.get("title") or data.get("interesting_headers"):
                    port_has_content = True
                    if data.get("title"):
                        port_lines.append(f"\n**Page title:** {data['title']}")
                    for h_name in data.get("interesting_headers", []):
                        port_lines.append(f"**{h_name}:** {data['headers'].get(h_name, '')}")

        if port_has_content:
            lines.extend(port_lines)
            lines.append("")
        else:
            empty_ports += 1

    if empty_ports:
        lines.append(f"_Plus {empty_ports} port(s) with generic/empty tool output._")

    return "\n".join(lines).strip()


def _update_host_note_autorecon(
    hosts_dir: Path,
    ip: str,
    hostname: str,
    target_data: dict,
    scan_source: str,
) -> tuple[str, str]:
    """Write or update the ## AutoRecon Enumeration section in a host note."""
    existing_path = _find_host_note_by_ip(hosts_dir, ip) if ip else None

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        if hostname and is_probable_fqdn(hostname):
            display = hostname
        else:
            display = ip or hostname or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports_section = ""
    existing_nessus_section = ""
    existing_burp_section = ""
    existing_preamble_lines: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes           = extract_operator_notes(old_body)
        existing_open_ports_section = extract_body_section(old_body, "## Open Ports")
        existing_nessus_section     = extract_body_section(old_body, "## Nessus Findings")
        existing_burp_section       = extract_body_section(old_body, "## Burp Suite Findings")
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating AutoRecon section in: {host_path.name}")
    else:
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if hostname and hostname != ip:
            existing_preamble_lines.append(f"**Hostname:** {hostname}")
        logging.debug(f"Creating AutoRecon-only host note: {host_path.name}")

    # Merge frontmatter
    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname and hostname not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname]

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    existing_tags: list = existing_fm.get("tags", [])
    if "autorecon" not in existing_tags:
        existing_tags = existing_tags + ["autorecon"]

    fm: dict = {
        "ip":                  ip or existing_fm.get("ip"),
        "hostnames":           existing_hostnames,
        "status":              existing_fm.get("status", "not-started"),
        "tags":                existing_tags,
        "sources":             existing_sources,
        "nessus_max_severity": existing_fm.get("nessus_max_severity", 0),
        "autorecon_tools_run": target_data.get("summary", {}).get("total_tools_run", 0),
    }

    # Build body
    preamble = "\n".join(existing_preamble_lines).rstrip()
    body_lines: list[str] = []

    if preamble:
        body_lines.append(preamble)

    if existing_open_ports_section:
        body_lines += ["", "## Open Ports", existing_open_ports_section]

    if existing_nessus_section:
        body_lines += ["", "## Nessus Findings", existing_nessus_section]

    if existing_burp_section:
        body_lines += ["", "## Burp Suite Findings", existing_burp_section]

    body_lines += ["", "## AutoRecon Enumeration", _render_autorecon_section(target_data)]

    body_lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        body_lines.append(f"- [[Scans/{src_stem}|{src}]]")

    body_lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        body_lines += ["", existing_op_notes]

    body = "\n".join(body_lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_autorecon_vault(
    vault_dir: Path,
    scan_name: str,
    autorecon_data: dict,
    analysis_text: str | None,
    model_name: str | None,
    validation_warnings: list[str] | None = None,
) -> dict:
    """Write AutoRecon host note sections and scan note."""
    logging.info(f"Writing AutoRecon vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    targets = autorecon_data.get("targets", [])
    scan_display = autorecon_data.get("report_name", scan_name)
    scan_stem    = safe_filename(f"{scan_display} - AutoRecon")
    scan_source  = f"{scan_display} - AutoRecon"
    scan_path    = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for target in targets:
        display, host_stem = _update_host_note_autorecon(
            hosts_dir,
            target.get("ip", ""),
            target.get("hostname", ""),
            target,
            scan_source,
        )
        host_entries.append((display, host_stem))

    # Build scan note
    total_tools = sum(t.get("summary", {}).get("total_tools_run", 0) for t in targets)
    tools_with_findings = sum(t.get("summary", {}).get("tools_with_findings", 0) for t in targets)

    scan_lines = [
        f"# {scan_display} - AutoRecon",
        "",
        f"- **Source:** {autorecon_data.get('source_file', scan_name)}",
        f"- **Parsed:** {autorecon_data.get('parsed_at', '')}",
        f"- **Targets:** {len(targets)}  ·  **Tools run:** {total_tools}  ·  **With findings:** {tools_with_findings}",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    scan_lines += ["", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    if validation_warnings:
        scan_lines += ["", "## Validation Warnings", ""]
        for w in validation_warnings:
            scan_lines.append(f"- {w}")

    # Include manual commands in collapsible section
    for target in targets:
        mc = target.get("manual_commands", "")
        if mc.strip():
            scan_lines += [
                "", f"## Suggested Manual Commands ({target.get('target', '')})",
                "",
                "<details>",
                "<summary>Click to expand</summary>",
                "",
                "```",
                mc.strip(),
                "```",
                "",
                "</details>",
            ]
        cl = target.get("commands_log", "")
        if cl.strip():
            scan_lines += [
                "", f"## Commands Log ({target.get('target', '')})",
                "",
                "<details>",
                "<summary>Click to expand</summary>",
                "",
                "```",
                cl.strip()[:5000],
                "```",
                "",
                "</details>",
            ]

    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote AutoRecon scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Canvas node / edge constructors
# ============================================================

def _text_node(nid: str, text: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "text", "text": text, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _file_node(nid: str, file: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "file", "file": file, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _group_node(nid: str, label: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "group", "label": label, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _edge(eid: str, from_id: str, to_id: str, label: str = "") -> dict:
    e: dict = {
        "id":       eid,
        "fromNode": from_id,
        "fromSide": "bottom",
        "toNode":   to_id,
        "toSide":   "top",
    }
    if label:
        e["label"] = label
    return e


# ============================================================
# Canvas content helpers
# ============================================================

def _read_host_frontmatter(host_path: Path) -> dict:
    try:
        fm, _ = read_frontmatter(host_path.read_text(encoding="utf-8"))
        return fm
    except Exception:
        return {}


def _build_campaign_overview(
    vault_dir: Path,
    scan_host_map: dict,
    all_analyses: list[tuple[str, str]],
) -> str:
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    total_hosts = len(host_notes)
    total_scans = len(scan_host_map)
    status_counts: dict[str, int] = {}
    all_tags: set[str] = set()
    total_ports = 0
    nessus_critical = 0
    nessus_high     = 0
    nessus_medium   = 0
    burp_high       = 0
    burp_total      = 0
    autorecon_targets = 0
    autorecon_writable = 0

    for hp in host_notes:
        text = ""
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue

        fm, body = read_frontmatter(text)
        status = fm.get("status", "not-started")
        status_counts[status] = status_counts.get(status, 0) + 1
        all_tags.update(fm.get("tags", []))

        # Count Nmap open ports
        total_ports += sum(
            1 for line in body.splitlines()
            if line.strip().startswith("- **") and "/" in line
            and line.strip().startswith("- **tcp") or
            (line.strip().startswith("- **") and "/" in line and not line.strip().startswith("- **tcp"))
        )

        # Count Nessus severity from frontmatter
        max_sev = fm.get("nessus_max_severity", 0)
        try:
            max_sev = int(max_sev)
        except (TypeError, ValueError):
            max_sev = 0
        if max_sev >= 4:
            nessus_critical += 1
        elif max_sev >= 3:
            nessus_high += 1
        elif max_sev >= 2:
            nessus_medium += 1

        # Count Burp issues from section
        burp_section = extract_body_section(body, "## Burp Suite Findings")
        if burp_section and "_No Burp" not in burp_section:
            burp_total += burp_section.count("#### [")
            burp_high  += burp_section.count("#### [High")

        # Count AutoRecon targets
        ar_section = extract_body_section(body, "## AutoRecon Enumeration")
        if ar_section and "_AutoRecon ran but" not in ar_section:
            autorecon_targets += 1
            if "READ/WRITE" in ar_section or "read/write" in ar_section:
                autorecon_writable += 1

    # Re-count ports more simply
    total_ports = 0
    for hp in host_notes:
        try:
            _, body = read_frontmatter(hp.read_text(encoding="utf-8"))
            ports_section = extract_body_section(body, "## Open Ports")
            total_ports += sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))
        except Exception:
            pass

    lines = [
        "# Campaign Overview",
        "",
        f"**Hosts:** {total_hosts}  ·  **Scans:** {total_scans}  ·  **Open Ports:** {total_ports}",
        "",
    ]

    if status_counts:
        parts = [f"{k}: {v}" for k, v in sorted(status_counts.items())]
        lines += [f"**Progress:** {' · '.join(parts)}", ""]

    display_tags = sorted(all_tags - {"rpc"})
    if display_tags:
        lines += [f"**Services seen:** {', '.join(display_tags)}", ""]

    nessus_parts = []
    if nessus_critical:
        nessus_parts.append(f"critical hosts: {nessus_critical}")
    if nessus_high:
        nessus_parts.append(f"high hosts: {nessus_high}")
    if nessus_medium:
        nessus_parts.append(f"medium hosts: {nessus_medium}")
    if nessus_parts:
        lines += [f"**Nessus:** {' · '.join(nessus_parts)}", ""]

    if burp_total:
        lines += [f"**Burp:** {burp_total} issues ({burp_high} high)", ""]

    if autorecon_targets:
        ar_parts = [f"targets: {autorecon_targets}"]
        if autorecon_writable:
            ar_parts.append(f"writable shares: {autorecon_writable}")
        lines += [f"**AutoRecon:** {' · '.join(ar_parts)}", ""]

    # Key Observations bullets from the most recent AI analysis
    if all_analyses:
        text = all_analyses[-1][1]
        match = re.search(
            r"##\s+Key Observations(.*?)(?=\n##\s|\Z)", text, re.DOTALL | re.IGNORECASE
        )
        if match:
            bullets = [
                l.strip() for l in match.group(1).splitlines()
                if l.strip().startswith(("-", "*", "•"))
            ][:4]
            if bullets:
                lines += ["**Key Findings:**"] + bullets

    return "\n".join(lines)


def build_priority_targets_prompt(vault_dir: Path, all_analyses: list[tuple[str, str]]) -> str:
    """
    Build an Ollama prompt asking it to rank all discovered hosts by exploitation
    priority, drawing on Nmap service tags, Nessus CVEs/CVSS, Burp findings, and AutoRecon enumeration.
    """
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    instructions = """
You are an experienced penetration tester. Given scan data about multiple hosts,
rank them from highest to lowest exploitation priority.
This ranking will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

GROUNDING RULES — follow these strictly:
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- Only rank a host as a top target if there is concrete scan evidence supporting
  exploitability. Do not rank based on port count alone.
- If evidence is insufficient to justify a high ranking, say so rather than
  padding the list.
- Tag each ranked entry as one of:
    [CONFIRMED]  — directly in the scan data (CVE with plugin output, confirmed vuln)
    [INFERRED]   — reasonable conclusion from the scan data (service fingerprint, role)
    [ASSUMED]    — requires verification; explain why

Output ONLY a numbered list — no preamble, no section headers, no trailing text.
Each entry must be a single line in this exact format:
  N. <hostname or IP> [TAG] — <primary evidence: CVE ID / finding name / service>; <brief impact>

Base your ranking on:
- CVEs with confirmed plugin output and known weaponized exploits ([MSF] first)
- CVSS scores (critical ≥ 9.0, high ≥ 7.0) — but only when plugin_output:yes
- Exposed management interfaces with evidence of weak auth (RDP, WinRM, SSH)
- Domain role evidence in hostnames, ports, or service banners (DC, CA, DB)
- Chaining potential supported by scan data (pivot hosts, credential stores)
- Web findings from Burp (Certain/Firm auth bypass, SQLi — not Tentative)
- AutoRecon enumeration: null sessions, writable shares, discovered users, weak TLS
- Unencrypted or unauthenticated services confirmed in scan data

Do not rank a host highly based solely on port count or service tags without
supporting vulnerability or misconfiguration evidence.
Maximum 10 entries.
""".strip()

    host_lines: list[str] = []
    for hp in host_notes:
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue
        fm, body = read_frontmatter(text)

        ip        = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        display   = hostnames[0] if hostnames else (ip or hp.stem)
        tags      = fm.get("tags", [])
        nessus_max = fm.get("nessus_max_severity", 0)
        try:
            nessus_max = int(nessus_max)
        except (TypeError, ValueError):
            nessus_max = 0

        ports_section  = extract_body_section(body, "## Open Ports")
        port_count     = sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))

        nessus_section = extract_body_section(body, "## Nessus Findings")
        crit_count     = nessus_section.count("[Critical]") if nessus_section else 0
        high_count_n   = nessus_section.count("[High]")     if nessus_section else 0
        cves           = re.findall(r"CVE-\d{4}-\d+", nessus_section)[:4] if nessus_section else []

        burp_section   = extract_body_section(body, "## Burp Suite Findings")
        burp_high      = burp_section.count("[High") if burp_section else 0
        burp_med       = burp_section.count("[Medium") if burp_section else 0

        ar_section     = extract_body_section(body, "## AutoRecon Enumeration")
        ar_null_sess   = bool(ar_section and "Null Session:** Yes" in ar_section)
        ar_writable    = bool(ar_section and ("READ/WRITE" in ar_section or "read/write" in ar_section))
        ar_users       = len(re.findall(r"`(\w+)`", ar_section)) if ar_section else 0

        parts = [f"{display} (IP: {ip})"]
        if tags:
            parts.append(f"services: {', '.join(tags)}")
        parts.append(f"open-ports: {port_count}")
        if nessus_max:
            parts.append(f"nessus-max: {severity_int_to_str(nessus_max)}")
        if crit_count or high_count_n:
            parts.append(f"critical/high findings: {crit_count}/{high_count_n}")
        if cves:
            parts.append(f"CVEs: {', '.join(cves)}")
        if burp_high or burp_med:
            parts.append(f"burp high/medium: {burp_high}/{burp_med}")
        if ar_null_sess:
            parts.append("null-session: yes")
        if ar_writable:
            parts.append("writable-shares: yes")
        if ar_users:
            parts.append(f"users-found: {ar_users}")

        host_lines.append("- " + " | ".join(parts))

    # Include Key Observations snippets from up to 3 most recent analyses for context
    analysis_ctx: list[str] = []
    for scan_display, text in all_analyses[-3:]:
        m = re.search(r"##\s+Key Observations(.*?)(?=\n##\s|\Z)", text, re.DOTALL | re.IGNORECASE)
        if m:
            bullets = [
                l.strip() for l in m.group(1).splitlines()
                if l.strip().startswith(("-", "*", "•"))
            ][:3]
            if bullets:
                analysis_ctx.append(f"[{scan_display}]")
                analysis_ctx.extend(bullets)

    prompt = f"{instructions}\n\nHosts to rank:\n" + "\n".join(host_lines)
    if analysis_ctx:
        prompt += "\n\nKey scan findings for context:\n" + "\n".join(analysis_ctx)
    logging.debug(f"Built priority targets prompt ({len(prompt)} chars)")
    return prompt


def _build_priority_targets_fallback(vault_dir: Path) -> str:
    """
    Build a static Priority Targets list (no AI) sorted by Nessus severity,
    Burp high count, and open port count.  Used when Ollama is skipped or
    no analyses have been generated yet.
    """
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    entries: list[dict] = []
    for hp in host_notes:
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue
        fm, body = read_frontmatter(text)

        ip        = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        display   = hostnames[0] if hostnames else (ip or hp.stem)
        tags      = fm.get("tags", [])
        nessus_max = fm.get("nessus_max_severity", 0)
        try:
            nessus_max = int(nessus_max)
        except (TypeError, ValueError):
            nessus_max = 0

        ports_section  = extract_body_section(body, "## Open Ports")
        port_count     = sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))

        nessus_section = extract_body_section(body, "## Nessus Findings")
        cves           = re.findall(r"CVE-\d{4}-\d+", nessus_section)[:2] if nessus_section else []

        burp_section   = extract_body_section(body, "## Burp Suite Findings")
        burp_high      = burp_section.count("[High") if burp_section else 0

        reasons: list[str] = []
        if cves:
            reasons.append(", ".join(cves))
        if "domain-controller" in tags:
            reasons.append("domain controller")
        if nessus_max >= 4:
            reasons.append("critical Nessus findings")
        elif nessus_max >= 3:
            reasons.append("high-severity Nessus findings")
        elif nessus_max >= 2:
            reasons.append("medium-severity Nessus findings")
        if burp_high:
            reasons.append(f"{burp_high} high web issue(s)")
        exposed = [t for t in tags if t in ("rdp", "winrm", "smb", "kerberos", "ldap", "mssql", "ftp", "snmp")]
        if exposed:
            reasons.append(f"exposed: {', '.join(exposed)}")
        if not reasons and port_count:
            service_str = ", ".join(tags[:4]) if tags else "unknown services"
            reasons.append(f"{port_count} open port(s) — {service_str}")

        entries.append({
            "display":  display,
            "sort_key": (-nessus_max, -burp_high, -port_count),
            "reason":   "; ".join(reasons) if reasons else "no notable findings",
        })

    entries.sort(key=lambda e: e["sort_key"])

    if not entries:
        return "## Priority Targets\n\n_No hosts found. Run scans to populate._"

    lines = ["## Priority Targets", ""]
    for i, entry in enumerate(entries[:10], 1):
        lines.append(f"{i}. {entry['display']} — {entry['reason']}")
    return "\n".join(lines)


def _build_next_steps(all_analyses: list[tuple[str, str]], max_items: int = 14) -> str:
    items: list[str] = []

    for _scan_display, text in all_analyses:
        for section in (
            "Potential Attack Paths",
            "Enumeration Suggestions",
            "Exploitation Priority",
            "Attack Chains",
        ):
            match = re.search(
                rf"##\s+{re.escape(section)}(.*?)(?=\n##\s|\Z)",
                text, re.DOTALL | re.IGNORECASE,
            )
            if not match:
                continue
            for line in match.group(1).splitlines():
                stripped = line.strip()
                if not stripped or stripped.startswith("#"):
                    continue
                cleaned = re.sub(r"^[-*•\d]+[.)]\s*", "", stripped).strip()
                if cleaned and len(cleaned) > 12 and cleaned not in items:
                    items.append(cleaned)
                if len(items) >= max_items:
                    break
            if len(items) >= max_items:
                break

    if not items:
        return "# Next Steps\n\n_Run scans with AI analysis enabled to populate next steps._"

    lines = ["# Next Steps", ""]
    for item in items[:max_items]:
        lines.append(f"- {item}")
    return "\n".join(lines)


# ============================================================
# Canvas layout engine
# ============================================================

def _group_dimensions(n_hosts: int, cols: int) -> tuple[int, int]:
    if n_hosts == 0:
        return 300, 200
    actual_cols = min(n_hosts, cols)
    rows = math.ceil(n_hosts / cols)
    inner_w = actual_cols * CARD_W + (actual_cols - 1) * CARD_GAP_X
    inner_h = rows * CARD_H + (rows - 1) * CARD_GAP_Y
    w = GROUP_PAD * 2 + inner_w
    h = GROUP_PAD + GROUP_LABEL_H + inner_h + GROUP_PAD
    return w, h


def build_canvas(
    vault_dir: Path,
    canvas_name: str,
    scan_host_map: dict,
    all_analyses: list[tuple[str, str]],
    canvas_cols: int = 2,
    max_groups_per_row: int = 3,
    priority_targets_text: str | None = None,
) -> Path:
    hosts_dir   = vault_dir / "Hosts"
    canvas_path = vault_dir / canvas_name

    existing_canvas: dict = {"nodes": [], "edges": []}
    if canvas_path.exists():
        try:
            existing_canvas = json.loads(canvas_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    managed_ids: set[str] = set()
    our_nodes:   list[dict] = []
    our_edges:   list[dict] = []

    # 1. Campaign Overview
    overview_id = stable_id("campaign_overview")
    managed_ids.add(overview_id)
    overview_x = -(OVERVIEW_W // 2)
    overview_y = -900
    our_nodes.append(_text_node(
        overview_id,
        _build_campaign_overview(vault_dir, scan_host_map, all_analyses),
        overview_x, overview_y, OVERVIEW_W, OVERVIEW_H,
    ))

    # 2. Scan note cards
    scan_stems = list(scan_host_map.keys())
    total_scan_row_w = (
        len(scan_stems) * SCAN_CARD_W + max(0, len(scan_stems) - 1) * CARD_GAP_X
    )
    scan_row_x0 = -(total_scan_row_w // 2)
    scan_row_y  = overview_y + OVERVIEW_H + 100

    scan_node_ids: dict[str, str] = {}

    for i, scan_stem in enumerate(scan_stems):
        sid = stable_id(f"scan_{scan_stem}")
        managed_ids.add(sid)
        scan_node_ids[scan_stem] = sid
        sx  = scan_row_x0 + i * (SCAN_CARD_W + CARD_GAP_X)
        rel = f"Scans/{ensure_md_suffix(scan_stem)}"
        our_nodes.append(_file_node(sid, rel, sx, scan_row_y, SCAN_CARD_W, SCAN_CARD_H))
        eid = stable_id(f"edge_overview_{sid}")
        managed_ids.add(eid)
        our_edges.append(_edge(eid, overview_id, sid))

    # 2b. Priority Targets node (centered, between scan cards and subnet groups)
    pt_id = stable_id("priority-targets")
    managed_ids.add(pt_id)
    pt_text = priority_targets_text or _build_priority_targets_fallback(vault_dir)
    # Ensure the markdown header is present
    if not pt_text.lstrip().startswith("## Priority Targets"):
        pt_text = "## Priority Targets\n\n" + pt_text.lstrip()
    pt_x = -(NEXT_STEPS_W // 2)
    pt_y = scan_row_y + SCAN_CARD_H + ROW_GAP
    our_nodes.append(_text_node(pt_id, pt_text, pt_x, pt_y, NEXT_STEPS_W, NEXT_STEPS_H, color="2"))

    # Overview → Priority Targets
    eid_ov_pt = stable_id("edge_overview_pt")
    managed_ids.add(eid_ov_pt)
    our_edges.append(_edge(eid_ov_pt, overview_id, pt_id))

    # 3. Subnet groups with host cards
    subnet_hosts: dict[str, list[tuple[str, Path]]] = {}
    if hosts_dir.exists():
        for hp in sorted(hosts_dir.glob("*.md")):
            fm     = _read_host_frontmatter(hp)
            ip     = fm.get("ip")
            subnet = get_subnet_label(ip) if ip and is_ipv4(str(ip)) else "unknown"
            subnet_hosts.setdefault(subnet, []).append((hp.stem, hp))

    subnets  = sorted(subnet_hosts.keys())
    n_groups = len(subnets)
    gdims    = [_group_dimensions(len(subnet_hosts[s]), canvas_cols) for s in subnets]

    # Groups start below the Priority Targets node
    cur_y = pt_y + NEXT_STEPS_H + ROW_GAP
    group_positions: list[tuple[int, int]] = []

    for row_i in range(math.ceil(n_groups / max_groups_per_row)):
        start = row_i * max_groups_per_row
        end   = min(start + max_groups_per_row, n_groups)
        row_indices = range(start, end)
        row_total_w = sum(gdims[j][0] for j in row_indices) + (end - start - 1) * GROUP_GAP
        row_height  = max((gdims[j][1] for j in row_indices), default=200)
        x = -(row_total_w // 2)
        for j in row_indices:
            group_positions.append((x, cur_y))
            x += gdims[j][0] + GROUP_GAP
        cur_y += row_height + ROW_GAP

    host_node_ids: dict[str, str] = {}

    for i, subnet in enumerate(subnets):
        if i >= len(group_positions):
            break
        gx, gy = group_positions[i]
        gw, gh = gdims[i]

        gid = stable_id(f"subnet_{subnet}")
        managed_ids.add(gid)
        our_nodes.append(_group_node(gid, subnet, gx, gy, gw, gh))

        # Priority Targets → subnet group edge
        eid_pt_g = stable_id(f"edge_pt_{gid}")
        managed_ids.add(eid_pt_g)
        our_edges.append(_edge(eid_pt_g, pt_id, gid))

        for j, (host_stem, hp) in enumerate(subnet_hosts[subnet]):
            col = j % canvas_cols
            row = j // canvas_cols
            hx  = gx + GROUP_PAD + col * (CARD_W + CARD_GAP_X)
            hy  = gy + GROUP_PAD + GROUP_LABEL_H + row * (CARD_H + CARD_GAP_Y)

            hid = stable_id(f"host_{host_stem}")
            managed_ids.add(hid)
            host_node_ids[host_stem] = hid

            fm     = _read_host_frontmatter(hp)
            status = fm.get("status", "not-started")
            color  = STATUS_COLORS.get(status)

            # If not-started, use Nessus severity for color hint
            if color is None and status == "not-started":
                nessus_max = fm.get("nessus_max_severity", 0)
                try:
                    nessus_max = int(nessus_max)
                except (TypeError, ValueError):
                    nessus_max = 0
                color = NESSUS_SEVERITY_CANVAS_COLORS.get(nessus_max)

            rel = f"Hosts/{ensure_md_suffix(host_stem)}"
            our_nodes.append(_file_node(hid, rel, hx, hy, CARD_W, CARD_H, color))

    # Scan → host edges
    for scan_stem, host_stems in scan_host_map.items():
        sid = scan_node_ids.get(scan_stem)
        if not sid:
            continue
        for host_stem in host_stems:
            hid = host_node_ids.get(host_stem)
            if not hid:
                continue
            eid = stable_id(f"edge_{sid}_{hid}")
            managed_ids.add(eid)
            our_edges.append(_edge(eid, sid, hid))

    # 4. Next Steps
    next_id = stable_id("next_steps")
    managed_ids.add(next_id)
    next_x = -(NEXT_STEPS_W // 2)
    next_y = cur_y + 40
    our_nodes.append(_text_node(
        next_id,
        _build_next_steps(all_analyses),
        next_x, next_y, NEXT_STEPS_W, NEXT_STEPS_H,
        color="5",
    ))

    preserved   = [n for n in existing_canvas.get("nodes", []) if n.get("id") not in managed_ids]
    final_nodes = preserved + our_nodes
    final_edges = our_edges

    canvas_path.write_text(
        json.dumps({"nodes": final_nodes, "edges": final_edges}, indent=2),
        encoding="utf-8",
    )
    logging.info(
        f"Canvas written: {canvas_path.name} "
        f"({len(final_nodes)} nodes, {len(final_edges)} edges, "
        f"{len(preserved)} preserved)"
    )
    return canvas_path


# ============================================================
# Excel export
# ============================================================

def _xl_header_row(ws, headers: list[str]) -> None:
    """Write bold header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=False)


def _xl_severity_fill(severity_str: str) -> "PatternFill | None":
    fills = {
        "critical": PatternFill("solid", fgColor="FF4444"),
        "high":     PatternFill("solid", fgColor="FF8C00"),
        "medium":   PatternFill("solid", fgColor="FFD700"),
        "low":      PatternFill("solid", fgColor="ADD8E6"),
    }
    return fills.get((severity_str or "").lower())


def _xl_autofit(ws, max_width: int = 60) -> None:
    """Approximate column auto-fit based on max content length."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=0,
        )
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _trunc(s: str | None, n: int) -> str:
    if not s:
        return ""
    s = s.strip()
    return s[:n - 3] + "..." if len(s) > n else s


def export_excel(
    vault_dir: Path,
    all_nmap_scans: list[tuple[str, str, dict]],      # (scan_stem, scan_display, scan_data)
    all_nessus_scans: list[tuple[str, str, dict]],    # (scan_stem, scan_display, nessus_data)
    all_burp_scans: list[tuple[str, str, dict]],      # (scan_stem, scan_display, burp_data)
    all_autorecon_scans: list[tuple[str, str, dict]] | None = None,
) -> Path | None:
    if not HAS_OPENPYXL:
        logging.error(
            "openpyxl is not installed. Run: pip install openpyxl"
        )
        return None

    wb = openpyxl.Workbook()

    # ---- Summary sheet (first tab) ----
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = [
        "IP", "Hostname", "Status", "Open Ports",
        "Highest Nessus Severity", "Critical", "High", "Medium", "Low",
        "Burp Issue Count", "Sources",
    ]
    _xl_header_row(ws_summary, summary_headers)

    hosts_dir = vault_dir / "Hosts"
    host_rows: list[list] = []

    if hosts_dir.exists():
        for hp in sorted(hosts_dir.glob("*.md")):
            try:
                text = hp.read_text(encoding="utf-8")
            except Exception:
                continue
            fm, body = read_frontmatter(text)

            ip       = fm.get("ip", "")
            hostname = ", ".join(fm.get("hostnames", [])) if fm.get("hostnames") else ""
            status   = fm.get("status", "not-started")
            sources  = ", ".join(fm.get("sources", []))

            # Open ports from ## Open Ports section
            ports_section = extract_body_section(body, "## Open Ports")
            open_port_strs = [
                l.strip().lstrip("- ").split(" — ")[0].strip("**")
                for l in ports_section.splitlines()
                if l.strip().startswith("- **")
            ]
            open_ports_str = ", ".join(open_port_strs)

            # Nessus severity counts from ## Nessus Findings section
            nessus_section = extract_body_section(body, "## Nessus Findings")
            sev_counts = {4: 0, 3: 0, 2: 0, 1: 0}
            if nessus_section:
                for line in nessus_section.splitlines():
                    m = re.match(r"####\s+\[(Critical|High|Medium|Low)\]", line)
                    if m:
                        sev_counts[severity_str_to_int(m.group(1))] += 1

            max_sev_int = fm.get("nessus_max_severity", 0)
            try:
                max_sev_int = int(max_sev_int)
            except (TypeError, ValueError):
                max_sev_int = 0
            max_sev_str = severity_int_to_str(max_sev_int) if max_sev_int > 0 else ""

            # Burp issue count
            burp_section = extract_body_section(body, "## Burp Suite Findings")
            burp_count   = burp_section.count("#### [") if burp_section else 0

            host_rows.append([
                ip, hostname, status, open_ports_str,
                max_sev_str,
                sev_counts[4], sev_counts[3], sev_counts[2], sev_counts[1],
                burp_count, sources,
            ])

    for row_data in host_rows:
        ws_summary.append(row_data)
        # Color severity cell (col 5 = Highest Nessus Severity)
        row_idx = ws_summary.max_row
        sev_cell = ws_summary.cell(row=row_idx, column=5)
        fill = _xl_severity_fill(str(sev_cell.value or ""))
        if fill:
            sev_cell.fill = fill

    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions
    _xl_autofit(ws_summary)

    # ---- Nmap sheet ----
    ws_nmap = wb.create_sheet("Nmap")
    nmap_headers = [
        "IP", "Hostname", "Port", "Protocol", "Service", "Version",
        "State", "Scripts/Banner", "Source Scan", "Status",
    ]
    _xl_header_row(ws_nmap, nmap_headers)

    # Build IP→frontmatter map for status lookup
    host_fm_by_ip: dict[str, dict] = {}
    if hosts_dir.exists():
        for hp in hosts_dir.glob("*.md"):
            try:
                fm, _ = read_frontmatter(hp.read_text(encoding="utf-8"))
                if fm.get("ip"):
                    host_fm_by_ip[fm["ip"]] = fm
            except Exception:
                pass

    for scan_stem, scan_display, scan_data in all_nmap_scans:
        for host in scan_data.get("hosts", []):
            ip       = get_primary_ipv4(host) or ""
            hostname = choose_host_display_name(host)
            h_fm     = host_fm_by_ip.get(ip, {})
            status   = h_fm.get("status", "not-started")

            open_ports = host.get("open_ports", [])
            if not open_ports:
                ws_nmap.append([ip, hostname, "", "", "", "", host.get("state", ""), "", scan_display, status])
                continue

            for p in open_ports:
                svc     = p.get("service", {})
                scripts = "; ".join(
                    f"{s['id']}: {' '.join(s['output'].split())[:80]}"
                    for s in p.get("scripts", [])[:3] if s.get("output")
                )
                version = " ".join(filter(None, [
                    svc.get("product", ""), svc.get("version", ""), svc.get("extrainfo", "")
                ])).strip()
                ws_nmap.append([
                    ip, hostname,
                    p["port"], p["protocol"],
                    svc.get("name", ""), version,
                    "open", scripts,
                    scan_display, status,
                ])

    ws_nmap.freeze_panes = "A2"
    ws_nmap.auto_filter.ref = ws_nmap.dimensions
    _xl_autofit(ws_nmap)

    # ---- Nessus sheet ----
    ws_nessus = wb.create_sheet("Nessus")
    nessus_headers = [
        "IP", "Hostname", "Port", "Protocol",
        "Plugin ID", "Plugin Name", "Severity",
        "CVSS v3", "CVSS v2", "CVE(s)",
        "Description", "Solution", "Plugin Output",
        "Source Scan",
    ]
    _xl_header_row(ws_nessus, nessus_headers)

    for scan_stem, scan_display, nessus_data in all_nessus_scans:
        for host in nessus_data.get("hosts", []):
            ip       = host.get("ip", "")
            hostname = host.get("hostname", "")
            for f in host.get("findings", []):
                sev_str  = severity_int_to_str(f["severity_int"])
                cves_str = ", ".join(f["cves"]) if f["cves"] else ""
                row = [
                    ip, hostname,
                    f["port"], f["protocol"],
                    f["plugin_id"], f["plugin_name"], sev_str,
                    f.get("cvss3_base", ""), f.get("cvss_base", ""),
                    cves_str,
                    _trunc(f.get("description"), 500),
                    _trunc(f.get("solution"), 300),
                    _trunc(f.get("plugin_output"), 300),
                    scan_display,
                ]
                ws_nessus.append(row)
                row_idx  = ws_nessus.max_row
                sev_cell = ws_nessus.cell(row=row_idx, column=7)
                fill     = _xl_severity_fill(sev_str)
                if fill:
                    sev_cell.fill = fill

    ws_nessus.freeze_panes = "A2"
    ws_nessus.auto_filter.ref = ws_nessus.dimensions
    _xl_autofit(ws_nessus)

    # ---- Burp sheet ----
    ws_burp = wb.create_sheet("Burp")
    burp_headers = [
        "Host", "Path", "Issue Name", "Severity", "Confidence",
        "Location", "Issue Detail", "Remediation", "Source Scan",
    ]
    _xl_header_row(ws_burp, burp_headers)

    for scan_stem, scan_display, burp_data in all_burp_scans:
        for host in burp_data.get("hosts", []):
            host_label = host.get("url") or host.get("ip") or ""
            for issue in host.get("issues", []):
                remediation = issue.get("remediation_detail") or issue.get("remediation_background") or ""
                row = [
                    host_label,
                    issue.get("path", ""),
                    issue.get("name", ""),
                    issue.get("severity", ""),
                    issue.get("confidence", ""),
                    issue.get("location", ""),
                    _trunc(issue.get("issue_detail"), 500),
                    _trunc(remediation, 300),
                    scan_display,
                ]
                ws_burp.append(row)
                row_idx  = ws_burp.max_row
                sev_cell = ws_burp.cell(row=row_idx, column=4)
                fill     = _xl_severity_fill(issue.get("severity", ""))
                if fill:
                    sev_cell.fill = fill

    ws_burp.freeze_panes = "A2"
    ws_burp.auto_filter.ref = ws_burp.dimensions
    _xl_autofit(ws_burp)

    # ---- AutoRecon sheet ----
    if all_autorecon_scans:
        ws_ar = wb.create_sheet("AutoRecon")
        ar_headers = [
            "Target IP", "Hostname", "Port", "Protocol", "Service",
            "Tool", "Finding Type", "Finding", "Detail", "Source Scan",
        ]
        _xl_header_row(ws_ar, ar_headers)

        ar_fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        ar_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for scan_stem, scan_display, ar_data in all_autorecon_scans:
            for target in ar_data.get("targets", []):
                t_ip = target.get("ip", "")
                t_host = target.get("hostname", "")
                for port_key, results in sorted(target.get("tool_results", {}).items()):
                    parts = port_key.split("/", 1)
                    proto = parts[0] if parts else ""
                    port = parts[1] if len(parts) > 1 else ""
                    service = ""
                    for r in results:
                        fn_m = AUTORECON_FILENAME_RE.match(r.get("filename", ""))
                        if fn_m:
                            service = fn_m.group(3)
                            break

                    for entry in results:
                        data = entry.get("data", {})
                        tool = data.get("tool", entry.get("tool", "unknown"))

                        rows_to_add: list[tuple[str, str, str, str | None]] = []

                        if tool == "dirbusting":
                            for p in data.get("interesting", [])[:30]:
                                detail = f"Status {p['status']}, Size {p['size']}"
                                fill_type = None
                                if p["status"] == 200 and re.search(
                                    r"admin|backup|config|upload|api|secret|private|\.bak|\.old",
                                    p["path"], re.IGNORECASE,
                                ):
                                    fill_type = "yellow"
                                rows_to_add.append(("Discovered Path", p["path"], detail, fill_type))
                        elif tool == "nikto":
                            for f in data.get("findings", [])[:20]:
                                osvdb = f"OSVDB-{f['osvdb']} " if f.get("osvdb") else ""
                                rows_to_add.append(("Nikto Finding", f["path"], f"{osvdb}{f['description'][:200]}", None))
                        elif tool in ("enum4linux", "smbmap", "smbclient"):
                            for s in data.get("shares", []):
                                access = s.get("permissions", s.get("access", s.get("type", "")))
                                ft = "orange" if "WRITE" in str(access).upper() else None
                                rows_to_add.append(("SMB Share", s["name"], access, ft))
                            for u in data.get("users", []):
                                rows_to_add.append(("User Account", u, "", "yellow"))
                        elif tool == "whatweb":
                            for t in data.get("technologies", []):
                                ver = f"/{t['version']}" if t.get("version") else ""
                                rows_to_add.append(("Technology", f"{t['name']}{ver}", "", None))
                        elif tool == "sslscan":
                            for wp in data.get("weak_protocols", []):
                                rows_to_add.append(("TLS Issue", f"Weak protocol: {wp}", "", "yellow"))
                            for wc in data.get("weak_ciphers", []):
                                rows_to_add.append(("TLS Issue", f"Weak cipher: {wc}", "", "yellow"))
                            if data.get("cert_expired"):
                                rows_to_add.append(("TLS Issue", "Certificate expired", data.get("cert_not_after", ""), "yellow"))
                        elif tool == "dnsrecon":
                            for r in data.get("records", [])[:20]:
                                rows_to_add.append(("DNS Record", f"{r['type']} {r['name']}", r["value"], None))
                            if data.get("zone_transfer_successful"):
                                rows_to_add.append(("DNS Record", "Zone Transfer", "SUCCESSFUL", "orange"))
                        elif tool == "snmpwalk":
                            if data.get("sys_descr"):
                                rows_to_add.append(("SNMP Info", "sysDescr", data["sys_descr"], None))
                        elif tool == "onesixtyone":
                            for cs in data.get("community_strings", []):
                                rows_to_add.append(("SNMP Community", cs["community"], cs.get("sys_descr", "")[:100], "yellow"))

                        for finding_type, finding, detail, fill_kind in rows_to_add:
                            ws_ar.append([
                                t_ip, t_host, port, proto, service,
                                tool, finding_type, finding, detail, scan_display,
                            ])
                            if fill_kind:
                                row_idx = ws_ar.max_row
                                fill = ar_fill_orange if fill_kind == "orange" else ar_fill_yellow
                                ws_ar.cell(row=row_idx, column=7).fill = fill

        ws_ar.freeze_panes = "A2"
        ws_ar.auto_filter.ref = ws_ar.dimensions
        _xl_autofit(ws_ar)

    out_path = vault_dir / "Export.xlsx"
    wb.save(str(out_path))
    logging.info(f"Excel export written: {out_path}")
    return out_path


# ============================================================
# Main
# ============================================================

def main() -> None:
    ap = argparse.ArgumentParser(
        description="mAIpper v0.8 — Pentest scan analysis and Obsidian vault generator"
    )
    ap.add_argument("--xml",        default=None,  help="Process a single Nmap XML file")
    ap.add_argument("--scans-dir",  default="scans", help="Base scans directory (default: ./scans)")
    ap.add_argument("--vault",      default="Obsidian", help="Obsidian vault output directory")
    ap.add_argument("--tool-name",  default="Nmap")
    ap.add_argument("--model",      default="qwen2.5:14b-instruct-q5_K_M")
    ap.add_argument("--ollama-url", default="http://localhost:11434")
    ap.add_argument("--no-ollama",  action="store_true", help="Skip AI analysis")
    ap.add_argument("--no-canvas",  action="store_true", help="Skip canvas generation")
    ap.add_argument("--no-nessus",  action="store_true", help="Skip Nessus scan processing")
    ap.add_argument("--no-burp",      action="store_true", help="Skip Burp Suite scan processing")
    ap.add_argument("--autorecon",    default=None, metavar="DIR",
                    help="AutoRecon results directory (contains per-target subdirs)")
    ap.add_argument("--no-autorecon", action="store_true", help="Skip AutoRecon processing")
    ap.add_argument("--excel",        action="store_true", help="Generate Export.xlsx (requires openpyxl)")
    ap.add_argument("--canvas-name",           default="Assessment Canvas.canvas")
    ap.add_argument("--canvas-cols",           type=int, default=2,
                    help="Host cards per row within each subnet group (default: 2)")
    ap.add_argument("--canvas-groups-per-row", type=int, default=3,
                    help="Subnet groups per canvas row (default: 3)")
    ap.add_argument("--temperature",     type=float, default=0.15,
                    help="Ollama sampling temperature (0.0–1.0, default: 0.15)")
    ap.add_argument("--skip-validation", action="store_true",
                    help="Skip post-processing AI output validation (faster, no hallucination checks)")
    ap.add_argument("-v", "--verbose", action="count", default=1,
                    help="Increase verbosity: -v = INFO, -vv = DEBUG")

    args = ap.parse_args()

    level = logging.WARNING
    if args.verbose >= 1: level = logging.INFO
    if args.verbose >= 2: level = logging.DEBUG
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")
    logging.debug(f"Args: {vars(args)}")

    vault_dir = Path(args.vault).resolve()
    base      = Path(args.scans_dir).resolve()

    scan_host_map: dict[str, list[str]]   = {}
    all_analyses:  list[tuple[str, str]]  = []

    # ------------------------------------------------------------------ #
    # Nmap                                                                 #
    # ------------------------------------------------------------------ #
    all_nmap_scans: list[tuple[str, str, dict]] = []

    xml_files: list[Path] = []
    if args.xml:
        xml_path = Path(args.xml).resolve()
        if not xml_path.exists():
            logging.error(f"File not found: {xml_path}")
            return
        xml_files = [xml_path]
    else:
        nmap_dir = base / "nmap"
        if not nmap_dir.exists():
            nmap_dir = base / "Nmap"
        if nmap_dir.exists():
            xml_files = sorted(nmap_dir.glob("*.xml"))
        else:
            logging.info("No scans/nmap/ directory found; skipping Nmap processing")

    if xml_files:
        logging.info(f"Processing {len(xml_files)} Nmap file(s)")
    for xml_path in xml_files:
        logging.info(f"Nmap → {xml_path.name}")
        scan_data = parse_nmap_xml(xml_path)

        analysis: str | None = None
        nmap_warnings: list[str] = []
        if not args.no_ollama:
            try:
                raw_analysis = ollama_chat(
                    args.ollama_url, args.model,
                    build_ollama_prompt(scan_data),
                    args.temperature,
                )
                if not args.skip_validation:
                    analysis, nmap_warnings = validate_ai_output(
                        raw_analysis, scan_data, "nmap"
                    )
                    for w in nmap_warnings:
                        logging.warning(f"[Nmap Validation] {w}")
                else:
                    analysis = raw_analysis
            except Exception as exc:
                logging.warning(f"Ollama failed ({xml_path.name}): {exc}")
                analysis = f"_Ollama failed: {exc}_"
        else:
            logging.info("Skipping Ollama (--no-ollama)")

        result = create_obsidian_vault(
            vault_dir,
            xml_path.stem,
            scan_data,
            args.tool_name,
            analysis,
            None if args.no_ollama else args.model,
            validation_warnings=nmap_warnings or None,
        )

        scan_host_map[result["scan_stem"]] = result["host_stems"]
        if analysis and not args.no_ollama:
            all_analyses.append((result["scan_display"], analysis))
        all_nmap_scans.append((result["scan_stem"], result["scan_display"], scan_data))

    # ------------------------------------------------------------------ #
    # Nessus                                                               #
    # ------------------------------------------------------------------ #
    all_nessus_scans: list[tuple[str, str, dict]] = []

    if not args.no_nessus:
        nessus_dir = base / "nessus"
        if not nessus_dir.exists():
            nessus_dir = base / "Nessus"
        nessus_files: list[Path] = sorted(nessus_dir.glob("*.nessus")) if nessus_dir.exists() else []

        if nessus_files:
            logging.info(f"Processing {len(nessus_files)} Nessus file(s)")
        for nessus_path in nessus_files:
            logging.info(f"Nessus → {nessus_path.name}")
            nessus_data = parse_nessus_xml(nessus_path)

            analysis = None
            nessus_warnings: list[str] = []
            if not args.no_ollama:
                try:
                    # Pass 1: fact extraction only
                    facts: str | None = None
                    try:
                        fact_prompt = _build_nessus_fact_extraction_prompt(nessus_data)
                        facts = ollama_chat(
                            args.ollama_url, args.model, fact_prompt, args.temperature
                        )
                        logging.info(
                            f"Nessus Pass 1 complete ({nessus_path.name}): "
                            f"{len(facts)} chars extracted"
                        )
                    except Exception as exc:
                        logging.warning(
                            f"Nessus Pass 1 failed ({nessus_path.name}): {exc}; "
                            "falling back to single-pass"
                        )

                    # Pass 2: analysis (optionally grounded in Pass 1 facts)
                    p2_prompt = build_nessus_ollama_prompt(nessus_data)
                    if facts:
                        p2_prompt = (
                            "The following facts have been extracted directly from the scan "
                            "data. Base your analysis ONLY on these confirmed facts:\n\n"
                            + facts + "\n\n" + p2_prompt
                        )
                    raw_analysis = ollama_chat(
                        args.ollama_url, args.model, p2_prompt, args.temperature
                    )
                    if not args.skip_validation:
                        analysis, nessus_warnings = validate_ai_output(
                            raw_analysis, nessus_data, "nessus"
                        )
                        for w in nessus_warnings:
                            logging.warning(f"[Nessus Validation] {w}")
                    else:
                        analysis = raw_analysis
                except Exception as exc:
                    logging.warning(f"Ollama failed ({nessus_path.name}): {exc}")
                    analysis = f"_Ollama failed: {exc}_"

            result = create_nessus_vault(
                vault_dir,
                nessus_path.stem,
                nessus_data,
                analysis,
                None if args.no_ollama else args.model,
                validation_warnings=nessus_warnings or None,
            )

            scan_host_map[result["scan_stem"]] = result["host_stems"]
            if analysis and not args.no_ollama:
                all_analyses.append((result["scan_display"], analysis))
            all_nessus_scans.append((result["scan_stem"], result["scan_display"], nessus_data))
    else:
        logging.info("Skipping Nessus (--no-nessus)")

    # ------------------------------------------------------------------ #
    # Burp Suite                                                           #
    # ------------------------------------------------------------------ #
    all_burp_scans: list[tuple[str, str, dict]] = []

    if not args.no_burp:
        burp_dir = base / "burp"
        if not burp_dir.exists():
            burp_dir = base / "Burp"
        burp_files: list[Path] = sorted(burp_dir.glob("*.xml")) if burp_dir.exists() else []

        if burp_files:
            logging.info(f"Processing {len(burp_files)} Burp file(s)")
        for burp_path in burp_files:
            logging.info(f"Burp → {burp_path.name}")
            # Distinguish Burp XML from Nmap XML by checking root tag
            try:
                root_tag = ET.parse(burp_path).getroot().tag
            except Exception as exc:
                logging.warning(f"Failed to parse {burp_path.name}: {exc}")
                continue
            if root_tag != "issues":
                logging.debug(f"Skipping non-Burp XML: {burp_path.name} (root tag: {root_tag})")
                continue

            burp_data = parse_burp_xml(burp_path)

            analysis = None
            burp_warnings: list[str] = []
            if not args.no_ollama:
                try:
                    raw_analysis = ollama_chat(
                        args.ollama_url, args.model,
                        build_burp_ollama_prompt(burp_data),
                        args.temperature,
                    )
                    if not args.skip_validation:
                        analysis, burp_warnings = validate_ai_output(
                            raw_analysis, burp_data, "burp"
                        )
                        for w in burp_warnings:
                            logging.warning(f"[Burp Validation] {w}")
                    else:
                        analysis = raw_analysis
                except Exception as exc:
                    logging.warning(f"Ollama failed ({burp_path.name}): {exc}")
                    analysis = f"_Ollama failed: {exc}_"

            result = create_burp_vault(
                vault_dir,
                burp_path.stem,
                burp_data,
                analysis,
                None if args.no_ollama else args.model,
                validation_warnings=burp_warnings or None,
            )

            scan_host_map[result["scan_stem"]] = result["host_stems"]
            if analysis and not args.no_ollama:
                all_analyses.append((result["scan_display"], analysis))
            all_burp_scans.append((result["scan_stem"], result["scan_display"], burp_data))
    else:
        logging.info("Skipping Burp Suite (--no-burp)")

    # ------------------------------------------------------------------ #
    # AutoRecon                                                            #
    # ------------------------------------------------------------------ #
    all_autorecon_scans: list[tuple[str, str, dict]] = []

    if not args.no_autorecon:
        autorecon_dir: Path | None = None
        if args.autorecon:
            autorecon_dir = Path(args.autorecon).resolve()
            if not autorecon_dir.exists():
                logging.error(f"AutoRecon directory not found: {autorecon_dir}")
                autorecon_dir = None
        else:
            candidate = base / "autorecon"
            if not candidate.exists():
                candidate = base / "AutoRecon"
            if candidate.exists():
                autorecon_dir = candidate

        if autorecon_dir:
            logging.info(f"Processing AutoRecon results: {autorecon_dir}")

            # Layer 1: Ingest nmap XMLs into existing Nmap pipeline
            for target_subdir in sorted(autorecon_dir.iterdir()):
                if not target_subdir.is_dir():
                    continue
                xml_dir = target_subdir / "scans" / "xml"
                if not xml_dir.exists():
                    continue
                ar_xml_files = sorted(xml_dir.glob("*.xml"))
                for xml_file in ar_xml_files:
                    logging.info(f"AutoRecon Layer 1 (Nmap) → {target_subdir.name}/{xml_file.name}")
                    try:
                        scan_data = parse_nmap_xml(xml_file)
                    except Exception as exc:
                        logging.warning(f"Failed to parse AutoRecon nmap XML {xml_file}: {exc}")
                        continue

                    nmap_analysis: str | None = None
                    nmap_warnings_ar: list[str] = []
                    if not args.no_ollama:
                        try:
                            raw = ollama_chat(
                                args.ollama_url, args.model,
                                build_ollama_prompt(scan_data), args.temperature,
                            )
                            if not args.skip_validation:
                                nmap_analysis, nmap_warnings_ar = validate_ai_output(
                                    raw, scan_data, "nmap"
                                )
                                for w in nmap_warnings_ar:
                                    logging.warning(f"[AutoRecon Nmap Validation] {w}")
                            else:
                                nmap_analysis = raw
                        except Exception as exc:
                            logging.warning(f"Ollama failed (AutoRecon Nmap {xml_file.name}): {exc}")
                            nmap_analysis = f"_Ollama failed: {exc}_"

                    result = create_obsidian_vault(
                        vault_dir,
                        xml_file.stem,
                        scan_data,
                        f"Nmap (AutoRecon {target_subdir.name})",
                        nmap_analysis,
                        None if args.no_ollama else args.model,
                        validation_warnings=nmap_warnings_ar or None,
                    )
                    scan_host_map[result["scan_stem"]] = result["host_stems"]
                    if nmap_analysis and not args.no_ollama:
                        all_analyses.append((result["scan_display"], nmap_analysis))
                    all_nmap_scans.append((result["scan_stem"], result["scan_display"], scan_data))

            # Layer 2: Parse tool outputs and run AutoRecon-specific analysis
            autorecon_data = parse_autorecon_results(autorecon_dir)

            for target in autorecon_data.get("targets", []):
                target_name = target.get("target", "unknown")
                logging.info(f"AutoRecon Layer 2 → {target_name}")

                ar_analysis: str | None = None
                ar_warnings: list[str] = []
                if not args.no_ollama:
                    try:
                        # Pass 1: fact extraction
                        facts: str | None = None
                        try:
                            fact_prompt = _build_autorecon_fact_extraction_prompt(target)
                            facts = ollama_chat(
                                args.ollama_url, args.model, fact_prompt, args.temperature,
                            )
                            logging.info(
                                f"AutoRecon Pass 1 complete ({target_name}): {len(facts)} chars"
                            )
                        except Exception as exc:
                            logging.warning(
                                f"AutoRecon Pass 1 failed ({target_name}): {exc}; "
                                "falling back to single-pass"
                            )

                        # Pass 2: analysis (grounded in Pass 1 facts if available)
                        p2_prompt = build_autorecon_ollama_prompt(target)
                        if facts:
                            p2_prompt = (
                                "The following facts have been extracted directly from the "
                                "enumeration data. Base your analysis ONLY on these confirmed "
                                "facts:\n\n" + facts + "\n\n" + p2_prompt
                            )
                        raw_analysis = ollama_chat(
                            args.ollama_url, args.model, p2_prompt, args.temperature,
                        )
                        if not args.skip_validation:
                            ar_analysis, ar_warnings = validate_ai_output(
                                raw_analysis, {"hosts": [target]}, "autorecon"
                            )
                            for w in ar_warnings:
                                logging.warning(f"[AutoRecon Validation] {w}")
                        else:
                            ar_analysis = raw_analysis
                    except Exception as exc:
                        logging.warning(f"Ollama failed (AutoRecon {target_name}): {exc}")
                        ar_analysis = f"_Ollama failed: {exc}_"

                ar_scan_data = {
                    "source_file": str(autorecon_dir),
                    "parsed_at": autorecon_data["parsed_at"],
                    "report_name": target_name,
                    "targets": [target],
                    "hosts": [target],
                }

                result = create_autorecon_vault(
                    vault_dir,
                    target_name,
                    ar_scan_data,
                    ar_analysis,
                    None if args.no_ollama else args.model,
                    validation_warnings=ar_warnings or None,
                )
                scan_host_map[result["scan_stem"]] = result["host_stems"]
                if ar_analysis and not args.no_ollama:
                    all_analyses.append((result["scan_display"], ar_analysis))
                all_autorecon_scans.append((result["scan_stem"], result["scan_display"], ar_scan_data))
    else:
        logging.info("Skipping AutoRecon (--no-autorecon)")

    # ------------------------------------------------------------------ #
    # Canvas                                                               #
    # ------------------------------------------------------------------ #
    if not args.no_canvas:
        # Build Priority Targets text: ask Ollama if we have any analyses,
        # otherwise fall back to a static severity-sorted list.
        priority_targets_text: str | None = None
        if not args.no_ollama and all_analyses:
            try:
                pt_prompt = build_priority_targets_prompt(vault_dir, all_analyses)
                raw_pt    = ollama_chat(
                    args.ollama_url, args.model, pt_prompt, args.temperature
                )
                if not args.skip_validation:
                    # Build a combined source for cross-reference (all Nmap hosts)
                    combined_hosts: list[dict] = []
                    for _, _, sd in all_nmap_scans:
                        combined_hosts.extend(sd.get("hosts", []))
                    raw_pt, pt_warnings = validate_ai_output(
                        raw_pt, {"hosts": combined_hosts}, "nmap"
                    )
                    for w in pt_warnings:
                        logging.warning(f"[Priority Targets Validation] {w}")
                # Prepend the standard header so the canvas node is self-labelled
                priority_targets_text = "## Priority Targets\n\n" + raw_pt.strip()
            except Exception as exc:
                logging.warning(f"Ollama priority targets failed: {exc}")
                priority_targets_text = _build_priority_targets_fallback(vault_dir)
        else:
            priority_targets_text = _build_priority_targets_fallback(vault_dir)

        build_canvas(
            vault_dir,
            args.canvas_name,
            scan_host_map,
            all_analyses,
            canvas_cols=args.canvas_cols,
            max_groups_per_row=args.canvas_groups_per_row,
            priority_targets_text=priority_targets_text,
        )
    else:
        logging.info("Skipping canvas (--no-canvas)")

    # ------------------------------------------------------------------ #
    # Excel export                                                         #
    # ------------------------------------------------------------------ #
    if args.excel:
        out = export_excel(vault_dir, all_nmap_scans, all_nessus_scans, all_burp_scans,
                           all_autorecon_scans or None)
        if out:
            logging.info(f"Excel export: {out}")
        else:
            logging.error("Excel export failed — is openpyxl installed?")

    logging.info("Done.")


if __name__ == "__main__":
    main()
