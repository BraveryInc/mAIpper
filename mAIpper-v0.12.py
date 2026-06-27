#!/usr/bin/env python3

"""
mAIpper v0.12 - Pentest Tool Analysis & Obsidian Export Tool

Changes from v0.11:
  - RAG (Retrieval Augmented Generation): index your PDF security books
    and a local HackTricks clone, then get cited references in all LLM
    analysis; --build-index to create/update, --no-rag to disable
  - Explicit index building: auto_build defaults to false so no surprise
    overhead during testing; run --build-index when ready
  - /rag interactive command to check index status
  - docs/ directory created by --init for PDF reference books

Changes from v0.10:
  - Users Canvas: new canvas (Users Canvas.canvas) visualizing credential
    → host → service relationships; grouped by source host; nodes show
    per-credential notes from Credentials.md; dynamically sized boxes;
    edge labels include operator notes context
  - Credential annotations: Loot/Credentials.md now has per-host operator
    notes sections and per-credential Notes column in the table, both
    preserved on re-run; annotations feed into Users Canvas
  - Smarter Misc analysis: auto-detects tool type (nikto, gobuster,
    linpeas, etc.) from filename and content signatures; scales LLM effort
    to content value — real tool output gets full analysis, random notes
    get minimal treatment (key facts only, no deep analysis)
  - Eat Me page: vault-root drop zone where assessors paste anything
    (arp tables, host lists, tool output, notes); processed on next run
    into host notes with next-step scan commands; archived to timestamped
    scan note; page resets for next use; works in batch and interactive
  - Campaign Targets note: auto-generated Hosts/Campaign Targets.md with
    copy-paste-ready lists of IPs, subnets, hostnames, and domains in
    fenced code blocks; extracts from all vault data + loot + misc sources
  - Vault change detection in interactive mode: watches host notes,
    Eat Me, and Credentials.md for operator edits; auto-triggers deep
    dives, operator notes reload, Users Canvas rebuild, and Campaign
    Targets update when vault files change in Obsidian
  - --no-users-canvas flag to skip Users Canvas generation

Changes from v0.9:
  - --init flag: creates scans/ directory tree with empty subdirectories
    for all supported parsers (nmap, nessus, burp, autorecon, loot, misc);
    auto-triggers when scans/ doesn't exist on a normal run
  - Interactive mode (-i / --interactive): watch for file changes, chat
    with the LLM about the assessment, drop credentials and host data
    inline; configurable poll interval via --watch-interval (default 30s)
  - Pre-LLM empty host filtering: Nmap hosts with no open ports, Nessus
    hosts with only informational (severity 0) findings, and AutoRecon
    targets with no scan data are now filtered out before LLM analysis
    and vault writing
  - Loot restructure: Loot/ is now a top-level output directory with
    Credentials.md and Hashes.md pages; host notes get lightweight
    ## Loot sections with summary stats and links to centralized pages;
    file listings stay inline in host notes
  - Misc input folder: scans/misc/ accepts arbitrary tool text output;
    each file gets Scans/<filename> - Misc.md with LLM interpretation;
    --misc DIR / --no-misc flags; host association via subdirectory or
    filename prefix
  - Deep dive checkboxes: Investigate checkboxes next to ports, Nessus
    findings, and Burp issues; check a box, next run generates detailed
    analysis in a collapsible callout under ## Deep Dives

Changes from v0.8:
  - Loot ingestion: assessors drop text files into scans/loot/ (or
    --loot DIR); Python pre-parses credentials, hashes, file listings
    before LLM analysis; associates loot with hosts via subdirectory
    name or filename prefix; unassociated loot goes to campaign level
  - ## Loot section in host notes with credential/hash tables
  - Operator Notes feedback loop: ## Operator Notes content is now read
    on re-run and injected into ALL prompt builders as [OPERATOR] context,
    enabling human-in-the-loop AI refinement
  - --loot DIR / --no-loot flags; auto-discovers scans/loot/
  - Loot sheet in Excel export
  - Priority Targets ranking now considers loot (confirmed creds rank highest)

Changes from v0.7:
  - AutoRecon integration: ingests AutoRecon results directories
    (per-target subdirs with scans/xml/, scans/tcp*/, scans/udp*/)
  - Layer 1: feeds AutoRecon's nmap XMLs through existing Nmap pipeline
  - Layer 2: Python extractors parse structured facts from tool outputs
    (gobuster, nikto, enum4linux, smbmap, smbclient, whatweb, snmpwalk,
    onesixtyone, sslscan, dnsrecon, curl) before LLM analysis
  - Two-pass LLM analysis for AutoRecon (fact extraction → analysis),
    matching the Nessus approach for accuracy
  - ## AutoRecon Enumeration section in host notes with structured tables
  - Scans/<target> - AutoRecon.md scan notes with collapsible manual
    commands and commands log
  - --autorecon DIR / --no-autorecon flags; auto-discovers scans/autorecon/
  - AutoRecon sheet in Excel export
  - Canvas Campaign Overview updated with AutoRecon stats

Changes from v0.6:
  - Nessus parser: ingests .nessus XML files from scans/nessus/; adds
    ## Nessus Findings sections to host notes (severity-sorted); creates
    Scans/<name> - Nessus.md; AI analysis focused on CVEs and exploitation
    priority; nessus_max_severity stored in frontmatter for canvas coloring
  - Burp Suite parser: ingests Burp Pro Scanner XML from scans/burp/; adds
    ## Burp Suite Findings sections to host notes; creates
    Scans/<name> - Burp.md; AI analysis focused on web attack surface
  - Excel export (--excel): generates <vault>/Export.xlsx with Summary,
    Nmap, Nessus, and Burp sheets; severity color-coding, filters, frozen
    headers, auto-fit columns
  - --no-nessus / --no-burp flags to skip those parsers
  - Canvas Campaign Overview and Next Steps updated with Nessus/Burp data
  - Nmap re-runs preserve existing Nessus/Burp sections in host notes
  - Canvas "Priority Targets" text node: AI-ranked list of highest-priority
    hosts (by CVSS, CVEs, service exposure, Burp findings); positioned between
    Campaign Overview and subnet groups; edges Overview→PriorityTargets and
    PriorityTargets→each subnet group; falls back to static severity sort when
    Ollama is skipped; stable node ID "priority-targets"

Requirements:
  pip install requests
  pip install openpyxl   # only needed for --excel

Author: Zachary Levine
"""

from __future__ import annotations

import argparse
import configparser
import datetime as dt
import hashlib
import json
import logging
import math
import re
import sys
import threading
import time
from pathlib import Path
import xml.etree.ElementTree as ET

import base64
import heapq
import sqlite3
import struct

import requests

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pypdf
    HAS_PYPDF = True
except ImportError:
    try:
        import PyPDF2 as pypdf
        HAS_PYPDF = True
    except ImportError:
        HAS_PYPDF = False


# ============================================================
# Constants
# ============================================================

IPV4_RE          = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$")
CVE_RE           = re.compile(r"\bCVE-\d{4}-\d+\b", re.IGNORECASE)
PORT_IN_TEXT_RE  = re.compile(r"(?:port\s+(\d{1,5})|(\d{1,5})/(?:tcp|udp))", re.IGNORECASE)
IP_IN_TEXT_RE    = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
FQDN_IN_TEXT_RE  = re.compile(
    r"\b(?:[a-zA-Z0-9](?:[a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,6}\b"
)
AUTORECON_FILENAME_RE = re.compile(
    r"^(tcp|udp)_(\d+)_(\w+)_(.+)\.(txt|html|xml|json|log)$"
)
AUTORECON_PORT_DIR_RE = re.compile(r"^(tcp|udp)_?(\d+)$")

LOOT_IP_PREFIX_RE = re.compile(r"^(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})[_\-](.+)$")
LOOT_FQDN_PREFIX_RE = re.compile(
    r"^((?:[a-zA-Z0-9](?:[a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,6})[_\-](.+)$"
)
CRED_KV_RE = re.compile(
    r"(?:user(?:name)?|login)\s*[=:]\s*(\S+)\s+(?:pass(?:word)?)\s*[=:]\s*(\S+)",
    re.IGNORECASE | re.MULTILINE,
)
HASHCAT_FORMAT_RE = re.compile(r"^([^\s:]{1,64}):([a-fA-F0-9]{32,128}|\$\S+)$", re.MULTILINE)
NTLM_HASH_RE     = re.compile(r"\b[a-fA-F0-9]{32}\b")
BCRYPT_HASH_RE   = re.compile(r"\$2[aby]?\$\d{1,2}\$[./A-Za-z0-9]{53}")
SHA256_HASH_RE   = re.compile(r"\b[a-fA-F0-9]{64}\b")
FTP_LISTING_RE   = re.compile(
    r"^([d\-rwxsStT]{10})\s+\d+\s+\S+\s+\S+\s+(\d+)\s+"
    r"(\w+\s+\d+\s+[\d:]+)\s+(.+)$",
    re.MULTILINE,
)
SMB_LISTING_RE = re.compile(
    r"^\s+(\S.+?)\s{2,}([ADHR]*)\s+(\d+)\s+(\w.+)$", re.MULTILINE
)
BINARY_EXTENSIONS = frozenset({
    ".exe", ".dll", ".bin", ".so", ".o", ".a", ".lib",
    ".zip", ".gz", ".tar", ".rar", ".7z",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico",
    ".pdf", ".doc", ".docx", ".xls", ".xlsx",
    ".pyc", ".class", ".db", ".sqlite",
})
LOOT_MAX_FILE_SIZE = 512 * 1024

STATUS_COLORS: dict[str, str | None] = {
    "not-started": None,
    "in-progress":  "3",   # yellow
    "done":         "4",   # green
    "exploited":    "1",   # red
    "blocked":      "6",   # purple
}

# Canvas colors for Nessus severity (applied when status == not-started)
NESSUS_SEVERITY_CANVAS_COLORS: dict[int, str] = {
    4: "1",   # Critical → red
    3: "5",   # High     → orange
    2: "3",   # Medium   → yellow
}

OPERATOR_NOTES_SENTINEL = "## Operator Notes"
OPERATOR_NOTES_HINT = "_Add your own findings, observations, and next steps below._"

DEEP_DIVE_SECTION = "## Deep Dives"
CROSS_SOURCE_SECTION = "## Cross-Source Analysis"
DEEP_DIVE_CHECKBOX_RE = re.compile(
    r"^\s*- \[( |x|/)\] Investigate: (.+)$", re.MULTILINE
)
ANALYZE_PENDING_RE = re.compile(
    r"^\s*- \[x\] Analyze: (.+)$", re.MULTILINE
)
DEEP_DIVE_PENDING_RE = re.compile(
    r"^\s*- \[x\] Investigate: (.+)$", re.MULTILINE
)

# RAG constants
RAG_INDEX_DB_FILENAME = ".maipper_rag_index.db"
RAG_INDEX_VERSION = 2
RAG_DEFAULT_EMBEDDING_MODEL = "nomic-embed-text"
RAG_DEFAULT_CHUNK_SIZE = 500
RAG_DEFAULT_CHUNK_OVERLAP = 50
RAG_DEFAULT_MAX_CHUNKS = 5
RAG_MIN_CHUNK_WORDS = 50
RAG_MD_HEADER_RE = re.compile(r"^(#{1,3})\s+(.+)$", re.MULTILINE)

# Module-level RAG state
_RAG_INDEX: dict | None = None
_RAG_BUILDER = None  # _RagIndexBuilder | None
_RAG_OLLAMA_URL: str = ""
_RAG_EMBEDDING_MODEL: str = RAG_DEFAULT_EMBEDDING_MODEL

# Canonical section order in host notes
BODY_SECTION_ORDER = [
    "## Open Ports",
    "## Nessus Findings",
    "## Burp Suite Findings",
    "## AutoRecon Enumeration",
    "## Loot",
    DEEP_DIVE_SECTION,
    CROSS_SOURCE_SECTION,
    "## Scan References",
    OPERATOR_NOTES_SENTINEL,
]

CARD_W          = 360
CARD_H          = 220
CARD_GAP_X      = 60
CARD_GAP_Y      = 60
GROUP_PAD       = 60
GROUP_LABEL_H   = 40
SCAN_CARD_W     = 360
SCAN_CARD_H     = 160
OVERVIEW_W      = 720
OVERVIEW_H      = 220
NEXT_STEPS_W    = 640
NEXT_STEPS_H    = 420
GROUP_GAP       = 120
ROW_GAP         = 120

# Users Canvas layout
UC_USER_W       = 300
UC_USER_H       = 80
UC_HOST_W       = 300
UC_HOST_H       = 80
UC_GAP_Y        = 40
UC_GROUP_PAD    = 50
UC_GROUP_LABEL_H = 40
UC_COL_GAP      = 500

# Misc tool detection signatures: tool_name → list of regex patterns
_MISC_TOOL_SIGNATURES: dict[str, list[str]] = {
    "nikto":         [r"^\+\s*Target IP:", r"Nikto", r"^\+\s*\d+ host\(s\) tested"],
    "gobuster":      [r"Gobuster v", r"^/\S+\s+\(Status:\s*\d+\)", r"==============="],
    "dirb":          [r"DIRB v", r"---- Scanning URL:", r"^\+\s*http"],
    "dirsearch":     [r"dirsearch", r"Extensions:", r"^\d{3}\s+\d+\w*\s+"],
    "feroxbuster":   [r"feroxbuster", r"Target Url", r"WLD"],
    "ffuf":          [r"::\s*Method\s*:", r"FUZZ", r"ffuf"],
    "wpscan":        [r"WPScan", r"WordPress", r"^\[\+\]\s*URL:"],
    "linpeas":       [r"linpeas", r"Linux Privesc", r"╔═"],
    "winpeas":       [r"winPEAS", r"Windows Privesc", r"════"],
    "enum4linux":    [r"enum4linux", r"Starting enum4linux", r"Target Information"],
    "smbmap":        [r"SMBMap", r"\[\+\]\s*IP:", r"READ ONLY|READ, WRITE"],
    "smbclient":     [r"smb:\s*\\\\>", r"Sharename\s+Type"],
    "nmap":          [r"Nmap scan report", r"Starting Nmap", r"PORT\s+STATE\s+SERVICE"],
    "masscan":       [r"Masscan", r"Discovered open port"],
    "hydra":         [r"\[DATA\]", r"Hydra.*starting", r"\[\d+\]\["],
    "crackmapexec":  [r"crackmapexec|CME|nxc|NetExec", r"SMB\s+\d+\.\d+"],
    "bloodhound":    [r"bloodhound|SharpHound", r"Resolved Collection Methods"],
    "responder":     [r"Responder", r"\[SMB\]|\[HTTP\]|\[LDAP\]"],
    "impacket":      [r"Impacket v", r"secretsdump|GetNPUsers|GetUserSPNs"],
    "kerbrute":      [r"kerbrute", r"VALID USERNAME"],
    "john":          [r"John the Ripper", r"Loaded \d+ password"],
    "hashcat":       [r"hashcat", r"Session\.+:", r"Status\.+:"],
    "whatweb":       [r"WhatWeb", r"http.*\[200\]"],
    "testssl":       [r"testssl", r"Testing protocols", r"TLS\s+\d"],
    "nuclei":        [r"\[nuclei\]", r"Templates loaded", r"\[\w+\]\s*\["],
    "snmpwalk":      [r"iso\.\d+\.\d+\.\d+", r"SNMPv2-MIB", r"STRING:"],
    "ldapsearch":    [r"^dn:\s+", r"ldapsearch", r"objectClass:"],
    "rpcclient":     [r"rpcclient", r"Domain Name:", r"querydispinfo"],
}


def _detect_misc_tool(filename: str, content: str) -> tuple[str, str]:
    """Detect the tool that produced misc output. Returns (tool_name, analysis_level).

    analysis_level is one of: "full", "standard", "minimal".
    """
    fname_lower = filename.lower()
    preview = content[:8000]

    # Check filename hints
    for tool in _MISC_TOOL_SIGNATURES:
        if tool in fname_lower:
            return tool, "full"

    # Check content signatures
    best_tool = ""
    best_score = 0
    for tool, patterns in _MISC_TOOL_SIGNATURES.items():
        score = sum(
            1 for p in patterns
            if re.search(p, preview, re.MULTILINE | re.IGNORECASE)
        )
        if score > best_score:
            best_score = score
            best_tool = tool

    if best_score >= 2:
        return best_tool, "full"
    if best_score == 1:
        return best_tool, "standard"

    # Heuristic: short prose-heavy files are probably notes/todo
    lines = [l for l in content.splitlines()[:60] if l.strip()]
    if not lines:
        return "unknown", "minimal"

    total = len(lines)
    prose = sum(
        1 for l in lines
        if len(l.strip()) > 30
        and not l.strip().startswith(("-", "#", "|", "+", "*", "/", "\\", "["))
        and ":" not in l[:20]
    )
    if total <= 15 and prose / max(total, 1) > 0.5:
        return "notes", "minimal"
    if total <= 30 and prose / max(total, 1) > 0.7:
        return "notes", "minimal"

    return "unknown", "standard"


# ============================================================
# Basic helpers
# ============================================================

def safe_filename(s: str) -> str:
    for ch in '<>:"/\\|?*\n\r\t':
        s = s.replace(ch, "_")
    return s.strip().strip(".")


def stable_id(key: str) -> str:
    return hashlib.md5(key.encode()).hexdigest()[:12]


def is_ipv4(value: str) -> bool:
    return bool(value and IPV4_RE.match(value.strip()))


def is_probable_fqdn(value: str) -> bool:
    v = (value or "").strip().rstrip(".")
    return bool(v and "." in v and not is_ipv4(v))


def ensure_md_suffix(name: str) -> str:
    return name if name.lower().endswith(".md") else f"{name}.md"


def get_subnet_label(ip: str) -> str:
    if not is_ipv4(ip):
        return "unknown"
    parts = ip.split(".")
    return f"{parts[0]}.{parts[1]}.{parts[2]}.0/24"


def severity_int_to_str(n: int) -> str:
    return {0: "Info", 1: "Low", 2: "Medium", 3: "High", 4: "Critical"}.get(n, "Unknown")


def severity_str_to_int(s: str) -> int:
    return {"information": 0, "info": 0, "low": 1, "medium": 2, "high": 3, "critical": 4}.get(
        (s or "").lower(), 0
    )


# ============================================================
# Frontmatter helpers
# ============================================================

def _fm_encode(v: object) -> str:
    if isinstance(v, list):
        return json.dumps(v, ensure_ascii=False)
    if v is None:
        return "null"
    return str(v)


def write_frontmatter(fm: dict) -> str:
    lines = ["---"]
    for k, v in fm.items():
        lines.append(f"{k}: {_fm_encode(v)}")
    lines.append("---")
    return "\n".join(lines) + "\n"


def read_frontmatter(text: str) -> tuple[dict, str]:
    if not text.startswith("---\n"):
        return {}, text
    end = text.find("\n---\n", 4)
    if end == -1:
        return {}, text
    fm_text = text[4:end]
    body = text[end + 5:]
    fm: dict = {}
    for line in fm_text.splitlines():
        if ": " not in line and not line.endswith(":"):
            continue
        key, _, raw = line.partition(": ")
        key = key.strip()
        raw = raw.strip()
        if not key:
            continue
        try:
            fm[key] = json.loads(raw)
        except (json.JSONDecodeError, ValueError):
            fm[key] = raw if raw else None
    return fm, body


def extract_operator_notes(body: str) -> str:
    idx = body.find(OPERATOR_NOTES_SENTINEL)
    if idx == -1:
        return ""
    after = body[idx + len(OPERATOR_NOTES_SENTINEL):]
    lines = after.splitlines()
    content_lines: list[str] = []
    past_hint = False
    for line in lines:
        if not past_hint and line.strip() in ("", OPERATOR_NOTES_HINT):
            continue
        past_hint = True
        content_lines.append(line)
    while content_lines and not content_lines[-1].strip():
        content_lines.pop()
    return "\n".join(content_lines)


def extract_body_section(body: str, section_header: str) -> str:
    """Extract content of a named ## section (without its header line)."""
    pattern = rf"(?:^|\n){re.escape(section_header)}\n(.*?)(?=\n## |\Z)"
    m = re.search(pattern, body, re.DOTALL)
    return m.group(1).strip() if m else ""


# ============================================================
# Port hints & tags
# ============================================================

def get_tags_from_ports(open_ports: list) -> list[str]:
    tags: set[str] = set()
    port_nums = {p.get("port", 0) for p in open_ports}
    svc_names = {(p.get("service", {}).get("name", "") or "").lower() for p in open_ports}

    if port_nums & {80, 81, 443, 8000, 8080, 8443} or svc_names & {"http", "https", "http-proxy"}:
        tags.add("web")
    if 445 in port_nums or svc_names & {"microsoft-ds", "smb"}:
        tags.add("smb")
    if port_nums & {88, 464} or any("kerberos" in s for s in svc_names):
        tags.add("kerberos")
    if port_nums & {389, 636, 3268, 3269} or svc_names & {"ldap", "ldaps", "globalcatldap", "globalcatldapssl"}:
        tags.add("ldap")
    if 135 in port_nums or "msrpc" in svc_names:
        tags.add("rpc")
    if port_nums & {1433, 1434} or any("mssql" in s for s in svc_names):
        tags.add("mssql")
    if 3306 in port_nums or "mysql" in svc_names:
        tags.add("mysql")
    if 5432 in port_nums or any("postgres" in s for s in svc_names):
        tags.add("postgres")
    if 22 in port_nums or "ssh" in svc_names:
        tags.add("ssh")
    if 3389 in port_nums or "ms-wbt-server" in svc_names:
        tags.add("rdp")
    if port_nums & {5985, 5986} or any("winrm" in s for s in svc_names):
        tags.add("winrm")
    if 21 in port_nums or "ftp" in svc_names:
        tags.add("ftp")
    if 53 in port_nums or "domain" in svc_names:
        tags.add("dns")
    if port_nums & {161, 162} or "snmp" in svc_names:
        tags.add("snmp")
    if port_nums & {25, 465, 587} or svc_names & {"smtp", "smtps", "submission"}:
        tags.add("smtp")
    if 2049 in port_nums or "nfs" in svc_names:
        tags.add("nfs")
    if 6379 in port_nums or any("redis" in s for s in svc_names):
        tags.add("redis")
    if {"kerberos", "ldap", "smb"} <= tags:
        tags.add("domain-controller")
    return sorted(tags)


def get_port_hints(port: int, service_name: str, product: str = "", extrainfo: str = "") -> list[str]:
    svc   = (service_name or "").lower()
    prod  = (product or "").lower()
    extra = (extrainfo or "").lower()
    hints: list[str] = []

    if port in (80, 81, 443, 8000, 8080, 8443) or svc in ("http", "https", "http-proxy"):
        hints.append("Web service detected; consider whatweb, nikto, gobuster/feroxbuster/ffuf, curl, and manual browsing.")
        hints.append("Check for default credentials, admin panels, exposed APIs, virtual hosts, and interesting headers.")
    if port == 445 or svc in ("microsoft-ds", "smb", "netbios-ssn"):
        hints.append("SMB detected; consider netexec smb, smbclient, smbmap, and enum4linux-ng.")
        hints.append("Check share access, null sessions, SMB signing, local admin reuse, and domain membership clues.")
    if port in (88, 464) or "kerberos" in svc or "kerberos" in prod:
        hints.append("Kerberos-related service detected; consider kerbrute, netexec, GetUserSPNs, and AS-REP roast checks.")
        hints.append("Look for domain naming clues, valid usernames, SPNs, and clock skew issues.")
    if port in (389, 636, 3268, 3269) or svc in ("ldap", "ldaps", "globalcatldap", "globalcatldapssl"):
        hints.append("LDAP detected; consider ldapsearch, netexec ldap, and directory enumeration.")
        hints.append("Look for anonymous bind, naming contexts, domain structure, users, groups, and password policy info.")
    if port == 135 or svc == "msrpc":
        hints.append("MSRPC detected; consider rpcclient, netexec, and enumerate Windows service exposure and domain clues.")
    if port == 139 or svc == "netbios-ssn":
        hints.append("NetBIOS detected; enumerate shares, names, and Windows host information.")
    if port in (5985, 5986) or "winrm" in svc:
        hints.append("WinRM detected; consider netexec winrm and evil-winrm if credentials are obtained.")
    if port in (1433, 1434) or "mssql" in svc or "sql server" in prod:
        hints.append("MSSQL detected; consider impacket-mssqlclient, netexec mssql, and SQL login enumeration.")
    if port == 3306 or "mysql" in svc:
        hints.append("MySQL detected; test authentication paths and enumerate version-specific exposure carefully.")
    if port == 5432 or "postgres" in svc:
        hints.append("PostgreSQL detected; test for weak/default credentials and accessible databases.")
    if port == 27017 or "mongodb" in svc:
        hints.append("MongoDB detected; verify whether authentication is required and whether remote access is overly exposed.")
    if port == 22 or svc == "ssh":
        hints.append("SSH detected; consider ssh-audit, banner review, key-based auth checks, and cautious credential testing.")
    if port == 21 or svc == "ftp":
        hints.append("FTP detected; check for anonymous access, writable locations, and plaintext credential exposure.")
    if port in (25, 465, 587) or svc in ("smtp", "smtps", "submission"):
        hints.append("SMTP detected; consider smtp-user-enum or swaks and look for open relay or user enumeration behavior.")
    if port == 53 or svc == "domain":
        hints.append("DNS detected; consider dig, nslookup, dnsrecon, and zone transfer testing where appropriate.")
    if port == 69 or svc == "tftp":
        hints.append("TFTP detected; check for unauthenticated file retrieval or upload opportunities.")
    if port == 111 or svc == "rpcbind":
        hints.append("RPCbind detected; consider rpcinfo and follow-on enumeration for NFS and related services.")
    if port == 2049 or svc == "nfs":
        hints.append("NFS detected; use showmount and test for accessible exports and weak export permissions.")
    if port in (161, 162) or svc == "snmp":
        hints.append("SNMP detected; try snmpwalk/onesixtyone and test common community strings carefully.")
    if port == 3389 or svc == "ms-wbt-server":
        hints.append("RDP detected; consider netexec rdp, xfreerdp, and review NLA / domain clues.")
    if port in (5900, 5901, 5902) or svc == "vnc":
        hints.append("VNC detected; verify authentication requirements and assess screenshot or password attack viability.")
    if port == 6379 or "redis" in svc:
        hints.append("Redis detected; verify bind/auth configuration and whether dangerous commands are exposed remotely.")
    if port == 11211 or "memcached" in svc:
        hints.append("Memcached detected; check whether it is exposed beyond localhost and assess information disclosure risk.")
    if port in (9200, 9300) or "elasticsearch" in svc:
        hints.append("Elasticsearch detected; check for unauthenticated cluster or index access.")
    if port in (2375, 2376) or "docker" in svc:
        hints.append("Docker API exposure may be high risk; verify remote daemon access and authentication/TLS posture.")
    if port == 6443 or "kubernetes" in svc:
        hints.append("Kubernetes-related service detected; look for exposed API endpoints and auth configuration issues.")
    if "ssl" in extra or "tls" in extra:
        hints.append("TLS-related context detected; review certificate details, protocol support, and any weak crypto indicators.")
    return hints


# ============================================================
# Nmap XML parsing
# ============================================================

def parse_nmap_xml(xml_path: Path) -> dict:
    logging.info(f"Parsing Nmap XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    scan_info = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "nmap_args":   root.get("args", ""),
        "nmap_version":root.get("version", ""),
        "hosts":       [],
    }

    for host in root.findall("host"):
        state_el = host.find("status")
        state = state_el.get("state") if state_el is not None else "unknown"

        addresses = [
            {"addr": a.get("addr", ""), "addrtype": a.get("addrtype", "")}
            for a in host.findall("address")
        ]

        hostnames = []
        hn_parent = host.find("hostnames")
        if hn_parent is not None:
            for hn in hn_parent.findall("hostname"):
                hostnames.append({"name": hn.get("name", ""), "type": hn.get("type", "")})

        ports = []
        ports_parent = host.find("ports")
        if ports_parent is not None:
            for port_el in ports_parent.findall("port"):
                state_el2 = port_el.find("state")
                if state_el2 is None or state_el2.get("state") != "open":
                    continue
                service_el = port_el.find("service")
                service = {}
                if service_el is not None:
                    service = {
                        "name":      service_el.get("name", ""),
                        "product":   service_el.get("product", ""),
                        "version":   service_el.get("version", ""),
                        "extrainfo": service_el.get("extrainfo", ""),
                        "tunnel":    service_el.get("tunnel", ""),
                    }
                scripts = [
                    {"id": s.get("id", ""), "output": s.get("output", "")}
                    for s in port_el.findall("script")
                ]
                ports.append({
                    "protocol": port_el.get("protocol", ""),
                    "port":     int(port_el.get("portid", "0")),
                    "service":  service,
                    "scripts":  scripts,
                })

        host_record = {
            "state":      state,
            "addresses":  addresses,
            "hostnames":  hostnames,
            "open_ports": sorted(ports, key=lambda p: (p["protocol"], p["port"])),
        }
        scan_info["hosts"].append(host_record)
        logging.debug(
            f"Nmap host: {choose_host_display_name(host_record)} "
            f"state={state} ports={len(host_record['open_ports'])}"
        )

    pre_filter = len(scan_info["hosts"])
    scan_info["hosts"] = [
        h for h in scan_info["hosts"]
        if h.get("state") == "up" and h.get("open_ports")
    ]
    if len(scan_info["hosts"]) < pre_filter:
        logging.info(
            f"Nmap filtered: {pre_filter - len(scan_info['hosts'])} host(s) "
            f"with no open ports removed"
        )

    logging.info(f"Nmap done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# Nessus XML parsing
# ============================================================

def parse_nessus_xml(xml_path: Path) -> dict:
    """Parse a .nessus (NessusClientData_v2) file."""
    logging.info(f"Parsing Nessus XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Report name from the Report element
    report_el = root.find("Report")
    report_name = report_el.get("name", xml_path.stem) if report_el is not None else xml_path.stem

    scan_info: dict = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "report_name": report_name,
        "hosts":       [],
    }

    report_hosts = root.findall(".//ReportHost")
    for rh in report_hosts:
        # Host properties
        props: dict[str, str] = {}
        host_props_el = rh.find("HostProperties")
        if host_props_el is not None:
            for tag_el in host_props_el.findall("tag"):
                props[tag_el.get("name", "")] = (tag_el.text or "").strip()

        ip       = props.get("host-ip", rh.get("name", ""))
        hostname = props.get("hostname", props.get("host-fqdn", ""))

        findings: list[dict] = []
        for item in rh.findall("ReportItem"):
            plugin_id   = item.get("pluginID", "")
            plugin_name = item.get("pluginName", "")
            severity    = int(item.get("severity", "0"))
            port        = int(item.get("port", "0"))
            protocol    = item.get("protocol", "tcp")

            def _txt(tag: str) -> str:
                el = item.find(tag)
                return (el.text or "").strip() if el is not None else ""

            cves = [el.text.strip() for el in item.findall("cve") if el.text]

            findings.append({
                "plugin_id":    plugin_id,
                "plugin_name":  plugin_name,
                "severity_int": severity,
                "port":         port,
                "protocol":     protocol,
                "description":  _txt("description"),
                "solution":     _txt("solution"),
                "cves":         cves,
                "cvss_base":    _txt("cvss_base_score"),
                "cvss3_base":   _txt("cvss3_base_score"),
                "plugin_output":_txt("plugin_output"),
            })

        # Sort findings: highest severity first, then plugin name
        findings.sort(key=lambda f: (-f["severity_int"], f["plugin_name"]))

        host_record = {
            "ip":       ip,
            "hostname": hostname,
            "findings": findings,
        }
        scan_info["hosts"].append(host_record)
        logging.debug(
            f"Nessus host: {ip} ({hostname}) findings={len(findings)}"
        )

    pre_filter = len(scan_info["hosts"])
    scan_info["hosts"] = [
        h for h in scan_info["hosts"]
        if any(f["severity_int"] > 0 for f in h.get("findings", []))
    ]
    if len(scan_info["hosts"]) < pre_filter:
        logging.info(
            f"Nessus filtered: {pre_filter - len(scan_info['hosts'])} host(s) "
            f"with only informational findings removed"
        )

    logging.info(f"Nessus done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# Burp Suite XML parsing
# ============================================================

def parse_burp_xml(xml_path: Path) -> dict:
    """Parse a Burp Suite Pro Scanner Issues XML export."""
    logging.info(f"Parsing Burp XML: {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    scan_info: dict = {
        "source_file": str(xml_path),
        "parsed_at":   dt.datetime.now().isoformat(timespec="seconds"),
        "hosts":       [],
    }

    # Group issues by host IP
    host_map: dict[str, dict] = {}  # ip → {ip, url, issues}

    for issue_el in root.findall("issue"):
        host_el = issue_el.find("host")
        if host_el is None:
            continue

        ip  = host_el.get("ip", "")
        url = (host_el.text or "").strip()

        # Derive a canonical host key (prefer IP, fall back to URL base)
        host_key = ip if ip else url

        def _txt(tag: str) -> str:
            el = issue_el.find(tag)
            return (el.text or "").strip() if el is not None else ""

        issue = {
            "name":                _txt("name"),
            "path":                _txt("path"),
            "location":            _txt("location"),
            "severity":            _txt("severity"),
            "confidence":          _txt("confidence"),
            "issue_detail":        _txt("issueDetail"),
            "issue_background":    _txt("issueBackground"),
            "remediation_detail":  _txt("remediationDetail"),
            "remediation_background": _txt("remediationBackground"),
        }

        if host_key not in host_map:
            host_map[host_key] = {"ip": ip, "url": url, "issues": []}
        host_map[host_key]["issues"].append(issue)

    sev_order = {"high": 0, "medium": 1, "low": 2, "information": 3}
    for host_rec in host_map.values():
        host_rec["issues"].sort(
            key=lambda i: (sev_order.get(i["severity"].lower(), 4), i["name"])
        )
        scan_info["hosts"].append(host_rec)
        logging.debug(
            f"Burp host: {host_rec['ip']} ({host_rec['url']}) issues={len(host_rec['issues'])}"
        )

    logging.info(f"Burp done: {xml_path.name} | hosts={len(scan_info['hosts'])}")
    return scan_info


# ============================================================
# AutoRecon tool-output extractors
# ============================================================

def _extract_dirbusting(text: str) -> dict:
    """Extract discovered paths from gobuster / feroxbuster / dirsearch output."""
    paths: list[dict] = []
    # gobuster: /path (Status: 200) [Size: 1234]
    for m in re.finditer(
        r"^(/\S+)\s+\(Status:\s*(\d+)\)\s+\[Size:\s*(\d+)\](?:\s+\[--> (\S+)\])?",
        text, re.MULTILINE,
    ):
        paths.append({"path": m.group(1), "status": int(m.group(2)),
                       "size": int(m.group(3)), "redirect": m.group(4)})
    # feroxbuster: 200  GET  10l  20w  1234c http://target/path
    if not paths:
        for m in re.finditer(
            r"^(\d{3})\s+\w+\s+\d+l\s+\d+w\s+(\d+)c\s+https?://[^/]+(/.*)$",
            text, re.MULTILINE,
        ):
            paths.append({"path": m.group(3), "status": int(m.group(1)),
                           "size": int(m.group(2)), "redirect": None})
    # dirsearch: 200  1234B  /path
    if not paths:
        for m in re.finditer(
            r"^\s*(\d{3})\s+[\d.]+[KMB]*\s+(/.+)$", text, re.MULTILINE,
        ):
            paths.append({"path": m.group(2).strip(), "status": int(m.group(1)),
                           "size": 0, "redirect": None})

    paths = [p for p in paths if p["status"] != 404]
    interest_order = {200: 0, 301: 1, 302: 1, 401: 2, 403: 2}
    paths.sort(key=lambda p: (interest_order.get(p["status"], 9), p["path"]))
    interesting = [p for p in paths if p["status"] in interest_order]
    return {
        "tool": "dirbusting",
        "paths": paths[:100],
        "interesting": interesting[:50],
        "total_found": len(paths),
    }


def _extract_nikto(text: str) -> dict:
    """Extract findings from nikto output."""
    server = ""
    headers: dict[str, str] = {}
    findings: list[dict] = []

    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("+"):
            continue
        body = line[1:].strip()

        sm = re.match(r"Server:\s+(.+)", body)
        if sm:
            server = sm.group(1).strip()
            continue

        hm = re.match(r"Retrieved\s+(\S+)\s+header:\s+(.+)", body, re.IGNORECASE)
        if hm:
            headers[hm.group(1).lower()] = hm.group(2).strip()
            continue

        fm = re.match(r"(?:OSVDB-(\d+):\s+)?(/\S*):\s+(.+)", body)
        if fm:
            findings.append({
                "osvdb": fm.group(1) or "",
                "path": fm.group(2),
                "description": fm.group(3).strip(),
            })

    return {
        "tool": "nikto",
        "server": server,
        "headers": headers,
        "findings": findings,
        "total_findings": len(findings),
    }


def _extract_enum4linux(text: str) -> dict:
    """Extract SMB/NetBIOS enumeration data from enum4linux output."""
    result: dict = {
        "tool": "enum4linux",
        "os_info": "",
        "domain": "",
        "workgroup": "",
        "null_session": False,
        "users": [],
        "shares": [],
        "groups": [],
        "password_policy": {},
    }

    os_m = re.search(r"OS information on \S+.*?:\s*(.+)", text)
    if not os_m:
        os_m = re.search(r"OS:\s+(.+)", text)
    if os_m:
        result["os_info"] = os_m.group(1).strip()

    dm = re.search(r"Domain Name:\s+(\S+)", text)
    if dm:
        result["domain"] = dm.group(1)
    wm = re.search(r"Workgroup:\s+(\S+)", text)
    if wm:
        result["workgroup"] = wm.group(1)

    if re.search(r"(?:Null session|anonymous.*(?:allowed|access|success))", text, re.IGNORECASE):
        result["null_session"] = True

    # RID cycling users: 500: DOMAIN\Administrator (Local User)
    for m in re.finditer(r"\d+:\s+\S+\\(\S+)\s+\(", text):
        u = m.group(1)
        if u not in result["users"]:
            result["users"].append(u)

    # Shares: //host/SHARE  Mapping: OK  Listing: OK
    for m in re.finditer(
        r"//\S+/(\S+)\s+Mapping:\s+(\w+)\s+Listing:\s+(\w+)", text
    ):
        access = "read" if m.group(2) == "OK" else "denied"
        if m.group(3) == "OK":
            access = "read/write" if access == "read" else "read"
        result["shares"].append({"name": m.group(1), "access": access})
    # Fallback: Disk share lines
    if not result["shares"]:
        for m in re.finditer(r"^\s*(\S+)\s+(Disk|IPC|Printer)", text, re.MULTILINE):
            result["shares"].append({"name": m.group(1), "type": m.group(2), "access": ""})

    for m in re.finditer(r"group:\[([^\]]+)\]", text):
        g = m.group(1).strip()
        if g and g not in result["groups"]:
            result["groups"].append(g)

    pp = result["password_policy"]
    ml = re.search(r"Minimum password length:\s*(\d+)", text)
    if ml:
        pp["min_length"] = int(ml.group(1))
    lt = re.search(r"Account Lockout Threshold:\s*(\d+)", text)
    if lt:
        pp["lockout_threshold"] = int(lt.group(1))
    if re.search(r"Password Complexity.*Enabled", text, re.IGNORECASE):
        pp["complexity"] = True

    return result


def _extract_smbmap(text: str) -> dict:
    """Extract share permissions from smbmap output."""
    shares: list[dict] = []
    for m in re.finditer(
        r"^\s+(\S+)\s+(READ ONLY|READ,\s*WRITE|READ/WRITE|NO ACCESS)\s*(.*?)$",
        text, re.MULTILINE,
    ):
        perm = m.group(2).strip().replace(",", "/").replace("  ", " ")
        shares.append({
            "name": m.group(1),
            "permissions": perm,
            "comment": m.group(3).strip(),
        })
    writable = [s["name"] for s in shares if "WRITE" in s["permissions"]]
    readable = [s["name"] for s in shares
                if "READ" in s["permissions"] and s["name"] not in writable]
    return {
        "tool": "smbmap",
        "shares": shares,
        "writable_shares": writable,
        "readable_shares": readable,
    }


def _extract_smbclient(text: str) -> dict:
    """Extract share listing from smbclient output."""
    shares: list[dict] = []
    for m in re.finditer(r"^\s+(\S+)\s+(Disk|IPC|Printer)\s*(.*?)$", text, re.MULTILINE):
        shares.append({
            "name": m.group(1),
            "type": m.group(2),
            "comment": m.group(3).strip(),
        })
    return {"tool": "smbclient", "shares": shares}


def _extract_whatweb(text: str) -> dict:
    """Extract technology fingerprints from whatweb output."""
    technologies: list[dict] = []
    title = ""
    status_code = 0

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # WhatWeb format: http://target [200 OK] Apache[2.4.41], PHP[7.4.3], ...
        sc_m = re.search(r"\[(\d{3})\s", line)
        if sc_m:
            status_code = int(sc_m.group(1))
        for seg in re.split(r",\s*", line):
            seg = seg.strip()
            tm = re.match(r"([\w][\w\s./-]*?)(?:\[([^\]]*)\])?$", seg)
            if tm:
                name = tm.group(1).strip()
                version = (tm.group(2) or "").strip()
                if name.lower() == "title":
                    title = version
                elif name and not re.match(r"^https?://", name) and not re.match(r"^\[\d{3}", name):
                    technologies.append({"name": name, "version": version})

    seen: set[str] = set()
    deduped: list[dict] = []
    for t in technologies:
        key = t["name"].lower()
        if key not in seen and len(key) > 1:
            seen.add(key)
            deduped.append(t)

    return {
        "tool": "whatweb",
        "technologies": deduped,
        "title": title,
        "status_code": status_code,
    }


def _extract_snmpwalk(text: str) -> dict:
    """Extract key SNMP data from snmpwalk output."""
    result: dict = {
        "tool": "snmpwalk",
        "sys_descr": "",
        "sys_name": "",
        "sys_contact": "",
        "sys_location": "",
        "interfaces": [],
        "ip_addresses": [],
        "running_processes": [],
        "raw_oid_count": 0,
    }
    result["raw_oid_count"] = sum(1 for line in text.splitlines() if "=" in line)

    oid_values: dict[str, str] = {}
    for m in re.finditer(r"^(\S+)\s+=\s+\S+:\s+(.+)$", text, re.MULTILINE):
        oid_values[m.group(1).lower()] = m.group(2).strip().strip('"')

    for oid_key, val in oid_values.items():
        if "sysdescr" in oid_key or "1.3.6.1.2.1.1.1.0" in oid_key:
            result["sys_descr"] = val
        elif "sysname" in oid_key or "1.3.6.1.2.1.1.5.0" in oid_key:
            result["sys_name"] = val
        elif "syscontact" in oid_key or "1.3.6.1.2.1.1.4.0" in oid_key:
            result["sys_contact"] = val
        elif "syslocation" in oid_key or "1.3.6.1.2.1.1.6.0" in oid_key:
            result["sys_location"] = val
        elif "ifdescr" in oid_key:
            if val and val not in result["interfaces"]:
                result["interfaces"].append(val)
        elif "ipadentaddr" in oid_key or "ipaddressaddr" in oid_key:
            if IPV4_RE.match(val) and val not in result["ip_addresses"]:
                result["ip_addresses"].append(val)
        elif "hrswrunname" in oid_key:
            if val and val not in result["running_processes"]:
                result["running_processes"].append(val)

    return result


def _extract_onesixtyone(text: str) -> dict:
    """Extract SNMP community strings from onesixtyone output."""
    community_strings: list[dict] = []
    for m in re.finditer(r"^(\S+)\s+\[(\S+)\]\s+(.+)$", text, re.MULTILINE):
        community_strings.append({
            "ip": m.group(1),
            "community": m.group(2),
            "sys_descr": m.group(3).strip(),
        })
    return {"tool": "onesixtyone", "community_strings": community_strings}


def _extract_sslscan(text: str) -> dict:
    """Extract TLS/SSL scan results from sslscan output."""
    protocols: dict[str, bool] = {}
    weak_ciphers: list[str] = []
    cert_subject = cert_issuer = ""
    cert_not_after = ""
    cert_expired = False
    key_size = 0
    heartbleed_vulnerable = False

    for m in re.finditer(
        r"(SSLv[23]|TLSv1\.[0-3])\s+(enabled|disabled)", text, re.IGNORECASE,
    ):
        protocols[m.group(1)] = m.group(2).lower() == "enabled"

    for m in re.finditer(
        r"Accepted\s+\S+\s+\d+\s+bits\s+(\S+)", text,
    ):
        cipher = m.group(1)
        if re.search(r"DES|RC4|NULL|EXPORT|MD5", cipher, re.IGNORECASE):
            if cipher not in weak_ciphers:
                weak_ciphers.append(cipher)

    sm = re.search(r"Subject:\s*(.+)", text)
    if sm:
        cert_subject = sm.group(1).strip()
    im = re.search(r"Issuer:\s*(.+)", text)
    if im:
        cert_issuer = im.group(1).strip()
    na = re.search(r"Not valid after:\s*(.+)", text)
    if na:
        cert_not_after = na.group(1).strip()
    if re.search(r"expired", text, re.IGNORECASE):
        cert_expired = True
    ks = re.search(r"RSA Key Strength:\s*(\d+)", text)
    if ks:
        key_size = int(ks.group(1))
    if re.search(r"vulnerable to heartbleed", text, re.IGNORECASE):
        heartbleed_vulnerable = True

    weak_protos = [p for p, enabled in protocols.items()
                   if enabled and p in ("SSLv2", "SSLv3", "TLSv1.0", "TLSv1.1")]

    return {
        "tool": "sslscan",
        "protocols": protocols,
        "weak_protocols": weak_protos,
        "weak_ciphers": weak_ciphers,
        "cert_subject": cert_subject,
        "cert_issuer": cert_issuer,
        "cert_not_after": cert_not_after,
        "cert_expired": cert_expired,
        "key_size": key_size,
        "heartbleed_vulnerable": heartbleed_vulnerable,
    }


def _extract_dnsrecon(text: str) -> dict:
    """Extract DNS records from dnsrecon output."""
    records: list[dict] = []
    zone_transfer = False

    for m in re.finditer(r"\[\*\]\s+(\w+)\s+(\S+)\s+(\S+)(?:\s+(.+))?$", text, re.MULTILINE):
        records.append({
            "type": m.group(1),
            "name": m.group(2),
            "value": m.group(3),
            "extra": (m.group(4) or "").strip(),
        })
    if re.search(r"Zone Transfer.*success|AXFR.*records", text, re.IGNORECASE):
        zone_transfer = True

    return {
        "tool": "dnsrecon",
        "records": records,
        "zone_transfer_successful": zone_transfer,
    }


def _extract_curl_headers(text: str) -> dict:
    """Extract useful info from curl output (HTML or header dump)."""
    title = ""
    headers: dict[str, str] = {}

    tm = re.search(r"<title[^>]*>([^<]+)</title>", text, re.IGNORECASE)
    if tm:
        title = tm.group(1).strip()

    for m in re.finditer(r"^([\w-]+):\s+(.+)$", text, re.MULTILINE):
        hname = m.group(1).lower()
        if hname in ("server", "x-powered-by", "x-aspnet-version",
                      "x-generator", "x-redirect-by", "www-authenticate",
                      "content-type", "set-cookie"):
            headers[hname] = m.group(2).strip()

    interesting = [h for h in headers if h != "content-type"]
    return {
        "tool": "curl",
        "title": title,
        "headers": headers,
        "interesting_headers": interesting,
    }


def _extract_generic(filename: str, text: str) -> dict:
    """Fallback extractor for unrecognized tool output."""
    return {
        "tool": "unknown",
        "filename": filename,
        "line_count": len(text.splitlines()),
        "first_lines": text[:500],
        "non_empty": bool(text.strip()),
    }


AUTORECON_EXTRACTORS: dict[str, any] = {
    r"gobuster|feroxbuster|dirsearch": _extract_dirbusting,
    r"nikto":        _extract_nikto,
    r"enum4linux":   _extract_enum4linux,
    r"smbmap":       _extract_smbmap,
    r"smbclient":    _extract_smbclient,
    r"whatweb":      _extract_whatweb,
    r"snmpwalk":     _extract_snmpwalk,
    r"onesixtyone":  _extract_onesixtyone,
    r"sslscan":      _extract_sslscan,
    r"dnsrecon":     _extract_dnsrecon,
    r"curl":         _extract_curl_headers,
}


def _get_extractor(tool_name: str):
    """Look up the extractor function for a given tool name."""
    for pattern, func in AUTORECON_EXTRACTORS.items():
        if re.search(pattern, tool_name, re.IGNORECASE):
            return func
    return None


# ============================================================
# AutoRecon directory parsing
# ============================================================

def _build_autorecon_target_summary(tool_results: dict) -> dict:
    """Compute a structured summary from parsed tool results (no LLM)."""
    summary: dict = {
        "total_tools_run": 0,
        "tools_with_findings": 0,
        "technologies": [],
        "writable_shares": [],
        "null_session": False,
        "users_found": [],
        "weak_tls": [],
        "community_strings": [],
    }
    seen_techs: set[str] = set()
    seen_users: set[str] = set()

    for port_key, results in tool_results.items():
        for entry in results:
            summary["total_tools_run"] += 1
            data = entry.get("data", {})
            tool = data.get("tool", "unknown")
            has_findings = False

            if tool == "dirbusting" and data.get("paths"):
                has_findings = True
            elif tool == "nikto" and data.get("findings"):
                has_findings = True
            elif tool in ("enum4linux", "smbmap", "smbclient") and data.get("shares"):
                has_findings = True
            elif tool == "whatweb" and data.get("technologies"):
                has_findings = True
                for t in data["technologies"]:
                    key = t["name"]
                    if t.get("version"):
                        key += f"/{t['version']}"
                    if key not in seen_techs:
                        seen_techs.add(key)
                        summary["technologies"].append(key)
            elif tool == "snmpwalk" and data.get("sys_descr"):
                has_findings = True
            elif tool == "onesixtyone" and data.get("community_strings"):
                has_findings = True
                for cs in data["community_strings"]:
                    c = cs.get("community", "")
                    if c and c not in summary["community_strings"]:
                        summary["community_strings"].append(c)
            elif tool == "sslscan" and (data.get("weak_protocols") or data.get("weak_ciphers")):
                has_findings = True
                summary["weak_tls"].extend(data.get("weak_protocols", []))
            elif tool == "dnsrecon" and data.get("records"):
                has_findings = True
            elif tool == "curl" and data.get("interesting_headers"):
                has_findings = True
            elif tool == "unknown" and data.get("non_empty"):
                has_findings = True

            if has_findings:
                summary["tools_with_findings"] += 1

            if tool == "enum4linux":
                if data.get("null_session"):
                    summary["null_session"] = True
                for u in data.get("users", []):
                    if u not in seen_users:
                        seen_users.add(u)
                        summary["users_found"].append(u)
                for s in data.get("shares", []):
                    if s.get("access") in ("read/write",):
                        if s["name"] not in summary["writable_shares"]:
                            summary["writable_shares"].append(s["name"])

            if tool == "smbmap":
                for s_name in data.get("writable_shares", []):
                    if s_name not in summary["writable_shares"]:
                        summary["writable_shares"].append(s_name)

    return summary


def parse_autorecon_results(results_dir: Path) -> dict:
    """Walk an AutoRecon results directory and extract structured data per target."""
    logging.info(f"Parsing AutoRecon results: {results_dir}")

    ar_data: dict = {
        "source_file": str(results_dir),
        "parsed_at": dt.datetime.now().isoformat(timespec="seconds"),
        "report_name": results_dir.name,
        "targets": [],
        "hosts": [],
    }

    for target_dir in sorted(results_dir.iterdir()):
        if not target_dir.is_dir():
            continue
        scans_dir = target_dir / "scans"
        if not scans_dir.exists():
            logging.debug(f"Skipping {target_dir.name}: no scans/ directory")
            continue

        target_name = target_dir.name
        ip = target_name if IPV4_RE.match(target_name) else ""
        hostname = "" if ip else target_name

        # Layer 1: collect nmap XML paths
        xml_dir = scans_dir / "xml"
        nmap_xml_files: list[Path] = []
        if xml_dir.exists():
            nmap_xml_files = sorted(xml_dir.glob("*.xml"))
            # Try to resolve IP/hostname from nmap XML
            if nmap_xml_files and (not ip):
                try:
                    nmap_data = parse_nmap_xml(nmap_xml_files[0])
                    for h in nmap_data.get("hosts", []):
                        for addr in h.get("addresses", []):
                            if addr.get("addrtype") == "ipv4" and addr.get("addr"):
                                ip = addr["addr"]
                                break
                        if not hostname:
                            for hn in h.get("hostnames", []):
                                if hn.get("name"):
                                    hostname = hn["name"]
                                    break
                        if ip:
                            break
                except Exception as exc:
                    logging.debug(f"Failed to pre-parse nmap XML for IP resolution: {exc}")

        # Layer 2: parse tool output files per port directory
        tool_results: dict[str, list[dict]] = {}

        for port_dir in sorted(scans_dir.iterdir()):
            if not port_dir.is_dir():
                continue
            dir_match = AUTORECON_PORT_DIR_RE.match(port_dir.name)
            if not dir_match:
                continue
            protocol = dir_match.group(1)
            port = int(dir_match.group(2))
            port_key = f"{protocol}/{port}"

            port_results: list[dict] = []
            for filepath in sorted(port_dir.iterdir()):
                if not filepath.is_file():
                    continue
                fname = filepath.name
                fn_match = AUTORECON_FILENAME_RE.match(fname)
                tool_name = fn_match.group(4) if fn_match else filepath.stem

                # Skip nmap text files (XML handled in Layer 1)
                if tool_name.lower().endswith("nmap") or tool_name.lower().startswith("nmap"):
                    continue

                try:
                    raw_text = filepath.read_text(encoding="utf-8", errors="replace")
                except Exception as exc:
                    logging.debug(f"Cannot read {filepath}: {exc}")
                    continue
                if not raw_text.strip():
                    logging.debug(f"Empty file: {filepath}")
                    continue

                extractor = _get_extractor(tool_name)
                if extractor:
                    data = extractor(raw_text)
                else:
                    data = _extract_generic(fname, raw_text)

                port_results.append({
                    "tool": data.get("tool", tool_name),
                    "filename": fname,
                    "data": data,
                })

            if port_results:
                tool_results[port_key] = port_results

        # Read _commands.log and _manual_commands.txt
        commands_log = ""
        manual_commands = ""
        cl_path = scans_dir / "_commands.log"
        if cl_path.exists():
            try:
                commands_log = cl_path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                pass
        mc_path = scans_dir / "_manual_commands.txt"
        if mc_path.exists():
            try:
                manual_commands = mc_path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                pass

        summary = _build_autorecon_target_summary(tool_results)

        target_record = {
            "target": target_name,
            "ip": ip,
            "hostname": hostname,
            "nmap_xml_files": nmap_xml_files,
            "commands_log": commands_log,
            "manual_commands": manual_commands,
            "tool_results": tool_results,
            "summary": summary,
        }
        ar_data["targets"].append(target_record)
        ar_data["hosts"].append(target_record)
        logging.info(
            f"AutoRecon target: {target_name} | ports={len(tool_results)} "
            f"tools={summary['total_tools_run']} with_findings={summary['tools_with_findings']}"
        )

    pre_filter = len(ar_data["targets"])
    ar_data["targets"] = [
        t for t in ar_data["targets"]
        if t.get("tool_results") or t.get("nmap_xml_files")
    ]
    ar_data["hosts"] = ar_data["targets"]
    if len(ar_data["targets"]) < pre_filter:
        logging.info(
            f"AutoRecon filtered: {pre_filter - len(ar_data['targets'])} target(s) "
            f"with no scan data removed"
        )

    logging.info(f"AutoRecon done: {results_dir.name} | targets={len(ar_data['targets'])}")
    return ar_data


# ============================================================
# Loot pre-processing extractors
# ============================================================

_INLINE_NOTE_RE = re.compile(r"\s+[-–—]\s*\(.*\)$|\s+[-–—]\s+\S.*$|\s+\((?:db |only|note|for ).*\)$", re.IGNORECASE)


def _strip_password_note(pw: str) -> tuple[str, str]:
    """Split 'password - (db only)' into ('password', 'db only')."""
    m = _INLINE_NOTE_RE.search(pw)
    if m:
        note = m.group(0).strip().lstrip("-–— ").strip("()")
        return pw[:m.start()].strip(), note
    return pw, ""


def _extract_credentials(text: str) -> list[dict]:
    """Extract credential pairs from loot text."""
    creds: list[dict] = []
    seen: set[tuple[str, str]] = set()

    for m in CRED_KV_RE.finditer(text):
        key = (m.group(1).lower(), m.group(2))
        if key not in seen:
            seen.add(key)
            creds.append({"username": m.group(1), "password": m.group(2),
                          "cred_type": "cleartext", "source_pattern": "key=value",
                          "inline_note": ""})

    for m in HASHCAT_FORMAT_RE.finditer(text):
        user, hash_val = m.group(1), m.group(2)
        if "://" in user or user.startswith("/"):
            continue
        key = (user.lower(), hash_val)
        if key not in seen:
            seen.add(key)
            ctype = "hash"
            if hash_val.startswith("$"):
                ctype = "bcrypt_hash"
            creds.append({"username": user, "password": hash_val,
                          "cred_type": ctype, "source_pattern": "hashcat",
                          "inline_note": ""})

    for line in text.splitlines():
        line = line.strip()
        if not line or "://" in line or line.startswith("#"):
            continue
        if ":" in line:
            parts = line.split(":", 1)
            if len(parts) == 2:
                user, pw_raw = parts[0].strip(), parts[1].strip()
                if (not user or not pw_raw or len(user) > 64 or len(pw_raw) > 128
                        or " " in user or user.startswith("/")):
                    continue
                if all(c in "0123456789abcdefABCDEF" for c in user) and len(user) > 20:
                    continue
                pw, inline_note = _strip_password_note(pw_raw)
                if not pw:
                    continue
                key = (user.lower(), pw)
                if key not in seen:
                    seen.add(key)
                    ctype = "cleartext"
                    if NTLM_HASH_RE.fullmatch(pw):
                        ctype = "NTLM_hash"
                    elif BCRYPT_HASH_RE.match(pw):
                        ctype = "bcrypt_hash"
                    elif SHA256_HASH_RE.fullmatch(pw):
                        ctype = "SHA256_hash"
                    creds.append({"username": user, "password": pw,
                                  "cred_type": ctype, "source_pattern": "colon",
                                  "inline_note": inline_note})

    return creds[:100]


_USERNAME_LINE_RE = re.compile(
    r"^(?:user(?:name)?|login|account|name)\s*[=:]\s*(\S+)",
    re.IGNORECASE | re.MULTILINE,
)

_USERNAME_CONTEXT_RE = re.compile(
    r"(?:username|user|account|login|name|member|employee|personnel|staff|people)",
    re.IGNORECASE,
)


def _extract_usernames(text: str, known_users: set[str] | None = None) -> list[str]:
    """Extract standalone usernames from text (not already in credential pairs).

    Handles:
    - key=value patterns: username=admin
    - Bare names on their own lines after a context line mentioning 'users'/'usernames'
    """
    known = {u.lower() for u in (known_users or set())}
    usernames: list[str] = []
    seen: set[str] = set()

    # Pattern 1: key=value
    for m in _USERNAME_LINE_RE.finditer(text):
        user = m.group(1).strip()
        if user.lower() not in known and user.lower() not in seen:
            if len(user) <= 64 and " " not in user and not user.startswith("/"):
                seen.add(user.lower())
                usernames.append(user)

    # Pattern 2: bare names after context lines
    lines = text.splitlines()
    in_username_block = False
    for line in lines:
        stripped = line.strip()
        if not stripped:
            if in_username_block:
                in_username_block = False
            continue

        if ":" in stripped:
            in_username_block = False
            continue

        if _USERNAME_CONTEXT_RE.search(stripped) and len(stripped) > 10:
            in_username_block = True
            continue

        if in_username_block:
            word = stripped.split()[0] if stripped.split() else ""
            if (word and len(word) <= 32 and len(stripped.split()) <= 2
                    and word.lower() not in known and word.lower() not in seen
                    and not word.startswith(("/", "#", "-", "*"))
                    and ":" not in word and "=" not in word):
                seen.add(word.lower())
                usernames.append(word)

    return usernames[:50]


def _extract_hashes(text: str) -> list[dict]:
    """Extract standalone hashes from loot text."""
    hashes: list[dict] = []
    seen: set[str] = set()

    # SAM format: user:RID:LM_HASH:NTLM_HASH:::
    for m in re.finditer(r"^(\S+):\d+:([a-fA-F0-9]{32}):([a-fA-F0-9]{32}):::\s*$",
                         text, re.MULTILINE):
        for h, htype in [(m.group(2), "LM"), (m.group(3), "NTLM")]:
            if h not in seen and h != "aad3b435b51404eeaad3b435b51404ee":
                seen.add(h)
                hashes.append({"hash": h, "hash_type": htype,
                               "username": m.group(1), "context_line": m.group(0).strip()})

    for m in BCRYPT_HASH_RE.finditer(text):
        h = m.group(0)
        if h not in seen:
            seen.add(h)
            line_start = text.rfind("\n", 0, m.start()) + 1
            line_end = text.find("\n", m.end())
            ctx = text[line_start:line_end if line_end != -1 else m.end() + 50].strip()
            hashes.append({"hash": h, "hash_type": "bcrypt", "username": None,
                           "context_line": ctx[:200]})

    return hashes[:100]


def _extract_file_listings(text: str) -> list[dict]:
    """Extract file/directory listing entries from loot text."""
    entries: list[dict] = []

    for m in FTP_LISTING_RE.finditer(text):
        entries.append({"filename": m.group(4).strip(), "permissions": m.group(1),
                        "size": int(m.group(2)), "listing_type": "ftp"})

    if not entries:
        for m in SMB_LISTING_RE.finditer(text):
            entries.append({"filename": m.group(1).strip(), "permissions": m.group(2),
                            "size": int(m.group(3)), "listing_type": "smb"})

    return entries[:200]


def _extract_network_refs(text: str) -> list[dict]:
    """Extract IP addresses and ports referenced in loot text."""
    refs: list[dict] = []
    seen_ips: set[str] = set()
    for ip in IP_IN_TEXT_RE.findall(text):
        if ip not in seen_ips:
            seen_ips.add(ip)
            refs.append({"ip": ip, "port": None})
    return refs


def _process_loot_file(filepath: Path) -> dict | None:
    """Process a single loot file, running all extractors."""
    if filepath.suffix.lower() in BINARY_EXTENSIONS:
        logging.debug(f"Skipping binary loot file: {filepath.name}")
        return None

    if filepath.stat().st_size > LOOT_MAX_FILE_SIZE:
        logging.warning(f"Loot file too large ({filepath.stat().st_size} bytes), truncating: {filepath.name}")

    try:
        raw = filepath.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        logging.debug(f"Cannot read loot file {filepath}: {exc}")
        return None

    if not raw.strip():
        return None

    if len(raw) > LOOT_MAX_FILE_SIZE:
        raw = raw[:LOOT_MAX_FILE_SIZE]

    credentials = _extract_credentials(raw)
    hashes = _extract_hashes(raw)
    file_listings = _extract_file_listings(raw)
    network_refs = _extract_network_refs(raw)

    # Extract standalone usernames not already in credential pairs
    known_cred_users = {c["username"] for c in credentials}
    # Also count hash-associated usernames as known
    for h in hashes:
        if h.get("username"):
            known_cred_users.add(h["username"])
    standalone_usernames = _extract_usernames(raw, known_cred_users)

    counts = {"credentials": len(credentials), "hashes": len(hashes),
              "file_listing": len(file_listings), "network": len(network_refs)}
    if not any(counts.values()) and not standalone_usernames:
        category = "notes"
    elif sum(1 for v in counts.values() if v > 0) > 2:
        category = "mixed"
    elif any(counts.values()):
        category = max(counts, key=counts.get)
    else:
        category = "usernames"

    return {
        "filename": filepath.name,
        "filepath": str(filepath),
        "size_bytes": filepath.stat().st_size,
        "credentials": credentials,
        "hashes": hashes,
        "file_listings": file_listings,
        "network_refs": network_refs,
        "standalone_usernames": standalone_usernames,
        "raw_preview": raw[:2000],
        "category": category,
    }


# ============================================================
# Loot directory parsing
# ============================================================

def _resolve_host_from_path(filepath: Path, loot_dir: Path) -> str | None:
    """Resolve which host a loot file belongs to, or None for campaign-level."""
    rel = filepath.relative_to(loot_dir)
    parts = rel.parts

    if len(parts) >= 2:
        subdir = parts[0]
        if IPV4_RE.match(subdir):
            return subdir
        if is_probable_fqdn(subdir):
            return subdir

    m = LOOT_IP_PREFIX_RE.match(filepath.name)
    if m:
        return m.group(1)
    m = LOOT_FQDN_PREFIX_RE.match(filepath.name)
    if m:
        return m.group(1)

    return None


_HOST_MENTION_RE = re.compile(
    r"(?:for|on|from|at|server|host|machine|box|target)\s+"
    r"([A-Za-z][A-Za-z0-9\-]{2,30}(?:\.[A-Za-z0-9\-]+)*)",
    re.IGNORECASE,
)


def _resolve_host_from_content(
    text: str,
    known_hosts: dict[str, str] | None = None,
) -> str | None:
    """Fallback: resolve host from file content when path gives no answer.

    Checks for:
    1. Known host identifiers (hostnames, stems, IPs from existing vault notes)
    2. Prose mentions like "for the server DANTE-WEB-NIX01"
    3. IP addresses in text (single or dominant)

    known_hosts: {identifier_lower: host_key} mapping from vault
    """
    text_lower = text.lower()

    # Check for known hostnames/stems from vault (case-insensitive)
    if known_hosts:
        matches: dict[str, int] = {}
        for identifier, host_key in known_hosts.items():
            if len(identifier) < 4:
                continue
            count = text_lower.count(identifier)
            if count > 0:
                matches[host_key] = matches.get(host_key, 0) + count
        if matches:
            if len(matches) == 1:
                return next(iter(matches))
            sorted_matches = sorted(matches.items(), key=lambda x: x[1], reverse=True)
            top_key, top_count = sorted_matches[0]
            second_count = sorted_matches[1][1] if len(sorted_matches) > 1 else 0
            if top_count >= second_count * 2:
                return top_key

    # Look for prose host mentions ("for the server X", "on host X", etc.)
    prose_hosts: dict[str, int] = {}
    for m in _HOST_MENTION_RE.finditer(text):
        name = m.group(1).strip().rstrip(".")
        # Filter out common English words that follow these prepositions
        if name.lower() in (
            "the", "this", "that", "each", "all", "any", "some", "our",
            "its", "their", "your", "following", "above", "below",
        ):
            continue
        if len(name) >= 3:
            prose_hosts[name] = prose_hosts.get(name, 0) + 1

    if prose_hosts:
        if len(prose_hosts) == 1:
            return next(iter(prose_hosts))
        sorted_prose = sorted(prose_hosts.items(), key=lambda x: x[1], reverse=True)
        top_name, top_count = sorted_prose[0]
        second_count = sorted_prose[1][1] if len(sorted_prose) > 1 else 0
        if top_count >= second_count * 2:
            return top_name

    # Fall back to IP matching
    ip_counts: dict[str, int] = {}
    for ip in IP_IN_TEXT_RE.findall(text):
        if _is_target_ip(ip):
            ip_counts[ip] = ip_counts.get(ip, 0) + 1

    if not ip_counts:
        return None

    if len(ip_counts) == 1:
        return next(iter(ip_counts))

    sorted_ips = sorted(ip_counts.items(), key=lambda x: x[1], reverse=True)
    top_ip, top_count = sorted_ips[0]
    second_count = sorted_ips[1][1] if len(sorted_ips) > 1 else 0
    if top_count >= 3 and top_count >= second_count * 2:
        return top_ip

    return None


def _build_known_hosts_lookup(vault_dir: Path) -> dict[str, str]:
    """Build a lookup of known host identifiers from vault host notes.

    Returns: {identifier_lower: host_key} where host_key is the IP or hostname
    used as the primary key in loot/misc data structures.
    """
    lookup: dict[str, str] = {}
    hosts_dir = vault_dir / "Hosts"
    if not hosts_dir.exists():
        return lookup

    for hp in hosts_dir.glob("*.md"):
        if hp.stem.startswith("_"):
            continue
        try:
            fm, _ = read_frontmatter(hp.read_text(encoding="utf-8"))
        except Exception:
            continue

        ip = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        host_key = ip or (hostnames[0] if hostnames else hp.stem)

        # Register the IP
        if ip:
            lookup[ip.lower()] = host_key
        # Register all hostnames
        for hn in hostnames:
            if hn:
                lookup[hn.lower()] = host_key
                # Also register short form (first part before dot)
                short = hn.split(".")[0].lower()
                if len(short) >= 4:
                    lookup[short] = host_key
        # Register the file stem (which may be the hostname or IP)
        if len(hp.stem) >= 4:
            lookup[hp.stem.lower()] = host_key

    return lookup


def parse_loot_dir(loot_dir: Path, known_hosts: dict[str, str] | None = None) -> dict:
    """Walk a loot directory and extract structured data, associating with hosts."""
    logging.info(f"Parsing loot directory: {loot_dir}")

    loot_data: dict = {
        "source_dir": str(loot_dir),
        "parsed_at": dt.datetime.now().isoformat(timespec="seconds"),
        "host_loot": {},
        "campaign_loot": [],
        "summary": {
            "total_files": 0,
            "host_files": 0,
            "campaign_files": 0,
            "total_credentials": 0,
            "total_hashes": 0,
            "hash_types_seen": [],
            "hosts_with_loot": [],
        },
    }

    hash_types: set[str] = set()

    for filepath in sorted(loot_dir.rglob("*")):
        if not filepath.is_file():
            continue

        result = _process_loot_file(filepath)
        if result is None:
            continue

        loot_data["summary"]["total_files"] += 1
        loot_data["summary"]["total_credentials"] += len(result["credentials"])
        loot_data["summary"]["total_hashes"] += len(result["hashes"])
        for h in result["hashes"]:
            hash_types.add(h["hash_type"])

        host_key = _resolve_host_from_path(filepath, loot_dir)
        if not host_key:
            host_key = _resolve_host_from_content(result.get("raw_preview", ""), known_hosts)
            if host_key:
                logging.info(f"Loot: content-based host association: {filepath.name} → {host_key}")
        if host_key:
            loot_data["summary"]["host_files"] += 1
            if host_key not in loot_data["host_loot"]:
                loot_data["host_loot"][host_key] = []
                loot_data["summary"]["hosts_with_loot"].append(host_key)
            loot_data["host_loot"][host_key].append(result)
        else:
            loot_data["summary"]["campaign_files"] += 1
            loot_data["campaign_loot"].append(result)

    loot_data["summary"]["hash_types_seen"] = sorted(hash_types)

    logging.info(
        f"Loot done: {loot_data['summary']['total_files']} files, "
        f"{loot_data['summary']['total_credentials']} credentials, "
        f"{loot_data['summary']['total_hashes']} hashes, "
        f"{len(loot_data['host_loot'])} hosts"
    )
    return loot_data


# ============================================================
# Misc tool output parsing
# ============================================================

def parse_misc_dir(misc_dir: Path, known_hosts: dict[str, str] | None = None) -> dict:
    """Walk a misc directory and read text files for LLM interpretation."""
    logging.info(f"Parsing misc directory: {misc_dir}")

    misc_data: dict = {
        "source_dir": str(misc_dir),
        "parsed_at": dt.datetime.now().isoformat(timespec="seconds"),
        "files": [],
        "summary": {
            "total_files": 0,
            "host_files": 0,
            "campaign_files": 0,
            "hosts_with_misc": [],
        },
    }

    hosts_seen: set[str] = set()

    for filepath in sorted(misc_dir.rglob("*")):
        if not filepath.is_file():
            continue
        if filepath.suffix.lower() in BINARY_EXTENSIONS:
            logging.debug(f"Skipping binary misc file: {filepath.name}")
            continue
        if filepath.name.startswith("."):
            continue

        try:
            raw = filepath.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            logging.debug(f"Cannot read misc file {filepath}: {exc}")
            continue

        if not raw.strip():
            continue

        if len(raw) > LOOT_MAX_FILE_SIZE:
            raw = raw[:LOOT_MAX_FILE_SIZE]
            logging.warning(f"Misc file truncated ({filepath.stat().st_size} bytes): {filepath.name}")

        host_key = _resolve_host_from_path(filepath, misc_dir)
        if not host_key:
            host_key = _resolve_host_from_content(raw[:2000], known_hosts)
            if host_key:
                logging.info(f"Misc: content-based host association: {filepath.name} → {host_key}")
        if host_key:
            hosts_seen.add(host_key)

        tool_type, analysis_level = _detect_misc_tool(filepath.name, raw)

        misc_data["files"].append({
            "filename": filepath.name,
            "filepath": str(filepath),
            "content": raw,
            "host_key": host_key,
            "size_bytes": filepath.stat().st_size,
            "tool_type": tool_type,
            "analysis_level": analysis_level,
        })

    misc_data["summary"]["total_files"] = len(misc_data["files"])
    misc_data["summary"]["host_files"] = sum(1 for f in misc_data["files"] if f["host_key"])
    misc_data["summary"]["campaign_files"] = sum(1 for f in misc_data["files"] if not f["host_key"])
    misc_data["summary"]["hosts_with_misc"] = sorted(hosts_seen)

    tool_breakdown: dict[str, int] = {}
    for f in misc_data["files"]:
        tool_breakdown[f["tool_type"]] = tool_breakdown.get(f["tool_type"], 0) + 1
    misc_data["summary"]["tool_types"] = tool_breakdown

    logging.info(
        f"Misc done: {misc_dir.name} | files={misc_data['summary']['total_files']} "
        f"host_files={misc_data['summary']['host_files']} "
        f"tools={tool_breakdown}"
    )
    return misc_data


# ============================================================
# Operator Notes lookup (feedback loop)
# ============================================================

def build_operator_notes_lookup(vault_dir: Path) -> dict[str, str]:
    """Read operator notes from all existing host notes for feedback into prompts."""
    lookup: dict[str, str] = {}
    hosts_dir = vault_dir / "Hosts"
    if not hosts_dir.exists():
        return lookup

    for hp in hosts_dir.glob("*.md"):
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue
        fm, body = read_frontmatter(text)
        notes = extract_operator_notes(body)
        if not notes.strip():
            continue

        ip = fm.get("ip", "")
        if ip:
            lookup[ip] = notes
        for hn in fm.get("hostnames", []):
            if hn:
                lookup[hn] = notes

    return lookup


def _format_operator_notes_for_prompt(notes: str, max_len: int = 500) -> str:
    """Format operator notes for injection into a prompt."""
    truncated = notes[:max_len]
    if len(notes) > max_len:
        truncated += "..."
    return (
        "\n  [OPERATOR] Pentester observations:\n"
        f"  {truncated}\n"
    )


# ============================================================
# Host naming & port rendering
# ============================================================

def choose_host_display_name(host: dict) -> str:
    hostnames = [h["name"].strip() for h in host.get("hostnames", []) if h.get("name")]
    for hn in hostnames:
        if is_probable_fqdn(hn):
            return hn.rstrip(".")
    for hn in hostnames:
        if hn:
            return hn.rstrip(".")
    for addr in host.get("addresses", []):
        if addr.get("addrtype") == "ipv4" and addr.get("addr"):
            return addr["addr"]
    for addr in host.get("addresses", []):
        if addr.get("addr"):
            return addr["addr"]
    return "unknown-host"


def get_primary_ipv4(host: dict) -> str | None:
    for addr in host.get("addresses", []):
        if addr.get("addrtype") == "ipv4" and addr.get("addr"):
            return addr["addr"]
    return None


_EXISTING_PORT_RE = re.compile(
    r"^- \*\*(\w+)/(\d+)\*\* — (.*)$"
)
_EXISTING_CHECKBOX_RE = re.compile(
    r"^\s+- \[( |x|/)\] Investigate: (.+)$"
)


def _parse_existing_ports(open_ports_section: str) -> tuple[dict, dict]:
    """Parse an existing ## Open Ports section back into structured data.

    Returns:
        (port_data, checkbox_states)
        port_data: {(protocol, port_num): service_string}
        checkbox_states: {(protocol, port_num): checkbox_char}  (' ', 'x', '/')
    """
    port_data: dict[tuple[str, int], str] = {}
    checkbox_states: dict[tuple[str, int], str] = {}
    current_key: tuple[str, int] | None = None

    for line in open_ports_section.splitlines():
        m = _EXISTING_PORT_RE.match(line.strip())
        if m:
            proto, port_str, svc_str = m.group(1), m.group(2), m.group(3)
            current_key = (proto, int(port_str))
            port_data[current_key] = svc_str.strip()
            continue

        m = _EXISTING_CHECKBOX_RE.match(line)
        if m and current_key:
            checkbox_states[current_key] = m.group(1)

    return port_data, checkbox_states


def _service_richness(svc: dict) -> int:
    """Score how much detail a service dict has — higher is richer."""
    score = 0
    if svc.get("name"):     score += 1
    if svc.get("product"):  score += 2
    if svc.get("version"):  score += 3
    if svc.get("extrainfo"): score += 1
    if svc.get("tunnel"):   score += 1
    return score


def _merge_port_lists(
    existing_ports_section: str,
    new_ports: list[dict],
) -> tuple[list[dict], dict]:
    """Merge existing port data with new scan ports.

    For each (protocol, port):
    - If only in new scan: use new
    - If only in existing: keep existing (as a synthetic port dict)
    - If in both: keep whichever has richer service detail

    Returns: (merged_ports, preserved_checkbox_states)
    """
    existing_data, checkbox_states = _parse_existing_ports(existing_ports_section)

    # Index new ports by key
    new_by_key: dict[tuple[str, int], dict] = {}
    for p in new_ports:
        key = (p["protocol"], p["port"])
        new_by_key[key] = p

    merged: dict[tuple[str, int], dict] = {}

    # Start with all existing ports
    for key, svc_str in existing_data.items():
        if key in new_by_key:
            new_p = new_by_key[key]
            new_richness = _service_richness(new_p.get("service", {}))
            # Existing is just a rendered string — estimate its richness by token count
            existing_richness = len([t for t in svc_str.split() if t != "—"])
            if new_richness >= existing_richness:
                merged[key] = new_by_key[key]
            else:
                # Keep existing as synthetic port dict
                merged[key] = {
                    "protocol": key[0],
                    "port": key[1],
                    "service": {"name": svc_str},
                    "scripts": [],
                    "_synthetic": True,
                }
        else:
            # Port only in existing note — preserve it
            merged[key] = {
                "protocol": key[0],
                "port": key[1],
                "service": {"name": svc_str},
                "scripts": [],
                "_synthetic": True,
            }

    # Add new ports not in existing
    for key, p in new_by_key.items():
        if key not in merged:
            merged[key] = p

    sorted_ports = [
        merged[k] for k in sorted(merged.keys())
    ]
    return sorted_ports, checkbox_states


def _summarize_open_ports_merged(
    merged_ports: list[dict],
    checkbox_states: dict[tuple[str, int], str],
) -> list[str]:
    """Render merged ports, preserving checkbox states from previous runs."""
    lines: list[str] = []
    for p in merged_ports:
        port_key = (p["protocol"], p["port"])
        port_label = f"{p['protocol']}/{p['port']}"

        if p.get("_synthetic"):
            svc = p["service"].get("name", "—")
        else:
            svc = render_service_string(p)

        svc_name = (p.get("service", {}).get("name", "") or "unknown")
        if not p.get("_synthetic"):
            svc_name = svc_name.upper()
        else:
            # For synthetic ports, extract the first word as service name
            svc_name = svc_name.split()[0].upper() if svc_name and svc_name != "—" else "UNKNOWN"

        lines.append(f"- **{port_label}** — {svc}")

        cb = checkbox_states.get(port_key, " ")
        lines.append(f"  - [{cb}] Investigate: {svc_name} ({port_label})")

    return lines


def render_service_string(port_info: dict) -> str:
    svc = port_info.get("service", {})
    parts = [svc.get("name", "")]
    if svc.get("product"):   parts.append(svc["product"])
    if svc.get("version"):   parts.append(svc["version"])
    if svc.get("extrainfo"): parts.append(f"({svc['extrainfo']})")
    if svc.get("tunnel"):    parts.append(f"[tunnel: {svc['tunnel']}]")
    return " ".join(p for p in parts if p).strip() or "—"


def summarize_open_ports(host: dict) -> list[str]:
    lines: list[str] = []
    for p in host.get("open_ports", []):
        svc = render_service_string(p)
        svc_name = (p.get("service", {}).get("name", "") or "unknown").upper()
        port_label = f"{p['protocol']}/{p['port']}"
        lines.append(f"- **{port_label}** — {svc}")
        lines.append(f"  - [ ] Investigate: {svc_name} ({port_label})")
    return lines


# ============================================================
# Ollama — Nmap
# ============================================================

def build_ollama_prompt(scan_data: dict, operator_notes_by_ip: dict[str, str] | None = None) -> str:
    hosts = scan_data.get("hosts", [])

    instructions = """
You are an experienced penetration tester analyzing Nmap scan results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.

Priorities:
- Identify notable exposed services and likely attack surface
- Suggest useful follow-up enumeration tools with a ready-to-run example command
  for each, substituting the actual target IP or hostname from the scan data
  (short, focused commands preferred)
- Highlight likely misconfigurations, risky exposures, or common weaknesses
- Infer likely technology stacks when reasonable, clearly tagging [INFERRED]
- Prefer real-world offensive security workflow recommendations over generic advice

Active Directory detection:
- If you observe Kerberos (88), LDAP (389/636/3268), DNS (53), and SMB (445)
  across one or more hosts, explicitly identify this as an Active Directory
  environment in the Environment Assessment section and tailor all attack path
  and enumeration suggestions to AD-specific methodology:
  AS-REP roasting, Kerberoasting, BloodHound, Pass-the-Hash, Pass-the-Ticket,
  DCSync, LDAP anonymous bind enumeration, GPO abuse, etc.

Cross-host correlation:
- Identify patterns across multiple hosts. If a finding applies network-wide
  (e.g., SMB signing disabled everywhere, same service version on all hosts),
  report it once as a network-level finding rather than repeating it per host.

When suggesting tools, favor practical enumeration tools such as:
- SMB: netexec, smbclient, smbmap, enum4linux-ng
- LDAP: ldapsearch, netexec, bloodyAD
- Kerberos: kerbrute, impacket tools (GetUserSPNs, GetNPUsers), netexec
- MSSQL: impacket-mssqlclient, netexec
- WinRM: evil-winrm, netexec
- RDP: netexec, xfreerdp
- SSH: ssh-audit, hydra
- Web: whatweb, nikto, gobuster, feroxbuster, ffuf, curl
- SNMP: snmpwalk, onesixtyone
- DNS: dig, nslookup, dnsrecon
- NFS: showmount, mount
- SMTP: swaks, smtp-user-enum
- FTP: ftp, Nmap NSE scripts
- RPC: rpcclient, netexec
- VNC: vncsnapshot, hydra
""".strip()

    lines = [
        f"Nmap args: {scan_data.get('nmap_args', '')}",
        f"Nmap version: {scan_data.get('nmap_version', '')}",
        f"Hosts count: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        name  = choose_host_display_name(host)
        state = host.get("state", "unknown")
        ip    = get_primary_ipv4(host) or "unknown"
        lines += [f"- Host: {name}", f"  State: {state}", f"  IP: {ip}"]
        open_ports = host.get("open_ports", [])
        if not open_ports:
            lines += ["  Open ports: none", ""]
            continue
        lines.append("  Open ports:")
        for p in open_ports:
            svc      = p.get("service", {})
            svc_name = svc.get("name", "") or "unknown"
            product  = svc.get("product", "")
            version  = svc.get("version", "")
            extrainfo= svc.get("extrainfo", "")
            tunnel   = svc.get("tunnel", "")
            parts = [svc_name]
            if product:   parts.append(product)
            if version:   parts.append(version)
            if extrainfo: parts.append(f"({extrainfo})")
            if tunnel:    parts.append(f"[tunnel: {tunnel}]")
            lines.append(f"    - {p['protocol']}/{p['port']}: {' '.join(parts).strip()}")
            for script in p.get("scripts", [])[:3]:
                out = " ".join((script.get("output", "") or "").split())
                if out:
                    if len(out) > 220:
                        out = out[:217] + "..."
                    lines.append(f"      script:{script.get('id', '')} -> {out}")
            for hint in get_port_hints(p["port"], svc_name, product, extrainfo):
                lines.append(f"      hint: {hint}")
        if operator_notes_by_ip and ip in operator_notes_by_ip:
            lines.append(_format_operator_notes_for_prompt(operator_notes_by_ip[ip]))
        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections in this order:

## Environment Assessment
- State the inferred environment type (Active Directory domain, standalone Linux
  hosts, mixed Windows/Linux, web application stack, OT/ICS, etc.)
- Provide a confidence level: HIGH / MEDIUM / LOW
- List the specific evidence from the scan data that supports the assessment
- If AD is detected, state the inferred domain name if visible in hostnames or
  DNS responses

## Key Observations
- Brief bullets of the most important findings
- Each bullet must include a [CONFIRMED], [INFERRED], or [ASSUMED] tag

## Enumeration Suggestions
- Group by service or host
- For each suggestion, include a ready-to-run example command with the actual
  target IP or hostname substituted in
- Tag each with [CONFIRMED], [INFERRED], or [ASSUMED]

## Potential Attack Paths
- Realistic next steps or attack paths with supporting scan evidence for each step
- Tag each path as [CONFIRMED], [INFERRED], or [ASSUMED]
- Do not include a path if the scan data does not support it

## Notable Risks or Misconfigurations
- Network-wide findings first, then per-host
- Tag each with [CONFIRMED], [INFERRED], or [ASSUMED]
""".strip()

    # RAG context injection
    services = []
    for h in hosts:
        for p in h.get("open_ports", []):
            svc = p.get("service", {})
            if svc.get("name"):
                services.append(svc["name"])
            if svc.get("product"):
                services.append(svc["product"])
    rag_context = _get_rag_context(_build_rag_query_from_services(services), top_k=5)

    rag_section = f"\n\n{rag_context}" if rag_context else ""
    prompt = f"{instructions}{rag_section}\n\nNmap Scan Summary\n=================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Nmap Ollama prompt ({len(prompt)} chars, RAG={'yes' if rag_context else 'no'})")
    return prompt


# ============================================================
# Ollama — Nessus
# ============================================================

def build_nessus_ollama_prompt(nessus_data: dict, operator_notes_by_ip: dict[str, str] | None = None) -> str:
    hosts = nessus_data.get("hosts", [])

    instructions = """
You are an experienced penetration tester analyzing Nessus vulnerability scan results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.
- Nessus plugin findings can produce false positives, particularly for
  version-based checks. If a finding relies solely on version detection without
  plugin output confirming exploitation, note it as
  [VERIFY — possible false positive].

Priorities:
- Identify the highest-impact vulnerabilities by CVSS score and exploitability
- For each CVE, tag its weaponization status:
    [MSF]         — known Metasploit module exists
    [POC]         — public PoC exists but may require adaptation
    [THEORETICAL] — no known public exploit at time of analysis
- Flag any findings suggesting credential exposure, default credentials,
  cleartext protocols, or password policy weaknesses — these often have
  outsized impact even when CVSS scores are moderate
- Version-based detection findings (where plugin_output is absent or says only
  "version X detected") must be flagged [VERIFY — version-based only]
- Cross-host correlation: if the same critical/high finding appears on multiple
  hosts, group them and call out the network-wide scope rather than listing each
  host separately
- Group findings by exploitation priority: critical/weaponized first, then
  high, then medium
""".strip()

    lines = [
        f"Report: {nessus_data.get('report_name', '')}",
        f"Hosts: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        ip       = host.get("ip", "unknown")
        hostname = host.get("hostname", "")
        findings = host.get("findings", [])

        label = f"{hostname} ({ip})" if hostname and hostname != ip else ip
        lines.append(f"- Host: {label}")

        # Skip info (severity 0) in AI prompt to reduce noise
        notable = [f for f in findings if f["severity_int"] >= 1]
        if not notable:
            lines.append("  Findings: none above informational")
            lines.append("")
            continue

        # Count by severity
        counts = {4: 0, 3: 0, 2: 0, 1: 0}
        for f in notable:
            counts[f["severity_int"]] = counts.get(f["severity_int"], 0) + 1
        lines.append(
            f"  Critical:{counts[4]} High:{counts[3]} Medium:{counts[2]} Low:{counts[1]}"
        )
        lines.append("  Top findings:")

        for f in notable[:15]:
            sev      = severity_int_to_str(f["severity_int"])
            cvss     = f.get("cvss3_base") or f.get("cvss_base") or "N/A"
            cves     = ", ".join(f["cves"]) if f["cves"] else "no CVE"
            has_out  = "plugin_output:yes" if f.get("plugin_output") else "plugin_output:none"
            lines.append(
                f"    [{sev}] {f['plugin_name']} | CVSS:{cvss} | {cves} | port {f['protocol']}/{f['port']} | {has_out}"
            )
            if f.get("solution"):
                sol = f["solution"][:200].replace("\n", " ")
                lines.append(f"      solution: {sol}")

        if operator_notes_by_ip and ip in operator_notes_by_ip:
            lines.append(_format_operator_notes_for_prompt(operator_notes_by_ip[ip]))
        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections:

## Key Observations
- Most impactful vulnerabilities and what they mean for the engagement
- Each bullet must include a [CONFIRMED], [INFERRED], or [ASSUMED] tag
- Group any finding that appears on multiple hosts as a single network-wide bullet

## Exploitation Priority
- Rank the top exploitable findings with reasoning
- For each CVE include its [MSF], [POC], or [THEORETICAL] tag
- Mark version-based-only detections with [VERIFY — version-based only]
- Flag credential/cleartext findings prominently

## Attack Chains
- How findings could be combined for privilege escalation or lateral movement
- Each step must cite specific scan evidence; tag with [CONFIRMED]/[INFERRED]/[ASSUMED]

## Remediation Focus
- Highest-priority patches or configuration fixes
- Network-wide issues first, then per-host
""".strip()

    prompt = f"{instructions}\n\nNessus Scan Summary\n===================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Nessus Ollama prompt ({len(prompt)} chars)")
    return prompt


# ============================================================
# Ollama — Nessus two-pass helpers
# ============================================================

def _build_nessus_fact_extraction_prompt(nessus_data: dict, operator_notes_by_ip: dict[str, str] | None = None) -> str:
    """
    Pass 1 prompt: instruct the model to extract ONLY facts explicitly present
    in the Nessus data — no inference, no recommendations.
    """
    hosts = nessus_data.get("hosts", [])
    lines = [
        "You are a data extraction assistant. From the following Nessus scan data, "
        "extract ONLY facts that are explicitly present. Do not infer, recommend, or analyze. "
        "Output a structured list: for each host, list confirmed CVEs with their CVSS scores, "
        "confirmed severity levels, and confirmed affected ports/services. "
        "Mark everything as [CONFIRMED]. Do not add any information not present in the data.",
        "",
        "Nessus data:",
    ]
    for host in hosts:
        ip       = host.get("ip", "unknown")
        hostname = host.get("hostname", "")
        label    = f"{hostname} ({ip})" if hostname and hostname != ip else ip
        lines.append(f"\nHost: {label}")
        notable = [f for f in host.get("findings", []) if f["severity_int"] >= 1]
        if not notable:
            lines.append("  No notable findings.")
            continue
        for f in notable[:20]:
            sev  = severity_int_to_str(f["severity_int"])
            cvss = f.get("cvss3_base") or f.get("cvss_base") or "N/A"
            cves = ", ".join(f["cves"]) if f["cves"] else "none"
            lines.append(
                f"  [{sev}] {f['plugin_name']} | CVSS:{cvss} | CVE:{cves} "
                f"| port:{f['protocol']}/{f['port']}"
            )
    return "\n".join(lines)


# ============================================================
# Ollama — Burp Suite
# ============================================================

def build_burp_ollama_prompt(burp_data: dict, operator_notes_by_ip: dict[str, str] | None = None) -> str:
    hosts = burp_data.get("hosts", [])

    instructions = """
You are an experienced web application penetration tester analyzing Burp Suite scanner results.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

Your job is to produce practical, concise, operator-focused analysis.

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- If evidence is insufficient to make a claim, say "insufficient data" rather
  than guessing.
- Tag each finding or suggestion as one of:
    [CONFIRMED]  — directly present in the scan data
    [INFERRED]   — reasonable conclusion from the scan data
    [ASSUMED]    — requires verification; explain why
- Do not suggest an attack path is viable unless the scan data provides
  supporting evidence for each step.
- Burp "Tentative" findings have a high false positive rate. Always flag these
  as [TENTATIVE — VERIFY MANUALLY] and do not build attack chains primarily on
  tentative findings.

Priorities:
- Identify the most exploitable web vulnerabilities (SQLi, XSS, SSRF, IDOR,
  auth bypass, etc.) and classify each by its OWASP Top 10 2021 category
  (e.g., A01:Broken Access Control, A03:Injection, A07:Auth Failures, etc.)
- For each Certain or Firm finding, suggest a specific manual verification step
  or payload to confirm exploitability beyond the scanner result
- Suggest how findings could be chained for greater impact
- Assess the web attack surface: authentication, input handling, business logic
- Highlight systemic issues (framework versions, CSP absence, insecure headers)
- Be explicit about what the scanner confirmed vs. what requires manual testing
""".strip()

    lines = [
        f"Hosts scanned: {len(hosts)}",
        "",
        "Host details:",
    ]

    for host in hosts:
        ip     = host.get("ip", "")
        url    = host.get("url", "")
        issues = host.get("issues", [])
        label  = url if url else ip
        if ip and url and ip not in url:
            label = f"{url} ({ip})"

        lines.append(f"- Host: {label}")

        notable = [i for i in issues if i["severity"].lower() not in ("information", "info")]
        if not notable:
            lines.append("  Issues: informational only")
            lines.append("")
            continue

        counts: dict[str, int] = {}
        for i in issues:
            counts[i["severity"]] = counts.get(i["severity"], 0) + 1
        lines.append(
            "  " + " ".join(f"{k}:{v}" for k, v in sorted(counts.items()))
        )
        lines.append("  Issues:")
        for i in notable[:15]:
            conf = i.get("confidence", "")
            loc  = i.get("path") or i.get("location") or ""
            lines.append(f"    [{i['severity']}|{conf}] {i['name']} @ {loc}")
            if i.get("issue_detail"):
                detail = i["issue_detail"][:200].replace("\n", " ")
                lines.append(f"      detail: {detail}")
        if operator_notes_by_ip and ip and ip in operator_notes_by_ip:
            lines.append(_format_operator_notes_for_prompt(operator_notes_by_ip[ip]))
        lines.append("")

    output_format = """
Return your response in markdown using exactly these sections:

## Key Observations
- Most significant web vulnerabilities and their potential impact
- Each bullet must include its OWASP Top 10 2021 category and a
  [CONFIRMED], [INFERRED], or [ASSUMED] tag
- Tentative findings must be tagged [TENTATIVE — VERIFY MANUALLY]

## Exploitation Priority
- Which issues to pursue first and why
- For each Certain/Firm finding, include a specific manual verification step
  or payload (e.g., exact parameter to test, HTTP method, expected response)
- Do not elevate Tentative findings above Certain/Firm without explanation

## Attack Chains
- How findings could be combined for greater impact
- Each step must cite specific scanner evidence; tag [CONFIRMED]/[INFERRED]/[ASSUMED]
- Do not build chains primarily on Tentative findings

## False Positive Risk Assessment
- List findings that are likely false positives, with reasoning
- Consider: finding type, confidence level (Tentative = higher FP risk), context,
  and whether the issue_detail provides concrete evidence or just a generic description

## Remediation Focus
- Top development/configuration fixes to eliminate the highest-impact issues
""".strip()

    prompt = f"{instructions}\n\nBurp Suite Scan Summary\n=======================\n{chr(10).join(lines)}\n\n{output_format}"
    logging.debug(f"Built Burp Ollama prompt ({len(prompt)} chars)")
    return prompt


# ============================================================
# Ollama — AutoRecon
# ============================================================

def _build_autorecon_fact_extraction_prompt(target_data: dict, operator_notes: str = "") -> str:
    """Pass 1: present Python-extracted structured data; ask LLM to organize only."""
    target = target_data.get("target", "unknown")
    ip = target_data.get("ip", target)
    hostname = target_data.get("hostname", "")
    label = f"{ip} ({hostname})" if hostname else ip

    lines: list[str] = [f"Target: {label}", ""]

    for port_key in sorted(target_data.get("tool_results", {})):
        results = target_data["tool_results"][port_key]
        lines.append(f"Port {port_key}:")
        for entry in results:
            data = entry.get("data", {})
            tool = data.get("tool", entry.get("tool", "unknown"))

            if tool == "dirbusting":
                interesting = data.get("interesting", [])[:20]
                if interesting:
                    lines.append(f"  Directory enumeration ({data.get('total_found', 0)} total):")
                    for p in interesting:
                        redir = f" -> {p['redirect']}" if p.get("redirect") else ""
                        lines.append(f"    {p['status']}  {p['path']}  ({p['size']}B){redir}")
            elif tool == "nikto":
                if data.get("server"):
                    lines.append(f"  Nikto server: {data['server']}")
                for h_name, h_val in data.get("headers", {}).items():
                    lines.append(f"  Header: {h_name}: {h_val}")
                for f in data.get("findings", [])[:15]:
                    osvdb = f"OSVDB-{f['osvdb']}: " if f.get("osvdb") else ""
                    lines.append(f"  Nikto: {osvdb}{f['path']} — {f['description'][:200]}")
            elif tool == "enum4linux":
                if data.get("os_info"):
                    lines.append(f"  OS: {data['os_info']}")
                if data.get("domain"):
                    lines.append(f"  Domain: {data['domain']}")
                lines.append(f"  Null session: {'YES' if data.get('null_session') else 'No'}")
                if data.get("users"):
                    lines.append(f"  Users ({len(data['users'])}): {', '.join(data['users'][:20])}")
                if data.get("shares"):
                    for s in data["shares"]:
                        lines.append(f"  Share: {s['name']} — {s.get('access', s.get('type', ''))}")
                if data.get("groups"):
                    lines.append(f"  Groups: {', '.join(data['groups'][:15])}")
                pp = data.get("password_policy", {})
                if pp:
                    parts = []
                    if "min_length" in pp:
                        parts.append(f"min_length={pp['min_length']}")
                    if "lockout_threshold" in pp:
                        parts.append(f"lockout={pp['lockout_threshold']}")
                    if pp.get("complexity"):
                        parts.append("complexity=on")
                    if parts:
                        lines.append(f"  Password policy: {', '.join(parts)}")
            elif tool == "smbmap":
                for s in data.get("shares", []):
                    lines.append(f"  Share: {s['name']} — {s['permissions']}")
            elif tool == "smbclient":
                for s in data.get("shares", []):
                    lines.append(f"  Share: {s['name']} ({s['type']})")
            elif tool == "whatweb":
                techs = data.get("technologies", [])
                if techs:
                    parts = [f"{t['name']}/{t['version']}" if t.get("version") else t["name"]
                             for t in techs]
                    lines.append(f"  Technologies: {', '.join(parts)}")
                if data.get("title"):
                    lines.append(f"  Title: {data['title']}")
            elif tool == "snmpwalk":
                if data.get("sys_descr"):
                    lines.append(f"  SNMP sysDescr: {data['sys_descr']}")
                if data.get("sys_name"):
                    lines.append(f"  SNMP sysName: {data['sys_name']}")
                if data.get("running_processes"):
                    lines.append(f"  Processes: {', '.join(data['running_processes'][:15])}")
                lines.append(f"  Total OIDs: {data.get('raw_oid_count', 0)}")
            elif tool == "onesixtyone":
                for cs in data.get("community_strings", []):
                    lines.append(f"  SNMP community: {cs['community']} — {cs['sys_descr'][:100]}")
            elif tool == "sslscan":
                if data.get("weak_protocols"):
                    lines.append(f"  Weak TLS: {', '.join(data['weak_protocols'])}")
                if data.get("weak_ciphers"):
                    lines.append(f"  Weak ciphers: {', '.join(data['weak_ciphers'][:5])}")
                if data.get("cert_subject"):
                    lines.append(f"  Cert: {data['cert_subject']}")
                if data.get("cert_expired"):
                    lines.append(f"  Cert expired: {data.get('cert_not_after', 'yes')}")
                if data.get("heartbleed_vulnerable"):
                    lines.append("  HEARTBLEED: VULNERABLE")
            elif tool == "dnsrecon":
                for r in data.get("records", [])[:20]:
                    lines.append(f"  DNS {r['type']}: {r['name']} -> {r['value']}")
                if data.get("zone_transfer_successful"):
                    lines.append("  ZONE TRANSFER: SUCCESSFUL")
            elif tool == "curl":
                if data.get("title"):
                    lines.append(f"  Page title: {data['title']}")
                for h_name, h_val in data.get("headers", {}).items():
                    lines.append(f"  Header: {h_name}: {h_val}")
        lines.append("")

    instructions = """You are a data extraction assistant. Your ONLY job is to organize
and deduplicate the following AutoRecon enumeration data into a clean, structured summary.

RULES:
- Do NOT analyze, infer, recommend, or draw conclusions.
- Do NOT add any information not present in the data below.
- Organize by port/service. Combine duplicate findings across tools.
- Tag every item as [CONFIRMED] — it comes directly from tool output.
- Output a structured list, not prose."""

    return f"{instructions}\n\n{chr(10).join(lines)}"


def build_autorecon_ollama_prompt(target_data: dict, operator_notes: str = "") -> str:
    """Pass 2: analyze structured AutoRecon findings with grounding rules."""
    target = target_data.get("target", "unknown")
    ip = target_data.get("ip", target)
    hostname = target_data.get("hostname", "")
    label = f"{ip} ({hostname})" if hostname else ip
    summary = target_data.get("summary", {})

    lines: list[str] = [
        f"Target: {label}",
        f"Tools run: {summary.get('total_tools_run', 0)}  |  "
        f"With findings: {summary.get('tools_with_findings', 0)}",
    ]
    if summary.get("technologies"):
        lines.append(f"Technologies: {', '.join(summary['technologies'][:15])}")
    if summary.get("null_session"):
        lines.append("Null session: YES")
    if summary.get("writable_shares"):
        lines.append(f"Writable shares: {', '.join(summary['writable_shares'])}")
    if summary.get("users_found"):
        lines.append(f"Users found ({len(summary['users_found'])}): "
                      f"{', '.join(summary['users_found'][:20])}")
    if summary.get("weak_tls"):
        lines.append(f"Weak TLS: {', '.join(summary['weak_tls'])}")
    if summary.get("community_strings"):
        lines.append(f"SNMP communities: {', '.join(summary['community_strings'])}")
    lines.append("")

    for port_key in sorted(target_data.get("tool_results", {})):
        results = target_data["tool_results"][port_key]
        lines.append(f"--- Port {port_key} ---")
        for entry in results:
            data = entry.get("data", {})
            tool = data.get("tool", entry.get("tool", "unknown"))

            if tool == "dirbusting":
                interesting = data.get("interesting", [])[:20]
                if interesting:
                    lines.append(f"[dirbusting] {data.get('total_found', 0)} paths found. Notable:")
                    for p in interesting[:15]:
                        lines.append(f"  {p['status']}  {p['path']}  ({p['size']}B)")
            elif tool == "nikto":
                lines.append(f"[nikto] {data.get('total_findings', 0)} findings "
                              f"(server: {data.get('server', 'unknown')})")
                for f in data.get("findings", [])[:10]:
                    osvdb = f"OSVDB-{f['osvdb']} " if f.get("osvdb") else ""
                    lines.append(f"  {osvdb}{f['path']}: {f['description'][:150]}")
            elif tool == "enum4linux":
                lines.append(f"[enum4linux] OS={data.get('os_info', '?')} "
                              f"Domain={data.get('domain', '?')} "
                              f"NullSession={'YES' if data.get('null_session') else 'no'}")
                if data.get("users"):
                    lines.append(f"  Users: {', '.join(data['users'][:20])}")
                if data.get("shares"):
                    for s in data["shares"]:
                        lines.append(f"  Share: {s['name']} [{s.get('access', s.get('type', ''))}]")
                pp = data.get("password_policy", {})
                if pp:
                    lines.append(f"  PwPolicy: {pp}")
            elif tool == "smbmap":
                for s in data.get("shares", []):
                    lines.append(f"[smbmap] {s['name']}: {s['permissions']}")
            elif tool == "smbclient":
                for s in data.get("shares", []):
                    lines.append(f"[smbclient] {s['name']} ({s['type']})")
            elif tool == "whatweb":
                techs = data.get("technologies", [])
                if techs:
                    parts = [f"{t['name']}/{t['version']}" if t.get("version") else t["name"]
                             for t in techs[:10]]
                    lines.append(f"[whatweb] {', '.join(parts)}")
            elif tool == "snmpwalk":
                if data.get("sys_descr"):
                    lines.append(f"[snmpwalk] sysDescr: {data['sys_descr']}")
                if data.get("running_processes"):
                    lines.append(f"  Processes: {', '.join(data['running_processes'][:10])}")
            elif tool == "onesixtyone":
                for cs in data.get("community_strings", []):
                    lines.append(f"[onesixtyone] community={cs['community']}")
            elif tool == "sslscan":
                parts = []
                if data.get("weak_protocols"):
                    parts.append(f"weak_proto={','.join(data['weak_protocols'])}")
                if data.get("cert_expired"):
                    parts.append("cert_expired")
                if data.get("heartbleed_vulnerable"):
                    parts.append("HEARTBLEED")
                if data.get("cert_subject"):
                    parts.append(f"cert={data['cert_subject']}")
                if parts:
                    lines.append(f"[sslscan] {' | '.join(parts)}")
            elif tool == "dnsrecon":
                lines.append(f"[dnsrecon] {len(data.get('records', []))} records")
                if data.get("zone_transfer_successful"):
                    lines.append("  ZONE TRANSFER SUCCESSFUL")
                for r in data.get("records", [])[:10]:
                    lines.append(f"  {r['type']} {r['name']} -> {r['value']}")
            elif tool == "curl":
                if data.get("title"):
                    lines.append(f"[curl] title={data['title']}")
                for h_name in data.get("interesting_headers", []):
                    lines.append(f"  {h_name}: {data['headers'].get(h_name, '')}")
        lines.append("")

    instructions = f"""You are an experienced penetration tester analyzing AutoRecon
enumeration results for target {label}.

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- All data below was pre-extracted from tool outputs by an automated parser.
  Every finding is [CONFIRMED] unless stated otherwise.
- Only reference paths, shares, users, technologies, and findings explicitly
  present in the data. Do not invent CVEs, paths, or services.
- If evidence is insufficient, say "insufficient data" — do not guess.
- Tag each claim: [CONFIRMED] = in the data, [INFERRED] = reasonable
  conclusion from confirmed data, [ASSUMED] = requires verification.
- Nikto has a high false-positive rate. Tag unverified nikto-only findings
  as [NIKTO — VERIFY MANUALLY].
- Provide ready-to-run commands with actual target IP {ip} substituted.
- Do not suggest an attack path unless enumeration data supports each step.
- Cross-correlate findings across tools (e.g., whatweb tech + nikto vuln +
  gobuster path = confirmed attack surface).
- Prioritize: writable shares, null sessions, admin panels, weak TLS,
  exposed credentials, directory listings, version-specific vulnerabilities."""

    output_format = """
Respond using EXACTLY these section headers:

## Key Observations
- Most actionable findings; cross-tool correlation where possible
- Each bullet: [CONFIRMED]/[INFERRED]/[ASSUMED] tag

## Service-Specific Findings
- Grouped by port/service
- What was confirmed vs needs manual verification
- Discovered paths with security implications (admin panels, backups, APIs)

## Credential & Access Findings
- Null sessions, user lists, password policies, writable shares
- Default credentials to test based on confirmed technologies

## Attack Paths
- How enumeration findings chain together for exploitation
- Each step must cite specific tool evidence
- Tag [CONFIRMED]/[INFERRED]/[ASSUMED]

## Manual Follow-Up
- Specific commands to run next, with actual IPs/hostnames
- Prioritized by likely impact
""".strip()

    if operator_notes:
        lines.append(_format_operator_notes_for_prompt(operator_notes))

    prompt = (f"{instructions}\n\nAutoRecon Enumeration Summary\n"
              f"{'=' * 30}\n{chr(10).join(lines)}\n\n{output_format}")
    logging.debug(f"Built AutoRecon Ollama prompt ({len(prompt)} chars)")
    return prompt


# ============================================================
# Ollama — Loot
# ============================================================

def _build_host_context_for_loot(hosts_dir: Path, host_key: str) -> dict | None:
    """Read existing host note to build context for loot correlation."""
    note_path = _find_host_note_by_ip(hosts_dir, host_key) if IPV4_RE.match(host_key) else None
    if not note_path:
        for hp in hosts_dir.glob("*.md"):
            if hp.stem.lower() == host_key.lower():
                note_path = hp
                break
    if not note_path or not note_path.exists():
        return None

    fm, body = read_frontmatter(note_path.read_text(encoding="utf-8"))
    ports_section = extract_body_section(body, "## Open Ports")
    port_lines = [l.strip() for l in ports_section.splitlines() if l.strip().startswith("- **")]

    return {
        "ip": fm.get("ip", ""),
        "hostnames": fm.get("hostnames", []),
        "tags": fm.get("tags", []),
        "open_ports_summary": port_lines[:20],
        "nessus_max_severity": fm.get("nessus_max_severity", 0),
    }


def build_loot_ollama_prompt(
    host_ip: str,
    loot_files: list[dict],
    host_context: dict | None = None,
    operator_notes: str = "",
) -> str:
    """Build a prompt for analyzing loot data associated with a host or campaign."""
    is_campaign = host_ip == "campaign"
    label = "Campaign-wide (not host-specific)" if is_campaign else host_ip

    lines: list[str] = [f"Target: {label}", ""]

    # Present credentials
    all_creds = []
    for lf in loot_files:
        for c in lf.get("credentials", []):
            all_creds.append({**c, "source": lf["filename"]})
    if all_creds:
        lines.append(f"Credentials found ({len(all_creds)}):")
        for c in all_creds[:30]:
            lines.append(f"  {c['username']}:{c['password']} [{c['cred_type']}] (from {c['source']})")
        lines.append("")
    else:
        lines.append("Credentials found: NONE")
        lines.append("")

    # Present hashes
    all_hashes = []
    for lf in loot_files:
        for h in lf.get("hashes", []):
            all_hashes.append({**h, "source": lf["filename"]})
    if all_hashes:
        lines.append(f"Hashes found ({len(all_hashes)}):")
        for h in all_hashes[:20]:
            user_part = f" (user: {h['username']})" if h.get("username") else ""
            lines.append(f"  [{h['hash_type']}]{user_part} {h['hash'][:40]}... (from {h['source']})")
        lines.append("")
    else:
        lines.append("Hashes found: NONE")
        lines.append("")

    # Present file listings
    for lf in loot_files:
        if lf.get("file_listings"):
            lines.append(f"File listing from {lf['filename']} ({len(lf['file_listings'])} entries):")
            for entry in lf["file_listings"][:15]:
                lines.append(f"  {entry['permissions']}  {entry['size']:>8}  {entry['filename']}")
            lines.append("")

    # Raw previews for context
    for lf in loot_files[:10]:
        if lf.get("raw_preview") and lf["category"] in ("notes", "mixed", "config"):
            lines.append(f"--- {lf['filename']} ({lf['category']}) ---")
            lines.append(lf["raw_preview"][:500])
            lines.append("")

    # Host context
    if host_context:
        lines.append("Existing scan data for this host:")
        if host_context.get("tags"):
            lines.append(f"  Services: {', '.join(host_context['tags'])}")
        if host_context.get("open_ports_summary"):
            lines.append(f"  Open ports ({len(host_context['open_ports_summary'])}):")
            for p in host_context["open_ports_summary"][:10]:
                lines.append(f"    {p}")
        lines.append("")

    if operator_notes:
        lines.append(_format_operator_notes_for_prompt(operator_notes))

    instructions = f"""You are an experienced penetration tester analyzing loot collected
during an engagement for target {label}.

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- ONLY reference credentials, hashes, and file listings that appear in the
  "Loot Summary" section below. If it says "NONE", there are NONE — do not
  invent example credentials, usernames, passwords, or hashes.
- All data below was extracted by automated parsers. Items listed are [CONFIRMED].
- Cross-correlate loot with the host's scan data (open ports, services) to
  identify which services these credentials could authenticate to.
- If no credentials or hashes are present, focus your analysis on the file
  listings and raw file content instead. If nothing useful is present, say so.
- Tag each claim: [CONFIRMED] = in the data, [INFERRED] = reasonable conclusion,
  [ASSUMED] = requires verification.
- Suggest specific commands to test discovered credentials against known services,
  with actual IP {host_ip} substituted.
- For hashes, suggest appropriate cracking approaches (hashcat modes, wordlists).
- Assess password reuse risk if the same username appears with different passwords."""

    output_format = """
Respond using EXACTLY these section headers:

## Loot Analysis
- Significance of discovered credentials and hashes
- Cross-correlation with scan findings (which services accept these creds?)
- File listing analysis (sensitive files, config files, potential data exposure)

## Credential Impact
- Which credentials might grant access to which services
- Password reuse risk across hosts
- Privilege level assessment (admin, service account, user)

## Recommended Next Steps
- Specific commands to test credentials against services (with actual IPs)
- Hash cracking suggestions with hashcat modes
- Files to examine further
""".strip()

    prompt = (f"{instructions}\n\nLoot Summary\n{'=' * 30}\n"
              f"{chr(10).join(lines)}\n\n{output_format}")
    logging.debug(f"Built Loot Ollama prompt ({len(prompt)} chars)")
    return prompt


def _build_misc_minimal_prompt(filename: str, content: str) -> str:
    """Build a short prompt for notes/unknown files — extract key facts only."""
    display_content = content[:20_000]
    if len(content) > 20_000:
        display_content += f"\n\n[... truncated, {len(content)} total chars ...]"
    return f"""You are an experienced penetration tester reviewing a file from an engagement.
This appears to be notes or non-tool output. Extract ONLY factual items.

GROUNDING RULES:
- Respond ONLY in English.
- NEVER fabricate data. Only reference what is explicitly present.
- Keep your response SHORT — bullet points only.

FILENAME: {filename}

CONTENT:
```
{display_content}
```

Extract any of the following if present (omit sections with nothing found):

## Key Facts
- IPs, hostnames, URLs, or network ranges mentioned
- Credentials, usernames, or passwords
- Service versions or software names
- Any actionable observations
""".strip()


_MISC_TOOL_HINTS: dict[str, str] = {
    "nikto":        "This is Nikto web server scanner output. Focus on server misconfigurations, outdated software, dangerous HTTP methods, and interesting paths.",
    "gobuster":     "This is Gobuster directory/DNS brute-force output. Focus on discovered paths, status codes, and interesting endpoints.",
    "dirb":         "This is DIRB directory brute-force output. Focus on discovered paths and status codes.",
    "dirsearch":    "This is Dirsearch directory brute-force output. Focus on discovered paths, status codes, and redirect targets.",
    "feroxbuster":  "This is Feroxbuster directory brute-force output. Focus on discovered paths, interesting status codes, and large responses.",
    "ffuf":         "This is ffuf fuzzer output. Focus on discovered paths, parameter values, and anomalous responses.",
    "wpscan":       "This is WPScan WordPress scanner output. Focus on vulnerable plugins/themes, user enumeration, and WordPress version.",
    "linpeas":      "This is LinPEAS Linux privilege escalation scanner output. Focus on SUID binaries, writable files, credentials, cron jobs, and kernel exploits.",
    "winpeas":      "This is WinPEAS Windows privilege escalation scanner output. Focus on unquoted service paths, writable services, stored credentials, and UAC bypass vectors.",
    "enum4linux":   "This is enum4linux SMB/NetBIOS enumeration output. Focus on shares, users, groups, password policy, and domain info.",
    "hydra":        "This is Hydra brute-force output. Focus on successful logins and valid credentials.",
    "crackmapexec": (
        "This is CrackMapExec/NetExec (nxc/cme) output. "
        "For each host line extract: IP, hostname, OS/build, domain, signing status, SMBv1 status. "
        "Flag signing:False hosts as SMB relay candidates. "
        "Flag SMBv1:True hosts as legacy/EternalBlue-risk. "
        "Extract share names and permissions (READ/WRITE) per host. "
        "Extract any successful auth lines ([+]) and Pwn3d! results. "
        "Extract domain, workgroup, and OS info. "
        "Present per-host findings as a markdown table."
    ),
    "bloodhound":   "This is BloodHound/SharpHound output. Focus on attack paths, high-privilege users, and delegation issues.",
    "impacket":     "This is Impacket tool output. Focus on extracted hashes, tickets, secrets, and successful authentications.",
    "kerbrute":     "This is Kerbrute output. Focus on valid usernames and AS-REP roastable accounts.",
    "nuclei":       "This is Nuclei vulnerability scanner output. Focus on confirmed vulnerabilities, severity levels, and affected endpoints.",
    "testssl":      "This is testssl.sh output. Focus on weak ciphers, expired certificates, and protocol vulnerabilities.",
    "ldapsearch":   "This is ldapsearch output. Focus on user accounts, groups, SPNs, and sensitive attributes.",
}


def build_misc_ollama_prompt(
    filename: str,
    content: str,
    host_context: str = "",
    operator_notes: str = "",
    tool_type: str = "unknown",
    analysis_level: str = "standard",
) -> str:
    """Build a prompt for interpreting miscellaneous tool output."""
    if analysis_level == "minimal":
        return _build_misc_minimal_prompt(filename, content)

    target_line = ""
    if host_context:
        target_line = f"\nHOST CONTEXT:\n{host_context}\n"

    op_line = ""
    if operator_notes:
        op_line = f"\nOPERATOR NOTES:\n{operator_notes}\n"

    tool_hint = ""
    if tool_type in _MISC_TOOL_HINTS:
        tool_hint = f"\nTOOL HINT: {_MISC_TOOL_HINTS[tool_type]}\n"

    max_content = 100_000
    display_content = content[:max_content]
    if len(content) > max_content:
        display_content += f"\n\n[... truncated, {len(content)} total chars ...]"

    tool_id_section = ""
    if tool_type == "unknown":
        tool_id_section = """
## Tool Identification
Identify what tool produced this output (e.g., nmap, gobuster, linpeas, enum4linux).
"""
    else:
        tool_id_section = f"""
## Tool Identification
Detected tool: **{tool_type}**
"""

    # Tool-specific output format sections
    if tool_type == "crackmapexec":
        output_sections = """
## Host Inventory
Create a markdown table with one row per host. Columns:
| IP | Hostname | OS | Domain | Signing | SMBv1 | Notes |
Extract every value directly from the output. Use ✓/✗ for Signing/SMBv1 boolean fields.
Put any [+] auth successes or Pwn3d! results in the Notes column.

## Risk Flags
- **SMB Relay Candidates** (signing:False): list IPs — these are targets for NTLM relay attacks
- **SMBv1 Enabled** (SMBv1:True): list IPs — legacy protocol, potential EternalBlue exposure
- **Successful Authentications**: list any [+] lines verbatim
- **Admin Access / Pwn3d!**: list any hosts where admin or code execution was confirmed

## Share Enumeration
For each host that had shares listed, show: host IP/name → share name → permissions.
Only include shares explicitly present in the output.

## Recommended Next Steps
Provide ready-to-run nxc/netexec commands using actual IPs from the output.
Prioritize: relay setup for signing:False hosts, further enumeration of readable shares,
credential spraying if usernames were found.
"""
    elif tool_type in ("enum4linux", "smbmap", "smbclient"):
        output_sections = """
## Host Inventory
Summarize per-host: IP, hostname, domain, shares found, users/groups enumerated.

## Key Findings
- Shares with READ or WRITE access
- User accounts and groups discovered
- Password policy details
- Domain/workgroup information

## Recommended Next Steps
Provide ready-to-run follow-up commands with actual target IPs/hostnames.
"""
    elif tool_type in ("bloodhound", "impacket", "kerbrute"):
        output_sections = """
## Key Findings
- Credentials, hashes, or tickets extracted
- Valid usernames confirmed
- Attack paths or delegation issues identified
- High-privilege accounts or groups

## Notable Items
List specific accounts, hashes, SPNs, or paths worth acting on.

## Recommended Next Steps
Provide ready-to-run follow-up commands with actual targets from the output.
"""
    else:
        output_sections = """
## Key Findings
Extract and summarize the most important findings. Prioritize:
- Credentials or secrets discovered
- Vulnerabilities or misconfigurations
- Interesting services, paths, or endpoints
- Privilege escalation vectors
- Network information useful for pivoting

## Notable Items
List specific items worth investigating further (paths, usernames, versions, etc.)

## Recommended Next Steps
Provide ready-to-run follow-up commands with actual target IPs/hostnames.
"""

    prompt = f"""You are an experienced penetration tester analyzing tool output.
This analysis will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- Only reference findings that are explicitly present in the tool output.
- Do not invent CVEs, version numbers, or service details not shown.
- Tag each finding as [CONFIRMED], [INFERRED], or [ASSUMED].
- If evidence is insufficient, say "insufficient data" rather than guessing.
{tool_hint}
FILENAME: {filename}

TOOL OUTPUT:
```
{display_content}
```
{target_line}{op_line}
Analyze this tool output and provide:
{tool_id_section}{output_sections}""".strip()

    logging.debug(f"Built Misc Ollama prompt ({len(prompt)} chars, tool={tool_type}, level={analysis_level})")
    return prompt


_SPINNER_CHARS = "⠋⠙⠹⠸⠼⠴⠦⠧⠇⠏"


class _Spinner:
    """Background spinner that shows elapsed time during LLM calls."""

    def __init__(self, label: str = "Querying Ollama"):
        self._label = label
        self._stop = threading.Event()
        self._thread: threading.Thread | None = None
        self._start_time = 0.0

    def start(self) -> None:
        self._start_time = time.time()
        self._stop.clear()
        self._thread = threading.Thread(target=self._spin, daemon=True)
        self._thread.start()

    def _spin(self) -> None:
        idx = 0
        while not self._stop.is_set():
            elapsed = int(time.time() - self._start_time)
            ch = _SPINNER_CHARS[idx % len(_SPINNER_CHARS)]
            sys.stderr.write(f"\r  {ch} {self._label}... {elapsed}s")
            sys.stderr.flush()
            idx += 1
            self._stop.wait(0.15)
        sys.stderr.write("\r" + " " * 60 + "\r")
        sys.stderr.flush()

    def stop(self) -> float:
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=2)
        return time.time() - self._start_time


def ollama_chat(base_url: str, model: str, prompt: str, temperature: float = 0.15) -> str:
    url = base_url.rstrip("/") + "/api/generate"
    payload = {"model": model, "prompt": prompt, "stream": False, "temperature": temperature}

    spinner = _Spinner("Querying Ollama")
    spinner.start()
    try:
        r = requests.post(url, json=payload, timeout=300)
        r.raise_for_status()
        response_text = r.json().get("response", "")
    except KeyboardInterrupt:
        spinner.stop()
        logging.warning("LLM call cancelled by user (Ctrl+C)")
        raise
    finally:
        elapsed = spinner.stop()

    logging.info(f"Ollama responded ({len(response_text):,} chars, {elapsed:.0f}s)")
    return response_text


# ============================================================
# RAG — Retrieval Augmented Generation
# ============================================================

def ollama_embed(base_url: str, model: str, text: str) -> list[float]:
    """Get an embedding vector from Ollama's /api/embeddings endpoint."""
    url = base_url.rstrip("/") + "/api/embeddings"
    payload = {"model": model, "prompt": text}
    r = requests.post(url, json=payload, timeout=120)
    r.raise_for_status()
    return r.json()["embedding"]


def _cosine_similarity(a: list[float], b: list[float]) -> float:
    """Cosine similarity between two vectors. Pure Python."""
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = math.sqrt(sum(x * x for x in a))
    norm_b = math.sqrt(sum(x * x for x in b))
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)


def _rag_db_path() -> Path:
    """Return the canonical path for the RAG SQLite database."""
    return Path.cwd() / RAG_INDEX_DB_FILENAME


def _rag_db_connect(db_path: str | Path) -> sqlite3.Connection:
    """Open a RAG SQLite database with appropriate pragmas."""
    conn = sqlite3.connect(str(db_path), timeout=10)
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    return conn


def _init_rag_db(db_path: Path) -> sqlite3.Connection:
    """Initialize or open the RAG SQLite database. Creates schema if needed."""
    conn = _rag_db_connect(db_path)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS metadata (key TEXT PRIMARY KEY, value TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS file_mtimes (file_path TEXT PRIMARY KEY, mtime REAL NOT NULL);
        CREATE TABLE IF NOT EXISTS chunks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            source_path TEXT NOT NULL,
            source_type TEXT NOT NULL,
            location TEXT NOT NULL,
            title TEXT NOT NULL,
            text TEXT NOT NULL,
            embedding BLOB
        );
        CREATE INDEX IF NOT EXISTS idx_chunks_source_path ON chunks(source_path);
    """)
    conn.commit()
    return conn


def _encode_embedding_f16(vec: list[float]) -> bytes:
    """Pack a float vector as float16 BLOB for SQLite storage."""
    return struct.pack(f"{len(vec)}e", *vec)


def _decode_embedding_f16(blob: bytes, dim: int) -> list[float]:
    """Unpack a float16 BLOB back to a float vector."""
    return list(struct.unpack(f"{dim}e", blob))


def _finalize_rag_db(conn: sqlite3.Connection, embedding_model: str, embedding_dim: int) -> dict:
    """Update metadata counts after build completion. Returns lightweight metadata dict."""
    chunk_count = conn.execute("SELECT COUNT(*) FROM chunks").fetchone()[0]
    source_count = conn.execute("SELECT COUNT(DISTINCT source_path) FROM chunks").fetchone()[0]
    built_at = dt.datetime.now().isoformat(timespec="seconds")
    for key, val in [
        ("version", str(RAG_INDEX_VERSION)),
        ("embedding_model", embedding_model),
        ("embedding_dim", str(embedding_dim)),
        ("built_at", built_at),
        ("chunk_count", str(chunk_count)),
        ("source_count", str(source_count)),
    ]:
        conn.execute("REPLACE INTO metadata (key, value) VALUES (?, ?)", (key, val))
    conn.commit()
    db_path = conn.execute("PRAGMA database_list").fetchone()[2]
    return {
        "version": RAG_INDEX_VERSION,
        "embedding_model": embedding_model,
        "embedding_dim": embedding_dim,
        "built_at": built_at,
        "chunk_count": chunk_count,
        "source_count": source_count,
        "db_path": db_path,
    }


def _extract_pdf_text(pdf_path: Path) -> list[dict]:
    """Extract text from a PDF, returning [{page: int, text: str}, ...]."""
    if not HAS_PYPDF:
        return []
    pages: list[dict] = []
    try:
        reader = pypdf.PdfReader(str(pdf_path))
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if len(text.strip()) >= 20:
                pages.append({"page": i + 1, "text": text})
    except Exception as exc:
        logging.warning(f"RAG: Failed to read PDF {pdf_path.name}: {exc}")
    return pages


def _extract_markdown_sections(md_path: Path) -> list[dict]:
    """Split a markdown file into sections at ## headers."""
    try:
        text = md_path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return []

    sections: list[dict] = []
    current_header = md_path.stem
    current_lines: list[str] = []

    for line in text.splitlines():
        m = RAG_MD_HEADER_RE.match(line)
        if m:
            if current_lines:
                body = "\n".join(current_lines).strip()
                if body:
                    sections.append({"header": current_header, "text": body})
            current_header = m.group(2).strip()
            current_lines = []
        else:
            current_lines.append(line)

    if current_lines:
        body = "\n".join(current_lines).strip()
        if body:
            sections.append({"header": current_header, "text": body})

    return sections


def _chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list[str]:
    """Split text into word-bounded chunks with overlap."""
    words = text.split()
    if len(words) <= chunk_size:
        return [text] if len(words) >= RAG_MIN_CHUNK_WORDS else []

    chunks: list[str] = []
    start = 0
    while start < len(words):
        end = min(start + chunk_size, len(words))
        chunk_words = words[start:end]
        if len(chunk_words) >= RAG_MIN_CHUNK_WORDS:
            chunks.append(" ".join(chunk_words))
        start += chunk_size - overlap

    return chunks


def _build_rag_index(
    docs_dir: Path | None,
    hacktricks_dir: Path | None,
    ollama_url: str,
    embedding_model: str,
    chunk_size: int = RAG_DEFAULT_CHUNK_SIZE,
    chunk_overlap: int = RAG_DEFAULT_CHUNK_OVERLAP,
    db_path: Path | None = None,
    progress_callback=None,
) -> dict:
    """Build or incrementally update the RAG index in SQLite."""
    if db_path is None:
        db_path = _rag_db_path()
    conn = _init_rag_db(db_path)

    try:
        conn.execute("REPLACE INTO metadata VALUES ('version', ?)", (str(RAG_INDEX_VERSION),))
        conn.execute("REPLACE INTO metadata VALUES ('embedding_model', ?)", (embedding_model,))
        conn.commit()

        old_mtimes = dict(conn.execute("SELECT file_path, mtime FROM file_mtimes").fetchall())

        new_mtimes: dict[str, float] = {}
        files_to_embed: list[dict] = []
        is_background = progress_callback is not None

        # Collect PDF files
        if docs_dir and docs_dir.exists() and HAS_PYPDF:
            all_pdfs = sorted(docs_dir.rglob("*.pdf"))
            new_pdfs = []
            for pdf in all_pdfs:
                fkey = str(pdf)
                mtime = pdf.stat().st_mtime
                new_mtimes[fkey] = mtime
                if fkey in old_mtimes and old_mtimes[fkey] == mtime:
                    continue
                new_pdfs.append(pdf)
            if new_pdfs and not is_background:
                print(f"  RAG: extracting text from {len(new_pdfs)} PDFs ({len(all_pdfs) - len(new_pdfs)} unchanged)...")
            for pi, pdf in enumerate(new_pdfs):
                fkey = str(pdf)
                conn.execute("DELETE FROM chunks WHERE source_path = ?", (fkey,))
                if not is_background:
                    print(f"  RAG: [{pi + 1}/{len(new_pdfs)}] chunking {pdf.name}")
                pages = _extract_pdf_text(pdf)
                for pg in pages:
                    for chunk_text in _chunk_text(pg["text"], chunk_size, chunk_overlap):
                        files_to_embed.append({
                            "source": pdf.name,
                            "source_path": fkey,
                            "source_type": "pdf",
                            "location": f"p.{pg['page']}",
                            "title": f"{pdf.stem} p.{pg['page']}",
                            "text": chunk_text,
                        })
        elif docs_dir and docs_dir.exists() and not HAS_PYPDF:
            logging.warning("RAG: pypdf not installed, skipping PDFs. Run: pip install pypdf")

        # Collect HackTricks markdown files
        if hacktricks_dir and hacktricks_dir.exists():
            all_mds = sorted(
                md for md in hacktricks_dir.rglob("*.md")
                if not md.name.startswith(".") and "SUMMARY" not in md.name
            )
            new_mds = []
            for md in all_mds:
                fkey = str(md)
                mtime = md.stat().st_mtime
                new_mtimes[fkey] = mtime
                if fkey in old_mtimes and old_mtimes[fkey] == mtime:
                    continue
                new_mds.append(md)
            if new_mds and not is_background:
                print(f"  RAG: chunking {len(new_mds)} markdown files ({len(all_mds) - len(new_mds)} unchanged)...")
            for md in new_mds:
                fkey = str(md)
                conn.execute("DELETE FROM chunks WHERE source_path = ?", (fkey,))
                try:
                    rel = md.relative_to(hacktricks_dir)
                except ValueError:
                    rel = md
                sections = _extract_markdown_sections(md)
                for sec in sections:
                    chunks = _chunk_text(sec["text"], chunk_size, chunk_overlap)
                    if not chunks and len(sec["text"].split()) >= RAG_MIN_CHUNK_WORDS:
                        chunks = [sec["text"]]
                    for chunk_text in chunks:
                        files_to_embed.append({
                            "source": f"HackTricks/{rel}",
                            "source_path": fkey,
                            "source_type": "markdown",
                            "location": str(rel).replace("\\", "/"),
                            "title": sec["header"],
                            "text": chunk_text,
                        })

        # Remove stale files (deleted from disk but still indexed)
        for stale in set(old_mtimes.keys()) - set(new_mtimes.keys()):
            conn.execute("DELETE FROM chunks WHERE source_path = ?", (stale,))
            conn.execute("DELETE FROM file_mtimes WHERE file_path = ?", (stale,))
        conn.commit()

        # Update file_mtimes for all current files
        for fkey, mtime in new_mtimes.items():
            conn.execute("REPLACE INTO file_mtimes VALUES (?, ?)", (fkey, mtime))
        conn.commit()

        embedding_dim = 0

        if files_to_embed:
            total = len(files_to_embed)
            is_background = progress_callback is not None
            if not is_background:
                print(f"  RAG: {total} chunks to embed...")
            start_time = time.time()
            last_source = ""
            checkpoint_interval = 200
            embedded_count = 0
            for i, chunk in enumerate(files_to_embed):
                src = chunk["source"]

                if progress_callback:
                    progress_callback(i + 1, total, src)

                if not is_background:
                    if src != last_source:
                        print(f"  RAG: [{i + 1}/{total}] {src}")
                        last_source = src
                    elif (i + 1) % 100 == 0:
                        elapsed = time.time() - start_time
                        rate = (i + 1) / elapsed if elapsed > 0 else 0
                        eta = int((total - i - 1) / rate) if rate > 0 else 0
                        print(f"  RAG: [{i + 1}/{total}] {rate:.1f} chunks/s, ~{eta}s remaining")

                try:
                    embedding = ollama_embed(ollama_url, embedding_model, chunk["text"][:8000])
                    if embedding_dim == 0:
                        embedding_dim = len(embedding)
                    blob = _encode_embedding_f16(embedding)
                    conn.execute(
                        "INSERT INTO chunks (source, source_path, source_type, location, title, text, embedding) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (chunk["source"], chunk["source_path"], chunk["source_type"],
                         chunk["location"], chunk["title"], chunk["text"], blob),
                    )
                    embedded_count += 1
                except Exception as exc:
                    logging.warning(f"RAG: Embedding failed for chunk {i}: {exc}")
                    if "model" in str(exc).lower() and i == 0:
                        if not is_background:
                            print(
                                f"  [!] Embedding model '{embedding_model}' not found. "
                                f"Run: ollama pull {embedding_model}"
                            )
                        break

                if (i + 1) % checkpoint_interval == 0:
                    conn.commit()
                    if not is_background:
                        print(f"  RAG: checkpoint saved ({embedded_count} new chunks)")

            conn.commit()

        if embedding_dim == 0:
            row = conn.execute("SELECT embedding FROM chunks WHERE embedding IS NOT NULL LIMIT 1").fetchone()
            if row and row[0]:
                embedding_dim = len(row[0]) // 2  # float16 = 2 bytes each

        meta = _finalize_rag_db(conn, embedding_model, embedding_dim)
    finally:
        conn.close()

    logging.info(
        f"RAG: Index complete — {meta['chunk_count']} chunks, "
        f"{meta['source_count']} source files"
    )
    return meta


def _load_rag_index(_vault_dir: Path | None = None) -> dict | None:
    """Load RAG index metadata from SQLite. Returns lightweight dict or None."""
    db_path = _rag_db_path()
    if not db_path.exists():
        return None
    try:
        conn = _rag_db_connect(db_path)
        rows = conn.execute("SELECT key, value FROM metadata").fetchall()
        conn.close()
    except Exception as exc:
        logging.warning(f"RAG: Failed to load index: {exc}")
        return None

    meta = {k: v for k, v in rows}
    if meta.get("version") != str(RAG_INDEX_VERSION):
        logging.warning("RAG: Index version mismatch, rebuild with --build-index")
        return None

    return {
        "version": int(meta["version"]),
        "embedding_model": meta.get("embedding_model", ""),
        "embedding_dim": int(meta.get("embedding_dim", 768)),
        "built_at": meta.get("built_at", ""),
        "chunk_count": int(meta.get("chunk_count", 0)),
        "source_count": int(meta.get("source_count", 0)),
        "db_path": str(db_path),
    }


def _load_rag_file_mtimes() -> dict[str, float]:
    """Load file_mtimes from the RAG SQLite database."""
    db_path = _rag_db_path()
    if not db_path.exists():
        return {}
    try:
        conn = _rag_db_connect(db_path)
        rows = conn.execute("SELECT file_path, mtime FROM file_mtimes").fetchall()
        conn.close()
        return {k: v for k, v in rows}
    except Exception:
        return {}


class _RagIndexBuilder:
    """Background thread that builds/updates the RAG index."""

    def __init__(self, docs_dir, hacktricks_dir, ollama_url, embedding_model,
                 chunk_size, chunk_overlap, db_path):
        self._ready = threading.Event()
        self._meta: dict | None = None
        self._error: str | None = None
        self.progress_done: int = 0
        self.progress_total: int = 0
        self.progress_source: str = ""
        self._params = {
            "docs_dir": docs_dir, "hacktricks_dir": hacktricks_dir,
            "ollama_url": ollama_url, "embedding_model": embedding_model,
            "chunk_size": chunk_size, "chunk_overlap": chunk_overlap,
            "db_path": db_path,
        }
        self._thread = threading.Thread(target=self._build, daemon=True)

    def start(self) -> None:
        self._thread.start()

    def is_ready(self) -> bool:
        return self._ready.is_set()

    def get_index(self) -> dict | None:
        self._ready.wait()
        return self._meta

    def get_index_nowait(self) -> dict | None:
        if self._ready.is_set():
            return self._meta
        return None

    def status_line(self) -> str:
        if self._ready.is_set():
            if self._error:
                return f"failed: {self._error}"
            return f"complete — {self._meta['chunk_count']} chunks"
        if self.progress_total:
            pct = int(self.progress_done / self.progress_total * 100)
            return f"embedding {self.progress_done}/{self.progress_total} ({pct}%) — {self.progress_source}"
        return "starting..."

    def _build(self) -> None:
        try:
            p = self._params
            self._meta = _build_rag_index(
                p["docs_dir"], p["hacktricks_dir"],
                p["ollama_url"], p["embedding_model"],
                p["chunk_size"], p["chunk_overlap"],
                db_path=p["db_path"],
                progress_callback=self._on_progress,
            )
        except Exception as exc:
            self._error = str(exc)
            logging.error(f"RAG: Background index build failed: {exc}")
        finally:
            self._ready.set()

    def _on_progress(self, done: int, total: int, source: str) -> None:
        self.progress_done = done
        self.progress_total = total
        self.progress_source = source


def _rag_retrieve(
    query_text: str,
    rag_meta: dict,
    ollama_url: str,
    embedding_model: str,
    top_k: int = 5,
) -> list[dict]:
    """Retrieve the top-k most relevant chunks by streaming from SQLite."""
    try:
        query_vec = ollama_embed(ollama_url, embedding_model, query_text[:2000])
    except Exception as exc:
        logging.warning(f"RAG: Query embedding failed: {exc}")
        return []

    db_path = rag_meta["db_path"]
    dim = rag_meta["embedding_dim"]

    qnorm = math.sqrt(sum(x * x for x in query_vec))
    if qnorm == 0:
        return []

    conn = _rag_db_connect(db_path)
    heap: list[tuple[float, int, dict]] = []

    cursor = conn.execute(
        "SELECT id, source, source_path, source_type, location, title, text, embedding "
        "FROM chunks WHERE embedding IS NOT NULL"
    )
    for row in cursor:
        blob = row[7]
        if not blob:
            continue
        emb = struct.unpack(f"{dim}e", blob)
        dot = sum(a * b for a, b in zip(query_vec, emb))
        norm_b = math.sqrt(sum(b * b for b in emb))
        if norm_b == 0:
            continue
        score = dot / (qnorm * norm_b)
        if score < 0.3:
            continue
        chunk = {
            "source": row[1], "source_path": row[2], "source_type": row[3],
            "location": row[4], "title": row[5], "text": row[6], "score": score,
        }
        if len(heap) < top_k:
            heapq.heappush(heap, (score, row[0], chunk))
        elif score > heap[0][0]:
            heapq.heapreplace(heap, (score, row[0], chunk))

    conn.close()
    heap.sort(key=lambda x: x[0], reverse=True)
    return [item[2] for item in heap]


def _format_rag_context(chunks: list[dict], max_chars: int = 4000) -> str:
    """Format retrieved RAG chunks as [REFERENCE] context for prompt injection."""
    if not chunks:
        return ""

    lines = [
        "[REFERENCE MATERIAL — from your cybersecurity library]",
        "The following excerpts are from trusted reference sources. "
        "Cite them using [REF: source] format when relevant.",
        "",
    ]

    chars_used = sum(len(l) for l in lines)
    for c in chunks:
        source = c.get("source", "unknown")
        location = c.get("location", "")
        title = c.get("title", "")
        text = c.get("text", "")

        if c["source_type"] == "pdf":
            cite = f"{source} {location}"
        else:
            cite = location or source

        header = f"[REF: {cite}]"
        if title and title != cite:
            header += f" — {title}"

        preview = text[:800]
        if len(text) > 800:
            preview += "..."

        block = f"{header}\n{preview}\n"
        if chars_used + len(block) > max_chars:
            break
        lines.append(block)
        chars_used += len(block)

    return "\n".join(lines)


def _build_rag_query_from_services(services: list[str], extra: str = "") -> str:
    """Build a RAG query from a list of service/product names."""
    parts = list(dict.fromkeys(services))[:15]
    query = " ".join(parts)
    if extra:
        query += " " + extra
    return query[:500]


def _get_rag_context(query_text: str, top_k: int = 5) -> str:
    """Get formatted RAG context for a prompt. Returns empty string if unavailable."""
    global _RAG_INDEX, _RAG_BUILDER

    if _RAG_BUILDER is not None:
        idx = _RAG_BUILDER.get_index_nowait()
        if idx:
            _RAG_INDEX = idx
            _RAG_BUILDER = None

    if _RAG_INDEX is None or not query_text.strip():
        return ""

    try:
        chunks = _rag_retrieve(
            query_text, _RAG_INDEX,
            _RAG_OLLAMA_URL, _RAG_EMBEDDING_MODEL,
            top_k=top_k,
        )
        if chunks:
            return _format_rag_context(chunks)
    except Exception as exc:
        logging.warning(f"RAG: Retrieval failed: {exc}")

    return ""


# ============================================================
# AI output validation
# ============================================================

def validate_ai_output(
    raw_text: str, source_data: dict, source_type: str
) -> tuple[str, list[str]]:
    """
    Post-process AI output to flag potential hallucinations.

    Returns (annotated_text, warnings_list).
    source_type: "nmap", "nessus", "burp", or "autorecon"
    """
    warnings: list[str] = []
    hosts = source_data.get("hosts", [])

    # --- Build source truth sets ---
    source_ips:       set[str] = set()
    source_ports:     set[int] = set()
    source_cves:      set[str] = set()
    source_hostnames: set[str] = set()

    if source_type == "nmap":
        for h in hosts:
            for addr in h.get("addresses", []):
                if addr.get("addrtype") == "ipv4" and addr.get("addr"):
                    source_ips.add(addr["addr"])
            for p in h.get("open_ports", []):
                source_ports.add(p["port"])
            for hn in h.get("hostnames", []):
                if hn.get("name"):
                    source_hostnames.add(hn["name"].lower().rstrip("."))

    elif source_type == "nessus":
        for h in hosts:
            if h.get("ip"):
                source_ips.add(h["ip"])
            if h.get("hostname"):
                source_hostnames.add(h["hostname"].lower().rstrip("."))
            for f in h.get("findings", []):
                if f.get("port"):
                    source_ports.add(int(f["port"]))
                for cve in f.get("cves", []):
                    source_cves.add(cve.strip().upper())

    elif source_type == "burp":
        for h in hosts:
            if h.get("ip"):
                source_ips.add(h["ip"])
            if h.get("url"):
                from urllib.parse import urlparse
                parsed = urlparse(h["url"])
                if parsed.netloc:
                    source_hostnames.add(parsed.netloc.split(":")[0].lower())

    elif source_type == "autorecon":
        for target in hosts:
            if target.get("ip"):
                source_ips.add(target["ip"])
            if target.get("hostname"):
                source_hostnames.add(target["hostname"].lower().rstrip("."))
            for nmap_scan in target.get("nmap_scans", []):
                for h in nmap_scan.get("hosts", []):
                    for hn in h.get("hostnames", []):
                        if hn.get("name"):
                            source_hostnames.add(hn["name"].lower().rstrip("."))
                    for p in h.get("open_ports", []):
                        source_ports.add(p["port"])
            for port_key in target.get("tool_results", {}):
                try:
                    port_num = int(port_key.split("/")[1])
                    source_ports.add(port_num)
                except (IndexError, ValueError):
                    pass

    elif source_type == "loot":
        pass

    annotated = raw_text

    # --- CVE validation (nessus only) ---
    if source_type == "nessus":
        found_cves = {m.upper() for m in CVE_RE.findall(annotated)}
        for cve in sorted(found_cves):
            if cve not in source_cves:
                inline_warn = (
                    f"\n> ⚠️ HALLUCINATION WARNING: {cve} was not found in the "
                    "scan data — verify before acting."
                )
                # Insert after the first occurrence of this CVE
                pos = 0
                while pos < len(annotated):
                    m = CVE_RE.search(annotated, pos)
                    if not m:
                        break
                    if m.group(0).upper() == cve:
                        end = m.end()
                        annotated = annotated[:end] + inline_warn + annotated[end:]
                        break
                    pos = m.end()
                warnings.append(
                    f"{cve} mentioned in analysis but not present in Nessus input"
                )

    # --- IP cross-reference (all types, only when we have source IPs to compare) ---
    if source_ips:
        found_ips = set(IP_IN_TEXT_RE.findall(annotated))
        for ip in sorted(found_ips):
            if ip not in source_ips:
                warnings.append(
                    f"IP {ip} appears in analysis but was not in input scan"
                )

    # --- Port cross-reference (all types, only when we have source ports) ---
    if source_ports:
        found_ports: set[int] = set()
        for m in PORT_IN_TEXT_RE.finditer(annotated):
            port_str = m.group(1) or m.group(2)
            try:
                found_ports.add(int(port_str))
            except (ValueError, TypeError):
                pass
        for port in sorted(found_ports):
            if port not in source_ports:
                warnings.append(
                    f"Port {port} mentioned in analysis but was not open on any host in the scan"
                )

    # --- Hostname cross-reference (all types, only when we have source hostnames) ---
    if source_hostnames:
        found_fqdns: set[str] = set()
        for m in FQDN_IN_TEXT_RE.finditer(annotated):
            val = m.group(0).lower().rstrip(".")
            if not is_ipv4(val):
                found_fqdns.add(val)
        for fqdn in sorted(found_fqdns):
            if not any(
                fqdn == sh
                or fqdn.endswith("." + sh)
                or sh.endswith("." + fqdn)
                for sh in source_hostnames
            ):
                warnings.append(
                    f"Hostname {fqdn} appears in analysis but was not in input scan"
                )

    # --- Prepend summary block if warnings exist ---
    if warnings:
        n = len(warnings)
        block_lines = [
            f"> ⚠️ **Validation Warnings** "
            f"({n} issue{'s' if n != 1 else ''} found — review before acting)"
        ]
        for w in warnings:
            block_lines.append(f"> - {w}")
        annotated = "\n".join(block_lines) + "\n\n" + annotated

    return annotated, warnings


# ============================================================
# Host note section rendering
# ============================================================

def _render_nessus_section(findings: list[dict]) -> str:
    """Render Nessus findings list as markdown body (no section header)."""
    notable = [f for f in findings if f["severity_int"] >= 1]
    info    = [f for f in findings if f["severity_int"] == 0]

    if not findings:
        return "_No Nessus findings._"

    lines: list[str] = []

    if notable:
        for f in notable:
            sev_str  = severity_int_to_str(f["severity_int"])
            cvss_val = f.get("cvss3_base") or f.get("cvss_base") or "N/A"
            cves_str = ", ".join(f["cves"]) if f["cves"] else "N/A"
            port_str = f"{f['protocol']}/{f['port']}" if f["port"] else "N/A"

            lines.append(f"#### [{sev_str}] {f['plugin_name']} (Plugin {f['plugin_id']})")
            if f["severity_int"] >= 2:
                lines.append(
                    f"- [ ] Investigate: Nessus {f['plugin_name']} "
                    f"(Plugin {f['plugin_id']}, {sev_str})"
                )
            lines.append(
                f"**Port:** {port_str}  ·  **CVSS v3:** {cvss_val}  ·  **CVE(s):** {cves_str}"
            )

            if f.get("description"):
                desc = f["description"].strip()
                if len(desc) > 500:
                    desc = desc[:497] + "..."
                lines.append(f"\n{desc}")

            if f.get("solution"):
                sol = f["solution"].strip()
                if len(sol) > 300:
                    sol = sol[:297] + "..."
                lines.append(f"\n**Solution:** {sol}")

            if f.get("plugin_output"):
                out = f["plugin_output"].strip()
                if len(out) > 300:
                    out = out[:297] + "..."
                lines.append(f"\n**Plugin Output:**\n```\n{out}\n```")

            lines.append("")

    if info:
        lines.append(f"_Plus {len(info)} informational finding(s) — omitted for brevity._")

    return "\n".join(lines).strip()


def _render_burp_section(issues: list[dict]) -> str:
    """Render Burp Suite issues list as markdown body (no section header)."""
    if not issues:
        return "_No Burp Suite findings._"

    sev_order = {"high": 0, "medium": 1, "low": 2, "information": 3}
    sorted_issues = sorted(
        issues, key=lambda i: (sev_order.get(i["severity"].lower(), 4), i["name"])
    )

    lines: list[str] = []
    for issue in sorted_issues:
        sev  = issue["severity"]
        conf = issue.get("confidence", "")
        conf_str = f" ({conf})" if conf else ""

        lines.append(f"#### [{sev}{conf_str}] {issue['name']}")
        if sev.lower() in ("high", "medium"):
            lines.append(
                f"- [ ] Investigate: Burp {issue['name']} ({sev})"
            )

        if issue.get("path"):
            lines.append(f"**Path:** `{issue['path']}`")
        if issue.get("location"):
            lines.append(f"**Location:** {issue['location']}")

        if issue.get("issue_detail"):
            detail = issue["issue_detail"].strip()
            if len(detail) > 500:
                detail = detail[:497] + "..."
            lines.append(f"\n{detail}")

        if issue.get("remediation_detail"):
            rem = issue["remediation_detail"].strip()
            if len(rem) > 300:
                rem = rem[:297] + "..."
            lines.append(f"\n**Remediation:** {rem}")
        elif issue.get("remediation_background"):
            rem = issue["remediation_background"].strip()
            if len(rem) > 300:
                rem = rem[:297] + "..."
            lines.append(f"\n**Remediation:** {rem}")

        lines.append("")

    return "\n".join(lines).strip()


# ============================================================
# Scan note — operator notes preservation helper
# ============================================================

def _preserve_scan_note_operator_notes(scan_path: Path, scan_lines: list[str]) -> list[str]:
    """Append ## Operator Notes to a scan note, preserving existing content."""
    existing_notes = ""
    if scan_path.exists():
        try:
            existing_text = scan_path.read_text(encoding="utf-8")
            existing_notes = extract_body_section(existing_text, OPERATOR_NOTES_SENTINEL).strip()
        except Exception:
            pass
    scan_lines += ["", OPERATOR_NOTES_SENTINEL, ""]
    scan_lines.append(existing_notes if existing_notes else OPERATOR_NOTES_HINT)
    return scan_lines


# ============================================================
# Analyze-full — check all pending boxes, then analyze
# ============================================================

_UNCHECKED_INVESTIGATE_RE = re.compile(r"^(\s*- )\[ \] (Investigate: .+)$", re.MULTILINE)
_UNCHECKED_ANALYZE_RE     = re.compile(r"^(\s*- )\[ \] (Analyze: .+)$",     re.MULTILINE)


def _check_all_pending_boxes(vault_dir: Path) -> tuple[int, int]:
    """Mark all unchecked [ ] Investigate and Analyze boxes as [x].

    Skips already-done [/] boxes. Returns (investigate_count, analyze_count).
    """
    investigate_checked = 0
    analyze_checked = 0

    hosts_dir = vault_dir / "Hosts"
    if hosts_dir.exists():
        for host_path in sorted(hosts_dir.glob("*.md")):
            text = host_path.read_text(encoding="utf-8")
            new_text, n = _UNCHECKED_INVESTIGATE_RE.subn(r"\1[x] \2", text)
            if n:
                host_path.write_text(new_text, encoding="utf-8")
                investigate_checked += n

    scans_dir = vault_dir / "Scans"
    if scans_dir.exists():
        for scan_path in sorted(scans_dir.glob("*.md")):
            text = scan_path.read_text(encoding="utf-8")
            new_text, n = _UNCHECKED_ANALYZE_RE.subn(r"\1[x] \2", text)
            if n:
                scan_path.write_text(new_text, encoding="utf-8")
                analyze_checked += n

    loot_overview = vault_dir / "Loot" / "Overview.md"
    if loot_overview.exists():
        text = loot_overview.read_text(encoding="utf-8")
        new_text, n = _UNCHECKED_ANALYZE_RE.subn(r"\1[x] \2", text)
        if n:
            loot_overview.write_text(new_text, encoding="utf-8")
            analyze_checked += n

    return investigate_checked, analyze_checked


# ============================================================
# Deep Dive — checkbox scanner, prompt, processor
# ============================================================

def _scan_host_note_for_deep_dives(host_path: Path) -> list[dict]:
    """Scan a host note for checked [x] Investigate checkboxes. Returns list of requests."""
    text = host_path.read_text(encoding="utf-8")
    requests: list[str] = []
    for m in DEEP_DIVE_PENDING_RE.finditer(text):
        requests.append(m.group(1).strip())
    if not requests:
        return []

    fm, body = read_frontmatter(text)

    results: list[dict] = []
    for topic in requests:
        results.append({
            "host_path": host_path,
            "host_stem": host_path.stem,
            "ip": fm.get("ip", ""),
            "topic": topic,
            "body": body,
        })
    return results


def _collect_deep_dive_context(body: str, topic: str) -> str:
    """Pull relevant sections from the host note body for a deep-dive prompt."""
    context_parts: list[str] = []

    for section_name in ["## Open Ports", "## Nessus Findings",
                         "## Burp Suite Findings", "## AutoRecon Enumeration",
                         "## Loot", "## Operator Notes"]:
        section = extract_body_section(body, section_name)
        if section:
            context_parts.append(f"{section_name}\n{section}")

    return "\n\n".join(context_parts)


def build_deep_dive_prompt(ip: str, topic: str, host_context: str,
                           operator_notes: str = "") -> str:
    """Build a focused deep-dive prompt for a specific topic on a host."""
    prompt = f"""You are an experienced penetration tester performing a deep-dive analysis
on a specific finding or service. This is a professional engagement under a signed
Rules of Engagement.

TARGET: {ip}
DEEP DIVE TOPIC: {topic}

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- Base your analysis ONLY on the host context provided below.
- Tag each finding as [CONFIRMED], [INFERRED], or [ASSUMED].
- Do not invent CVEs, version numbers, or service details not in the data.
- Every command must use the actual target IP ({ip}).

Provide a thorough, operator-focused analysis covering:

1. **Attack Surface Analysis** — What does this service/finding expose? What is the
   risk in context of the other services on this host?

2. **Enumeration Commands** — Provide ready-to-run commands to enumerate this service
   in depth. Include multiple tools where relevant. Use the actual target IP.

3. **Known Vulnerabilities & Exploits** — Based on the version/config data available,
   what known vulnerabilities apply? Include CVE IDs only if the version data supports
   them. Tag exploitability: [MSF] for Metasploit modules, [POC] for public PoC,
   [THEORETICAL] for version-plausible but unconfirmed.

4. **Attack Paths** — How could this service be leveraged to gain access or pivot?
   Consider credential reuse, chaining with other services on this host, and lateral
   movement opportunities.

5. **Recommended Next Steps** — Prioritized checklist of what the operator should do
   next, in order of likely impact.

HOST CONTEXT:
{host_context}
"""
    if operator_notes:
        prompt += f"\nOPERATOR NOTES:\n{operator_notes}\n"

    # RAG context for deep dives (higher top_k for thorough analysis)
    rag_context = _get_rag_context(f"{topic} {ip} pentest exploitation enumeration", top_k=8)
    if rag_context:
        prompt += f"\n{rag_context}\n"
        prompt += ("\nWhen reference material is provided, cite relevant sources "
                   "using [REF: source] format.\n")

    return prompt


def _process_deep_dives(
    vault_dir: Path,
    ollama_url: str,
    model: str,
    temperature: float,
    skip_validation: bool,
    operator_notes_lookup: dict[str, str],
) -> int:
    """Scan all host notes for checked deep-dive boxes, run analysis, write results."""
    hosts_dir = vault_dir / "Hosts"
    if not hosts_dir.exists():
        return 0

    all_requests: list[dict] = []
    for host_path in sorted(hosts_dir.glob("*.md")):
        all_requests.extend(_scan_host_note_for_deep_dives(host_path))

    if not all_requests:
        return 0

    logging.info(f"[Deep Dive] Found {len(all_requests)} checked investigation(s)")

    processed = 0
    for req in all_requests:
        topic = req["topic"]
        ip = req["ip"]
        host_path: Path = req["host_path"]
        logging.info(f"[Deep Dive] {req['host_stem']}: {topic}")

        host_context = _collect_deep_dive_context(req["body"], topic)
        op_notes = operator_notes_lookup.get(ip, "")

        try:
            prompt = build_deep_dive_prompt(ip, topic, host_context, op_notes)
            raw = ollama_chat(ollama_url, model, prompt, temperature)

            if not skip_validation:
                raw, warnings = validate_ai_output(
                    raw, {"hosts": []}, "nmap"
                )
                for w in warnings:
                    logging.warning(f"[Deep Dive Validation] {w}")

            _write_deep_dive_result(host_path, topic, raw)
            processed += 1
        except Exception as exc:
            logging.warning(f"[Deep Dive] Failed for {topic} on {ip}: {exc}")
            _write_deep_dive_result(host_path, topic, f"_Deep dive failed: {exc}_")
            processed += 1

    return processed


def _write_deep_dive_result(host_path: Path, topic: str, analysis: str) -> None:
    """Mark the checkbox as done and append the result to the Deep Dives section."""
    text = host_path.read_text(encoding="utf-8")

    # Mark the checkbox as investigated: [x] → [/] (green in CSS)
    escaped_topic = re.escape(topic)
    text = re.sub(
        rf"^(\s*)- \[x\] Investigate: {escaped_topic}$",
        rf"\1- [/] Investigate: {topic}",
        text,
        flags=re.MULTILINE,
    )

    # Build the callout block
    callout_lines = [
        f"> [!info]- Deep Dive: {topic}",
    ]
    for line in analysis.strip().splitlines():
        callout_lines.append(f"> {line}")
    callout_block = "\n".join(callout_lines)

    # Insert into Deep Dives section, or create it
    if DEEP_DIVE_SECTION in text:
        section_content = extract_body_section(text, DEEP_DIVE_SECTION)
        old_section = f"{DEEP_DIVE_SECTION}\n{section_content}" if section_content else DEEP_DIVE_SECTION
        new_section = f"{old_section}\n\n{callout_block}" if section_content else f"{DEEP_DIVE_SECTION}\n{callout_block}"
        text = text.replace(old_section, new_section, 1)
    else:
        # Insert before Scan References or Operator Notes
        for marker in ["## Scan References", OPERATOR_NOTES_SENTINEL]:
            if marker in text:
                text = text.replace(marker, f"{DEEP_DIVE_SECTION}\n{callout_block}\n\n{marker}", 1)
                break
        else:
            text += f"\n\n{DEEP_DIVE_SECTION}\n{callout_block}"

    host_path.write_text(text, encoding="utf-8")
    logging.info(f"[Deep Dive] Wrote result for: {topic}")


# ============================================================
# Cross-source deepdive — per-host synthesis across all sources
# ============================================================

def build_cross_source_prompt(
    ip: str,
    hostnames: list[str],
    sections: dict[str, str],
    operator_notes: str = "",
) -> str:
    """Build a unified cross-source correlation prompt from all host note sections."""
    host_label = ", ".join([ip] + hostnames) if hostnames else ip

    parts = [
        "You are an experienced penetration tester performing cross-source synthesis analysis.",
        "You have data from MULTIPLE independent scan sources for a single host.",
        "Your job is to correlate findings across sources and identify relationships,",
        "contradictions, and attack paths that only emerge when viewing all data together.",
        "",
        "GROUNDING RULES:",
        "- Respond ONLY in English.",
        "- NEVER fabricate, invent, or hallucinate data not present in the sections below.",
        "- Tag each finding as [CONFIRMED], [INFERRED], or [ASSUMED].",
        "- Only reference ports, CVEs, credentials, or services that appear in the data.",
        "- If a section is empty or missing, note the coverage gap — do not invent content.",
        "",
        f"HOST: {host_label}",
        "",
    ]

    section_map = {
        "## Open Ports":            "PORT SCAN (NMAP)",
        "## Nessus Findings":       "VULNERABILITY SCAN (NESSUS)",
        "## Burp Suite Findings":   "WEB APPLICATION SCAN (BURP)",
        "## AutoRecon Enumeration": "DEEP ENUMERATION (AUTORECON)",
        "## Loot":                  "LOOT / CREDENTIALS",
    }

    has_data = False
    for section_key, label in section_map.items():
        content = sections.get(section_key, "").strip()
        parts.append(f"### {label}")
        if content:
            parts.append(content[:4000])
            has_data = True
        else:
            parts.append("_(no data from this source)_")
        parts.append("")

    if not has_data:
        return ""

    if operator_notes and operator_notes.strip():
        parts += ["### OPERATOR OBSERVATIONS", operator_notes.strip()[:500], ""]

    parts += [
        "---",
        "",
        "Return your response using EXACTLY these sections:",
        "",
        "## Correlated Findings",
        "Findings that are more significant when viewed across multiple sources.",
        "Cross-reference specific evidence (e.g., 'CVE-XXXX from Nessus aligns with port 445",
        "from Nmap and SMB hash in Loot'). Each bullet must cite which source(s) support it.",
        "",
        "## Attack Path Synthesis",
        "Complete attack chains supported by evidence from 2+ sources.",
        "Show how findings chain together. Tag each step [CONFIRMED]/[INFERRED]/[ASSUMED].",
        "",
        "## Coverage Gaps",
        "What is missing, contradictory, or unclear across the sources.",
        "Flag any source that is absent or provides insufficient data.",
        "",
        "## Priority Actions",
        "Top 3-5 next steps ranked by impact, grounded in the cross-source evidence above.",
    ]

    return "\n".join(parts)


def _write_cross_source_result(host_path: Path, analysis: str) -> None:
    """Write/replace the ## Cross-Source Analysis section in a host note."""
    text = host_path.read_text(encoding="utf-8")
    new_section = f"{CROSS_SOURCE_SECTION}\n\n{analysis.strip()}"
    pattern = re.escape(CROSS_SOURCE_SECTION) + r"\n.*?(?=\n## |\Z)"

    if re.search(pattern, text, re.DOTALL):
        text = re.sub(pattern, new_section, text, count=1, flags=re.DOTALL)
    else:
        for anchor in ["## Scan References", OPERATOR_NOTES_SENTINEL]:
            idx = text.find(f"\n{anchor}")
            if idx != -1:
                text = text[:idx] + f"\n\n{new_section}" + text[idx:]
                break
        else:
            text += f"\n\n{new_section}"

    host_path.write_text(text, encoding="utf-8")
    logging.info(f"[DeepDive] Wrote cross-source analysis for: {host_path.stem}")


def _run_cross_source_deepdive(
    vault_dir: Path,
    args,
    operator_notes_lookup: dict[str, str],
) -> int:
    """Run cross-source correlation analysis for every host note in the vault."""
    hosts_dir = vault_dir / "Hosts"
    if not hosts_dir.exists():
        return 0

    processed = 0
    host_paths = sorted(p for p in hosts_dir.glob("*.md") if not p.stem.startswith("_"))

    for host_path in host_paths:
        try:
            text = host_path.read_text(encoding="utf-8")
            fm, body = read_frontmatter(text)
        except Exception:
            continue

        ip = fm.get("ip", "")
        hostnames = fm.get("hostnames", []) or []

        sections = {
            header: extract_body_section(body, header)
            for header in [
                "## Open Ports",
                "## Nessus Findings",
                "## Burp Suite Findings",
                "## AutoRecon Enumeration",
                "## Loot",
            ]
        }

        op_notes = operator_notes_lookup.get(ip, "")
        if not op_notes and hostnames:
            for hn in hostnames:
                op_notes = operator_notes_lookup.get(hn, "")
                if op_notes:
                    break

        prompt = build_cross_source_prompt(ip, hostnames, sections, op_notes)
        if not prompt:
            logging.debug(f"[DeepDive] Skipping {host_path.stem} — no scan data to correlate")
            continue

        logging.info(f"[DeepDive] Cross-source analysis: {host_path.stem}")
        print(f"  [*] Cross-source: {host_path.stem}...")
        try:
            raw = ollama_chat(args.ollama_url, args.model, prompt, args.temperature)
            if not args.skip_validation:
                raw, warnings = validate_ai_output(raw, {"hosts": []}, "nmap")
                for w in warnings:
                    logging.warning(f"[DeepDive Validation] {w}")
            _write_cross_source_result(host_path, raw)
            processed += 1
        except KeyboardInterrupt:
            raise
        except Exception as exc:
            logging.warning(f"[DeepDive] Failed for {host_path.stem}: {exc}")

    return processed


# ============================================================
# On-demand scan analysis (/analyze command)
# ============================================================

def _scan_notes_for_analysis_requests(vault_dir: Path) -> list[dict]:
    """Find all checked [x] Analyze: checkboxes in scan notes and Loot/Overview."""
    requests: list[dict] = []
    scans_dir = vault_dir / "Scans"

    if scans_dir.exists():
        for sp in sorted(scans_dir.glob("*.md")):
            try:
                text = sp.read_text(encoding="utf-8")
            except Exception:
                continue
            for m in ANALYZE_PENDING_RE.finditer(text):
                label = m.group(1).strip()
                requests.append({
                    "scan_path": sp,
                    "label": label,
                })

    loot_overview = vault_dir / "Loot" / "Overview.md"
    if loot_overview.exists():
        try:
            text = loot_overview.read_text(encoding="utf-8")
        except Exception:
            text = ""
        for m in ANALYZE_PENDING_RE.finditer(text):
            label = m.group(1).strip()
            requests.append({
                "scan_path": loot_overview,
                "label": label,
            })

    return requests


def _extract_source_path(scan_note_text: str) -> Path | None:
    """Extract the source file/dir path from a scan note's metadata."""
    m = re.search(r"- \*\*Source:\*\*\s+(.+)$", scan_note_text, re.MULTILINE)
    if m:
        return Path(m.group(1).strip())
    return None


def _update_scan_note_analysis(
    scan_path: Path,
    analysis: str,
    model_name: str,
    label: str,
    warnings: list[str] | None = None,
) -> None:
    """Replace ## Analysis content in a scan note and mark [x] Analyze → [/]."""
    text = scan_path.read_text(encoding="utf-8")

    escaped_label = re.escape(label)
    text = re.sub(
        rf"^(\s*)- \[x\] Analyze: {escaped_label}$",
        rf"\1- [/] Analyze: {label}",
        text,
        flags=re.MULTILINE,
    )

    # For loot per-host analysis, the section header varies
    if label.startswith("Loot"):
        parts = label.split("—", 1)
        if len(parts) > 1:
            host_key = parts[1].strip()
            if host_key == "campaign":
                section_header = "## Campaign-Level Loot Analysis"
            else:
                section_header = f"## Analysis — {host_key}"
        else:
            section_header = "## Analysis"
    else:
        section_header = "## Analysis"

    new_section = f"{section_header}\n\nModel: `{model_name}`\n\n{analysis}"
    if warnings:
        new_section += "\n\n## Validation Warnings\n\n"
        new_section += "\n".join(f"- {w}" for w in warnings)

    pattern = re.escape(section_header) + r"\n.*?(?=\n## |\Z)"
    text = re.sub(pattern, new_section, text, count=1, flags=re.DOTALL)

    scan_path.write_text(text, encoding="utf-8")
    logging.info(f"[Analyze] Wrote analysis for: {label} in {scan_path.name}")


def _reparse_and_analyze_scan(
    scan_path: Path,
    label: str,
    args,
    vault_dir: Path,
    operator_notes_lookup: dict[str, str],
) -> bool:
    """Re-parse a source file and run LLM analysis for a checked Analyze checkbox."""
    text = scan_path.read_text(encoding="utf-8")
    source_path = _extract_source_path(text)

    # Read scan note operator notes and prepend to prompts as context
    scan_op_notes = extract_body_section(text, OPERATOR_NOTES_SENTINEL).strip()

    def _wrap_prompt_with_scan_notes(prompt: str) -> str:
        if not scan_op_notes or scan_op_notes == OPERATOR_NOTES_HINT.strip():
            return prompt
        return (
            f"[OPERATOR SCAN NOTES — from the assessor's notes on this scan]\n"
            f"{scan_op_notes}\n\n---\n\n" + prompt
        )

    label_lower = label.lower().strip()
    analysis: str | None = None
    warnings: list[str] = []

    if label_lower == "nmap":
        if not source_path or not source_path.exists():
            logging.warning(f"[Analyze] Source not found for Nmap: {source_path}")
            return False
        scan_data = parse_nmap_xml(source_path)
        if not scan_data["hosts"]:
            return False
        raw = ollama_chat(
            args.ollama_url, args.model,
            _wrap_prompt_with_scan_notes(
                build_ollama_prompt(scan_data, operator_notes_by_ip=operator_notes_lookup)
            ),
            args.temperature,
        )
        if not args.skip_validation:
            analysis, warnings = validate_ai_output(raw, scan_data, "nmap")
        else:
            analysis = raw

    elif label_lower == "nessus":
        if not source_path or not source_path.exists():
            logging.warning(f"[Analyze] Source not found for Nessus: {source_path}")
            return False
        nessus_data = parse_nessus_xml(source_path)
        if not nessus_data["hosts"]:
            return False
        facts: str | None = None
        try:
            fact_prompt = _build_nessus_fact_extraction_prompt(
                nessus_data, operator_notes_by_ip=operator_notes_lookup
            )
            facts = ollama_chat(args.ollama_url, args.model, fact_prompt, args.temperature)
        except Exception:
            pass
        p2_prompt = build_nessus_ollama_prompt(nessus_data, operator_notes_by_ip=operator_notes_lookup)
        if facts:
            p2_prompt = (
                "The following facts have been extracted directly from the scan "
                "data. Base your analysis ONLY on these confirmed facts:\n\n"
                + facts + "\n\n" + p2_prompt
            )
        raw = ollama_chat(args.ollama_url, args.model,
                          _wrap_prompt_with_scan_notes(p2_prompt), args.temperature)
        if not args.skip_validation:
            analysis, warnings = validate_ai_output(raw, nessus_data, "nessus")
        else:
            analysis = raw

    elif label_lower == "burp":
        if not source_path or not source_path.exists():
            logging.warning(f"[Analyze] Source not found for Burp: {source_path}")
            return False
        burp_data = parse_burp_xml(source_path)
        raw = ollama_chat(
            args.ollama_url, args.model,
            _wrap_prompt_with_scan_notes(
                build_burp_ollama_prompt(burp_data, operator_notes_by_ip=operator_notes_lookup)
            ),
            args.temperature,
        )
        if not args.skip_validation:
            analysis, warnings = validate_ai_output(raw, burp_data, "burp")
        else:
            analysis = raw

    elif label_lower == "autorecon":
        if not source_path or not source_path.exists():
            logging.warning(f"[Analyze] Source not found for AutoRecon: {source_path}")
            return False
        ar_data = parse_autorecon_results(source_path)
        targets = ar_data.get("targets", [])
        if not targets:
            return False
        target = targets[0]
        facts = None
        try:
            fact_prompt = _build_autorecon_fact_extraction_prompt(
                target, operator_notes=operator_notes_lookup.get(target.get("ip", ""), "")
            )
            facts = ollama_chat(args.ollama_url, args.model, fact_prompt, args.temperature)
        except Exception:
            pass
        p2_prompt = build_autorecon_ollama_prompt(
            target, operator_notes=operator_notes_lookup.get(target.get("ip", ""), "")
        )
        if facts:
            p2_prompt = (
                "The following facts have been extracted directly from the "
                "enumeration data. Base your analysis ONLY on these confirmed "
                "facts:\n\n" + facts + "\n\n" + p2_prompt
            )
        raw = ollama_chat(args.ollama_url, args.model,
                          _wrap_prompt_with_scan_notes(p2_prompt), args.temperature)
        if not args.skip_validation:
            analysis, warnings = validate_ai_output(raw, {"hosts": [target]}, "autorecon")
        else:
            analysis = raw

    elif label_lower == "misc":
        # Re-read content from the Raw Output section in the scan note
        raw_match = re.search(r"```\n(.*?)\n```", text, re.DOTALL)
        if not raw_match:
            if source_path and source_path.exists():
                try:
                    content = source_path.read_text(encoding="utf-8", errors="replace")
                except Exception:
                    return False
            else:
                return False
        else:
            content = raw_match.group(1)

        fname = scan_path.stem.replace(" - Misc", "")
        # Extract tool type/level from scan note metadata
        tool_type = "unknown"
        level = "standard"
        tt_match = re.search(r"Detected tool:\*\*\s*(\S+)", text)
        if tt_match:
            tool_type = tt_match.group(1)
        al_match = re.search(r"Analysis level:\*\*\s*(\S+)", text)
        if al_match:
            level = al_match.group(1)

        raw = ollama_chat(
            args.ollama_url, args.model,
            _wrap_prompt_with_scan_notes(
                build_misc_ollama_prompt(fname, content, tool_type=tool_type, analysis_level=level)
            ),
            args.temperature,
        )
        if not args.skip_validation:
            analysis, warnings = validate_ai_output(raw, {"hosts": []}, "misc")
        else:
            analysis = raw

    elif label_lower.startswith("loot"):
        parts = label.split("—", 1)
        host_key = parts[1].strip() if len(parts) > 1 else "campaign"

        # Find the loot source directory
        loot_dir: Path | None = None
        for candidate_name in ["loot", "Loot"]:
            for base_candidate in [Path("scans"), Path(args.scans_dir)]:
                c = base_candidate / candidate_name
                if c.exists():
                    loot_dir = c.resolve()
                    break
            if loot_dir:
                break

        if not loot_dir or not loot_dir.exists():
            logging.warning("[Analyze] Loot source directory not found")
            return False

        known = _build_known_hosts_lookup(vault_dir)
        loot_data = parse_loot_dir(loot_dir, known_hosts=known)
        is_campaign = host_key == "campaign"

        if is_campaign:
            loot_files = loot_data.get("campaign_loot", [])
        else:
            loot_files = loot_data.get("host_loot", {}).get(host_key, [])

        if not loot_files:
            return False

        hosts_dir = vault_dir / "Hosts"
        host_context = _build_host_context_for_loot(hosts_dir, host_key) if not is_campaign else None
        op_notes = operator_notes_lookup.get(host_key, "")

        prompt = build_loot_ollama_prompt(
            host_key if not is_campaign else "campaign",
            loot_files,
            host_context=host_context,
            operator_notes=op_notes,
        )
        raw = ollama_chat(args.ollama_url, args.model, prompt, args.temperature)
        if not args.skip_validation:
            analysis, warnings = validate_ai_output(raw, {"hosts": []}, "loot")
        else:
            analysis = raw
    else:
        logging.warning(f"[Analyze] Unknown analysis type: {label}")
        return False

    if analysis:
        _update_scan_note_analysis(scan_path, analysis, args.model, label, warnings or None)
        return True
    return False


def _process_analyze_requests(
    vault_dir: Path,
    args,
    operator_notes_lookup: dict[str, str],
) -> int:
    """Process all checked [x] Analyze: checkboxes across scan notes."""
    requests = _scan_notes_for_analysis_requests(vault_dir)
    processed = 0

    for req in requests:
        logging.info(f"[Analyze] {req['scan_path'].stem}: {req['label']}")
        try:
            if _reparse_and_analyze_scan(
                req["scan_path"], req["label"], args,
                vault_dir, operator_notes_lookup,
            ):
                processed += 1
        except KeyboardInterrupt:
            logging.warning("[Analyze] Cancelled by user")
            break
        except Exception as exc:
            logging.warning(f"[Analyze] Failed for {req['label']}: {exc}")

    return processed


# ============================================================
# Vault writing — Nmap
# ============================================================

def _find_host_note_by_ip(hosts_dir: Path, ip: str) -> Path | None:
    """Find an existing host note whose frontmatter ip matches."""
    if not hosts_dir.exists() or not ip:
        return None
    for hp in hosts_dir.glob("*.md"):
        try:
            fm, _ = read_frontmatter(hp.read_text(encoding="utf-8"))
            if fm.get("ip") == ip:
                return hp
        except Exception:
            pass
    return None


_IPV4_STEM_RE = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")


def _merge_two_host_notes(keep_path: Path, delete_path: Path, vault_dir: Path) -> str:
    """Merge delete_path into keep_path, update vault references, delete duplicate."""
    keep_text   = keep_path.read_text(encoding="utf-8")
    delete_text = delete_path.read_text(encoding="utf-8")
    keep_fm,   keep_body   = read_frontmatter(keep_text)
    delete_fm, delete_body = read_frontmatter(delete_text)

    # Merge frontmatter
    merged_ip = (keep_fm.get("ip") or delete_fm.get("ip") or "").strip()

    all_hostnames: list[str] = list(dict.fromkeys(
        (keep_fm.get("hostnames") or []) + (delete_fm.get("hostnames") or [])
    ))
    for extra in [delete_path.stem, keep_path.stem]:
        if extra and extra not in all_hostnames and extra != merged_ip:
            all_hostnames.append(extra)

    merged_tags    = list(dict.fromkeys((keep_fm.get("tags") or []) + (delete_fm.get("tags") or [])))
    merged_sources = list(dict.fromkeys((keep_fm.get("sources") or []) + (delete_fm.get("sources") or [])))

    merged_fm: dict = dict(keep_fm)
    merged_fm["ip"]        = merged_ip
    merged_fm["hostnames"] = all_hostnames
    merged_fm["tags"]      = merged_tags
    merged_fm["sources"]   = merged_sources
    merged_fm["status"]    = keep_fm.get("status") or delete_fm.get("status") or "not-started"
    for field in ["nessus_max_severity", "autorecon_tools_run",
                  "loot_file_count", "loot_credential_count", "loot_hash_count"]:
        kv = keep_fm.get(field) or 0
        dv = delete_fm.get(field) or 0
        if kv or dv:
            merged_fm[field] = max(int(kv), int(dv))

    # Merge body sections — prefer the non-empty side; if both have content, append delete's
    MERGE_SECTIONS = [
        "## Open Ports",
        "## Nessus Findings",
        "## Burp Suite Findings",
        "## AutoRecon Enumeration",
        "## Loot",
        DEEP_DIVE_SECTION,
        CROSS_SOURCE_SECTION,
    ]
    body_lines: list[str] = []

    # Preserve preamble from keep note (lines before first ##)
    for line in keep_body.splitlines():
        if line.startswith("## "):
            break
        body_lines.append(line)
    while body_lines and not body_lines[-1].strip():
        body_lines.pop()

    for section in MERGE_SECTIONS:
        k = extract_body_section(keep_body, section).strip()
        d = extract_body_section(delete_body, section).strip()
        if k and d and k != d:
            body_lines += ["", section, k, "", d]
        elif k:
            body_lines += ["", section, k]
        elif d:
            body_lines += ["", section, d]

    # Scan references — union, deduplicated
    k_refs = [l for l in extract_body_section(keep_body,   "## Scan References").splitlines() if l.strip().startswith("-")]
    d_refs = [l for l in extract_body_section(delete_body, "## Scan References").splitlines() if l.strip().startswith("-")]
    all_refs = list(dict.fromkeys(k_refs + d_refs))
    body_lines += ["", "## Scan References"] + all_refs

    # Operator notes — concatenate both if they differ
    k_op = extract_operator_notes(keep_body).strip()
    d_op = extract_operator_notes(delete_body).strip()
    body_lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    combined_op = "\n\n".join(p for p in [k_op, d_op] if p and p != OPERATOR_NOTES_HINT)
    if combined_op:
        body_lines += ["", combined_op]

    keep_path.write_text(write_frontmatter(merged_fm) + "\n" + "\n".join(body_lines), encoding="utf-8")

    # Update all vault files that link to the deleted note
    _replace_vault_host_refs(vault_dir, delete_path.stem, keep_path.stem)

    delete_path.unlink()
    return f"Merged {delete_path.stem} → {keep_path.stem}"


def _replace_vault_host_refs(vault_dir: Path, old_stem: str, new_stem: str) -> None:
    """Replace Obsidian [[Hosts/old_stem|...]] links with the new stem in all vault files."""
    old_link_pat = re.compile(
        rf"\[\[Hosts/{re.escape(old_stem)}(\|[^\]]+)?\]\]"
    )
    new_link = f"[[Hosts/{new_stem}|{new_stem}]]"

    for md_path in vault_dir.rglob("*.md"):
        if md_path == vault_dir / "Hosts" / ensure_md_suffix(new_stem):
            continue
        try:
            text = md_path.read_text(encoding="utf-8")
            if old_stem in text:
                new_text = old_link_pat.sub(new_link, text)
                if new_text != text:
                    md_path.write_text(new_text, encoding="utf-8")
        except Exception:
            pass

    # Also update canvas files
    for canvas_path in vault_dir.glob("*.canvas"):
        try:
            text = canvas_path.read_text(encoding="utf-8")
            if old_stem in text:
                new_text = text.replace(
                    f"Hosts/{old_stem}.md", f"Hosts/{new_stem}.md"
                ).replace(
                    f'"file": "Hosts/{old_stem}"', f'"file": "Hosts/{new_stem}"'
                )
                if new_text != text:
                    canvas_path.write_text(new_text, encoding="utf-8")
        except Exception:
            pass


def _detect_and_merge_host_notes(vault_dir: Path) -> list[str]:
    """Detect and merge host notes that refer to the same host.

    Detection signals (any one is sufficient):
    - Two notes share the same ip: field in frontmatter
    - Note A's ip: matches note B's filename (B is named after the IP)
    - Note A's hostnames: list contains note B's stem
    Returns list of human-readable merge descriptions.
    """
    hosts_dir = vault_dir / "Hosts"
    if not hosts_dir.exists():
        return []

    # Load all non-campaign host notes
    notes: dict[Path, tuple[dict, str]] = {}
    for hp in sorted(hosts_dir.glob("*.md")):
        if hp.stem.startswith("_"):
            continue
        try:
            text = hp.read_text(encoding="utf-8")
            fm, body = read_frontmatter(text)
            notes[hp] = (fm, body)
        except Exception:
            continue

    # Build detection index: representative key -> set of paths
    # Use a union-find style: group paths by the signals they share
    groups: dict[str, set[Path]] = {}

    for path, (fm, _) in notes.items():
        ip = (fm.get("ip") or "").strip()
        if ip:
            groups.setdefault(f"ip:{ip}", set()).add(path)
        # If the stem itself looks like an IP, add it under that key too
        if _IPV4_STEM_RE.match(path.stem):
            groups.setdefault(f"ip:{path.stem}", set()).add(path)
        # Add hostname-based signals
        for hn in (fm.get("hostnames") or []):
            hn = hn.strip()
            if hn:
                groups.setdefault(f"hn:{hn.lower()}", set()).add(path)

    # Also: if any note's stem matches a hostname key
    for path in notes:
        groups.setdefault(f"hn:{path.stem.lower()}", set()).add(path)

    # Find groups with 2+ notes → those are duplicates
    seen: set[frozenset] = set()
    merge_pairs: list[tuple[Path, Path]] = []  # (keep, delete)

    for key, paths in groups.items():
        if len(paths) < 2:
            continue
        paths_list = sorted(paths, key=lambda p: (
            1 if _IPV4_STEM_RE.match(p.stem) else 0,  # prefer hostname over IP
            p.stem
        ))
        keep = paths_list[0]
        for delete in paths_list[1:]:
            pair = frozenset([keep, delete])
            if pair not in seen:
                seen.add(pair)
                merge_pairs.append((keep, delete))

    # Execute merges
    reports: list[str] = []
    for keep_path, delete_path in merge_pairs:
        if not keep_path.exists() or not delete_path.exists():
            continue
        try:
            msg = _merge_two_host_notes(keep_path, delete_path, vault_dir)
            reports.append(msg)
            logging.info(f"[Merge] {msg}")
        except Exception as exc:
            logging.warning(f"[Merge] Failed {keep_path.stem}+{delete_path.stem}: {exc}")

    return reports


def _write_host_note(
    hosts_dir: Path,
    host: dict,
    scan_stem: str,
    scan_display: str,
    tool_name: str,
) -> tuple[str, str]:
    """
    Write (or merge) a host note for an Nmap host.
    Preserves existing Nessus Findings, Burp Suite Findings, and AutoRecon Enumeration sections.
    Returns (display_name, host_stem).
    """
    display    = choose_host_display_name(host)
    host_stem  = safe_filename(display)
    host_path  = hosts_dir / ensure_md_suffix(host_stem)
    primary_ip = get_primary_ipv4(host)
    open_ports = host.get("open_ports", [])

    # If the display-name file doesn't exist, look up by IP to find an
    # existing note created by a different scan (e.g., hostname vs IP only)
    if not host_path.exists() and primary_ip:
        existing_by_ip = _find_host_note_by_ip(hosts_dir, primary_ip)
        if existing_by_ip:
            host_path = existing_by_ip
            host_stem = existing_by_ip.stem
            display = host_stem
            logging.debug(f"Found existing host note by IP lookup: {host_path.name}")

    existing_fm: dict  = {}
    existing_op_notes  = ""
    existing_nessus    = ""
    existing_burp      = ""
    existing_autorecon = ""
    existing_loot      = ""
    existing_deep_dives = ""
    existing_cross_source = ""
    existing_open_ports = ""

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes  = extract_operator_notes(old_body)
        existing_open_ports = extract_body_section(old_body, "## Open Ports")
        existing_nessus    = extract_body_section(old_body, "## Nessus Findings")
        existing_burp      = extract_body_section(old_body, "## Burp Suite Findings")
        existing_autorecon = extract_body_section(old_body, "## AutoRecon Enumeration")
        existing_loot      = extract_body_section(old_body, "## Loot")
        existing_deep_dives = extract_body_section(old_body, DEEP_DIVE_SECTION)
        existing_cross_source = extract_body_section(old_body, CROSS_SOURCE_SECTION)
        logging.debug(f"Merging Nmap host note: {host_path.name}")
    else:
        logging.debug(f"Creating Nmap host note: {host_path.name}")

    # Merge frontmatter
    existing_hostnames: list = existing_fm.get("hostnames", [])
    new_hostnames = [h["name"].strip() for h in host.get("hostnames", []) if h.get("name")]
    merged_hostnames = existing_hostnames + [h for h in new_hostnames if h not in existing_hostnames]

    scan_source = f"{scan_display} - {tool_name}"
    existing_sources: list = existing_fm.get("sources", [])
    merged_sources = existing_sources if scan_source in existing_sources else existing_sources + [scan_source]

    # Merge open ports: combine existing + new, keep richest service info per port
    if existing_open_ports.strip():
        merged_ports, checkbox_states = _merge_port_lists(existing_open_ports, open_ports)
        # Re-derive tags from merged port set
        all_tags_from_ports = get_tags_from_ports(merged_ports)
    else:
        merged_ports = open_ports
        checkbox_states = {}
        all_tags_from_ports = get_tags_from_ports(open_ports)

    merged_tags = sorted(set(existing_fm.get("tags", [])) | set(all_tags_from_ports))

    fm: dict = {
        "ip":        primary_ip or existing_fm.get("ip"),
        "hostnames": merged_hostnames,
        "status":    existing_fm.get("status", "not-started"),
        "tags":      merged_tags,
        "sources":   merged_sources,
    }
    if "nessus_max_severity" in existing_fm:
        fm["nessus_max_severity"] = existing_fm["nessus_max_severity"]
    if "autorecon_tools_run" in existing_fm:
        fm["autorecon_tools_run"] = existing_fm["autorecon_tools_run"]

    # Build body
    lines: list[str] = [f"**State:** {host['state']}"]
    if primary_ip:
        lines.append(f"**IP:** {primary_ip}")
    lines.append(f"**Open Ports:** {len(merged_ports)}")

    lines += ["", "## Open Ports"]
    if merged_ports:
        if checkbox_states:
            lines.extend(_summarize_open_ports_merged(merged_ports, checkbox_states))
        else:
            lines.extend(summarize_open_ports({"open_ports": merged_ports}))
    else:
        lines.append("_No open ports detected._")

    if existing_nessus:
        lines += ["", "## Nessus Findings", existing_nessus]

    if existing_burp:
        lines += ["", "## Burp Suite Findings", existing_burp]

    if existing_autorecon:
        lines += ["", "## AutoRecon Enumeration", existing_autorecon]

    if existing_loot:
        lines += ["", "## Loot", existing_loot]

    if existing_deep_dives:
        lines += ["", DEEP_DIVE_SECTION, existing_deep_dives]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_stem


def create_obsidian_vault(
    vault_dir: Path,
    scan_name: str,
    scan_data: dict,
    tool_name: str,
    analysis_text: str | None,
    model_name: str | None,
    validation_warnings: list[str] | None = None,
) -> dict:
    """Write all host notes and the scan note for one Nmap scan file."""
    logging.info(f"Writing Nmap vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = scan_name
    if len(scan_data["hosts"]) == 1:
        scan_display = choose_host_display_name(scan_data["hosts"][0])

    scan_stem = safe_filename(f"{scan_display} - {tool_name}")
    scan_path = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in scan_data["hosts"]:
        display, host_stem = _write_host_note(
            hosts_dir, host, scan_stem, scan_display, tool_name
        )
        host_entries.append((display, host_stem))

    scan_lines = [
        f"# {scan_display} - {tool_name}",
        "",
        f"- **Source:** {scan_data['source_file']}",
        f"- **Parsed:** {scan_data['parsed_at']}",
        f"- **Nmap Args:** `{scan_data.get('nmap_args', '')}`",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    _acb = "/" if analysis_text else " "
    scan_lines += ["", f"- [{_acb}] Analyze: Nmap", "", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    if validation_warnings:
        scan_lines += ["", "## Validation Warnings", ""]
        for w in validation_warnings:
            scan_lines.append(f"- {w}")

    scan_lines = _preserve_scan_note_operator_notes(scan_path, scan_lines)
    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Nmap scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — Nessus
# ============================================================

def _update_host_note_nessus(
    hosts_dir: Path,
    ip: str,
    hostname: str,
    findings: list[dict],
    scan_source: str,
) -> tuple[str, str]:
    """
    Write or update the ## Nessus Findings section in a host note.
    Returns (display_name, host_stem).
    """
    # Try to find an existing note by IP first
    existing_path = _find_host_note_by_ip(hosts_dir, ip)

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        # Determine display name
        if hostname and is_probable_fqdn(hostname):
            display = hostname
        else:
            display = ip or hostname or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports_section = ""
    existing_burp_section = ""
    existing_autorecon_section = ""
    existing_loot_section = ""
    existing_deep_dives = ""
    existing_cross_source = ""
    existing_preamble_lines: list[str] = []
    existing_scan_refs: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes           = extract_operator_notes(old_body)
        existing_open_ports_section = extract_body_section(old_body, "## Open Ports")
        existing_burp_section       = extract_body_section(old_body, "## Burp Suite Findings")
        existing_autorecon_section  = extract_body_section(old_body, "## AutoRecon Enumeration")
        existing_loot_section       = extract_body_section(old_body, "## Loot")
        existing_deep_dives         = extract_body_section(old_body, DEEP_DIVE_SECTION)
        existing_cross_source       = extract_body_section(old_body, CROSS_SOURCE_SECTION)
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating Nessus section in: {host_path.name}")
    else:
        # Minimal preamble for Nessus-only host
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if hostname and hostname != ip:
            existing_preamble_lines.append(f"**Hostname:** {hostname}")
        logging.debug(f"Creating Nessus-only host note: {host_path.name}")

    # Update frontmatter
    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname and hostname not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname]

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    max_sev = max((f["severity_int"] for f in findings), default=0)
    existing_max_sev = existing_fm.get("nessus_max_severity", 0)
    try:
        existing_max_sev = int(existing_max_sev)
    except (TypeError, ValueError):
        existing_max_sev = 0
    merged_max_sev = max(max_sev, existing_max_sev)

    fm: dict = {
        "ip":                  ip or existing_fm.get("ip"),
        "hostnames":           existing_hostnames,
        "status":              existing_fm.get("status", "not-started"),
        "tags":                existing_fm.get("tags", []),
        "sources":             existing_sources,
        "nessus_max_severity": merged_max_sev,
    }

    # Build body
    preamble = "\n".join(existing_preamble_lines).rstrip()
    lines: list[str] = []

    if preamble:
        lines.append(preamble)

    if existing_open_ports_section:
        lines += ["", "## Open Ports", existing_open_ports_section]

    lines += ["", "## Nessus Findings", _render_nessus_section(findings)]

    if existing_burp_section:
        lines += ["", "## Burp Suite Findings", existing_burp_section]

    if existing_autorecon_section:
        lines += ["", "## AutoRecon Enumeration", existing_autorecon_section]

    if existing_loot_section:
        lines += ["", "## Loot", existing_loot_section]

    if existing_deep_dives:
        lines += ["", DEEP_DIVE_SECTION, existing_deep_dives]

    if existing_cross_source:
        lines += ["", CROSS_SOURCE_SECTION, existing_cross_source]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_nessus_vault(
    vault_dir: Path,
    scan_name: str,
    nessus_data: dict,
    analysis_text: str | None,
    model_name: str | None,
    validation_warnings: list[str] | None = None,
) -> dict:
    """Write Nessus host note sections and scan note."""
    logging.info(f"Writing Nessus vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = nessus_data.get("report_name", scan_name)
    scan_stem    = safe_filename(f"{scan_display} - Nessus")
    scan_source  = f"{scan_display} - Nessus"
    scan_path    = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in nessus_data.get("hosts", []):
        display, host_stem = _update_host_note_nessus(
            hosts_dir,
            host["ip"],
            host.get("hostname", ""),
            host["findings"],
            scan_source,
        )
        host_entries.append((display, host_stem))

    # Build scan note
    total_findings = sum(len(h["findings"]) for h in nessus_data.get("hosts", []))
    critical_count = sum(
        sum(1 for f in h["findings"] if f["severity_int"] == 4)
        for h in nessus_data.get("hosts", [])
    )
    high_count = sum(
        sum(1 for f in h["findings"] if f["severity_int"] == 3)
        for h in nessus_data.get("hosts", [])
    )

    scan_lines = [
        f"# {scan_display} - Nessus",
        "",
        f"- **Source:** {nessus_data['source_file']}",
        f"- **Parsed:** {nessus_data['parsed_at']}",
        f"- **Total Findings:** {total_findings}  ·  **Critical:** {critical_count}  ·  **High:** {high_count}",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    _acb = "/" if analysis_text else " "
    scan_lines += ["", f"- [{_acb}] Analyze: Nessus", "", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    if validation_warnings:
        scan_lines += ["", "## Validation Warnings", ""]
        for w in validation_warnings:
            scan_lines.append(f"- {w}")

    scan_lines = _preserve_scan_note_operator_notes(scan_path, scan_lines)
    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Nessus scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — Burp Suite
# ============================================================

def _update_host_note_burp(
    hosts_dir: Path,
    ip: str,
    url: str,
    issues: list[dict],
    scan_source: str,
) -> tuple[str, str]:
    """
    Write or update the ## Burp Suite Findings section in a host note.
    Returns (display_name, host_stem).
    """
    existing_path = _find_host_note_by_ip(hosts_dir, ip) if ip else None

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        # Derive display name from URL host or IP
        if url:
            from urllib.parse import urlparse
            parsed = urlparse(url)
            netloc = parsed.netloc or url
            display = netloc.split(":")[0] if netloc else (ip or url)
        else:
            display = ip or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports_section = ""
    existing_nessus_section = ""
    existing_autorecon_section = ""
    existing_loot_section = ""
    existing_deep_dives = ""
    existing_cross_source = ""
    existing_preamble_lines: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes            = extract_operator_notes(old_body)
        existing_open_ports_section  = extract_body_section(old_body, "## Open Ports")
        existing_nessus_section      = extract_body_section(old_body, "## Nessus Findings")
        existing_autorecon_section   = extract_body_section(old_body, "## AutoRecon Enumeration")
        existing_loot_section        = extract_body_section(old_body, "## Loot")
        existing_deep_dives          = extract_body_section(old_body, DEEP_DIVE_SECTION)
        existing_cross_source        = extract_body_section(old_body, CROSS_SOURCE_SECTION)
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating Burp section in: {host_path.name}")
    else:
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if url:
            existing_preamble_lines.append(f"**URL:** {url}")
        logging.debug(f"Creating Burp-only host note: {host_path.name}")

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    # Derive hostname from URL for frontmatter
    hostname_for_fm = ""
    if url:
        from urllib.parse import urlparse
        parsed = urlparse(url)
        hostname_for_fm = parsed.netloc.split(":")[0] if parsed.netloc else ""

    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname_for_fm and hostname_for_fm not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname_for_fm]

    fm: dict = {
        "ip":        ip or existing_fm.get("ip"),
        "hostnames": existing_hostnames,
        "status":    existing_fm.get("status", "not-started"),
        "tags":      existing_fm.get("tags", []),
        "sources":   existing_sources,
    }
    if "nessus_max_severity" in existing_fm:
        fm["nessus_max_severity"] = existing_fm["nessus_max_severity"]

    preamble = "\n".join(existing_preamble_lines).rstrip()
    lines: list[str] = []

    if preamble:
        lines.append(preamble)

    if existing_open_ports_section:
        lines += ["", "## Open Ports", existing_open_ports_section]

    if existing_nessus_section:
        lines += ["", "## Nessus Findings", existing_nessus_section]

    lines += ["", "## Burp Suite Findings", _render_burp_section(issues)]

    if existing_autorecon_section:
        lines += ["", "## AutoRecon Enumeration", existing_autorecon_section]

    if existing_loot_section:
        lines += ["", "## Loot", existing_loot_section]

    if existing_deep_dives:
        lines += ["", DEEP_DIVE_SECTION, existing_deep_dives]

    if existing_cross_source:
        lines += ["", CROSS_SOURCE_SECTION, existing_cross_source]

    lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        lines.append(f"- [[Scans/{src_stem}|{src}]]")

    lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        lines += ["", existing_op_notes]

    body = "\n".join(lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_burp_vault(
    vault_dir: Path,
    scan_name: str,
    burp_data: dict,
    analysis_text: str | None,
    model_name: str | None,
    validation_warnings: list[str] | None = None,
) -> dict:
    """Write Burp Suite host note sections and scan note."""
    logging.info(f"Writing Burp vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    scan_display = scan_name
    scan_stem    = safe_filename(f"{scan_display} - Burp")
    scan_source  = f"{scan_display} - Burp"
    scan_path    = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for host in burp_data.get("hosts", []):
        display, host_stem = _update_host_note_burp(
            hosts_dir,
            host.get("ip", ""),
            host.get("url", ""),
            host["issues"],
            scan_source,
        )
        host_entries.append((display, host_stem))

    total_issues = sum(len(h["issues"]) for h in burp_data.get("hosts", []))
    high_count   = sum(
        sum(1 for i in h["issues"] if i["severity"].lower() == "high")
        for h in burp_data.get("hosts", [])
    )

    scan_lines = [
        f"# {scan_display} - Burp",
        "",
        f"- **Source:** {burp_data['source_file']}",
        f"- **Parsed:** {burp_data['parsed_at']}",
        f"- **Total Issues:** {total_issues}  ·  **High:** {high_count}",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    _acb = "/" if analysis_text else " "
    scan_lines += ["", f"- [{_acb}] Analyze: Burp", "", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    if validation_warnings:
        scan_lines += ["", "## Validation Warnings", ""]
        for w in validation_warnings:
            scan_lines.append(f"- {w}")

    scan_lines = _preserve_scan_note_operator_notes(scan_path, scan_lines)
    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote Burp scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — AutoRecon
# ============================================================

def _render_autorecon_section(target_data: dict) -> str:
    """Render AutoRecon enumeration data as markdown body (no section header)."""
    summary = target_data.get("summary", {})
    tool_results = target_data.get("tool_results", {})

    if not tool_results:
        return "_AutoRecon ran but produced no additional findings._"

    lines: list[str] = [
        f"**AutoRecon target:** {target_data.get('target', '?')} "
        f"| **Tools run:** {summary.get('total_tools_run', 0)} "
        f"| **With findings:** {summary.get('tools_with_findings', 0)}",
        "",
    ]

    empty_ports = 0

    for port_key in sorted(tool_results):
        results = tool_results[port_key]
        proto, port = port_key.split("/", 1)
        # Derive service hint from first result filename
        service_hint = ""
        for r in results:
            fn_match = AUTORECON_FILENAME_RE.match(r.get("filename", ""))
            if fn_match:
                service_hint = fn_match.group(3)
                break

        port_has_content = False
        port_lines: list[str] = [f"### {port_key} — {service_hint or proto.upper()}"]

        for entry in results:
            data = entry.get("data", {})
            tool = data.get("tool", "unknown")

            if tool == "dirbusting" and data.get("interesting"):
                port_has_content = True
                interesting = data["interesting"][:20]
                port_lines.append(f"\n#### Discovered Paths ({data.get('total_found', 0)} total)")
                port_lines.append("| Status | Path | Size |")
                port_lines.append("|--------|------|------|")
                for p in interesting:
                    redir = f" → {p['redirect']}" if p.get("redirect") else ""
                    port_lines.append(f"| {p['status']} | `{p['path']}` | {p['size']}{redir} |")

            elif tool == "nikto" and data.get("findings"):
                port_has_content = True
                port_lines.append("\n#### Nikto Findings")
                if data.get("server"):
                    port_lines.append(f"**Server:** {data['server']}")
                for f in data["findings"][:15]:
                    osvdb = f"OSVDB-{f['osvdb']}: " if f.get("osvdb") else ""
                    port_lines.append(f"- {osvdb}`{f['path']}` — {f['description'][:200]}")

            elif tool == "whatweb" and data.get("technologies"):
                port_has_content = True
                techs = [f"{t['name']}/{t['version']}" if t.get("version") else t["name"]
                         for t in data["technologies"]]
                port_lines.append(f"\n**Technologies:** {', '.join(techs)}")
                if data.get("title"):
                    port_lines.append(f"**Title:** {data['title']}")

            elif tool == "enum4linux":
                has_data = any([data.get("os_info"), data.get("users"),
                                data.get("shares"), data.get("null_session")])
                if has_data:
                    port_has_content = True
                    if data.get("os_info"):
                        port_lines.append(f"\n**OS:** {data['os_info']}")
                    if data.get("domain"):
                        port_lines.append(f"**Domain:** {data['domain']}")
                    if data.get("null_session"):
                        port_lines.append("**Null Session:** Yes")
                    if data.get("shares"):
                        port_lines.append("\n#### Shares")
                        port_lines.append("| Share | Access |")
                        port_lines.append("|-------|--------|")
                        for s in data["shares"]:
                            access = s.get("access", s.get("type", ""))
                            bold = "**" if "write" in access.lower() else ""
                            port_lines.append(f"| {s['name']} | {bold}{access}{bold} |")
                    if data.get("users"):
                        port_lines.append("\n#### Users Found")
                        port_lines.append(", ".join(f"`{u}`" for u in data["users"][:30]))
                    pp = data.get("password_policy", {})
                    if pp:
                        parts = []
                        if "min_length" in pp:
                            parts.append(f"Min length: {pp['min_length']}")
                        if "lockout_threshold" in pp:
                            val = pp["lockout_threshold"]
                            parts.append(f"Lockout threshold: {val}" + (" (no lockout)" if val == 0 else ""))
                        if pp.get("complexity"):
                            parts.append("Complexity: on")
                        if parts:
                            port_lines.append(f"\n**Password Policy:** {' | '.join(parts)}")

            elif tool == "smbmap" and data.get("shares"):
                port_has_content = True
                port_lines.append("\n#### SMB Shares (smbmap)")
                port_lines.append("| Share | Permissions |")
                port_lines.append("|-------|------------|")
                for s in data["shares"]:
                    bold = "**" if "WRITE" in s["permissions"] else ""
                    port_lines.append(f"| {s['name']} | {bold}{s['permissions']}{bold} |")

            elif tool == "smbclient" and data.get("shares"):
                port_has_content = True
                port_lines.append("\n#### Shares (smbclient)")
                for s in data["shares"]:
                    port_lines.append(f"- {s['name']} ({s['type']})")

            elif tool == "sslscan":
                has_data = any([data.get("weak_protocols"), data.get("weak_ciphers"),
                                data.get("cert_expired"), data.get("heartbleed_vulnerable")])
                if has_data:
                    port_has_content = True
                    port_lines.append("\n#### TLS/SSL")
                    if data.get("weak_protocols"):
                        port_lines.append(f"- **Weak protocols:** {', '.join(data['weak_protocols'])}")
                    if data.get("weak_ciphers"):
                        port_lines.append(f"- **Weak ciphers:** {', '.join(data['weak_ciphers'][:5])}")
                    if data.get("cert_subject"):
                        port_lines.append(f"- **Certificate:** {data['cert_subject']}")
                    if data.get("cert_expired"):
                        port_lines.append(f"- **Expired:** {data.get('cert_not_after', 'yes')}")
                    if data.get("heartbleed_vulnerable"):
                        port_lines.append("- **HEARTBLEED: VULNERABLE**")

            elif tool == "snmpwalk" and data.get("sys_descr"):
                port_has_content = True
                port_lines.append(f"\n**SNMP sysDescr:** {data['sys_descr']}")
                if data.get("sys_name"):
                    port_lines.append(f"**sysName:** {data['sys_name']}")
                if data.get("running_processes"):
                    port_lines.append(f"**Processes:** {', '.join(data['running_processes'][:15])}")

            elif tool == "onesixtyone" and data.get("community_strings"):
                port_has_content = True
                port_lines.append("\n#### SNMP Community Strings")
                for cs in data["community_strings"]:
                    port_lines.append(f"- `{cs['community']}` — {cs['sys_descr'][:100]}")

            elif tool == "dnsrecon" and data.get("records"):
                port_has_content = True
                port_lines.append("\n#### DNS Records")
                if data.get("zone_transfer_successful"):
                    port_lines.append("**ZONE TRANSFER SUCCESSFUL**")
                for r in data["records"][:20]:
                    port_lines.append(f"- {r['type']} `{r['name']}` → {r['value']}")

            elif tool == "curl":
                if data.get("title") or data.get("interesting_headers"):
                    port_has_content = True
                    if data.get("title"):
                        port_lines.append(f"\n**Page title:** {data['title']}")
                    for h_name in data.get("interesting_headers", []):
                        port_lines.append(f"**{h_name}:** {data['headers'].get(h_name, '')}")

        if port_has_content:
            lines.extend(port_lines)
            lines.append("")
        else:
            empty_ports += 1

    if empty_ports:
        lines.append(f"_Plus {empty_ports} port(s) with generic/empty tool output._")

    return "\n".join(lines).strip()


def _update_host_note_autorecon(
    hosts_dir: Path,
    ip: str,
    hostname: str,
    target_data: dict,
    scan_source: str,
) -> tuple[str, str]:
    """Write or update the ## AutoRecon Enumeration section in a host note."""
    existing_path = _find_host_note_by_ip(hosts_dir, ip) if ip else None

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        if hostname and is_probable_fqdn(hostname):
            display = hostname
        else:
            display = ip or hostname or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports_section = ""
    existing_nessus_section = ""
    existing_burp_section = ""
    existing_loot = ""
    existing_deep_dives = ""
    existing_cross_source = ""
    existing_preamble_lines: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes           = extract_operator_notes(old_body)
        existing_open_ports_section = extract_body_section(old_body, "## Open Ports")
        existing_nessus_section     = extract_body_section(old_body, "## Nessus Findings")
        existing_burp_section       = extract_body_section(old_body, "## Burp Suite Findings")
        existing_loot               = extract_body_section(old_body, "## Loot")
        existing_deep_dives         = extract_body_section(old_body, DEEP_DIVE_SECTION)
        existing_cross_source       = extract_body_section(old_body, CROSS_SOURCE_SECTION)
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating AutoRecon section in: {host_path.name}")
    else:
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if hostname and hostname != ip:
            existing_preamble_lines.append(f"**Hostname:** {hostname}")
        logging.debug(f"Creating AutoRecon-only host note: {host_path.name}")

    # Merge frontmatter
    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname and hostname not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname]

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    existing_tags: list = existing_fm.get("tags", [])
    if "autorecon" not in existing_tags:
        existing_tags = existing_tags + ["autorecon"]

    fm: dict = {
        "ip":                  ip or existing_fm.get("ip"),
        "hostnames":           existing_hostnames,
        "status":              existing_fm.get("status", "not-started"),
        "tags":                existing_tags,
        "sources":             existing_sources,
        "nessus_max_severity": existing_fm.get("nessus_max_severity", 0),
        "autorecon_tools_run": target_data.get("summary", {}).get("total_tools_run", 0),
    }

    # Build body
    preamble = "\n".join(existing_preamble_lines).rstrip()
    body_lines: list[str] = []

    if preamble:
        body_lines.append(preamble)

    if existing_open_ports_section:
        body_lines += ["", "## Open Ports", existing_open_ports_section]

    if existing_nessus_section:
        body_lines += ["", "## Nessus Findings", existing_nessus_section]

    if existing_burp_section:
        body_lines += ["", "## Burp Suite Findings", existing_burp_section]

    body_lines += ["", "## AutoRecon Enumeration", _render_autorecon_section(target_data)]

    if existing_loot:
        body_lines += ["", "## Loot", existing_loot]

    if existing_deep_dives:
        body_lines += ["", DEEP_DIVE_SECTION, existing_deep_dives]

    if existing_cross_source:
        body_lines += ["", CROSS_SOURCE_SECTION, existing_cross_source]

    body_lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        body_lines.append(f"- [[Scans/{src_stem}|{src}]]")

    body_lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        body_lines += ["", existing_op_notes]

    body = "\n".join(body_lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_autorecon_vault(
    vault_dir: Path,
    scan_name: str,
    autorecon_data: dict,
    analysis_text: str | None,
    model_name: str | None,
    validation_warnings: list[str] | None = None,
) -> dict:
    """Write AutoRecon host note sections and scan note."""
    logging.info(f"Writing AutoRecon vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    targets = autorecon_data.get("targets", [])
    scan_display = autorecon_data.get("report_name", scan_name)
    scan_stem    = safe_filename(f"{scan_display} - AutoRecon")
    scan_source  = f"{scan_display} - AutoRecon"
    scan_path    = scans_dir / ensure_md_suffix(scan_stem)

    host_entries: list[tuple[str, str]] = []
    for target in targets:
        display, host_stem = _update_host_note_autorecon(
            hosts_dir,
            target.get("ip", ""),
            target.get("hostname", ""),
            target,
            scan_source,
        )
        host_entries.append((display, host_stem))

    # Build scan note
    total_tools = sum(t.get("summary", {}).get("total_tools_run", 0) for t in targets)
    tools_with_findings = sum(t.get("summary", {}).get("tools_with_findings", 0) for t in targets)

    scan_lines = [
        f"# {scan_display} - AutoRecon",
        "",
        f"- **Source:** {autorecon_data.get('source_file', scan_name)}",
        f"- **Parsed:** {autorecon_data.get('parsed_at', '')}",
        f"- **Targets:** {len(targets)}  ·  **Tools run:** {total_tools}  ·  **With findings:** {tools_with_findings}",
        "",
        "## Hosts",
        "",
    ]
    for display, host_stem in host_entries:
        scan_lines.append(f"- [[Hosts/{host_stem}|{display}]]")

    _acb = "/" if analysis_text else " "
    scan_lines += ["", f"- [{_acb}] Analyze: AutoRecon", "", "## Analysis", ""]
    if model_name:
        scan_lines.append(f"Model: `{model_name}`\n")
    scan_lines.append(analysis_text or "_No AI analysis generated._")

    if validation_warnings:
        scan_lines += ["", "## Validation Warnings", ""]
        for w in validation_warnings:
            scan_lines.append(f"- {w}")

    # Include manual commands in collapsible section
    for target in targets:
        mc = target.get("manual_commands", "")
        if mc.strip():
            scan_lines += [
                "", f"## Suggested Manual Commands ({target.get('target', '')})",
                "",
                "<details>",
                "<summary>Click to expand</summary>",
                "",
                "```",
                mc.strip(),
                "```",
                "",
                "</details>",
            ]
        cl = target.get("commands_log", "")
        if cl.strip():
            scan_lines += [
                "", f"## Commands Log ({target.get('target', '')})",
                "",
                "<details>",
                "<summary>Click to expand</summary>",
                "",
                "```",
                cl.strip()[:5000],
                "```",
                "",
                "</details>",
            ]

    scan_lines = _preserve_scan_note_operator_notes(scan_path, scan_lines)
    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Wrote AutoRecon scan note: {scan_path.name}")

    return {
        "scan_note":    str(scan_path),
        "scan_stem":    scan_stem,
        "scan_display": scan_display,
        "host_stems":   [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — Loot
# ============================================================

def _render_loot_section(loot_files: list[dict]) -> str:
    """Render loot data as markdown body (no section header)."""
    if not loot_files:
        return "_No loot collected for this host._"

    lines: list[str] = [f"**Loot files:** {len(loot_files)}", ""]

    all_creds = []
    all_hashes = []
    for lf in loot_files:
        for c in lf.get("credentials", []):
            all_creds.append({**c, "source": lf["filename"]})
        for h in lf.get("hashes", []):
            all_hashes.append({**h, "source": lf["filename"]})

    if all_creds:
        lines.append("#### Credentials Found")
        lines.append("| Username | Password/Hash | Type | Source |")
        lines.append("|----------|--------------|------|--------|")
        for c in all_creds[:50]:
            pw_display = c["password"][:40] + "..." if len(c["password"]) > 40 else c["password"]
            lines.append(f"| `{c['username']}` | `{pw_display}` | {c['cred_type']} | {c['source']} |")
        lines.append("")

    if all_hashes:
        lines.append("#### Hashes")
        lines.append("| Hash | Type | Username | Source |")
        lines.append("|------|------|----------|--------|")
        for h in all_hashes[:30]:
            user = f"`{h['username']}`" if h.get("username") else "—"
            h_display = h["hash"][:32] + "..." if len(h["hash"]) > 32 else h["hash"]
            lines.append(f"| `{h_display}` | {h['hash_type']} | {user} | {h['source']} |")
        lines.append("")

    for lf in loot_files:
        listings = lf.get("file_listings", [])
        if listings:
            lines.append(f"#### File Listing — {lf['filename']} ({len(listings)} entries)")
            lines.append("<details>")
            lines.append("<summary>Click to expand</summary>")
            lines.append("")
            lines.append("| Permissions | Size | Filename |")
            lines.append("|-------------|------|----------|")
            for entry in listings[:50]:
                lines.append(f"| {entry['permissions']} | {entry['size']} | `{entry['filename']}` |")
            lines.append("")
            lines.append("</details>")
            lines.append("")

    other_files = [lf for lf in loot_files
                   if not lf.get("credentials") and not lf.get("hashes") and not lf.get("file_listings")]
    if other_files:
        lines.append("#### Other Loot Files")
        for lf in other_files[:10]:
            lines.append(f"- **{lf['filename']}** — {lf['category']}, {lf['size_bytes']} bytes")

    return "\n".join(lines).strip()


def _render_loot_section_lightweight(loot_files: list[dict]) -> str:
    """Render a lightweight ## Loot section with summary stats and links to centralized pages."""
    if not loot_files:
        return "_No loot collected for this host._"

    total_creds = sum(len(lf.get("credentials", [])) for lf in loot_files)
    total_hashes = sum(len(lf.get("hashes", [])) for lf in loot_files)
    total_listings = sum(len(lf.get("file_listings", [])) for lf in loot_files)

    parts = [f"**Loot files:** {len(loot_files)}"]
    if total_creds:
        parts.append(f"**Credentials:** {total_creds}")
    if total_hashes:
        parts.append(f"**Hashes:** {total_hashes}")

    lines: list[str] = ["  ·  ".join(parts), ""]

    links: list[str] = []
    if total_creds:
        links.append("[[Loot/Credentials|Credentials]]")
    if total_hashes:
        links.append("[[Loot/Hashes|Hashes]]")
    if links:
        lines.append("See " + " and ".join(links) + " for full details.")
        lines.append("")

    # File listings stay inline (host-specific, not worth centralizing)
    for lf in loot_files:
        listings = lf.get("file_listings", [])
        if listings:
            lines.append(f"#### File Listing — {lf['filename']} ({len(listings)} entries)")
            lines.append("<details>")
            lines.append("<summary>Click to expand</summary>")
            lines.append("")
            lines.append("| Permissions | Size | Filename |")
            lines.append("|-------------|------|----------|")
            for entry in listings[:50]:
                lines.append(f"| {entry['permissions']} | {entry['size']} | `{entry['filename']}` |")
            lines.append("")
            lines.append("</details>")
            lines.append("")

    other_files = [lf for lf in loot_files
                   if not lf.get("credentials") and not lf.get("hashes") and not lf.get("file_listings")]
    if other_files:
        lines.append("#### Other Loot Files")
        for lf in other_files[:10]:
            lines.append(f"- **{lf['filename']}** — {lf['category']}, {lf['size_bytes']} bytes")

    return "\n".join(lines).strip()


_CRED_NOTES_SENTINEL = "### Operator Notes"
_CRED_NOTES_HINT = "_Note which hosts/services these credentials work on, and any other context._"

_CRED_CONFIRMED_RE = re.compile(
    r"^\s*-\s*\[x?\]\s*(?:Confirmed|Works)\s+on\s+(\S+)",
    re.MULTILINE | re.IGNORECASE,
)


_CRED_TABLE_ROW_RE = re.compile(
    r"^\|\s*`([^`]*)`\s*\|"    # Username
    r"\s*(.*?)\s*\|"           # Password
    r"\s*(.*?)\s*\|"           # Hash
    r"\s*(.*?)\s*\|"           # Hash Type
    r"\s*(.*?)\s*\|"           # Source
    r"\s*(.*?)\s*\|$"          # Notes
)


def _read_credential_annotations(creds_path: Path) -> dict[str, str]:
    """Read existing per-host operator notes from Credentials.md.

    Returns: {host_key_or_"Campaign-Level": notes_text}
    """
    annotations: dict[str, str] = {}
    if not creds_path.exists():
        return annotations

    try:
        text = creds_path.read_text(encoding="utf-8")
    except Exception:
        return annotations

    current_host: str | None = None
    in_notes = False
    notes_lines: list[str] = []

    for line in text.splitlines():
        if line.startswith("## "):
            if current_host and notes_lines:
                annotations[current_host] = "\n".join(notes_lines).strip()
            notes_lines = []
            in_notes = False
            header = line[3:].strip()
            if header.startswith("[["):
                pipe = header.find("|")
                bracket = header.find("]]")
                if pipe != -1 and bracket != -1:
                    current_host = header[pipe + 1:bracket]
                else:
                    current_host = header
            else:
                current_host = header
        elif line.strip() == _CRED_NOTES_SENTINEL:
            in_notes = True
        elif in_notes:
            if line.strip() == _CRED_NOTES_HINT:
                continue
            notes_lines.append(line)

    if current_host and notes_lines:
        annotations[current_host] = "\n".join(notes_lines).strip()

    return annotations


def _read_credential_row_notes(creds_path: Path) -> dict[str, dict[str, str]]:
    """Read per-credential inline notes from the table's Notes column.

    Returns: {host_key: {username: notes_text}}
    """
    row_notes: dict[str, dict[str, str]] = {}
    if not creds_path.exists():
        return row_notes

    try:
        text = creds_path.read_text(encoding="utf-8")
    except Exception:
        return row_notes

    current_host: str | None = None

    for line in text.splitlines():
        if line.startswith("## "):
            header = line[3:].strip()
            if header.startswith("[["):
                pipe = header.find("|")
                bracket = header.find("]]")
                if pipe != -1 and bracket != -1:
                    current_host = header[pipe + 1:bracket]
                else:
                    current_host = header
            else:
                current_host = header
            continue

        if not current_host:
            continue

        m = _CRED_TABLE_ROW_RE.match(line)
        if m:
            username = m.group(1)
            note = m.group(6).strip()
            if note and note != "—":
                row_notes.setdefault(current_host, {})[username] = note

    return row_notes


def _parse_confirmed_hosts(annotations: dict[str, str]) -> dict[str, list[str]]:
    """Extract confirmed access targets from credential annotations.

    Returns: {host_key: [target_host_or_service, ...]}
    """
    confirmed: dict[str, list[str]] = {}
    for host_key, notes in annotations.items():
        matches = _CRED_CONFIRMED_RE.findall(notes)
        if matches:
            confirmed[host_key] = matches
    return confirmed


def _classify_hash_type(value: str) -> str:
    """Identify hash type from the hash value string."""
    if not value:
        return ""
    if BCRYPT_HASH_RE.match(value):
        return "bcrypt"
    if value.startswith("$6$"):
        return "SHA-512 (Unix)"
    if value.startswith("$5$"):
        return "SHA-256 (Unix)"
    if value.startswith("$1$"):
        return "MD5 (Unix)"
    if value.startswith("$apr1$"):
        return "APR1-MD5"
    if NTLM_HASH_RE.fullmatch(value) and len(value) == 32:
        return "NTLM"
    if SHA256_HASH_RE.fullmatch(value) and len(value) == 64:
        return "SHA-256"
    if len(value) == 40 and all(c in "0123456789abcdefABCDEF" for c in value):
        return "SHA-1"
    if len(value) == 128 and all(c in "0123456789abcdefABCDEF" for c in value):
        return "SHA-512"
    if value.startswith("$"):
        return "crypt"
    return "unknown"


def _is_hash(value: str) -> bool:
    """Check if a credential password value is actually a hash."""
    if not value:
        return False
    if value.startswith("$"):
        return True
    if NTLM_HASH_RE.fullmatch(value) and len(value) == 32:
        return True
    if SHA256_HASH_RE.fullmatch(value) and len(value) == 64:
        return True
    if len(value) >= 32 and all(c in "0123456789abcdefABCDEF" for c in value):
        return True
    return False


def _build_cred_rows(
    loot_files: list[dict],
) -> list[dict]:
    """Build normalized credential rows from loot files.

    Each row: {username, password, hash, hash_type, source, inline_note}
    Separates passwords from hashes. Adds standalone usernames with blanks.
    """
    rows: list[dict] = []
    seen_users: set[str] = set()

    for lf in loot_files:
        source = lf["filename"]
        for c in lf.get("credentials", []):
            user = c["username"]
            value = c["password"]
            inline_note = c.get("inline_note", "")
            seen_users.add(user.lower())

            if _is_hash(value):
                rows.append({
                    "username": user,
                    "password": "",
                    "hash": value,
                    "hash_type": _classify_hash_type(value),
                    "source": source,
                    "inline_note": inline_note,
                })
            else:
                rows.append({
                    "username": user,
                    "password": value,
                    "hash": "",
                    "hash_type": "",
                    "source": source,
                    "inline_note": inline_note,
                })

        # Add hash-associated usernames from the hashes extractor
        for h in lf.get("hashes", []):
            if h.get("username"):
                user = h["username"]
                if user.lower() not in seen_users:
                    seen_users.add(user.lower())
                    rows.append({
                        "username": user,
                        "password": "",
                        "hash": h["hash"],
                        "hash_type": _classify_hash_type(h["hash"]),
                        "source": source,
                        "inline_note": "",
                    })

        # Add standalone usernames (no cred pair found)
        for user in lf.get("standalone_usernames", []):
            if user.lower() not in seen_users:
                seen_users.add(user.lower())
                rows.append({
                    "username": user,
                    "password": "",
                    "hash": "",
                    "hash_type": "",
                    "source": source,
                    "inline_note": "",
                })

    return rows


def _write_loot_credentials_page(
    loot_dir_out: Path,
    loot_data: dict,
    host_stem_map: dict[str, str],
) -> None:
    """Write Loot/Credentials.md with all credentials organized by host."""
    page_path = loot_dir_out / "Credentials.md"

    existing_annotations = _read_credential_annotations(page_path)
    existing_row_notes = _read_credential_row_notes(page_path)

    lines: list[str] = [
        "# Credentials & Users",
        "",
        f"_Last updated: {loot_data['parsed_at']}_",
        "",
    ]

    total_creds = loot_data["summary"].get("total_credentials", 0)
    lines.append(f"**Total credential entries:** {total_creds}")
    lines.append("")

    has_content = False

    for host_key, loot_files in sorted(loot_data["host_loot"].items()):
        rows = _build_cred_rows(loot_files)
        if not rows:
            continue
        has_content = True
        host_stem = host_stem_map.get(host_key, safe_filename(host_key))
        host_row_notes = existing_row_notes.get(host_key, {})
        lines.append(f"## [[Hosts/{host_stem}|{host_key}]]")
        lines.append("")
        lines.append("| Username | Password | Hash | Hash Type | Source | Notes |")
        lines.append("|----------|----------|------|-----------|--------|-------|")
        for r in rows[:50]:
            pw = f"`{r['password'][:50]}`" if r["password"] else "—"
            hsh = f"`{r['hash'][:40]}...`" if len(r.get("hash", "")) > 40 else (f"`{r['hash']}`" if r["hash"] else "—")
            ht = r["hash_type"] or "—"
            note = host_row_notes.get(r["username"], "") or r.get("inline_note", "") or "—"
            lines.append(f"| `{r['username']}` | {pw} | {hsh} | {ht} | {r['source']} | {note} |")
        lines.append("")

        lines.append(_CRED_NOTES_SENTINEL)
        lines.append(_CRED_NOTES_HINT)
        existing_notes = existing_annotations.get(host_key, "")
        if existing_notes:
            lines.append("")
            lines.append(existing_notes)
        lines.append("")

    campaign_rows = _build_cred_rows(loot_data.get("campaign_loot", []))
    if campaign_rows:
        has_content = True
        campaign_row_notes = existing_row_notes.get("Campaign-Level", {})
        lines.append("## Campaign-Level")
        lines.append("")
        lines.append("| Username | Password | Hash | Hash Type | Source | Notes |")
        lines.append("|----------|----------|------|-----------|--------|-------|")
        for r in campaign_rows[:50]:
            pw = f"`{r['password'][:50]}`" if r["password"] else "—"
            hsh = f"`{r['hash'][:40]}...`" if len(r.get("hash", "")) > 40 else (f"`{r['hash']}`" if r["hash"] else "—")
            ht = r["hash_type"] or "—"
            note = campaign_row_notes.get(r["username"], "") or r.get("inline_note", "") or "—"
            lines.append(f"| `{r['username']}` | {pw} | {hsh} | {ht} | {r['source']} | {note} |")
        lines.append("")

        lines.append(_CRED_NOTES_SENTINEL)
        lines.append(_CRED_NOTES_HINT)
        existing_notes = existing_annotations.get("Campaign-Level", "")
        if existing_notes:
            lines.append("")
            lines.append(existing_notes)
        lines.append("")

    if not has_content:
        lines.append("_No credentials or users found._")

    page_path.write_text("\n".join(lines), encoding="utf-8")
    logging.info(f"Wrote Loot/Credentials.md ({total_creds} entries)")


def _write_loot_hashes_page(
    loot_dir_out: Path,
    loot_data: dict,
    host_stem_map: dict[str, str],
) -> None:
    """Write Loot/Hashes.md with all hashes organized by host."""
    lines: list[str] = [
        "# Hashes",
        "",
        f"_Last updated: {loot_data['parsed_at']}_",
        "",
    ]

    total_hashes = loot_data["summary"].get("total_hashes", 0)
    lines.append(f"**Total hashes:** {total_hashes}")
    lines.append("")

    has_content = False

    for host_key, loot_files in sorted(loot_data["host_loot"].items()):
        all_hashes = []
        for lf in loot_files:
            for h in lf.get("hashes", []):
                all_hashes.append({**h, "source": lf["filename"]})
        if not all_hashes:
            continue
        has_content = True
        host_stem = host_stem_map.get(host_key, safe_filename(host_key))
        lines.append(f"## [[Hosts/{host_stem}|{host_key}]]")
        lines.append("")
        lines.append("| Hash | Type | Username | Source |")
        lines.append("|------|------|----------|--------|")
        for h in all_hashes[:50]:
            user = f"`{h['username']}`" if h.get("username") else "—"
            h_display = h["hash"][:40] + "..." if len(h["hash"]) > 40 else h["hash"]
            lines.append(f"| `{h_display}` | {h['hash_type']} | {user} | {h['source']} |")
        lines.append("")

    campaign_hashes = []
    for lf in loot_data.get("campaign_loot", []):
        for h in lf.get("hashes", []):
            campaign_hashes.append({**h, "source": lf["filename"]})
    if campaign_hashes:
        has_content = True
        lines.append("## Campaign-Level")
        lines.append("")
        lines.append("| Hash | Type | Username | Source |")
        lines.append("|------|------|----------|--------|")
        for h in campaign_hashes[:50]:
            user = f"`{h['username']}`" if h.get("username") else "—"
            h_display = h["hash"][:40] + "..." if len(h["hash"]) > 40 else h["hash"]
            lines.append(f"| `{h_display}` | {h['hash_type']} | {user} | {h['source']} |")
        lines.append("")

    if not has_content:
        lines.append("_No hashes found._")

    page_path = loot_dir_out / "Hashes.md"
    page_path.write_text("\n".join(lines), encoding="utf-8")
    logging.info(f"Wrote Loot/Hashes.md ({total_hashes} hashes)")


def _update_host_note_loot(
    hosts_dir: Path,
    ip: str,
    hostname: str,
    loot_files: list[dict],
    scan_source: str,
) -> tuple[str, str]:
    """Write or update the ## Loot section in a host note."""
    existing_path = _find_host_note_by_ip(hosts_dir, ip) if ip and IPV4_RE.match(ip) else None
    if not existing_path:
        for hp in hosts_dir.glob("*.md"):
            if hp.stem.lower() == (hostname or ip or "").lower():
                existing_path = hp
                break

    if existing_path:
        host_path = existing_path
        display   = host_path.stem
    else:
        if hostname and is_probable_fqdn(hostname):
            display = hostname
        else:
            display = ip or hostname or "unknown-host"
        host_path = hosts_dir / ensure_md_suffix(safe_filename(display))

    existing_fm: dict = {}
    existing_op_notes = ""
    existing_open_ports = ""
    existing_nessus = ""
    existing_burp = ""
    existing_autorecon = ""
    existing_loot = ""
    existing_deep_dives = ""
    existing_cross_source = ""
    existing_preamble_lines: list[str] = []

    if host_path.exists():
        old_text = host_path.read_text(encoding="utf-8")
        existing_fm, old_body = read_frontmatter(old_text)
        existing_op_notes  = extract_operator_notes(old_body)
        existing_open_ports = extract_body_section(old_body, "## Open Ports")
        existing_nessus    = extract_body_section(old_body, "## Nessus Findings")
        existing_burp      = extract_body_section(old_body, "## Burp Suite Findings")
        existing_autorecon = extract_body_section(old_body, "## AutoRecon Enumeration")
        existing_loot      = extract_body_section(old_body, "## Loot")
        existing_deep_dives = extract_body_section(old_body, DEEP_DIVE_SECTION)
        existing_cross_source = extract_body_section(old_body, CROSS_SOURCE_SECTION)
        for line in old_body.splitlines():
            if line.startswith("## "):
                break
            existing_preamble_lines.append(line)
        logging.debug(f"Updating Loot section in: {host_path.name}")
    else:
        if ip:
            existing_preamble_lines.append(f"**IP:** {ip}")
        if hostname and hostname != ip:
            existing_preamble_lines.append(f"**Hostname:** {hostname}")
        logging.debug(f"Creating Loot-only host note: {host_path.name}")

    existing_hostnames: list = existing_fm.get("hostnames", [])
    if hostname and hostname not in existing_hostnames:
        existing_hostnames = existing_hostnames + [hostname]

    existing_sources: list = existing_fm.get("sources", [])
    if scan_source not in existing_sources:
        existing_sources = existing_sources + [scan_source]

    existing_tags: list = existing_fm.get("tags", [])
    if "loot" not in existing_tags:
        existing_tags = existing_tags + ["loot"]

    fm: dict = {
        "ip":                  ip or existing_fm.get("ip"),
        "hostnames":           existing_hostnames,
        "status":              existing_fm.get("status", "not-started"),
        "tags":                existing_tags,
        "sources":             existing_sources,
        "nessus_max_severity":  existing_fm.get("nessus_max_severity", 0),
        "loot_file_count":      len(loot_files),
        "loot_credential_count": sum(len(lf.get("credentials", [])) for lf in loot_files),
        "loot_hash_count":       sum(len(lf.get("hashes", [])) for lf in loot_files),
    }
    if "autorecon_tools_run" in existing_fm:
        fm["autorecon_tools_run"] = existing_fm["autorecon_tools_run"]

    preamble = "\n".join(existing_preamble_lines).rstrip()
    body_lines: list[str] = []
    if preamble:
        body_lines.append(preamble)
    if existing_open_ports:
        body_lines += ["", "## Open Ports", existing_open_ports]
    if existing_nessus:
        body_lines += ["", "## Nessus Findings", existing_nessus]
    if existing_burp:
        body_lines += ["", "## Burp Suite Findings", existing_burp]
    if existing_autorecon:
        body_lines += ["", "## AutoRecon Enumeration", existing_autorecon]

    body_lines += ["", "## Loot", _render_loot_section_lightweight(loot_files)]

    if existing_deep_dives:
        body_lines += ["", DEEP_DIVE_SECTION, existing_deep_dives]

    if existing_cross_source:
        body_lines += ["", CROSS_SOURCE_SECTION, existing_cross_source]

    body_lines += ["", "## Scan References"]
    for src in fm["sources"]:
        src_stem = safe_filename(src)
        body_lines.append(f"- [[Scans/{src_stem}|{src}]]")

    body_lines += ["", OPERATOR_NOTES_SENTINEL, OPERATOR_NOTES_HINT]
    if existing_op_notes:
        body_lines += ["", existing_op_notes]

    body = "\n".join(body_lines)
    host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return display, host_path.stem


def create_loot_vault(
    vault_dir: Path,
    loot_data: dict,
    analysis_by_host: dict[str, str] | None,
    model_name: str | None,
) -> dict:
    """Write centralized Loot/ pages and lightweight host note sections."""
    logging.info(f"Writing Loot vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir    = vault_dir / "Hosts"
    loot_dir_out = vault_dir / "Loot"
    hosts_dir.mkdir(exist_ok=True)
    loot_dir_out.mkdir(exist_ok=True)

    scan_stem   = "Loot"
    scan_source = "Loot"

    # Update host notes with lightweight loot sections
    host_entries: list[tuple[str, str]] = []
    host_stem_map: dict[str, str] = {}

    for host_key, loot_files in loot_data.get("host_loot", {}).items():
        ip = host_key if IPV4_RE.match(host_key) else ""
        hostname = "" if ip else host_key
        display, host_stem = _update_host_note_loot(
            hosts_dir, ip, hostname, loot_files, scan_source,
        )
        host_entries.append((display, host_stem))
        host_stem_map[host_key] = host_stem

    # Write centralized Loot pages
    _write_loot_credentials_page(loot_dir_out, loot_data, host_stem_map)
    _write_loot_hashes_page(loot_dir_out, loot_data, host_stem_map)

    # Write Loot/Overview.md (replaces old Scans/Loot.md)
    summary = loot_data.get("summary", {})
    overview_lines = [
        "# Loot Overview",
        "",
        f"- **Source:** {loot_data.get('source_dir', '')}",
        f"- **Parsed:** {loot_data.get('parsed_at', '')}",
        f"- **Total files:** {summary.get('total_files', 0)}  ·  "
        f"**Host-associated:** {summary.get('host_files', 0)}  ·  "
        f"**Campaign-level:** {summary.get('campaign_files', 0)}",
        f"- **Credentials:** {summary.get('total_credentials', 0)}  ·  "
        f"**Hashes:** {summary.get('total_hashes', 0)}",
        "",
        "## Pages",
        "",
        "- [[Loot/Credentials|Credentials]]",
        "- [[Loot/Hashes|Hashes]]",
        "",
    ]

    if host_entries:
        overview_lines += ["## Hosts with Loot", ""]
        for display, host_stem in host_entries:
            overview_lines.append(f"- [[Hosts/{host_stem}|{display}]]")
        overview_lines.append("")

    # Per-host loot analysis sections with checkboxes
    loot_host_keys = list(loot_data.get("host_loot", {}).keys())
    has_campaign = bool(loot_data.get("campaign_loot"))
    analysis_dict = analysis_by_host or {}

    for host_key in sorted(loot_host_keys):
        analysis = analysis_dict.get(host_key, "")
        _acb = "/" if analysis else " "
        overview_lines += [
            f"- [{_acb}] Analyze: Loot — {host_key}",
            "",
            f"## Analysis — {host_key}",
            "",
        ]
        if model_name and analysis:
            overview_lines.append(f"Model: `{model_name}`\n")
        overview_lines.append(analysis or "_No AI analysis generated._")
        overview_lines.append("")

    if has_campaign:
        campaign_analysis = analysis_dict.get("_campaign", "")
        _acb = "/" if campaign_analysis else " "
        overview_lines += [
            f"- [{_acb}] Analyze: Loot — campaign",
            "",
            "## Campaign-Level Loot Analysis",
            "",
        ]
        if model_name and campaign_analysis:
            overview_lines.append(f"Model: `{model_name}`\n")
        overview_lines.append(campaign_analysis or "_No AI analysis generated._")
        overview_lines.append("")

    campaign_loot = loot_data.get("campaign_loot", [])
    if campaign_loot:
        overview_lines += ["## Campaign-Level Loot Files", ""]
        for lf in campaign_loot:
            overview_lines.append(f"- **{lf['filename']}** — {lf['category']}, {lf['size_bytes']} bytes")
            if lf.get("credentials"):
                overview_lines.append(f"  Credentials: {len(lf['credentials'])}")
            if lf.get("hashes"):
                overview_lines.append(f"  Hashes: {len(lf['hashes'])}")

    overview_path = loot_dir_out / "Overview.md"
    overview_lines = _preserve_scan_note_operator_notes(overview_path, overview_lines)
    overview_path.write_text("\n".join(overview_lines), encoding="utf-8")
    logging.info(f"Wrote Loot/Overview.md")

    return {
        "scan_note":     str(overview_path),
        "scan_stem":     scan_stem,
        "scan_display":  "Loot",
        "scan_path_rel": "Loot/Overview.md",
        "host_stems":    [stem for _, stem in host_entries],
    }


# ============================================================
# Vault writing — Misc
# ============================================================

def _add_scan_reference_to_host(
    hosts_dir: Path, host_key: str, scan_source: str,
) -> str | None:
    """Add a scan reference to an existing host note. Returns host_stem or None."""
    ip = host_key if IPV4_RE.match(host_key) else ""
    existing_path = _find_host_note_by_ip(hosts_dir, ip) if ip else None
    if not existing_path and not ip:
        candidate = hosts_dir / ensure_md_suffix(safe_filename(host_key))
        if candidate.exists():
            existing_path = candidate
    if not existing_path:
        return None

    text = existing_path.read_text(encoding="utf-8")
    fm, body = read_frontmatter(text)

    sources: list = fm.get("sources", [])
    if scan_source in sources:
        return existing_path.stem

    sources = sources + [scan_source]
    fm["sources"] = sources

    ref_line = f"- [[Scans/{safe_filename(scan_source)}|{scan_source}]]"
    if ref_line not in body:
        body = body.replace(
            OPERATOR_NOTES_SENTINEL,
            f"{ref_line}\n\n{OPERATOR_NOTES_SENTINEL}",
        )

    existing_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
    return existing_path.stem


def create_misc_vault(
    vault_dir: Path,
    misc_data: dict,
    analysis_by_file: dict[str, str] | None,
    model_name: str | None,
) -> list[dict]:
    """Write Scans/<filename> - Misc.md for each misc file."""
    logging.info(f"Writing Misc vault content: {vault_dir}")
    vault_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    hosts_dir.mkdir(exist_ok=True)
    scans_dir.mkdir(exist_ok=True)

    results: list[dict] = []

    for file_info in misc_data.get("files", []):
        fname = file_info["filename"]
        stem = Path(fname).stem
        scan_display = f"{stem} - Misc"
        scan_stem = safe_filename(scan_display)
        scan_path = scans_dir / ensure_md_suffix(scan_stem)

        host_key = file_info.get("host_key")
        host_stems: list[str] = []

        tool_type = file_info.get("tool_type", "unknown")
        analysis_level = file_info.get("analysis_level", "standard")

        lines = [
            f"# {scan_display}",
            "",
            f"- **Source:** {file_info['filepath']}",
            f"- **Parsed:** {misc_data.get('parsed_at', '')}",
            f"- **Size:** {file_info['size_bytes']} bytes",
            f"- **Detected tool:** {tool_type}",
            f"- **Analysis level:** {analysis_level}",
        ]

        if host_key:
            lines.append(f"- **Host:** {host_key}")
            host_stem = _add_scan_reference_to_host(hosts_dir, host_key, scan_display)
            if host_stem:
                host_stems.append(host_stem)
                lines[-1] = f"- **Host:** [[Hosts/{host_stem}|{host_key}]]"

        # Raw content in collapsible block
        content = file_info.get("content", "")
        preview = content[:10000]
        if len(content) > 10000:
            preview += f"\n\n[... truncated, {len(content)} total chars ...]"

        lines += [
            "",
            "## Raw Output",
            "",
            "<details>",
            "<summary>Click to expand</summary>",
            "",
            "```",
            preview,
            "```",
            "",
            "</details>",
        ]

        # AI analysis
        analysis = (analysis_by_file or {}).get(fname, "")
        _acb = "/" if analysis else " "
        lines += ["", f"- [{_acb}] Analyze: Misc", "", "## Analysis", ""]
        if model_name:
            lines.append(f"Model: `{model_name}`\n")
        lines.append(analysis or "_No AI analysis generated._")

        lines = _preserve_scan_note_operator_notes(scan_path, lines)
        scan_path.write_text("\n".join(lines), encoding="utf-8")
        logging.info(f"Wrote misc scan note: {scan_path.name}")

        results.append({
            "scan_note":    str(scan_path),
            "scan_stem":    scan_stem,
            "scan_display": scan_display,
            "host_stems":   host_stems,
        })

    return results


# ============================================================
# Canvas node / edge constructors
# ============================================================

def _text_node(nid: str, text: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "text", "text": text, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _file_node(nid: str, file: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "file", "file": file, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _group_node(nid: str, label: str, x: int, y: int, w: int, h: int, color: str | None = None) -> dict:
    n: dict = {"id": nid, "type": "group", "label": label, "x": x, "y": y, "width": w, "height": h}
    if color:
        n["color"] = color
    return n


def _edge(eid: str, from_id: str, to_id: str, label: str = "") -> dict:
    e: dict = {
        "id":       eid,
        "fromNode": from_id,
        "fromSide": "bottom",
        "toNode":   to_id,
        "toSide":   "top",
    }
    if label:
        e["label"] = label
    return e


# ============================================================
# Canvas content helpers
# ============================================================

def _read_host_frontmatter(host_path: Path) -> dict:
    try:
        fm, _ = read_frontmatter(host_path.read_text(encoding="utf-8"))
        return fm
    except Exception:
        return {}


def _build_campaign_overview(
    vault_dir: Path,
    scan_host_map: dict,
    all_analyses: list[tuple[str, str]],
) -> str:
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    total_hosts = len(host_notes)
    total_scans = len(scan_host_map)
    status_counts: dict[str, int] = {}
    all_tags: set[str] = set()
    total_ports = 0
    nessus_critical = 0
    nessus_high     = 0
    nessus_medium   = 0
    burp_high       = 0
    burp_total      = 0
    autorecon_targets = 0
    autorecon_writable = 0
    loot_hosts = 0
    loot_creds = 0

    for hp in host_notes:
        text = ""
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue

        fm, body = read_frontmatter(text)
        status = fm.get("status", "not-started")
        status_counts[status] = status_counts.get(status, 0) + 1
        all_tags.update(fm.get("tags", []))

        # Count Nmap open ports
        total_ports += sum(
            1 for line in body.splitlines()
            if line.strip().startswith("- **") and "/" in line
            and line.strip().startswith("- **tcp") or
            (line.strip().startswith("- **") and "/" in line and not line.strip().startswith("- **tcp"))
        )

        # Count Nessus severity from frontmatter
        max_sev = fm.get("nessus_max_severity", 0)
        try:
            max_sev = int(max_sev)
        except (TypeError, ValueError):
            max_sev = 0
        if max_sev >= 4:
            nessus_critical += 1
        elif max_sev >= 3:
            nessus_high += 1
        elif max_sev >= 2:
            nessus_medium += 1

        # Count Burp issues from section
        burp_section = extract_body_section(body, "## Burp Suite Findings")
        if burp_section and "_No Burp" not in burp_section:
            burp_total += burp_section.count("#### [")
            burp_high  += burp_section.count("#### [High")

        # Count AutoRecon targets
        ar_section = extract_body_section(body, "## AutoRecon Enumeration")
        if ar_section and "_AutoRecon ran but" not in ar_section:
            autorecon_targets += 1
            if "READ/WRITE" in ar_section or "read/write" in ar_section:
                autorecon_writable += 1

        # Count loot from frontmatter
        lc = fm.get("loot_credential_count", 0)
        try:
            lc = int(lc)
        except (TypeError, ValueError):
            lc = 0
        if lc:
            loot_hosts += 1
            loot_creds += lc

    # Re-count ports more simply
    total_ports = 0
    for hp in host_notes:
        try:
            _, body = read_frontmatter(hp.read_text(encoding="utf-8"))
            ports_section = extract_body_section(body, "## Open Ports")
            total_ports += sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))
        except Exception:
            pass

    lines = [
        "# Campaign Overview",
        "",
        f"**Hosts:** {total_hosts}  ·  **Scans:** {total_scans}  ·  **Open Ports:** {total_ports}",
        "",
    ]

    if status_counts:
        parts = [f"{k}: {v}" for k, v in sorted(status_counts.items())]
        lines += [f"**Progress:** {' · '.join(parts)}", ""]

    display_tags = sorted(all_tags - {"rpc"})
    if display_tags:
        lines += [f"**Services seen:** {', '.join(display_tags)}", ""]

    nessus_parts = []
    if nessus_critical:
        nessus_parts.append(f"critical hosts: {nessus_critical}")
    if nessus_high:
        nessus_parts.append(f"high hosts: {nessus_high}")
    if nessus_medium:
        nessus_parts.append(f"medium hosts: {nessus_medium}")
    if nessus_parts:
        lines += [f"**Nessus:** {' · '.join(nessus_parts)}", ""]

    if burp_total:
        lines += [f"**Burp:** {burp_total} issues ({burp_high} high)", ""]

    if autorecon_targets:
        ar_parts = [f"targets: {autorecon_targets}"]
        if autorecon_writable:
            ar_parts.append(f"writable shares: {autorecon_writable}")
        lines += [f"**AutoRecon:** {' · '.join(ar_parts)}", ""]

    if loot_hosts:
        loot_parts = [f"hosts: {loot_hosts}"]
        if loot_creds:
            loot_parts.append(f"credentials: {loot_creds}")
        lines += [f"**Loot:** {' · '.join(loot_parts)}", ""]

    # Key Observations bullets from the most recent AI analysis
    if all_analyses:
        text = all_analyses[-1][1]
        match = re.search(
            r"##\s+Key Observations(.*?)(?=\n##\s|\Z)", text, re.DOTALL | re.IGNORECASE
        )
        if match:
            bullets = [
                l.strip() for l in match.group(1).splitlines()
                if l.strip().startswith(("-", "*", "•"))
            ][:4]
            if bullets:
                lines += ["**Key Findings:**"] + bullets

    return "\n".join(lines)


def build_priority_targets_prompt(vault_dir: Path, all_analyses: list[tuple[str, str]],
                                  operator_notes_by_ip: dict[str, str] | None = None) -> str:
    """
    Build an Ollama prompt asking it to rank all discovered hosts by exploitation
    priority, drawing on Nmap service tags, Nessus CVEs/CVSS, Burp findings, and AutoRecon enumeration.
    """
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    instructions = """
You are an experienced penetration tester. Given scan data about multiple hosts,
rank them from highest to lowest exploitation priority.
This ranking will be used in a professional penetration testing engagement
under a signed Rules of Engagement. Accuracy matters more than completeness.

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- Only reference ports, services, CVEs, and findings that are explicitly present
  in the data provided. Do not invent or infer CVE IDs, version numbers, or
  service details not shown.
- Only rank a host as a top target if there is concrete scan evidence supporting
  exploitability. Do not rank based on port count alone.
- If evidence is insufficient to justify a high ranking, say so rather than
  padding the list.
- Tag each ranked entry as one of:
    [CONFIRMED]  — directly in the scan data (CVE with plugin output, confirmed vuln)
    [INFERRED]   — reasonable conclusion from the scan data (service fingerprint, role)
    [ASSUMED]    — requires verification; explain why

Output ONLY a numbered list — no preamble, no section headers, no trailing text.
Each entry must be a single line in this exact format:
  N. <hostname or IP> [TAG] — <primary evidence: CVE ID / finding name / service>; <brief impact>

Base your ranking on:
- CVEs with confirmed plugin output and known weaponized exploits ([MSF] first)
- CVSS scores (critical ≥ 9.0, high ≥ 7.0) — but only when plugin_output:yes
- Exposed management interfaces with evidence of weak auth (RDP, WinRM, SSH)
- Domain role evidence in hostnames, ports, or service banners (DC, CA, DB)
- Chaining potential supported by scan data (pivot hosts, credential stores)
- Web findings from Burp (Certain/Firm auth bypass, SQLi — not Tentative)
- AutoRecon enumeration: null sessions, writable shares, discovered users, weak TLS
- Loot: discovered credentials, cracked hashes (hosts with valid creds rank highest)
- Operator notes indicating confirmed access or exploitation progress
- Unencrypted or unauthenticated services confirmed in scan data

Do not rank a host highly based solely on port count or service tags without
supporting vulnerability or misconfiguration evidence.
Maximum 10 entries.
""".strip()

    host_lines: list[str] = []
    for hp in host_notes:
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue
        fm, body = read_frontmatter(text)

        ip        = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        display   = hostnames[0] if hostnames else (ip or hp.stem)
        tags      = fm.get("tags", [])
        nessus_max = fm.get("nessus_max_severity", 0)
        try:
            nessus_max = int(nessus_max)
        except (TypeError, ValueError):
            nessus_max = 0

        ports_section  = extract_body_section(body, "## Open Ports")
        port_count     = sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))

        nessus_section = extract_body_section(body, "## Nessus Findings")
        crit_count     = nessus_section.count("[Critical]") if nessus_section else 0
        high_count_n   = nessus_section.count("[High]")     if nessus_section else 0
        cves           = re.findall(r"CVE-\d{4}-\d+", nessus_section)[:4] if nessus_section else []

        burp_section   = extract_body_section(body, "## Burp Suite Findings")
        burp_high      = burp_section.count("[High") if burp_section else 0
        burp_med       = burp_section.count("[Medium") if burp_section else 0

        ar_section     = extract_body_section(body, "## AutoRecon Enumeration")
        ar_null_sess   = bool(ar_section and "Null Session:** Yes" in ar_section)
        ar_writable    = bool(ar_section and ("READ/WRITE" in ar_section or "read/write" in ar_section))
        ar_users       = len(re.findall(r"`(\w+)`", ar_section)) if ar_section else 0

        parts = [f"{display} (IP: {ip})"]
        if tags:
            parts.append(f"services: {', '.join(tags)}")
        parts.append(f"open-ports: {port_count}")
        if nessus_max:
            parts.append(f"nessus-max: {severity_int_to_str(nessus_max)}")
        if crit_count or high_count_n:
            parts.append(f"critical/high findings: {crit_count}/{high_count_n}")
        if cves:
            parts.append(f"CVEs: {', '.join(cves)}")
        if burp_high or burp_med:
            parts.append(f"burp high/medium: {burp_high}/{burp_med}")
        if ar_null_sess:
            parts.append("null-session: yes")
        if ar_writable:
            parts.append("writable-shares: yes")
        if ar_users:
            parts.append(f"users-found: {ar_users}")

        loot_creds = fm.get("loot_credential_count", 0)
        loot_hashes = fm.get("loot_hash_count", 0)
        try:
            loot_creds = int(loot_creds)
        except (TypeError, ValueError):
            loot_creds = 0
        try:
            loot_hashes = int(loot_hashes)
        except (TypeError, ValueError):
            loot_hashes = 0
        if loot_creds:
            parts.append(f"loot-credentials: {loot_creds}")
        if loot_hashes:
            parts.append(f"loot-hashes: {loot_hashes}")

        if operator_notes_by_ip and ip and ip in operator_notes_by_ip:
            parts.append("operator-notes: yes")

        host_lines.append("- " + " | ".join(parts))

    # Include Key Observations snippets from up to 3 most recent analyses for context
    analysis_ctx: list[str] = []
    for scan_display, text in all_analyses[-3:]:
        m = re.search(r"##\s+Key Observations(.*?)(?=\n##\s|\Z)", text, re.DOTALL | re.IGNORECASE)
        if m:
            bullets = [
                l.strip() for l in m.group(1).splitlines()
                if l.strip().startswith(("-", "*", "•"))
            ][:3]
            if bullets:
                analysis_ctx.append(f"[{scan_display}]")
                analysis_ctx.extend(bullets)

    prompt = f"{instructions}\n\nHosts to rank:\n" + "\n".join(host_lines)
    if analysis_ctx:
        prompt += "\n\nKey scan findings for context:\n" + "\n".join(analysis_ctx)
    logging.debug(f"Built priority targets prompt ({len(prompt)} chars)")
    return prompt


def _build_priority_targets_fallback(vault_dir: Path) -> str:
    """
    Build a static Priority Targets list (no AI) sorted by Nessus severity,
    Burp high count, and open port count.  Used when Ollama is skipped or
    no analyses have been generated yet.
    """
    hosts_dir  = vault_dir / "Hosts"
    host_notes = sorted(hosts_dir.glob("*.md")) if hosts_dir.exists() else []

    entries: list[dict] = []
    for hp in host_notes:
        try:
            text = hp.read_text(encoding="utf-8")
        except Exception:
            continue
        fm, body = read_frontmatter(text)

        ip        = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        display   = hostnames[0] if hostnames else (ip or hp.stem)
        tags      = fm.get("tags", [])
        nessus_max = fm.get("nessus_max_severity", 0)
        try:
            nessus_max = int(nessus_max)
        except (TypeError, ValueError):
            nessus_max = 0

        ports_section  = extract_body_section(body, "## Open Ports")
        port_count     = sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))

        nessus_section = extract_body_section(body, "## Nessus Findings")
        cves           = re.findall(r"CVE-\d{4}-\d+", nessus_section)[:2] if nessus_section else []

        burp_section   = extract_body_section(body, "## Burp Suite Findings")
        burp_high      = burp_section.count("[High") if burp_section else 0

        reasons: list[str] = []
        if cves:
            reasons.append(", ".join(cves))
        if "domain-controller" in tags:
            reasons.append("domain controller")
        if nessus_max >= 4:
            reasons.append("critical Nessus findings")
        elif nessus_max >= 3:
            reasons.append("high-severity Nessus findings")
        elif nessus_max >= 2:
            reasons.append("medium-severity Nessus findings")
        if burp_high:
            reasons.append(f"{burp_high} high web issue(s)")
        exposed = [t for t in tags if t in ("rdp", "winrm", "smb", "kerberos", "ldap", "mssql", "ftp", "snmp")]
        if exposed:
            reasons.append(f"exposed: {', '.join(exposed)}")
        if not reasons and port_count:
            service_str = ", ".join(tags[:4]) if tags else "unknown services"
            reasons.append(f"{port_count} open port(s) — {service_str}")

        entries.append({
            "display":  display,
            "sort_key": (-nessus_max, -burp_high, -port_count),
            "reason":   "; ".join(reasons) if reasons else "no notable findings",
        })

    entries.sort(key=lambda e: e["sort_key"])

    if not entries:
        return "## Priority Targets\n\n_No hosts found. Run scans to populate._"

    lines = ["## Priority Targets", ""]
    for i, entry in enumerate(entries[:10], 1):
        lines.append(f"{i}. {entry['display']} — {entry['reason']}")
    return "\n".join(lines)


def _build_next_steps(all_analyses: list[tuple[str, str]], max_items: int = 14) -> str:
    items: list[str] = []

    for _scan_display, text in all_analyses:
        for section in (
            "Potential Attack Paths",
            "Enumeration Suggestions",
            "Exploitation Priority",
            "Attack Chains",
        ):
            match = re.search(
                rf"##\s+{re.escape(section)}(.*?)(?=\n##\s|\Z)",
                text, re.DOTALL | re.IGNORECASE,
            )
            if not match:
                continue
            for line in match.group(1).splitlines():
                stripped = line.strip()
                if not stripped or stripped.startswith("#"):
                    continue
                cleaned = re.sub(r"^[-*•\d]+[.)]\s*", "", stripped).strip()
                if cleaned and len(cleaned) > 12 and cleaned not in items:
                    items.append(cleaned)
                if len(items) >= max_items:
                    break
            if len(items) >= max_items:
                break

    if not items:
        return "# Next Steps\n\n_Run scans with AI analysis enabled to populate next steps._"

    lines = ["# Next Steps", ""]
    for item in items[:max_items]:
        lines.append(f"- {item}")
    return "\n".join(lines)


# ============================================================
# Canvas layout engine
# ============================================================

def _group_dimensions(n_hosts: int, cols: int) -> tuple[int, int]:
    if n_hosts == 0:
        return 300, 200
    actual_cols = min(n_hosts, cols)
    rows = math.ceil(n_hosts / cols)
    inner_w = actual_cols * CARD_W + (actual_cols - 1) * CARD_GAP_X
    inner_h = rows * CARD_H + (rows - 1) * CARD_GAP_Y
    w = GROUP_PAD * 2 + inner_w
    h = GROUP_PAD + GROUP_LABEL_H + inner_h + GROUP_PAD
    return w, h


def build_canvas(
    vault_dir: Path,
    canvas_name: str,
    scan_host_map: dict,
    all_analyses: list[tuple[str, str]],
    canvas_cols: int = 2,
    max_groups_per_row: int = 3,
    priority_targets_text: str | None = None,
    scan_path_overrides: dict[str, str] | None = None,
) -> Path:
    hosts_dir   = vault_dir / "Hosts"
    canvas_path = vault_dir / canvas_name

    existing_canvas: dict = {"nodes": [], "edges": []}
    if canvas_path.exists():
        try:
            existing_canvas = json.loads(canvas_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    managed_ids: set[str] = set()
    our_nodes:   list[dict] = []
    our_edges:   list[dict] = []

    # 1. Campaign Overview
    overview_id = stable_id("campaign_overview")
    managed_ids.add(overview_id)
    overview_x = -(OVERVIEW_W // 2)
    overview_y = -900
    our_nodes.append(_text_node(
        overview_id,
        _build_campaign_overview(vault_dir, scan_host_map, all_analyses),
        overview_x, overview_y, OVERVIEW_W, OVERVIEW_H,
    ))

    # 2. Scan note cards
    scan_stems = list(scan_host_map.keys())
    total_scan_row_w = (
        len(scan_stems) * SCAN_CARD_W + max(0, len(scan_stems) - 1) * CARD_GAP_X
    )
    scan_row_x0 = -(total_scan_row_w // 2)
    scan_row_y  = overview_y + OVERVIEW_H + 100

    scan_node_ids: dict[str, str] = {}

    for i, scan_stem in enumerate(scan_stems):
        sid = stable_id(f"scan_{scan_stem}")
        managed_ids.add(sid)
        scan_node_ids[scan_stem] = sid
        sx  = scan_row_x0 + i * (SCAN_CARD_W + CARD_GAP_X)
        _overrides = scan_path_overrides or {}
        rel = _overrides.get(scan_stem, f"Scans/{ensure_md_suffix(scan_stem)}")
        our_nodes.append(_file_node(sid, rel, sx, scan_row_y, SCAN_CARD_W, SCAN_CARD_H))
        eid = stable_id(f"edge_overview_{sid}")
        managed_ids.add(eid)
        our_edges.append(_edge(eid, overview_id, sid))

    # 2b. Priority Targets node (centered, between scan cards and subnet groups)
    pt_id = stable_id("priority-targets")
    managed_ids.add(pt_id)
    pt_text = priority_targets_text or _build_priority_targets_fallback(vault_dir)
    # Ensure the markdown header is present
    if not pt_text.lstrip().startswith("## Priority Targets"):
        pt_text = "## Priority Targets\n\n" + pt_text.lstrip()
    pt_x = -(NEXT_STEPS_W // 2)
    pt_y = scan_row_y + SCAN_CARD_H + ROW_GAP
    our_nodes.append(_text_node(pt_id, pt_text, pt_x, pt_y, NEXT_STEPS_W, NEXT_STEPS_H, color="2"))

    # Overview → Priority Targets
    eid_ov_pt = stable_id("edge_overview_pt")
    managed_ids.add(eid_ov_pt)
    our_edges.append(_edge(eid_ov_pt, overview_id, pt_id))

    # 3. Subnet groups with host cards
    subnet_hosts: dict[str, list[tuple[str, Path]]] = {}
    if hosts_dir.exists():
        for hp in sorted(hosts_dir.glob("*.md")):
            fm     = _read_host_frontmatter(hp)
            ip     = fm.get("ip")
            subnet = get_subnet_label(ip) if ip and is_ipv4(str(ip)) else "unknown"
            subnet_hosts.setdefault(subnet, []).append((hp.stem, hp))

    subnets  = sorted(subnet_hosts.keys())
    n_groups = len(subnets)
    gdims    = [_group_dimensions(len(subnet_hosts[s]), canvas_cols) for s in subnets]

    # Groups start below the Priority Targets node
    cur_y = pt_y + NEXT_STEPS_H + ROW_GAP
    group_positions: list[tuple[int, int]] = []

    for row_i in range(math.ceil(n_groups / max_groups_per_row)):
        start = row_i * max_groups_per_row
        end   = min(start + max_groups_per_row, n_groups)
        row_indices = range(start, end)
        row_total_w = sum(gdims[j][0] for j in row_indices) + (end - start - 1) * GROUP_GAP
        row_height  = max((gdims[j][1] for j in row_indices), default=200)
        x = -(row_total_w // 2)
        for j in row_indices:
            group_positions.append((x, cur_y))
            x += gdims[j][0] + GROUP_GAP
        cur_y += row_height + ROW_GAP

    host_node_ids: dict[str, str] = {}

    for i, subnet in enumerate(subnets):
        if i >= len(group_positions):
            break
        gx, gy = group_positions[i]
        gw, gh = gdims[i]

        gid = stable_id(f"subnet_{subnet}")
        managed_ids.add(gid)
        our_nodes.append(_group_node(gid, subnet, gx, gy, gw, gh))

        # Priority Targets → subnet group edge
        eid_pt_g = stable_id(f"edge_pt_{gid}")
        managed_ids.add(eid_pt_g)
        our_edges.append(_edge(eid_pt_g, pt_id, gid))

        for j, (host_stem, hp) in enumerate(subnet_hosts[subnet]):
            col = j % canvas_cols
            row = j // canvas_cols
            hx  = gx + GROUP_PAD + col * (CARD_W + CARD_GAP_X)
            hy  = gy + GROUP_PAD + GROUP_LABEL_H + row * (CARD_H + CARD_GAP_Y)

            hid = stable_id(f"host_{host_stem}")
            managed_ids.add(hid)
            host_node_ids[host_stem] = hid

            fm     = _read_host_frontmatter(hp)
            status = fm.get("status", "not-started")
            color  = STATUS_COLORS.get(status)

            # If not-started, use Nessus severity for color hint
            if color is None and status == "not-started":
                nessus_max = fm.get("nessus_max_severity", 0)
                try:
                    nessus_max = int(nessus_max)
                except (TypeError, ValueError):
                    nessus_max = 0
                color = NESSUS_SEVERITY_CANVAS_COLORS.get(nessus_max)

            rel = f"Hosts/{ensure_md_suffix(host_stem)}"
            our_nodes.append(_file_node(hid, rel, hx, hy, CARD_W, CARD_H, color))

    # Scan → host edges
    for scan_stem, host_stems in scan_host_map.items():
        sid = scan_node_ids.get(scan_stem)
        if not sid:
            continue
        for host_stem in host_stems:
            hid = host_node_ids.get(host_stem)
            if not hid:
                continue
            eid = stable_id(f"edge_{sid}_{hid}")
            managed_ids.add(eid)
            our_edges.append(_edge(eid, sid, hid))

    # 4. Next Steps
    next_id = stable_id("next_steps")
    managed_ids.add(next_id)
    next_x = -(NEXT_STEPS_W // 2)
    next_y = cur_y + 40
    our_nodes.append(_text_node(
        next_id,
        _build_next_steps(all_analyses),
        next_x, next_y, NEXT_STEPS_W, NEXT_STEPS_H,
        color="5",
    ))

    preserved   = [n for n in existing_canvas.get("nodes", []) if n.get("id") not in managed_ids]
    final_nodes = preserved + our_nodes
    final_edges = our_edges

    canvas_path.write_text(
        json.dumps({"nodes": final_nodes, "edges": final_edges}, indent=2),
        encoding="utf-8",
    )
    logging.info(
        f"Canvas written: {canvas_path.name} "
        f"({len(final_nodes)} nodes, {len(final_edges)} edges, "
        f"{len(preserved)} preserved)"
    )
    return canvas_path


# ============================================================
# Eat Me — operator drop zone
# ============================================================

_EAT_ME_TEMPLATE = """\
# Eat Me

_Paste anything here — IPs, tool output, notes, arp tables, host lists._
_On next mAIpper run, content is processed and moved to a scan note._

---

"""

_ARP_LINE_RE = re.compile(
    r"[?\s]*\(?(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\)?\s+"
    r"(?:at\s+)?([0-9a-fA-F:.\-]{11,17})",
)

_EAT_ME_PROCESSED_SENTINEL = "---\n"


def _install_eat_me_page(vault_dir: Path) -> None:
    """Create the Eat Me page if it doesn't exist."""
    eat_me_path = vault_dir / "Eat Me.md"
    if not eat_me_path.exists():
        eat_me_path.write_text(_EAT_ME_TEMPLATE, encoding="utf-8")
        logging.info("Created Eat Me page")


def _append_eat_me_to_credentials_md(
    vault_dir: Path,
    creds: list[dict],
    usernames: list[str],
    source_label: str,
) -> int:
    """Append Eat Me credentials and potential usernames to Loot/Credentials.md.

    Returns the count of new entries added. Skips duplicates by username (case-insensitive).
    """
    if not creds and not usernames:
        return 0

    loot_dir = vault_dir / "Loot"
    loot_dir.mkdir(exist_ok=True)
    page_path = loot_dir / "Credentials.md"

    existing_text = page_path.read_text(encoding="utf-8") if page_path.exists() else ""

    # Collect already-known usernames to skip duplicates
    known_users: set[str] = set()
    for line in existing_text.splitlines():
        m = re.match(r"\|\s*`([^`]+)`\s*\|", line)
        if m:
            known_users.add(m.group(1).lower())

    # Build new table rows
    new_rows: list[str] = []
    for c in creds:
        if c["username"].lower() in known_users:
            continue
        known_users.add(c["username"].lower())
        if c["cred_type"] in ("NTLM_hash", "bcrypt_hash", "SHA256_hash", "hash", "bcrypt_hash"):
            pw_col = "—"
            hsh_col = f"`{c['password'][:40]}`"
            ht_col = c["cred_type"]
        else:
            pw_col = f"`{c['password'][:50]}`"
            hsh_col = "—"
            ht_col = "—"
        new_rows.append(f"| `{c['username']}` | {pw_col} | {hsh_col} | {ht_col} | {source_label} | — |")

    for user in usernames:
        if user.lower() in known_users:
            continue
        known_users.add(user.lower())
        new_rows.append(f"| `{user}` | — | — | — | {source_label} (potential) | — |")

    if not new_rows:
        return 0

    table_header = [
        "| Username | Password | Hash | Hash Type | Source | Notes |",
        "|----------|----------|------|-----------|--------|-------|",
    ]

    if not existing_text.strip():
        lines = [
            "# Credentials & Users",
            "",
            f"_Last updated: {dt.datetime.now().isoformat(timespec='seconds')}_",
            "",
            "## Campaign-Level",
            "",
        ] + table_header + new_rows + ["", _CRED_NOTES_SENTINEL, _CRED_NOTES_HINT, ""]
        page_path.write_text("\n".join(lines), encoding="utf-8")
        return len(new_rows)

    # File exists — find the Campaign-Level section and insert rows before its Operator Notes
    lines = existing_text.splitlines()
    in_campaign = False
    table_header_found = False
    insert_idx: int | None = None

    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped == "## Campaign-Level":
            in_campaign = True
            table_header_found = False
            continue
        if in_campaign and stripped.startswith("## ") and stripped != "## Campaign-Level":
            insert_idx = i
            break
        if in_campaign and stripped.startswith("|---"):
            table_header_found = True
            continue
        if in_campaign and table_header_found and stripped == _CRED_NOTES_SENTINEL:
            insert_idx = i
            break

    if insert_idx is not None:
        new_lines = lines[:insert_idx] + new_rows + lines[insert_idx:]
    elif in_campaign:
        # Reached end of file inside Campaign-Level with no sentinel
        new_lines = lines + new_rows
    else:
        # No Campaign-Level section — append one
        new_lines = lines + [
            "",
            "## Campaign-Level",
            "",
        ] + table_header + new_rows + ["", _CRED_NOTES_SENTINEL, _CRED_NOTES_HINT, ""]

    page_path.write_text("\n".join(new_lines), encoding="utf-8")
    return len(new_rows)


def _process_eat_me_page(vault_dir: Path, args=None) -> dict | None:
    """Process the Eat Me page: extract IPs, hostnames, creds, etc.

    Returns a dict of extracted data, or None if nothing to process.
    Moves processed content to Scans/ and resets the page.
    """
    eat_me_path = vault_dir / "Eat Me.md"
    if not eat_me_path.exists():
        return None

    try:
        raw = eat_me_path.read_text(encoding="utf-8")
    except Exception:
        return None

    # Strip template boilerplate to find user content
    content = raw
    for prefix in [
        "# Eat Me\n",
        "_Paste anything here",
        "_On next mAIpper run",
        "---",
    ]:
        idx = content.find(prefix)
        if idx != -1:
            newline = content.find("\n", idx)
            if newline != -1:
                content = content[newline + 1:]

    content = content.strip()
    if not content:
        return None

    logging.info(f"Eat Me: processing {len(content)} chars of operator input")

    # ── Extract data ──
    discovered_ips: set[str] = set()
    discovered_hostnames: set[str] = set()
    discovered_urls: set[str] = set()
    discovered_creds: list[dict] = []
    ip_mac_map: dict[str, str] = {}

    # Parse arp -a output (Linux and Windows formats)
    for m in _ARP_LINE_RE.finditer(content):
        ip = m.group(1)
        mac = m.group(2)
        if _is_target_ip(ip):
            discovered_ips.add(ip)
            ip_mac_map[ip] = mac

    # Parse plain IPs
    for ip in IP_IN_TEXT_RE.findall(content):
        if _is_target_ip(ip):
            discovered_ips.add(ip)

    # Parse FQDNs
    for m in FQDN_IN_TEXT_RE.finditer(content):
        fqdn = m.group(0).lower().rstrip(".")
        if is_probable_fqdn(fqdn) and "." in fqdn:
            discovered_hostnames.add(fqdn)

    # Parse URLs (extract domain)
    url_re = re.compile(r"https?://([a-zA-Z0-9\-]+(?:\.[a-zA-Z0-9\-]+)+)(?:[:/]|$)", re.IGNORECASE)
    for m in url_re.finditer(content):
        domain = m.group(1).lower()
        if not IPV4_RE.match(domain):
            discovered_urls.add(domain)
            discovered_hostnames.add(domain)

    # Parse credentials and standalone usernames
    discovered_creds = _extract_credentials(content)
    known_cred_users = {c["username"] for c in discovered_creds}
    discovered_usernames = _extract_usernames(content, known_cred_users)

    if not discovered_ips and not discovered_hostnames and not discovered_creds and not discovered_usernames:
        logging.info("Eat Me: no structured data found, saving as misc")

    # ── Create host notes for new IPs ──
    hosts_dir = vault_dir / "Hosts"
    hosts_dir.mkdir(exist_ok=True)
    new_hosts: list[str] = []

    for ip in sorted(discovered_ips):
        existing = _find_host_note_by_ip(hosts_dir, ip)
        if existing:
            logging.debug(f"Eat Me: host {ip} already exists ({existing.stem})")
            continue

        new_hosts.append(ip)
        host_stem = safe_filename(ip)
        host_path = hosts_dir / ensure_md_suffix(host_stem)

        # Check if any discovered hostname maps to this IP
        # (we can't resolve, but if the content has "hostname (IP)" patterns, link them)
        associated_hostnames: list[str] = []
        for hn in discovered_hostnames:
            # Check if hostname appears near this IP in the content
            for m in re.finditer(
                rf"(?:{re.escape(hn)}.*?{re.escape(ip)}|{re.escape(ip)}.*?{re.escape(hn)})",
                content[:5000],
            ):
                associated_hostnames.append(hn)
                break

        tags: list[str] = ["eat-me"]
        mac = ip_mac_map.get(ip, "")

        fm = {
            "ip": ip,
            "hostnames": associated_hostnames,
            "status": "not-started",
            "tags": tags,
            "sources": ["Eat Me"],
        }

        body_lines = [
            f"**IP:** {ip}",
        ]
        if mac:
            body_lines.append(f"**MAC:** {mac}")
        if associated_hostnames:
            body_lines.append(f"**Hostname:** {', '.join(associated_hostnames)}")

        body_lines += [
            "",
            "## Next Steps",
            "",
            f"- [ ] Port scan: `nmap -sC -sV -oX scans/nmap/{ip}.xml {ip}`",
            f"- [ ] Full TCP: `nmap -p- -T4 -oX scans/nmap/{ip}_full.xml {ip}`",
            f"- [ ] UDP top ports: `nmap -sU --top-ports 50 -oX scans/nmap/{ip}_udp.xml {ip}`",
        ]
        if mac:
            body_lines.append(f"- [ ] Identify OS/vendor from MAC `{mac}`")

        body_lines += [
            "",
            OPERATOR_NOTES_SENTINEL,
            OPERATOR_NOTES_HINT,
            "",
            "_Discovered via Eat Me page._",
        ]

        body = "\n".join(body_lines)
        host_path.write_text(write_frontmatter(fm) + "\n" + body, encoding="utf-8")
        logging.info(f"Eat Me: created host note for {ip}")

    # ── Write scan note with the original content ──
    scans_dir = vault_dir / "Scans"
    scans_dir.mkdir(exist_ok=True)

    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")

    # Name the note — try LLM for accuracy, fall back to Python heuristic
    content_label = None
    if args and not getattr(args, "no_ollama", True):
        try:
            preview = content[:3000]
            naming_prompt = (
                "You are naming a file for a penetration testing engagement notebook. "
                "Read the content below and respond with ONLY a short descriptive title "
                "(2-5 words, no quotes, no punctuation). Examples: 'NetExec SMB Scan', "
                "'ARP Table', 'Domain User List', 'Gobuster Results', 'FTP Credentials', "
                "'Nmap Service Scan', 'BloodHound Output'.\n\n"
                f"CONTENT:\n```\n{preview}\n```\n\nTITLE:"
            )
            raw_name = ollama_chat(
                args.ollama_url, args.model, naming_prompt, 0.1,
            ).strip().strip("'\"").strip()
            if raw_name and len(raw_name) <= 60 and "\n" not in raw_name:
                content_label = raw_name
                logging.info(f"Eat Me: LLM named note '{content_label}'")
        except Exception as exc:
            logging.debug(f"Eat Me: LLM naming failed, using heuristic: {exc}")

    if not content_label:
        content_label = "Notes"
        arp_lines = len(_ARP_LINE_RE.findall(content))
        if arp_lines >= 3:
            content_label = "ARP Table"
        elif discovered_creds and not discovered_ips:
            content_label = "Credentials"
        elif len(discovered_ips) >= 3 and not discovered_creds:
            content_label = "Host List"
        elif discovered_creds and discovered_ips:
            content_label = "Recon Data"
        else:
            tool_type, _ = _detect_misc_tool("eat_me", content)
            if tool_type not in ("unknown", "notes"):
                content_label = f"{tool_type.title()} Output"
            elif len(discovered_ips) > 0:
                content_label = "Host Discovery"
            elif discovered_hostnames:
                content_label = "DNS Info"

    scan_display = f"{content_label} - {timestamp}"
    scan_stem = safe_filename(scan_display)
    scan_path = scans_dir / ensure_md_suffix(scan_stem)

    eat_me_tool_type, eat_me_analysis_level = _detect_misc_tool(content_label, content)

    scan_lines = [
        f"# {scan_display}",
        "",
        f"- **Processed:** {dt.datetime.now().isoformat(timespec='seconds')}",
        f"- **IPs found:** {len(discovered_ips)}",
        f"- **New hosts created:** {len(new_hosts)}",
        f"- **Hostnames found:** {len(discovered_hostnames)}",
        f"- **Credentials found:** {len(discovered_creds)}",
        f"- **Usernames found:** {len(discovered_usernames)}",
        f"- **Detected tool:** {eat_me_tool_type}",
        f"- **Analysis level:** {eat_me_analysis_level}",
        "",
    ]

    if new_hosts:
        scan_lines += ["## New Hosts", ""]
        for ip in new_hosts:
            host_stem = safe_filename(ip)
            scan_lines.append(f"- [[Hosts/{host_stem}|{ip}]]")
        scan_lines.append("")

    if discovered_ips - set(new_hosts):
        scan_lines += ["## Existing Hosts Referenced", ""]
        for ip in sorted(discovered_ips - set(new_hosts)):
            existing = _find_host_note_by_ip(hosts_dir, ip)
            if existing:
                scan_lines.append(f"- [[Hosts/{existing.stem}|{ip}]]")
            else:
                scan_lines.append(f"- {ip}")
        scan_lines.append("")

    if discovered_hostnames:
        scan_lines += ["## Hostnames", ""]
        for hn in sorted(discovered_hostnames):
            scan_lines.append(f"- {hn}")
        scan_lines.append("")

    if discovered_creds:
        scan_lines += ["## Credentials Found", ""]
        scan_lines.append("| Username | Password/Hash | Type |")
        scan_lines.append("|----------|--------------|------|")
        for c in discovered_creds[:30]:
            pw = c["password"][:40] + "..." if len(c["password"]) > 40 else c["password"]
            scan_lines.append(f"| `{c['username']}` | `{pw}` | {c['cred_type']} |")
        scan_lines.append("")

    if discovered_usernames:
        scan_lines += ["## Potential Usernames", ""]
        scan_lines.append("_These names were found without an associated password. Added to Loot/Credentials.md._")
        scan_lines.append("")
        for u in discovered_usernames[:50]:
            scan_lines.append(f"- `{u}`")
        scan_lines.append("")

    # Analysis checkbox + analysis section + original content
    scan_lines += [
        "",
        "- [ ] Analyze: Misc",
        "",
        "## Analysis",
        "",
        "_No AI analysis generated._",
        "",
        "## Original Input",
        "",
        "<details>",
        "<summary>Click to expand</summary>",
        "",
        "```",
        content,
        "```",
        "",
        "</details>",
    ]

    scan_path.write_text("\n".join(scan_lines), encoding="utf-8")
    logging.info(f"Eat Me: wrote scan note {scan_path.name}")

    # ── Save credentials and usernames directly to Loot/Credentials.md ──
    added = _append_eat_me_to_credentials_md(
        vault_dir, discovered_creds, discovered_usernames, scan_display
    )
    if added:
        logging.info(f"Eat Me: added {added} entries to Loot/Credentials.md")

    # ── Reset the Eat Me page ──
    eat_me_path.write_text(_EAT_ME_TEMPLATE, encoding="utf-8")
    logging.info("Eat Me: page reset for next use")

    return {
        "scan_stem": scan_stem,
        "scan_display": scan_display,
        "discovered_ips": sorted(discovered_ips),
        "new_hosts": new_hosts,
        "discovered_hostnames": sorted(discovered_hostnames),
        "discovered_creds": discovered_creds,
        "discovered_usernames": discovered_usernames,
        "host_stems": [safe_filename(ip) for ip in new_hosts],
    }


# ============================================================
# Campaign Targets Note
# ============================================================

_URL_RE = re.compile(r"https?://([a-zA-Z0-9\-]+(?:\.[a-zA-Z0-9\-]+)+)(?:[:/]|$)", re.IGNORECASE)

# RFC 5737 documentation, loopback, link-local, multicast, broadcast
_BOGUS_IP_PREFIXES = (
    "192.0.2.", "198.51.100.", "203.0.113.",  # RFC 5737 TEST-NET
    "127.", "0.", "255.",                       # loopback, unspecified, broadcast
    "169.254.",                                 # link-local
    "224.", "225.", "226.", "227.", "228.",      # multicast
    "229.", "230.", "231.", "232.", "233.",
    "234.", "235.", "236.", "237.", "238.", "239.",
)


def _is_target_ip(ip: str) -> bool:
    """Return True if ip looks like a real assessment target, not documentation/noise."""
    if not ip or not IPV4_RE.match(ip):
        return False
    if ip.startswith(_BOGUS_IP_PREFIXES):
        return False
    # Filter IPs where any octet > 255 (malformed)
    try:
        octets = [int(o) for o in ip.split(".")]
        if any(o > 255 for o in octets):
            return False
    except ValueError:
        return False
    return True


def _write_campaign_targets_note(
    vault_dir: Path,
    loot_data: dict | None = None,
    misc_data: dict | None = None,
) -> None:
    """Write Hosts/Campaign Targets.md — copy-paste-ready target lists."""
    hosts_dir = vault_dir / "Hosts"
    scans_dir = vault_dir / "Scans"
    loot_dir = vault_dir / "Loot"

    ips: set[str] = set()
    hostnames: set[str] = set()
    subnets: set[str] = set()
    hosts_entries: dict[str, list[str]] = {}  # ip → [hostname, ...]

    # Scan all host notes for IPs and hostnames
    if hosts_dir.exists():
        for hp in sorted(hosts_dir.glob("*.md")):
            if hp.stem == "_Campaign Targets":
                continue
            try:
                fm, body = read_frontmatter(hp.read_text(encoding="utf-8"))
            except Exception:
                continue

            ip = fm.get("ip", "")
            if ip and _is_target_ip(str(ip)):
                ips.add(ip)
                subnets.add(get_subnet_label(ip))

            host_fqdns: list[str] = []
            for hn in fm.get("hostnames", []):
                if hn and is_probable_fqdn(str(hn)):
                    clean = str(hn).rstrip(".")
                    hostnames.add(clean)
                    host_fqdns.append(clean)

            if ip and _is_target_ip(str(ip)) and host_fqdns:
                hosts_entries[ip] = host_fqdns

    if not ips and not hostnames:
        return

    lines: list[str] = [
        "# Campaign Targets",
        "",
        "_Auto-generated target lists for copy-paste into tools._",
        f"_Last updated: {dt.datetime.now().isoformat(timespec='seconds')}_",
        "",
    ]

    if ips:
        lines += [
            "## IPs",
            "",
            "```",
        ]
        for ip in sorted(ips, key=lambda x: tuple(int(p) for p in x.split("."))):
            lines.append(ip)
        lines += ["```", ""]

    if subnets:
        lines += [
            "## Subnets",
            "",
            "```",
        ]
        for s in sorted(subnets):
            lines.append(s)
        lines += ["```", ""]

    if hostnames:
        lines += [
            "## Hostnames",
            "",
            "```",
        ]
        for hn in sorted(hostnames):
            lines.append(hn)
        lines += ["```", ""]

    if hosts_entries:
        lines += [
            "## /etc/hosts",
            "",
            "```",
        ]
        for ip in sorted(hosts_entries, key=lambda x: tuple(int(p) for p in x.split("."))):
            names = " ".join(hosts_entries[ip])
            lines.append(f"{ip}\t{names}")
        lines += ["```", ""]

    # Quick stats
    lines += [
        "## Summary",
        "",
        f"- **IPs:** {len(ips)}",
        f"- **Subnets:** {len(subnets)}",
        f"- **Hostnames:** {len(hostnames)}",
    ]

    targets_path = hosts_dir / "_Campaign Targets.md"
    targets_path.write_text("\n".join(lines), encoding="utf-8")
    logging.info(
        f"Wrote Campaign Targets note: {len(ips)} IPs, "
        f"{len(subnets)} subnets, {len(hostnames)} hostnames"
    )


# ============================================================
# Users Canvas
# ============================================================

def _collect_credential_access_map(
    loot_data: dict | None,
    vault_dir: Path,
) -> dict[str, list[dict]]:
    """Build a map of username → list of access entries.

    Each access entry: {
        "username", "password", "hash", "hash_type", "source_file",
        "host_key", "host_stem", "note", "confirmed_targets": [str],
    }
    Includes standalone usernames (no password or hash).
    """
    access_map: dict[str, list[dict]] = {}
    if not loot_data:
        return access_map

    hosts_dir = vault_dir / "Hosts"
    loot_dir = vault_dir / "Loot"

    creds_path = loot_dir / "Credentials.md"
    annotations = _read_credential_annotations(creds_path)
    confirmed = _parse_confirmed_hosts(annotations)
    row_notes = _read_credential_row_notes(creds_path)

    host_stem_lookup: dict[str, str] = {}
    if hosts_dir.exists():
        for hp in hosts_dir.glob("*.md"):
            fm = _read_host_frontmatter(hp)
            ip = fm.get("ip", "")
            if ip:
                host_stem_lookup[ip] = hp.stem
            for hn in fm.get("hostnames", []):
                if hn:
                    host_stem_lookup[hn] = hp.stem

    for host_key, loot_files in loot_data.get("host_loot", {}).items():
        host_stem = host_stem_lookup.get(host_key, safe_filename(host_key))
        confirmed_targets = confirmed.get(host_key, [])
        host_row = row_notes.get(host_key, {})
        rows = _build_cred_rows(loot_files)
        for r in rows:
            username = r["username"]
            has_pw = bool(r["password"])
            has_hash = bool(r["hash"])
            if has_hash:
                cred_type = r["hash_type"] or "hash"
            elif has_pw:
                cred_type = "cleartext"
            else:
                cred_type = "username only"
            entry = {
                "username": username,
                "password": r["password"],
                "hash": r["hash"],
                "hash_type": r["hash_type"],
                "cred_type": cred_type,
                "source_file": r["source"],
                "host_key": host_key,
                "host_stem": host_stem,
                "note": host_row.get(username, ""),
                "confirmed_targets": confirmed_targets,
            }
            access_map.setdefault(username, []).append(entry)

    campaign_row = row_notes.get("Campaign-Level", {})
    campaign_rows = _build_cred_rows(loot_data.get("campaign_loot", []))
    confirmed_targets = confirmed.get("Campaign-Level", [])
    for r in campaign_rows:
        username = r["username"]
        has_pw = bool(r["password"])
        has_hash = bool(r["hash"])
        if has_hash:
            cred_type = r["hash_type"] or "hash"
        elif has_pw:
            cred_type = "cleartext"
        else:
            cred_type = "username only"
        entry = {
            "username": username,
            "password": r["password"],
            "hash": r["hash"],
            "hash_type": r["hash_type"],
            "cred_type": cred_type,
            "source_file": r["source"],
            "host_key": "_campaign",
            "host_stem": "",
            "note": campaign_row.get(username, ""),
            "confirmed_targets": confirmed_targets,
        }
        access_map.setdefault(username, []).append(entry)

    return access_map


def _uc_node_height(text: str, width: int = UC_USER_W) -> int:
    """Estimate canvas node height based on text content."""
    lines = text.split("\n")
    chars_per_line = max(width // 8, 20)
    wrapped = 0
    for line in lines:
        if not line.strip():
            wrapped += 1
        else:
            wrapped += max(1, math.ceil(len(line) / chars_per_line))
    return max(UC_USER_H, min(60 + wrapped * 22, 500))


def build_users_canvas(
    vault_dir: Path,
    loot_data: dict | None,
    canvas_name: str = "Users Canvas.canvas",
) -> Path | None:
    """Build the Users Canvas showing credential → host → service relationships."""
    access_map = _collect_credential_access_map(loot_data, vault_dir)
    if not access_map:
        logging.info("Users Canvas: no credentials found, skipping")
        return None

    hosts_dir = vault_dir / "Hosts"
    canvas_path = vault_dir / canvas_name

    existing_canvas: dict = {"nodes": [], "edges": []}
    if canvas_path.exists():
        try:
            existing_canvas = json.loads(canvas_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    managed_ids: set[str] = set()
    our_nodes: list[dict] = []
    our_edges: list[dict] = []

    # Title node
    title_id = stable_id("uc_title")
    managed_ids.add(title_id)

    total_users = len(access_map)
    total_entries = sum(len(v) for v in access_map.values())
    all_hosts: set[str] = set()
    for entries in access_map.values():
        for e in entries:
            if e["host_key"] != "_campaign":
                all_hosts.add(e["host_key"])

    title_text = (
        "# Users & Access Map\n\n"
        f"**Unique users:** {total_users}  ·  "
        f"**Credential entries:** {total_entries}  ·  "
        f"**Hosts with creds:** {len(all_hosts)}\n\n"
        "_Edges show where credentials were found or verified.\n"
        "Green = operator-confirmed access._"
    )
    our_nodes.append(_text_node(title_id, title_text, -350, -280, 700, 160))

    # ── Group credentials by source host, not account type ──
    # Each source host becomes a group on the left.
    # Campaign-level creds get their own group.
    source_groups: dict[str, list[str]] = {}
    for username, entries in sorted(access_map.items()):
        host_keys = sorted({e["host_key"] for e in entries})
        group_key = host_keys[0] if len(host_keys) == 1 else "|".join(host_keys)
        source_groups.setdefault(group_key, []).append(username)

    # Collect all target hosts for the right side
    host_keys_seen: set[str] = set()
    for entries in access_map.values():
        for e in entries:
            if e["host_key"] != "_campaign":
                host_keys_seen.add(e["host_key"])
        for e in entries:
            for t in e.get("confirmed_targets", []):
                if IPV4_RE.match(t) or is_probable_fqdn(t):
                    host_keys_seen.add(t)

    # ── Build user node text with full context ──
    user_texts: dict[str, str] = {}
    for username, entries in access_map.items():
        lines: list[str] = [f"**{username}**", ""]

        seen: set[tuple[str, str]] = set()
        for e in entries:
            key = (e["host_key"], e["source_file"])
            if key in seen:
                continue
            seen.add(key)
            host_label = e["host_key"] if e["host_key"] != "_campaign" else "campaign-level"
            # Show what we have for this credential
            detail_parts: list[str] = []
            if e.get("password"):
                detail_parts.append("password")
            if e.get("hash"):
                ht = e.get("hash_type", "hash")
                detail_parts.append(ht if ht else "hash")
            if not detail_parts:
                detail_parts.append("username only")
            detail = ", ".join(detail_parts)
            line = f"- `{detail}` from `{e['source_file']}` ({host_label})"
            lines.append(line)

        notes = [e["note"] for e in entries if e.get("note")]
        if notes:
            unique_notes = list(dict.fromkeys(notes))
            lines.append("")
            for n in unique_notes:
                lines.append(f"> {n}")

        user_texts[username] = "\n".join(lines)

    # ── Layout: source groups on the left, hosts on the right ──
    left_x = -550
    right_x = left_x + UC_USER_W + UC_COL_GAP
    cur_y = 0

    user_node_ids: dict[str, str] = {}

    group_color_cycle = ["6", "3", None, "2"]
    group_idx = 0

    for group_key in sorted(source_groups.keys()):
        usernames = source_groups[group_key]

        if group_key == "_campaign":
            group_label = "Campaign-Level Credentials"
        elif "|" in group_key:
            parts = group_key.split("|")
            labels = [p if p != "_campaign" else "campaign" for p in parts]
            group_label = f"Found across: {', '.join(labels)}"
        else:
            group_label = f"Found on {group_key}"

        # Pre-compute node heights so the group can be sized correctly
        node_heights: list[int] = []
        for username in usernames:
            h = _uc_node_height(user_texts[username])
            node_heights.append(h)

        total_nodes_h = sum(node_heights) + max(0, len(usernames) - 1) * (UC_GAP_Y // 2)
        group_h = UC_GROUP_PAD * 2 + UC_GROUP_LABEL_H + total_nodes_h

        color = group_color_cycle[group_idx % len(group_color_cycle)]
        group_idx += 1

        gid = stable_id(f"uc_group_{group_key}")
        managed_ids.add(gid)
        our_nodes.append(_group_node(
            gid, group_label,
            left_x - UC_GROUP_PAD, cur_y - UC_GROUP_PAD,
            UC_USER_W + UC_GROUP_PAD * 2, group_h,
            color=color,
        ))

        uy = cur_y + UC_GROUP_LABEL_H
        for i, username in enumerate(usernames):
            uid = stable_id(f"uc_user_{username}")
            managed_ids.add(uid)
            user_node_ids[username] = uid

            nh = node_heights[i]
            our_nodes.append(_text_node(uid, user_texts[username], left_x, uy, UC_USER_W, nh))
            uy += nh + UC_GAP_Y // 2

        cur_y += group_h + ROW_GAP

    # ── Host nodes on the right side ──
    host_node_ids: dict[str, str] = {}
    sorted_hosts = sorted(host_keys_seen)

    host_subnets: dict[str, list[str]] = {}
    for hk in sorted_hosts:
        subnet = get_subnet_label(hk) if is_ipv4(hk) else "other"
        host_subnets.setdefault(subnet, []).append(hk)

    host_y = 0
    for subnet, host_list in sorted(host_subnets.items()):
        n_hosts = len(host_list)
        subnet_group_h = (
            UC_GROUP_PAD * 2 + UC_GROUP_LABEL_H
            + n_hosts * UC_HOST_H
            + max(0, n_hosts - 1) * (UC_GAP_Y // 2)
        )

        sgid = stable_id(f"uc_subnet_{subnet}")
        managed_ids.add(sgid)
        our_nodes.append(_group_node(
            sgid, subnet,
            right_x - UC_GROUP_PAD, host_y - UC_GROUP_PAD,
            UC_HOST_W + UC_GROUP_PAD * 2, subnet_group_h,
        ))

        for j, hk in enumerate(host_list):
            hy = host_y + UC_GROUP_LABEL_H + j * (UC_HOST_H + UC_GAP_Y // 2)
            hid = stable_id(f"uc_host_{hk}")
            managed_ids.add(hid)
            host_node_ids[hk] = hid

            host_stem = None
            if hosts_dir.exists():
                hp = _find_host_note_by_ip(hosts_dir, hk) if IPV4_RE.match(hk) else None
                if not hp:
                    candidate = hosts_dir / ensure_md_suffix(safe_filename(hk))
                    if candidate.exists():
                        hp = candidate
                if hp:
                    host_stem = hp.stem

            if host_stem:
                rel = f"Hosts/{ensure_md_suffix(host_stem)}"
                our_nodes.append(_file_node(hid, rel, right_x, hy, UC_HOST_W, UC_HOST_H))
            else:
                our_nodes.append(_text_node(hid, f"**{hk}**", right_x, hy, UC_HOST_W, UC_HOST_H))

        host_y += subnet_group_h + ROW_GAP

    # ── Edges: user → host ──
    for username, entries in access_map.items():
        uid = user_node_ids.get(username)
        if not uid:
            continue

        seen_edges: set[str] = set()

        # Group entries by host for richer edge labels
        host_entries: dict[str, list[dict]] = {}
        for e in entries:
            if e["host_key"] != "_campaign":
                host_entries.setdefault(e["host_key"], []).append(e)

        for hk, hk_entries in host_entries.items():
            hid = host_node_ids.get(hk)
            if not hid:
                continue
            edge_key = f"{username}→{hk}"
            if edge_key in seen_edges:
                continue
            seen_edges.add(edge_key)

            eid = stable_id(f"uc_edge_{username}_{hk}")
            managed_ids.add(eid)

            # Build edge label from cred type + note context
            label_parts: list[str] = []
            cred_types = sorted({ent["cred_type"] for ent in hk_entries})
            label_parts.append(", ".join(cred_types))
            notes = [ent["note"] for ent in hk_entries if ent.get("note")]
            if notes:
                label_parts.append(notes[0][:40])

            edge: dict = {
                "id": eid,
                "fromNode": uid,
                "fromSide": "right",
                "toNode": hid,
                "toSide": "left",
            }
            if label_parts:
                edge["label"] = " · ".join(label_parts)
            our_edges.append(edge)

        # Confirmed access edges (from bottom Operator Notes annotations)
        for e in entries:
            for target in e.get("confirmed_targets", []):
                target_hid = host_node_ids.get(target)
                if not target_hid:
                    continue
                edge_key = f"{username}→{target}_confirmed"
                if edge_key in seen_edges:
                    continue
                seen_edges.add(edge_key)

                eid = stable_id(f"uc_confirmed_{username}_{target}")
                managed_ids.add(eid)
                our_edges.append({
                    "id": eid,
                    "fromNode": uid,
                    "fromSide": "right",
                    "toNode": target_hid,
                    "toSide": "left",
                    "color": "4",  # green
                    "label": "confirmed access",
                })

    # Preserve user-created nodes
    preserved = [n for n in existing_canvas.get("nodes", []) if n.get("id") not in managed_ids]
    final_nodes = preserved + our_nodes
    final_edges = our_edges

    canvas_path.write_text(
        json.dumps({"nodes": final_nodes, "edges": final_edges}, indent=2),
        encoding="utf-8",
    )
    logging.info(
        f"Users Canvas written: {canvas_path.name} "
        f"({len(final_nodes)} nodes, {len(final_edges)} edges, "
        f"{total_users} users, {len(host_keys_seen)} hosts)"
    )
    return canvas_path


# ============================================================
# Excel export
# ============================================================

def _xl_header_row(ws, headers: list[str]) -> None:
    """Write bold header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=False)


def _xl_severity_fill(severity_str: str) -> "PatternFill | None":
    fills = {
        "critical": PatternFill("solid", fgColor="FF4444"),
        "high":     PatternFill("solid", fgColor="FF8C00"),
        "medium":   PatternFill("solid", fgColor="FFD700"),
        "low":      PatternFill("solid", fgColor="ADD8E6"),
    }
    return fills.get((severity_str or "").lower())


def _xl_autofit(ws, max_width: int = 60) -> None:
    """Approximate column auto-fit based on max content length."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=0,
        )
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _trunc(s: str | None, n: int) -> str:
    if not s:
        return ""
    s = s.strip()
    return s[:n - 3] + "..." if len(s) > n else s


def export_excel(
    vault_dir: Path,
    all_nmap_scans: list[tuple[str, str, dict]],      # (scan_stem, scan_display, scan_data)
    all_nessus_scans: list[tuple[str, str, dict]],    # (scan_stem, scan_display, nessus_data)
    all_burp_scans: list[tuple[str, str, dict]],      # (scan_stem, scan_display, burp_data)
    all_autorecon_scans: list[tuple[str, str, dict]] | None = None,
    all_loot_data: dict | None = None,
) -> Path | None:
    if not HAS_OPENPYXL:
        logging.error(
            "openpyxl is not installed. Run: pip install openpyxl"
        )
        return None

    wb = openpyxl.Workbook()

    # ---- Summary sheet (first tab) ----
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = [
        "IP", "Hostname", "Status", "Open Ports",
        "Highest Nessus Severity", "Critical", "High", "Medium", "Low",
        "Burp Issue Count", "Sources",
    ]
    _xl_header_row(ws_summary, summary_headers)

    hosts_dir = vault_dir / "Hosts"
    host_rows: list[list] = []

    if hosts_dir.exists():
        for hp in sorted(hosts_dir.glob("*.md")):
            try:
                text = hp.read_text(encoding="utf-8")
            except Exception:
                continue
            fm, body = read_frontmatter(text)

            ip       = fm.get("ip", "")
            hostname = ", ".join(fm.get("hostnames", [])) if fm.get("hostnames") else ""
            status   = fm.get("status", "not-started")
            sources  = ", ".join(fm.get("sources", []))

            # Open ports from ## Open Ports section
            ports_section = extract_body_section(body, "## Open Ports")
            open_port_strs = [
                l.strip().lstrip("- ").split(" — ")[0].strip("**")
                for l in ports_section.splitlines()
                if l.strip().startswith("- **")
            ]
            open_ports_str = ", ".join(open_port_strs)

            # Nessus severity counts from ## Nessus Findings section
            nessus_section = extract_body_section(body, "## Nessus Findings")
            sev_counts = {4: 0, 3: 0, 2: 0, 1: 0}
            if nessus_section:
                for line in nessus_section.splitlines():
                    m = re.match(r"####\s+\[(Critical|High|Medium|Low)\]", line)
                    if m:
                        sev_counts[severity_str_to_int(m.group(1))] += 1

            max_sev_int = fm.get("nessus_max_severity", 0)
            try:
                max_sev_int = int(max_sev_int)
            except (TypeError, ValueError):
                max_sev_int = 0
            max_sev_str = severity_int_to_str(max_sev_int) if max_sev_int > 0 else ""

            # Burp issue count
            burp_section = extract_body_section(body, "## Burp Suite Findings")
            burp_count   = burp_section.count("#### [") if burp_section else 0

            host_rows.append([
                ip, hostname, status, open_ports_str,
                max_sev_str,
                sev_counts[4], sev_counts[3], sev_counts[2], sev_counts[1],
                burp_count, sources,
            ])

    for row_data in host_rows:
        ws_summary.append(row_data)
        # Color severity cell (col 5 = Highest Nessus Severity)
        row_idx = ws_summary.max_row
        sev_cell = ws_summary.cell(row=row_idx, column=5)
        fill = _xl_severity_fill(str(sev_cell.value or ""))
        if fill:
            sev_cell.fill = fill

    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions
    _xl_autofit(ws_summary)

    # ---- Nmap sheet ----
    ws_nmap = wb.create_sheet("Nmap")
    nmap_headers = [
        "IP", "Hostname", "Port", "Protocol", "Service", "Version",
        "State", "Scripts/Banner", "Source Scan", "Status",
    ]
    _xl_header_row(ws_nmap, nmap_headers)

    # Build IP→frontmatter map for status lookup
    host_fm_by_ip: dict[str, dict] = {}
    if hosts_dir.exists():
        for hp in hosts_dir.glob("*.md"):
            try:
                fm, _ = read_frontmatter(hp.read_text(encoding="utf-8"))
                if fm.get("ip"):
                    host_fm_by_ip[fm["ip"]] = fm
            except Exception:
                pass

    for scan_stem, scan_display, scan_data in all_nmap_scans:
        for host in scan_data.get("hosts", []):
            ip       = get_primary_ipv4(host) or ""
            hostname = choose_host_display_name(host)
            h_fm     = host_fm_by_ip.get(ip, {})
            status   = h_fm.get("status", "not-started")

            open_ports = host.get("open_ports", [])
            if not open_ports:
                ws_nmap.append([ip, hostname, "", "", "", "", host.get("state", ""), "", scan_display, status])
                continue

            for p in open_ports:
                svc     = p.get("service", {})
                scripts = "; ".join(
                    f"{s['id']}: {' '.join(s['output'].split())[:80]}"
                    for s in p.get("scripts", [])[:3] if s.get("output")
                )
                version = " ".join(filter(None, [
                    svc.get("product", ""), svc.get("version", ""), svc.get("extrainfo", "")
                ])).strip()
                ws_nmap.append([
                    ip, hostname,
                    p["port"], p["protocol"],
                    svc.get("name", ""), version,
                    "open", scripts,
                    scan_display, status,
                ])

    ws_nmap.freeze_panes = "A2"
    ws_nmap.auto_filter.ref = ws_nmap.dimensions
    _xl_autofit(ws_nmap)

    # ---- Nessus sheet ----
    ws_nessus = wb.create_sheet("Nessus")
    nessus_headers = [
        "IP", "Hostname", "Port", "Protocol",
        "Plugin ID", "Plugin Name", "Severity",
        "CVSS v3", "CVSS v2", "CVE(s)",
        "Description", "Solution", "Plugin Output",
        "Source Scan",
    ]
    _xl_header_row(ws_nessus, nessus_headers)

    for scan_stem, scan_display, nessus_data in all_nessus_scans:
        for host in nessus_data.get("hosts", []):
            ip       = host.get("ip", "")
            hostname = host.get("hostname", "")
            for f in host.get("findings", []):
                sev_str  = severity_int_to_str(f["severity_int"])
                cves_str = ", ".join(f["cves"]) if f["cves"] else ""
                row = [
                    ip, hostname,
                    f["port"], f["protocol"],
                    f["plugin_id"], f["plugin_name"], sev_str,
                    f.get("cvss3_base", ""), f.get("cvss_base", ""),
                    cves_str,
                    _trunc(f.get("description"), 500),
                    _trunc(f.get("solution"), 300),
                    _trunc(f.get("plugin_output"), 300),
                    scan_display,
                ]
                ws_nessus.append(row)
                row_idx  = ws_nessus.max_row
                sev_cell = ws_nessus.cell(row=row_idx, column=7)
                fill     = _xl_severity_fill(sev_str)
                if fill:
                    sev_cell.fill = fill

    ws_nessus.freeze_panes = "A2"
    ws_nessus.auto_filter.ref = ws_nessus.dimensions
    _xl_autofit(ws_nessus)

    # ---- Burp sheet ----
    ws_burp = wb.create_sheet("Burp")
    burp_headers = [
        "Host", "Path", "Issue Name", "Severity", "Confidence",
        "Location", "Issue Detail", "Remediation", "Source Scan",
    ]
    _xl_header_row(ws_burp, burp_headers)

    for scan_stem, scan_display, burp_data in all_burp_scans:
        for host in burp_data.get("hosts", []):
            host_label = host.get("url") or host.get("ip") or ""
            for issue in host.get("issues", []):
                remediation = issue.get("remediation_detail") or issue.get("remediation_background") or ""
                row = [
                    host_label,
                    issue.get("path", ""),
                    issue.get("name", ""),
                    issue.get("severity", ""),
                    issue.get("confidence", ""),
                    issue.get("location", ""),
                    _trunc(issue.get("issue_detail"), 500),
                    _trunc(remediation, 300),
                    scan_display,
                ]
                ws_burp.append(row)
                row_idx  = ws_burp.max_row
                sev_cell = ws_burp.cell(row=row_idx, column=4)
                fill     = _xl_severity_fill(issue.get("severity", ""))
                if fill:
                    sev_cell.fill = fill

    ws_burp.freeze_panes = "A2"
    ws_burp.auto_filter.ref = ws_burp.dimensions
    _xl_autofit(ws_burp)

    # ---- AutoRecon sheet ----
    if all_autorecon_scans:
        ws_ar = wb.create_sheet("AutoRecon")
        ar_headers = [
            "Target IP", "Hostname", "Port", "Protocol", "Service",
            "Tool", "Finding Type", "Finding", "Detail", "Source Scan",
        ]
        _xl_header_row(ws_ar, ar_headers)

        ar_fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        ar_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for scan_stem, scan_display, ar_data in all_autorecon_scans:
            for target in ar_data.get("targets", []):
                t_ip = target.get("ip", "")
                t_host = target.get("hostname", "")
                for port_key, results in sorted(target.get("tool_results", {}).items()):
                    parts = port_key.split("/", 1)
                    proto = parts[0] if parts else ""
                    port = parts[1] if len(parts) > 1 else ""
                    service = ""
                    for r in results:
                        fn_m = AUTORECON_FILENAME_RE.match(r.get("filename", ""))
                        if fn_m:
                            service = fn_m.group(3)
                            break

                    for entry in results:
                        data = entry.get("data", {})
                        tool = data.get("tool", entry.get("tool", "unknown"))

                        rows_to_add: list[tuple[str, str, str, str | None]] = []

                        if tool == "dirbusting":
                            for p in data.get("interesting", [])[:30]:
                                detail = f"Status {p['status']}, Size {p['size']}"
                                fill_type = None
                                if p["status"] == 200 and re.search(
                                    r"admin|backup|config|upload|api|secret|private|\.bak|\.old",
                                    p["path"], re.IGNORECASE,
                                ):
                                    fill_type = "yellow"
                                rows_to_add.append(("Discovered Path", p["path"], detail, fill_type))
                        elif tool == "nikto":
                            for f in data.get("findings", [])[:20]:
                                osvdb = f"OSVDB-{f['osvdb']} " if f.get("osvdb") else ""
                                rows_to_add.append(("Nikto Finding", f["path"], f"{osvdb}{f['description'][:200]}", None))
                        elif tool in ("enum4linux", "smbmap", "smbclient"):
                            for s in data.get("shares", []):
                                access = s.get("permissions", s.get("access", s.get("type", "")))
                                ft = "orange" if "WRITE" in str(access).upper() else None
                                rows_to_add.append(("SMB Share", s["name"], access, ft))
                            for u in data.get("users", []):
                                rows_to_add.append(("User Account", u, "", "yellow"))
                        elif tool == "whatweb":
                            for t in data.get("technologies", []):
                                ver = f"/{t['version']}" if t.get("version") else ""
                                rows_to_add.append(("Technology", f"{t['name']}{ver}", "", None))
                        elif tool == "sslscan":
                            for wp in data.get("weak_protocols", []):
                                rows_to_add.append(("TLS Issue", f"Weak protocol: {wp}", "", "yellow"))
                            for wc in data.get("weak_ciphers", []):
                                rows_to_add.append(("TLS Issue", f"Weak cipher: {wc}", "", "yellow"))
                            if data.get("cert_expired"):
                                rows_to_add.append(("TLS Issue", "Certificate expired", data.get("cert_not_after", ""), "yellow"))
                        elif tool == "dnsrecon":
                            for r in data.get("records", [])[:20]:
                                rows_to_add.append(("DNS Record", f"{r['type']} {r['name']}", r["value"], None))
                            if data.get("zone_transfer_successful"):
                                rows_to_add.append(("DNS Record", "Zone Transfer", "SUCCESSFUL", "orange"))
                        elif tool == "snmpwalk":
                            if data.get("sys_descr"):
                                rows_to_add.append(("SNMP Info", "sysDescr", data["sys_descr"], None))
                        elif tool == "onesixtyone":
                            for cs in data.get("community_strings", []):
                                rows_to_add.append(("SNMP Community", cs["community"], cs.get("sys_descr", "")[:100], "yellow"))

                        for finding_type, finding, detail, fill_kind in rows_to_add:
                            ws_ar.append([
                                t_ip, t_host, port, proto, service,
                                tool, finding_type, finding, detail, scan_display,
                            ])
                            if fill_kind:
                                row_idx = ws_ar.max_row
                                fill = ar_fill_orange if fill_kind == "orange" else ar_fill_yellow
                                ws_ar.cell(row=row_idx, column=7).fill = fill

        ws_ar.freeze_panes = "A2"
        ws_ar.auto_filter.ref = ws_ar.dimensions
        _xl_autofit(ws_ar)

    # ---- Loot sheet ----
    if all_loot_data and all_loot_data.get("summary", {}).get("total_files", 0):
        ws_loot = wb.create_sheet("Loot")
        loot_headers = [
            "Host IP", "Hostname", "Source File", "Category",
            "Username", "Password/Hash", "Credential Type",
            "Hash Type", "Detail",
        ]
        _xl_header_row(ws_loot, loot_headers)

        loot_fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        loot_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for host_key, loot_files in all_loot_data.get("host_loot", {}).items():
            t_ip = host_key if IPV4_RE.match(host_key) else ""
            t_host = "" if t_ip else host_key
            for lf in loot_files:
                for c in lf.get("credentials", []):
                    ws_loot.append([
                        t_ip, t_host, lf["filename"], lf["category"],
                        c["username"], c["password"], c["cred_type"],
                        "", "",
                    ])
                    row_idx = ws_loot.max_row
                    if c["cred_type"] == "cleartext":
                        ws_loot.cell(row=row_idx, column=7).fill = loot_fill_orange
                for h in lf.get("hashes", []):
                    ws_loot.append([
                        t_ip, t_host, lf["filename"], lf["category"],
                        h.get("username", ""), h["hash"][:40], "",
                        h["hash_type"], h.get("context_line", "")[:100],
                    ])
                    row_idx = ws_loot.max_row
                    ws_loot.cell(row=row_idx, column=8).fill = loot_fill_yellow

        for lf in all_loot_data.get("campaign_loot", []):
            for c in lf.get("credentials", []):
                ws_loot.append([
                    "", "", lf["filename"], lf["category"],
                    c["username"], c["password"], c["cred_type"], "", "",
                ])
                row_idx = ws_loot.max_row
                if c["cred_type"] == "cleartext":
                    ws_loot.cell(row=row_idx, column=7).fill = loot_fill_orange

        ws_loot.freeze_panes = "A2"
        ws_loot.auto_filter.ref = ws_loot.dimensions
        _xl_autofit(ws_loot)

    out_path = vault_dir / "Export.xlsx"
    wb.save(str(out_path))
    logging.info(f"Excel export written: {out_path}")
    return out_path


# ============================================================
_SCAN_SUBDIRS = ["nmap", "nessus", "burp", "nikto", "autorecon", "loot", "misc"]

STATE_FILENAME = ".maipper_state.json"


def _load_analysis_state(vault_dir: Path) -> dict[str, float]:
    """Load the analysis state file. Returns {filepath: mtime} for previously analyzed files."""
    state_path = vault_dir / STATE_FILENAME
    if not state_path.exists():
        return {}
    try:
        return json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_analysis_state(vault_dir: Path, state: dict[str, float]) -> None:
    """Save the analysis state file."""
    state_path = vault_dir / STATE_FILENAME
    state_path.write_text(json.dumps(state, indent=2), encoding="utf-8")


def _file_needs_analysis(filepath: Path, state: dict[str, float]) -> bool:
    """Check if a file needs (re)analysis based on its mtime."""
    key = str(filepath)
    if key not in state:
        return True
    try:
        return filepath.stat().st_mtime != state[key]
    except OSError:
        return True


def _mark_analyzed(filepath: Path, state: dict[str, float]) -> None:
    """Mark a file as analyzed in the state dict."""
    try:
        state[str(filepath)] = filepath.stat().st_mtime
    except OSError:
        pass

_SCAN_EXTENSIONS = {
    "nmap": ["*.xml"],
    "nessus": ["*.nessus"],
    "burp": ["*.xml"],
}


def _snapshot_vault_files(vault_dir: Path) -> dict[str, float]:
    """Snapshot modification times of vault files we watch for operator edits."""
    snapshot: dict[str, float] = {}
    hosts_dir = vault_dir / "Hosts"
    if hosts_dir.exists():
        for hp in hosts_dir.glob("*.md"):
            if hp.stem == "_Campaign Targets":
                continue
            try:
                snapshot[str(hp)] = hp.stat().st_mtime
            except OSError:
                pass

    eat_me = vault_dir / "Eat Me.md"
    if eat_me.exists():
        try:
            snapshot[str(eat_me)] = eat_me.stat().st_mtime
        except OSError:
            pass

    creds = vault_dir / "Loot" / "Credentials.md"
    if creds.exists():
        try:
            snapshot[str(creds)] = creds.stat().st_mtime
        except OSError:
            pass

    return snapshot


def _snapshot_scan_files(base: Path, args) -> dict[str, float]:
    """Collect all scan files under base and return {path: mtime} map."""
    files: dict[str, float] = {}

    if getattr(args, "xml", None):
        p = Path(args.xml).resolve()
        if p.exists():
            files[str(p)] = p.stat().st_mtime
        return files

    for subdir, globs in _SCAN_EXTENSIONS.items():
        skip_flag = f"no_{subdir}"
        if getattr(args, skip_flag, False):
            continue
        d = base / subdir
        if not d.exists():
            d = base / subdir.capitalize()
        if not d.exists():
            continue
        for g in globs:
            for f in d.glob(g):
                files[str(f)] = f.stat().st_mtime

    if not getattr(args, "no_autorecon", False):
        ar = Path(args.autorecon).resolve() if getattr(args, "autorecon", None) else None
        if ar is None:
            for name in ["autorecon", "AutoRecon"]:
                if (base / name).exists():
                    ar = base / name
                    break
        if ar and ar.exists():
            for f in ar.rglob("*"):
                if f.is_file():
                    files[str(f)] = f.stat().st_mtime

    if not getattr(args, "no_loot", False):
        ld = Path(args.loot).resolve() if getattr(args, "loot", None) else None
        if ld is None:
            for name in ["loot", "Loot"]:
                if (base / name).exists():
                    ld = base / name
                    break
        if ld and ld.exists():
            for f in ld.rglob("*"):
                if f.is_file():
                    files[str(f)] = f.stat().st_mtime

    if not getattr(args, "no_misc", False):
        md = Path(args.misc).resolve() if getattr(args, "misc", None) else None
        if md is None:
            for name in ["misc", "Misc"]:
                if (base / name).exists():
                    md = base / name
                    break
        if md and md.exists():
            for f in md.rglob("*"):
                if f.is_file():
                    files[str(f)] = f.stat().st_mtime

    return files


def _init_scan_dirs(base: Path) -> None:
    """Create the scans/ directory tree with empty subdirectories for each parser."""
    base.mkdir(parents=True, exist_ok=True)
    for sub in _SCAN_SUBDIRS:
        (base / sub).mkdir(exist_ok=True)

    print(f"\n[+] Initialized scan directory structure at: {base}\n")
    print("    Drop your scan files into the appropriate folders:\n")
    print(f"      {base / 'nmap/'}          — Nmap XML output  (nmap -oX)")
    print(f"      {base / 'nessus/'}        — Nessus .nessus exports")
    print(f"      {base / 'burp/'}          — Burp Suite Issues XML exports")
    print(f"      {base / 'nikto/'}         — Nikto scan output (text/XML)")
    print(f"      {base / 'autorecon/'}     — AutoRecon results (per-target subdirs)")
    print(f"      {base / 'loot/'}          — Loot (credentials, hashes, keys, sensitive files)")
    print(f"      {base / 'misc/'}          — Miscellaneous tool output (text files)")
    print("\n    Then run mAIpper again to process.\n")


def _build_assessment_context(vault_dir: Path) -> str:
    """Build a condensed summary of the current assessment for LLM context."""
    hosts_dir = vault_dir / "Hosts"
    if not hosts_dir.exists():
        return "No assessment data yet. Run mAIpper or drop scan files into the scans directory."

    host_notes = sorted(hosts_dir.glob("*.md"))
    if not host_notes:
        return "No hosts discovered yet."

    lines: list[str] = [f"ASSESSMENT STATE ({len(host_notes)} hosts discovered)", ""]

    for hp in host_notes:
        try:
            text = hp.read_text(encoding="utf-8")
            fm, body = read_frontmatter(text)
        except Exception:
            continue

        ip = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        status = fm.get("status", "not-started")
        tags = fm.get("tags", [])
        nessus_sev = fm.get("nessus_max_severity", 0)
        loot_creds = fm.get("loot_credential_count", 0)
        loot_hashes = fm.get("loot_hash_count", 0)

        host_label = ip or hp.stem
        if hostnames:
            host_label += f" ({', '.join(hostnames[:2])})"

        parts = [f"Host: {host_label} | Status: {status}"]
        if tags:
            parts.append(f"  Tags: {', '.join(tags[:10])}")

        ports_section = extract_body_section(body, "## Open Ports")
        if ports_section:
            port_lines = [l.strip() for l in ports_section.splitlines()
                          if l.strip().startswith("- **")]
            if port_lines:
                port_summary = ", ".join(
                    l.split("**")[1] if "**" in l else l[:20] for l in port_lines[:8]
                )
                parts.append(f"  Open ports: {port_summary}")

        if nessus_sev:
            sev_labels = {4: "Critical", 3: "High", 2: "Medium", 1: "Low"}
            parts.append(f"  Nessus max severity: {nessus_sev} ({sev_labels.get(int(nessus_sev), '?')})")

        if loot_creds or loot_hashes:
            loot_parts = []
            if loot_creds:
                loot_parts.append(f"{loot_creds} credentials")
            if loot_hashes:
                loot_parts.append(f"{loot_hashes} hashes")
            parts.append(f"  Loot: {', '.join(loot_parts)}")

        op_notes = extract_operator_notes(body)
        if op_notes:
            preview = op_notes.replace("\n", " ").strip()[:150]
            parts.append(f"  Operator notes: {preview}")

        lines.append("\n".join(parts))
        lines.append("")

    return "\n".join(lines)


def _build_chat_prompt(question: str, context: str, chat_history: list[dict]) -> str:
    """Build a prompt for interactive chat with assessment context."""
    history_text = ""
    if chat_history:
        history_lines = []
        for entry in chat_history[-6:]:
            history_lines.append(f"Operator: {entry['q']}")
            history_lines.append(f"Assistant: {entry['a'][:300]}...")
            history_lines.append("")
        history_text = "\nRECENT CONVERSATION:\n" + "\n".join(history_lines)

    return f"""You are an experienced penetration tester acting as an interactive assistant
during a live engagement. You have access to the current assessment state below.

GROUNDING RULES — follow these strictly:
- Respond ONLY in English. Do not use any other language.
- NEVER fabricate, invent, or hallucinate data. If the provided data is insufficient, say so explicitly rather than making up examples.
- Only reference hosts, ports, services, credentials, and findings that appear
  in the assessment state below. Do not invent data.
- Tag claims: [CONFIRMED] = in the data, [INFERRED] = reasonable conclusion,
  [ASSUMED] = needs verification.
- When suggesting commands, use actual IPs from the assessment data.
- Be concise and operator-focused. This is a live engagement, not a report.

{context}
{history_text}
OPERATOR QUESTION: {question}

Respond concisely with actionable information."""


def _handle_cred_drop(text: str, vault_dir: Path) -> str:
    """Parse a credential drop and save to loot files."""
    text = text.strip()
    if text.startswith("+cred"):
        text = text[5:].strip()

    parts = text.split()
    host_key = None
    cred_part = parts[0] if parts else ""

    if len(parts) >= 2 and (IPV4_RE.match(parts[-1]) or is_probable_fqdn(parts[-1])):
        host_key = parts[-1]
        cred_part = " ".join(parts[:-1])

    if ":" not in cred_part:
        return f"Could not parse credential. Use format: +cred user:pass [host_ip]"

    username, _, password = cred_part.partition(":")

    loot_dir = vault_dir.parent / "scans" / "loot"
    if not loot_dir.exists():
        loot_dir.mkdir(parents=True, exist_ok=True)

    if host_key:
        host_loot_dir = loot_dir / host_key
        host_loot_dir.mkdir(exist_ok=True)
        target = host_loot_dir / "interactive_creds.txt"
    else:
        target = loot_dir / "interactive_creds.txt"

    with open(target, "a", encoding="utf-8") as f:
        f.write(f"{username}:{password}\n")

    host_msg = f" for {host_key}" if host_key else ""
    return f"[+] Saved credential {username}:{password}{host_msg} → {target.name}"


def _handle_paste(vault_dir: Path) -> str:
    """Read multiline paste input until empty line, detect content type, save."""
    print("  Paste your data, then press Enter twice to submit:")
    lines: list[str] = []
    while True:
        try:
            line = input()
        except EOFError:
            break
        if not line.strip() and lines:
            break
        lines.append(line)

    if not lines:
        return "No data received."

    text = "\n".join(lines)

    # Detect IPs
    ips_found = list(set(IP_IN_TEXT_RE.findall(text)))

    # Detect credentials
    creds_found = _extract_credentials(text)

    results: list[str] = []

    if ips_found:
        loot_dir = vault_dir.parent / "scans" / "loot"
        loot_dir.mkdir(parents=True, exist_ok=True)
        target = loot_dir / "interactive_hosts.txt"
        with open(target, "a", encoding="utf-8") as f:
            for ip in sorted(ips_found):
                f.write(f"{ip}\n")
        results.append(f"[+] Saved {len(ips_found)} IP(s) → {target.name}")

    if creds_found:
        loot_dir = vault_dir.parent / "scans" / "loot"
        loot_dir.mkdir(parents=True, exist_ok=True)
        target = loot_dir / "interactive_creds.txt"
        with open(target, "a", encoding="utf-8") as f:
            for c in creds_found:
                f.write(f"{c['username']}:{c['password']}\n")
        results.append(f"[+] Saved {len(creds_found)} credential(s) → {target.name}")

    if not results:
        misc_dir = vault_dir.parent / "scans" / "misc"
        misc_dir.mkdir(parents=True, exist_ok=True)
        target = misc_dir / "interactive_paste.txt"
        with open(target, "a", encoding="utf-8") as f:
            f.write(text + "\n\n")
        results.append(f"[+] Saved paste data → {target.name} (will be analyzed on next run)")

    return "\n".join(results)


def _interactive_hosts(vault_dir: Path) -> str:
    """List all discovered hosts."""
    hosts_dir = vault_dir / "Hosts"
    if not hosts_dir.exists():
        return "No hosts discovered yet."
    lines: list[str] = []
    for hp in sorted(hosts_dir.glob("*.md")):
        try:
            fm, _ = read_frontmatter(hp.read_text(encoding="utf-8"))
        except Exception:
            continue
        ip = fm.get("ip", "")
        hostnames = fm.get("hostnames", [])
        status = fm.get("status", "not-started")
        label = ip or hp.stem
        if hostnames:
            label += f" ({', '.join(hostnames[:2])})"
        lines.append(f"  [{status}] {label}")
    return f"Hosts ({len(lines)}):\n" + "\n".join(lines) if lines else "No hosts discovered."


def _interactive_status(vault_dir: Path) -> str:
    """Show assessment status summary."""
    hosts_dir = vault_dir / "Hosts"
    if not hosts_dir.exists():
        return "No assessment data yet."
    counts = {"not-started": 0, "in-progress": 0, "done": 0, "exploited": 0, "blocked": 0}
    total_ports = 0
    total_creds = 0
    total_hashes = 0
    for hp in hosts_dir.glob("*.md"):
        try:
            fm, body = read_frontmatter(hp.read_text(encoding="utf-8"))
        except Exception:
            continue
        status = fm.get("status", "not-started")
        counts[status] = counts.get(status, 0) + 1
        ports_section = extract_body_section(body, "## Open Ports")
        total_ports += sum(1 for l in ports_section.splitlines() if l.strip().startswith("- **"))
        try:
            total_creds += int(fm.get("loot_credential_count", 0))
            total_hashes += int(fm.get("loot_hash_count", 0))
        except (TypeError, ValueError):
            pass
    total = sum(counts.values())
    lines = [
        f"Assessment Status:",
        f"  Hosts: {total} total",
    ]
    for status, count in counts.items():
        if count:
            lines.append(f"    {status}: {count}")
    lines.append(f"  Open ports: {total_ports}")
    if total_creds:
        lines.append(f"  Credentials: {total_creds}")
    if total_hashes:
        lines.append(f"  Hashes: {total_hashes}")
    return "\n".join(lines)


def _process_vault_changes(
    vault_dir: Path,
    changed_files: set[str],
    args,
    operator_notes_lookup: dict[str, str],
    all_loot_data: dict | None,
) -> dict[str, str]:
    """React to vault file changes. Returns updated operator_notes_lookup."""
    eat_me_path = str(vault_dir / "Eat Me.md")
    creds_path = str(vault_dir / "Loot" / "Credentials.md")
    hosts_dir = vault_dir / "Hosts"
    any_action = False

    # ── Eat Me ──
    if eat_me_path in changed_files:
        em_result = _process_eat_me_page(vault_dir, args=args)
        if em_result:
            any_action = True
            if em_result["new_hosts"]:
                print(f"  [Vault] Eat Me: created {len(em_result['new_hosts'])} new host(s)")
                for ip in em_result["new_hosts"]:
                    print(f"      → {ip}")
            if em_result["discovered_hostnames"]:
                print(f"  [Vault] Eat Me: found {len(em_result['discovered_hostnames'])} hostname(s)")
            if em_result["discovered_creds"]:
                print(f"  [Vault] Eat Me: found {len(em_result['discovered_creds'])} credential(s)")
            print(f"  [Vault] Eat Me: saved to Scans/{em_result['scan_stem']}.md")

    # ── Host notes: check for operator notes + investigation boxes ──
    host_files_changed = [
        f for f in changed_files
        if f.startswith(str(hosts_dir)) and f.endswith(".md")
    ]
    if host_files_changed:
        # Reload operator notes (they may have changed)
        operator_notes_lookup = build_operator_notes_lookup(vault_dir)
        changed_names = [Path(f).stem for f in host_files_changed]
        print(f"  [Vault] Host notes changed: {', '.join(changed_names[:5])}"
              + (f" +{len(changed_names) - 5} more" if len(changed_names) > 5 else ""))

        # Count pending investigation checkboxes (don't auto-run — user runs /analyze)
        pending = 0
        if hosts_dir.exists():
            for hp in hosts_dir.glob("*.md"):
                pending += len(_scan_host_note_for_deep_dives(hp))
        if pending:
            print(f"  [Vault] {pending} pending investigation(s) — run /analyze to process")

    # ── Credentials.md: rebuild Users Canvas ──
    if creds_path in changed_files:
        any_action = True
        print("  [Vault] Credentials.md changed")
        if not args.no_canvas and not args.no_users_canvas:
            build_users_canvas(vault_dir, all_loot_data)
            print("  [Vault] Users Canvas rebuilt")

    # ── Update Campaign Targets if anything changed ──
    if any_action or host_files_changed:
        _write_campaign_targets_note(vault_dir)

    return operator_notes_lookup


def _watch_loop(args, base: Path) -> None:
    """Interactive watch loop — polls for changes and accepts operator input."""
    global _RAG_INDEX, _RAG_BUILDER
    interval = args.watch_interval
    last_scan_snapshot: dict[str, float] = {}
    last_vault_snapshot: dict[str, float] = {}
    vault_dir = Path(args.vault).resolve()
    chat_history: list[dict] = []
    all_loot_data: dict | None = None

    print(f"\n[*] mAIpper Interactive Mode  (Ctrl+C to exit)")
    print(f"    Scans dir : {base}")
    print(f"    Vault dir : {vault_dir}")
    print(f"    Polling   : every {interval}s")
    print(f"    Watching  : scans/ + vault files")
    print()
    print("    Commands:")
    print("      /hosts      — list discovered hosts")
    print("      /status     — assessment summary")
    print("      /analyze    — analyze checked [x] items (investigation + scan boxes)")
    print("      /analyze-full — check all pending boxes then analyze (skips done [/] items)")
    print("      /deepdive   — cross-source correlation per host (Nmap+Nessus+Burp+AutoRecon+Loot)")
    print("      /merge      — detect and merge duplicate host notes (same IP or hostname)")
    print("      /refresh    — re-process all scan files (no analysis)")
    print("      /rag        — show RAG index status")
    print("      /build-index — build or update RAG index")
    print("      /paste      — paste multiline data (hosts, creds, tool output)")
    print("      +cred u:p   — add credential (+cred user:pass [host_ip])")
    print("      <question>  — ask the LLM about the assessment")
    print()

    if not base.exists():
        _init_scan_dirs(base)

    # Load operator notes for deep dive context
    operator_notes_lookup: dict[str, str] = {}
    if vault_dir.exists():
        operator_notes_lookup = build_operator_notes_lookup(vault_dir)

    # Initial processing run (no LLM — use /analyze or /deepdive for analysis)
    current = _snapshot_scan_files(base, args)
    if current:
        logging.info("[Watch] Initial processing (no analysis)...")
        try:
            _run_processing(args, base, skip_llm=True)
        except KeyboardInterrupt:
            print("\n  [!] Processing interrupted — returning to prompt")
        except Exception as exc:
            logging.error(f"[Watch] Processing failed: {exc}")
        last_scan_snapshot = _snapshot_scan_files(base, args)
    else:
        last_scan_snapshot = current

    # Take initial vault snapshot AFTER processing (so we don't re-trigger on our own writes)
    last_vault_snapshot = _snapshot_vault_files(vault_dir)

    try:
        while True:
            # ── Check for scan file changes ──
            current_scans = _snapshot_scan_files(base, args)
            if current_scans != last_scan_snapshot:
                new_files = set(current_scans) - set(last_scan_snapshot)
                modified = {
                    f for f in set(current_scans) & set(last_scan_snapshot)
                    if current_scans[f] != last_scan_snapshot[f]
                }
                if new_files or modified:
                    for f in sorted(new_files):
                        print(f"  [Watch] New: {Path(f).name}")
                    for f in sorted(modified):
                        print(f"  [Watch] Modified: {Path(f).name}")
                    print("  [Watch] Processing scans (no analysis)...")
                    try:
                        _run_processing(args, base, skip_llm=True)
                        operator_notes_lookup = build_operator_notes_lookup(vault_dir)
                    except KeyboardInterrupt:
                        print("\n  [!] Processing interrupted — returning to prompt")
                    except Exception as exc:
                        logging.error(f"[Watch] Processing failed: {exc}")
                last_scan_snapshot = _snapshot_scan_files(base, args)
                last_vault_snapshot = _snapshot_vault_files(vault_dir)

            # ── Check for vault file changes ──
            current_vault = _snapshot_vault_files(vault_dir)
            if current_vault != last_vault_snapshot:
                changed_vault = set()
                for f in set(current_vault) | set(last_vault_snapshot):
                    old_mtime = last_vault_snapshot.get(f)
                    new_mtime = current_vault.get(f)
                    if old_mtime != new_mtime:
                        changed_vault.add(f)

                if changed_vault:
                    operator_notes_lookup = _process_vault_changes(
                        vault_dir, changed_vault, args,
                        operator_notes_lookup, all_loot_data,
                    )

                last_vault_snapshot = _snapshot_vault_files(vault_dir)

            # Check if background RAG build completed
            if _RAG_BUILDER is not None:
                idx = _RAG_BUILDER.get_index_nowait()
                if idx:
                    _RAG_INDEX = idx
                    print(f"  [+] RAG index updated: {idx['chunk_count']} chunks")
                    _RAG_BUILDER = None

            # Interactive prompt
            try:
                print("\033[90m  Enter to check · type a question · /help for commands\033[0m")
                user_input = input("mAIpper> ").strip()
            except EOFError:
                break

            if not user_input:
                # On empty Enter: check vault changes (no LLM), report pending items
                any_change = False
                current_vault = _snapshot_vault_files(vault_dir)
                if current_vault != last_vault_snapshot:
                    changed_vault = set()
                    for f in set(current_vault) | set(last_vault_snapshot):
                        old_mtime = last_vault_snapshot.get(f)
                        new_mtime = current_vault.get(f)
                        if old_mtime != new_mtime:
                            changed_vault.add(f)

                    if changed_vault:
                        any_change = True
                        operator_notes_lookup = _process_vault_changes(
                            vault_dir, changed_vault, args,
                            operator_notes_lookup, all_loot_data,
                        )
                    last_vault_snapshot = _snapshot_vault_files(vault_dir)

                # Count pending items across host notes and scan notes
                pending_investigate = 0
                hosts_dir = vault_dir / "Hosts"
                if hosts_dir.exists():
                    for hp in hosts_dir.glob("*.md"):
                        pending_investigate += len(_scan_host_note_for_deep_dives(hp))
                pending_analyze = len(_scan_notes_for_analysis_requests(vault_dir))

                if pending_investigate or pending_analyze:
                    parts = []
                    if pending_investigate:
                        parts.append(f"{pending_investigate} investigation(s)")
                    if pending_analyze:
                        parts.append(f"{pending_analyze} scan analysis request(s)")
                    print(f"  [*] Pending: {', '.join(parts)}")
                    if args.no_ollama:
                        print("  LLM disabled — run /analyze manually after enabling Ollama.")
                    else:
                        try:
                            answer = input("  Analyze now? [Y/n]: ").strip().lower()
                        except (EOFError, KeyboardInterrupt):
                            answer = "n"
                        if answer in ("", "y", "yes"):
                            operator_notes_lookup = build_operator_notes_lookup(vault_dir)
                            dd_count = 0
                            sa_count = 0
                            try:
                                dd_count = _process_deep_dives(
                                    vault_dir, args.ollama_url, args.model,
                                    args.temperature, args.skip_validation,
                                    operator_notes_lookup,
                                )
                                if dd_count:
                                    print(f"  [+] Processed {dd_count} deep dive(s)")
                                sa_count = _process_analyze_requests(
                                    vault_dir, args, operator_notes_lookup,
                                )
                                if sa_count:
                                    print(f"  [+] Processed {sa_count} scan analysis request(s)")
                            except KeyboardInterrupt:
                                print("\n  [!] Analysis interrupted")
                            last_vault_snapshot = _snapshot_vault_files(vault_dir)
                elif not any_change:
                    print("  No changes detected.")
                continue

            # Commands
            if user_input.lower() in ("/quit", "/exit", "/q"):
                break
            elif user_input.lower() in ("/help", "/?"):
                print("  /hosts      — list discovered hosts")
                print("  /status     — assessment summary")
                print("  /analyze    — analyze checked [x] items (investigation + scan boxes)")
                print("  /analyze-full — check all pending boxes then analyze (skips done [/] items)")
                print("  /deepdive   — cross-source correlation per host (Nmap+Nessus+Burp+AutoRecon+Loot)")
                print("  /merge      — detect and merge duplicate host notes (same IP or hostname)")
                print("  /refresh    — re-process all scan files (no analysis)")
                print("  /rag        — show RAG index status")
                print("  /build-index — build or update RAG index")
                print("  /paste      — paste multiline data (hosts, creds, tool output)")
                print("  +cred u:p   — add credential (+cred user:pass [host_ip])")
                print("  Ctrl+C      — cancel current operation / exit at prompt")
                print("  <question>  — ask the LLM about the assessment")
            elif user_input.lower() == "/hosts":
                print(_interactive_hosts(vault_dir))
            elif user_input.lower() == "/status":
                print(_interactive_status(vault_dir))
            elif user_input.lower() == "/analyze":
                if args.no_ollama:
                    print("  LLM is disabled (--no-ollama). Cannot analyze.")
                else:
                    print("  Scanning for checked analysis/investigation boxes...")
                    operator_notes_lookup = build_operator_notes_lookup(vault_dir)
                    dd_count = 0
                    sa_count = 0
                    try:
                        dd_count = _process_deep_dives(
                            vault_dir, args.ollama_url, args.model,
                            args.temperature, args.skip_validation,
                            operator_notes_lookup,
                        )
                        if dd_count:
                            print(f"  [+] Processed {dd_count} deep dive(s)")
                        sa_count = _process_analyze_requests(
                            vault_dir, args, operator_notes_lookup,
                        )
                        if sa_count:
                            print(f"  [+] Processed {sa_count} scan analysis request(s)")
                    except KeyboardInterrupt:
                        print("\n  [!] Analysis interrupted")
                    if not dd_count and not sa_count:
                        print("  No checked boxes found. Check [x] items in Obsidian, then try again.")
                    last_vault_snapshot = _snapshot_vault_files(vault_dir)
            elif user_input.lower() == "/analyze-full":
                if args.no_ollama:
                    print("  LLM is disabled (--no-ollama). Cannot analyze.")
                else:
                    inv_n, ana_n = _check_all_pending_boxes(vault_dir)
                    total_checked = inv_n + ana_n
                    if total_checked == 0:
                        print("  No unchecked boxes found — everything already analyzed or no scan notes exist.")
                    else:
                        print(f"  Checked {inv_n} investigation box(es) and {ana_n} scan analysis box(es). Running analysis...")
                        operator_notes_lookup = build_operator_notes_lookup(vault_dir)
                        dd_count = 0
                        sa_count = 0
                        try:
                            dd_count = _process_deep_dives(
                                vault_dir, args.ollama_url, args.model,
                                args.temperature, args.skip_validation,
                                operator_notes_lookup,
                            )
                            if dd_count:
                                print(f"  [+] Processed {dd_count} deep dive(s)")
                            sa_count = _process_analyze_requests(
                                vault_dir, args, operator_notes_lookup,
                            )
                            if sa_count:
                                print(f"  [+] Processed {sa_count} scan analysis request(s)")
                        except KeyboardInterrupt:
                            print("\n  [!] Analysis interrupted")
                    last_vault_snapshot = _snapshot_vault_files(vault_dir)
            elif user_input.lower() == "/merge":
                reports = _detect_and_merge_host_notes(vault_dir)
                if reports:
                    for msg in reports:
                        print(f"  [+] {msg}")
                    operator_notes_lookup = build_operator_notes_lookup(vault_dir)
                    last_vault_snapshot = _snapshot_vault_files(vault_dir)
                else:
                    print("  No duplicate host notes found.")
            elif user_input.lower() == "/deepdive":
                if args.no_ollama:
                    print("  LLM is disabled (--no-ollama). Cannot run deep dive.")
                else:
                    print("  Running cross-source correlation for all hosts...")
                    operator_notes_lookup = build_operator_notes_lookup(vault_dir)
                    try:
                        count = _run_cross_source_deepdive(vault_dir, args, operator_notes_lookup)
                        if count:
                            print(f"  [+] Cross-source analysis written for {count} host(s).")
                        else:
                            print("  No host notes with scan data found.")
                    except KeyboardInterrupt:
                        print("\n  [!] Deep dive interrupted")
                    except Exception as exc:
                        logging.error(f"Deep dive failed: {exc}")
                    last_vault_snapshot = _snapshot_vault_files(vault_dir)
            elif user_input.lower() == "/rag":
                if _RAG_BUILDER and not _RAG_BUILDER.is_ready():
                    print(f"  RAG: {_RAG_BUILDER.status_line()}")
                    if _RAG_INDEX:
                        print(f"  Current index: {_RAG_INDEX['chunk_count']} chunks (updating)")
                elif _RAG_INDEX:
                    print(f"  RAG: {_RAG_INDEX['chunk_count']} chunks indexed")
                    print(f"  Model: {_RAG_INDEX.get('embedding_model', '?')}")
                    print(f"  Built: {_RAG_INDEX.get('built_at', '?')}")
                    db_size = Path(_RAG_INDEX["db_path"]).stat().st_size / (1024 * 1024)
                    print(f"  Size: {db_size:.1f} MB")
                    try:
                        conn = _rag_db_connect(_RAG_INDEX["db_path"])
                        sources = [r[0] for r in conn.execute(
                            "SELECT DISTINCT source FROM chunks"
                        ).fetchall()]
                        conn.close()
                        top_sources = sorted(set(s.split("/")[0] for s in sources))[:10]
                        print(f"  Sources: {', '.join(top_sources)}")
                    except Exception:
                        pass
                else:
                    print("  RAG: no index. Run /build-index to create one.")
            elif user_input.lower() == "/build-index":
                if args.no_ollama:
                    print("  LLM is disabled (--no-ollama). Cannot build index.")
                else:
                    docs_dir = Path(args.rag_docs_dir).resolve() if args.rag_docs_dir else None
                    ht_dir = Path(args.rag_hacktricks_dir).resolve() if args.rag_hacktricks_dir else None
                    if not docs_dir and not ht_dir:
                        print("  No RAG source directories configured.")
                        print("  Set docs_dir or hacktricks_dir in [rag] section of maipper.conf.")
                    else:
                        source_count = 0
                        if docs_dir and docs_dir.exists():
                            source_count += len(list(docs_dir.rglob("*.pdf")))
                        if ht_dir and ht_dir.exists():
                            source_count += len(list(ht_dir.rglob("*.md")))
                        if source_count == 0:
                            print("  No source files found. Add PDFs to docs/ or clone HackTricks.")
                        else:
                            print(f"  Building RAG index ({source_count} source files)...")
                            try:
                                meta = _build_rag_index(
                                    docs_dir if docs_dir and docs_dir.exists() else None,
                                    ht_dir if ht_dir and ht_dir.exists() else None,
                                    args.ollama_url, args.rag_embedding_model,
                                    db_path=_rag_db_path(),
                                )
                                _RAG_INDEX = meta
                                print(f"  [+] RAG index built: {meta['chunk_count']} chunks")
                            except KeyboardInterrupt:
                                _RAG_INDEX = _load_rag_index()
                                print("\n  [!] Index build cancelled (progress saved)")
                            except Exception as exc:
                                print(f"  [!] Index build failed: {exc}")
            elif user_input.lower() == "/refresh":
                print("  Re-processing all scan files (no analysis)...")
                args.reanalyze = True
                try:
                    _run_processing(args, base, skip_llm=True)
                except KeyboardInterrupt:
                    print("\n  [!] Refresh interrupted")
                except Exception as exc:
                    logging.error(f"Refresh failed: {exc}")
                args.reanalyze = False
                last_scan_snapshot = _snapshot_scan_files(base, args)
                last_vault_snapshot = _snapshot_vault_files(vault_dir)
                print("  Done.")
            elif user_input.lower() == "/paste":
                result = _handle_paste(vault_dir)
                print(result)
            elif user_input.lower().startswith("+cred"):
                result = _handle_cred_drop(user_input, vault_dir)
                print(result)

            # LLM question
            elif not args.no_ollama:
                context = _build_assessment_context(vault_dir)
                prompt = _build_chat_prompt(user_input, context, chat_history)
                try:
                    response = ollama_chat(
                        args.ollama_url, args.model, prompt, args.temperature,
                    )
                    print()
                    print(response.strip())
                    chat_history.append({"q": user_input, "a": response.strip()})
                    if len(chat_history) > 20:
                        chat_history = chat_history[-20:]
                except KeyboardInterrupt:
                    print("\n  [!] LLM query cancelled")
                except Exception as exc:
                    print(f"  LLM query failed: {exc}")
            else:
                print("  LLM is disabled (--no-ollama). Use commands or enable Ollama.")

    except KeyboardInterrupt:
        print("\n[*] Interactive mode stopped.")


CONF_FILENAME = "maipper.conf"

_DEFAULT_CONF = """\
# mAIpper Configuration File
# -------------------------
# Settings here are used as defaults. Command-line flags override these values.
# Lines starting with # are comments. Blank lines are ignored.

[general]
# Base directory containing scan subdirectories (nmap, nessus, burp, autorecon, loot, misc)
scans_dir = scans

# Obsidian vault output directory
vault = Obsidian

# Default tool name label for Nmap scans
tool_name = Nmap

# Verbosity level: 0 = WARNING, 1 = INFO, 2 = DEBUG
verbose = 1

[ollama]
# Ollama API URL
ollama_url = http://localhost:11434

# LLM model to use
model = qwen2.5:14b-instruct-q5_K_M

# Sampling temperature (0.0-1.0, lower = more deterministic)
temperature = 0.15

[canvas]
# Canvas filename
canvas_name = Assessment Canvas.canvas

# Host cards per row within each subnet group
canvas_cols = 2

# Subnet groups per canvas row
canvas_groups_per_row = 3

[interactive]
# Start in interactive mode by default (true/false)
interactive = false

# Seconds between file change polls in interactive mode
watch_interval = 30

[parsers]
# Set to true to skip specific parsers by default
no_ollama = false
no_canvas = false
no_nessus = false
no_burp = false
no_autorecon = false
no_loot = false
no_misc = false
skip_validation = false

[rag]
# RAG (Retrieval Augmented Generation) — reference your cybersecurity library
# Requires: pip install pypdf (for PDF books)

# Directory containing PDF reference books (RTFM, PNPT, etc.)
docs_dir = docs

# Directory containing HackTricks markdown clone
# git clone https://github.com/HackTricks-wiki/hacktricks
# hacktricks_dir = hacktricks

# Ollama embedding model (auto-pulled on first use)
embedding_model = nomic-embed-text

# Maximum reference chunks injected per analysis
max_chunks = 5

# Auto-build index on first run if missing (false = explicit --build-index only)
auto_build = true
"""


def _load_config() -> dict:
    """Load maipper.conf from the current directory. Returns a flat dict of defaults."""
    conf_path = Path(CONF_FILENAME)
    if not conf_path.exists():
        return {}

    cp = configparser.ConfigParser()
    cp.read(conf_path, encoding="utf-8")

    defaults: dict = {}
    key_map = {
        ("general", "scans_dir"):          ("scans_dir",          str),
        ("general", "vault"):              ("vault",              str),
        ("general", "tool_name"):          ("tool_name",          str),
        ("general", "verbose"):            ("verbose",            int),
        ("ollama",  "ollama_url"):         ("ollama_url",         str),
        ("ollama",  "model"):             ("model",              str),
        ("ollama",  "temperature"):        ("temperature",        float),
        ("canvas",  "canvas_name"):        ("canvas_name",        str),
        ("canvas",  "canvas_cols"):        ("canvas_cols",        int),
        ("canvas",  "canvas_groups_per_row"): ("canvas_groups_per_row", int),
        ("interactive", "interactive"):     ("interactive",        bool),
        ("interactive", "watch_interval"):  ("watch_interval",     int),
        ("parsers", "no_ollama"):          ("no_ollama",          bool),
        ("parsers", "no_canvas"):          ("no_canvas",          bool),
        ("parsers", "no_nessus"):          ("no_nessus",          bool),
        ("parsers", "no_burp"):            ("no_burp",            bool),
        ("parsers", "no_autorecon"):       ("no_autorecon",       bool),
        ("parsers", "no_loot"):            ("no_loot",            bool),
        ("parsers", "no_misc"):            ("no_misc",            bool),
        ("parsers", "no_excel"):           ("no_excel",           bool),
        ("parsers", "skip_validation"):    ("skip_validation",    bool),
        ("rag", "docs_dir"):               ("rag_docs_dir",       str),
        ("rag", "hacktricks_dir"):         ("rag_hacktricks_dir", str),
        ("rag", "embedding_model"):        ("rag_embedding_model", str),
        ("rag", "max_chunks"):             ("rag_max_chunks",     int),
        ("rag", "auto_build"):             ("rag_auto_build",     bool),
    }

    for (section, key), (arg_name, typ) in key_map.items():
        if cp.has_option(section, key):
            raw = cp.get(section, key).strip()
            try:
                if typ is bool:
                    defaults[arg_name] = raw.lower() in ("true", "yes", "1")
                elif typ is int:
                    defaults[arg_name] = int(raw)
                elif typ is float:
                    defaults[arg_name] = float(raw)
                else:
                    defaults[arg_name] = raw
            except (ValueError, TypeError):
                pass

    return defaults


# Main
# ============================================================

def main() -> None:
    config_defaults = _load_config()

    ap = argparse.ArgumentParser(
        description="mAIpper v0.12 — Pentest scan analysis and Obsidian vault generator"
    )
    ap.add_argument("--config",     default=CONF_FILENAME,
                    help=f"Config file path (default: ./{CONF_FILENAME})")
    ap.add_argument("--init",       action="store_true",
                    help="Initialize scan directory structure and exit")
    ap.add_argument("-i", "--interactive", action="store_true",
                    default=config_defaults.get("interactive", False),
                    help="Interactive mode: watch for changes, chat with LLM, drop data")
    ap.add_argument("--watch-interval", type=int,
                    default=config_defaults.get("watch_interval", 30), metavar="SEC",
                    help="Seconds between file change polls in interactive mode (default: 30)")
    ap.add_argument("--xml",        default=None,  help="Process a single Nmap XML file")
    ap.add_argument("--scans-dir",
                    default=config_defaults.get("scans_dir", "scans"),
                    help="Base scans directory (default: ./scans)")
    ap.add_argument("--vault",
                    default=config_defaults.get("vault", "Obsidian"),
                    help="Obsidian vault output directory")
    ap.add_argument("--tool-name",
                    default=config_defaults.get("tool_name", "Nmap"))
    ap.add_argument("--model",
                    default=config_defaults.get("model", "qwen2.5:14b-instruct-q5_K_M"))
    ap.add_argument("--ollama-url",
                    default=config_defaults.get("ollama_url", "http://localhost:11434"))
    ap.add_argument("--no-ollama",  action="store_true",
                    default=config_defaults.get("no_ollama", False),
                    help="Skip AI analysis")
    ap.add_argument("--no-canvas",  action="store_true",
                    default=config_defaults.get("no_canvas", False),
                    help="Skip canvas generation")
    ap.add_argument("--no-users-canvas", action="store_true",
                    default=config_defaults.get("no_users_canvas", False),
                    help="Skip Users Canvas generation")
    ap.add_argument("--no-nessus",  action="store_true",
                    default=config_defaults.get("no_nessus", False),
                    help="Skip Nessus scan processing")
    ap.add_argument("--no-burp",      action="store_true",
                    default=config_defaults.get("no_burp", False),
                    help="Skip Burp Suite scan processing")
    ap.add_argument("--autorecon",    default=None, metavar="DIR",
                    help="AutoRecon results directory (contains per-target subdirs)")
    ap.add_argument("--no-autorecon", action="store_true",
                    default=config_defaults.get("no_autorecon", False),
                    help="Skip AutoRecon processing")
    ap.add_argument("--loot",         default=None, metavar="DIR",
                    help="Loot directory (default: auto-discover scans/loot/)")
    ap.add_argument("--no-loot",      action="store_true",
                    default=config_defaults.get("no_loot", False),
                    help="Skip loot processing")
    ap.add_argument("--misc",          default=None, metavar="DIR",
                    help="Misc tool output directory (default: auto-discover scans/misc/)")
    ap.add_argument("--no-misc",      action="store_true",
                    default=config_defaults.get("no_misc", False),
                    help="Skip misc processing")
    ap.add_argument("--excel",        action="store_true", help="Generate Export.xlsx (requires openpyxl)")
    ap.add_argument("--canvas-name",
                    default=config_defaults.get("canvas_name", "Assessment Canvas.canvas"))
    ap.add_argument("--canvas-cols",           type=int,
                    default=config_defaults.get("canvas_cols", 2),
                    help="Host cards per row within each subnet group (default: 2)")
    ap.add_argument("--canvas-groups-per-row", type=int,
                    default=config_defaults.get("canvas_groups_per_row", 3),
                    help="Subnet groups per canvas row (default: 3)")
    ap.add_argument("--temperature",     type=float,
                    default=config_defaults.get("temperature", 0.15),
                    help="Ollama sampling temperature (0.0–1.0, default: 0.15)")
    ap.add_argument("--skip-validation", action="store_true",
                    default=config_defaults.get("skip_validation", False),
                    help="Skip post-processing AI output validation (faster, no hallucination checks)")
    ap.add_argument("--reanalyze", action="store_true",
                    help="Force re-analysis of all files (ignore previous state)")
    ap.add_argument("--build-index", action="store_true",
                    help="Build or update the RAG index from docs and HackTricks, then exit")
    ap.add_argument("--no-rag", action="store_true",
                    default=config_defaults.get("no_rag", False),
                    help="Disable RAG context injection")
    ap.add_argument("--rag-docs-dir",
                    default=config_defaults.get("rag_docs_dir", "docs"),
                    metavar="DIR", help="Directory containing PDF reference books (default: docs)")
    ap.add_argument("--rag-hacktricks-dir",
                    default=config_defaults.get("rag_hacktricks_dir", "docs/hacktricks"),
                    metavar="DIR", help="Directory containing HackTricks markdown clone (default: docs/hacktricks)")
    ap.add_argument("--rag-embedding-model",
                    default=config_defaults.get("rag_embedding_model", RAG_DEFAULT_EMBEDDING_MODEL),
                    help=f"Ollama embedding model (default: {RAG_DEFAULT_EMBEDDING_MODEL})")
    ap.add_argument("-v", "--verbose", action="count",
                    default=config_defaults.get("verbose", 1),
                    help="Increase verbosity: -v = INFO, -vv = DEBUG")

    args = ap.parse_args()

    level = logging.WARNING
    if args.verbose >= 1: level = logging.INFO
    if args.verbose >= 2: level = logging.DEBUG
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")
    if config_defaults:
        logging.info(f"Loaded config from {CONF_FILENAME}")
    logging.debug(f"Args: {vars(args)}")

    base = Path(args.scans_dir).resolve()

    # --init: create directory structure and config file, then exit
    if args.init:
        _init_scan_dirs(base)
        # Create docs/ directory for RAG reference books
        docs_path = Path(args.rag_docs_dir).resolve() if args.rag_docs_dir else base.parent / "docs"
        docs_path.mkdir(parents=True, exist_ok=True)
        print(f"\n      {docs_path}/          — PDF reference books for RAG")
        print(f"\n    RAG (optional — cybersecurity reference library):")
        print(f"      1. pip install pypdf")
        print(f"      2. Drop PDF books into {docs_path}/")
        print(f"      3. Optionally clone HackTricks: git clone https://github.com/HackTricks-wiki/hacktricks")
        print(f"      4. Set paths in [rag] section of maipper.conf")
        print(f"      5. python mAIpper-v0.12.py --build-index")
        conf_path = Path(CONF_FILENAME)
        if not conf_path.exists():
            conf_path.write_text(_DEFAULT_CONF, encoding="utf-8")
            print(f"\n[+] Created default config: {conf_path.resolve()}")
        else:
            print(f"\n[*] Config already exists: {conf_path.resolve()}")
        return

    # --build-index: build/update RAG index and exit
    if args.build_index:
        docs_dir = Path(args.rag_docs_dir).resolve() if args.rag_docs_dir else None
        ht_dir = Path(args.rag_hacktricks_dir).resolve() if args.rag_hacktricks_dir else None
        if not docs_dir and not ht_dir:
            print("[!] No RAG source directories configured.")
            print("    Set docs_dir or hacktricks_dir in [rag] section of maipper.conf,")
            print("    or use --rag-docs-dir / --rag-hacktricks-dir flags.")
            return
        source_count = 0
        if docs_dir and docs_dir.exists():
            pdf_count = len(list(docs_dir.rglob("*.pdf")))
            print(f"[*] docs_dir: {docs_dir} ({pdf_count} PDFs)")
            source_count += pdf_count
        elif docs_dir:
            print(f"[!] docs_dir not found: {docs_dir}")
        if ht_dir and ht_dir.exists():
            md_count = len(list(ht_dir.rglob("*.md")))
            print(f"[*] hacktricks_dir: {ht_dir} ({md_count} markdown files)")
            source_count += md_count
        elif ht_dir:
            print(f"[!] hacktricks_dir not found: {ht_dir}")
        if source_count == 0:
            print("[!] No source files found. Add PDFs to docs/ or clone HackTricks.")
            return
        existing = _load_rag_index()
        if existing:
            print(f"[*] Existing index found: {existing['chunk_count']} chunks — updating incrementally")
        print(f"[*] Building RAG index (model: {args.rag_embedding_model})...")
        try:
            meta = _build_rag_index(
                docs_dir if docs_dir and docs_dir.exists() else None,
                ht_dir if ht_dir and ht_dir.exists() else None,
                args.ollama_url,
                args.rag_embedding_model,
                db_path=_rag_db_path(),
            )
            print(f"[+] RAG index built: {meta['chunk_count']} chunks")
        except KeyboardInterrupt:
            print("\n[!] Index build cancelled (progress saved)")
        except Exception as exc:
            print(f"[!] Index build failed: {exc}")
        return

    # Auto-generate config if missing
    conf_path = Path(CONF_FILENAME)
    if not conf_path.exists():
        conf_path.write_text(_DEFAULT_CONF, encoding="utf-8")
        logging.info(f"Generated default config: {conf_path.resolve()}")

    # Auto-initialize when scans/ doesn't exist (unless --xml points to a specific file)
    if not base.exists() and not args.xml:
        logging.info(f"Scan directory not found: {base}")
        _init_scan_dirs(base)
        return

    # Ensure all subdirectories exist even if scans/ was partially created
    if not args.xml:
        for sub in _SCAN_SUBDIRS:
            (base / sub).mkdir(exist_ok=True)

    # Ensure RAG directories exist
    if args.rag_docs_dir:
        Path(args.rag_docs_dir).resolve().mkdir(parents=True, exist_ok=True)
    if args.rag_hacktricks_dir:
        Path(args.rag_hacktricks_dir).resolve().mkdir(parents=True, exist_ok=True)

    if args.interactive:
        _watch_loop(args, base)
    else:
        _run_processing(args, base)


_VAULT_SUBDIRS = ["Hosts", "Scans", "Loot"]

_MAIPPER_CSS = """\
/* mAIpper Investigation Checkboxes */
/* Enable in Obsidian: Settings > Appearance > CSS snippets > maipper */

/* ── Remove strikethrough from all checked items ── */
.markdown-rendered .task-list-item.is-checked,
.markdown-rendered .task-list-item.is-checked *,
.markdown-source-view .task-list-item.is-checked,
.markdown-source-view .task-list-item.is-checked *,
.task-list-item[data-task="x"],
.task-list-item[data-task="x"] *,
.task-list-item[data-task="/"],
.task-list-item[data-task="/"] * {
    text-decoration: none !important;
    color: inherit !important;
}

/* ── Pending investigation [x] — yellow highlight ── */
.task-list-item[data-task="x"] {
    background-color: rgba(255, 193, 7, 0.18) !important;
    border-left: 3px solid #f5a623;
    padding-left: 6px;
    border-radius: 3px;
    margin-bottom: 2px;
}

/* ── Investigated complete [/] — green highlight ── */
.task-list-item[data-task="/"] {
    background-color: rgba(76, 175, 80, 0.18) !important;
    border-left: 3px solid #4caf50;
    padding-left: 6px;
    border-radius: 3px;
    margin-bottom: 2px;
}

.task-list-item[data-task="/"] input[type="checkbox"] {
    accent-color: #4caf50;
}
"""


def _install_vault_css(vault_dir: Path) -> None:
    """Install and auto-enable the mAIpper CSS snippet in the Obsidian vault."""
    obsidian_dir = vault_dir / ".obsidian"
    snippets_dir = obsidian_dir / "snippets"
    snippets_dir.mkdir(parents=True, exist_ok=True)

    css_path = snippets_dir / "maipper.css"
    if not css_path.exists() or css_path.read_text(encoding="utf-8") != _MAIPPER_CSS:
        css_path.write_text(_MAIPPER_CSS, encoding="utf-8")
        logging.info("Installed mAIpper CSS snippet")

    appearance_path = obsidian_dir / "appearance.json"
    try:
        appearance = json.loads(appearance_path.read_text(encoding="utf-8")) if appearance_path.exists() else {}
    except Exception:
        appearance = {}
    enabled = appearance.get("enabledCssSnippets", [])
    if "maipper" not in enabled:
        enabled.append("maipper")
        appearance["enabledCssSnippets"] = enabled
        appearance_path.write_text(json.dumps(appearance, indent=2), encoding="utf-8")
        logging.info("Auto-enabled mAIpper CSS snippet in Obsidian")


def _run_processing(args, base: Path, *, skip_llm: bool = False) -> None:
    """Run the full scan-processing pipeline once.

    When skip_llm=True, all LLM calls are skipped but Python parsing,
    vault writing, and canvas generation still run. Used in interactive
    mode for fast initial processing.
    """
    vault_dir = Path(args.vault).resolve()
    vault_dir.mkdir(parents=True, exist_ok=True)
    for sub in _VAULT_SUBDIRS:
        (vault_dir / sub).mkdir(exist_ok=True)
    _install_vault_css(vault_dir)
    _install_eat_me_page(vault_dir)

    # Merge duplicate host notes before processing (e.g., 10.10.110.4.md + dante-web-nix01.md)
    merge_reports = _detect_and_merge_host_notes(vault_dir)
    for msg in merge_reports:
        logging.info(f"[Merge] {msg}")
        print(f"  [*] Auto-merged: {msg}")

    scan_host_map: dict[str, list[str]]   = {}
    all_analyses:  list[tuple[str, str]]  = []
    scan_path_overrides: dict[str, str]   = {}

    # Load analysis state for incremental processing
    analysis_state: dict[str, float] = {} if args.reanalyze else _load_analysis_state(vault_dir)
    if args.reanalyze:
        logging.info("Reanalyze mode: all files will be re-processed")

    # Load operator notes from existing vault for feedback into prompts
    operator_notes_lookup: dict[str, str] = {}
    if vault_dir.exists():
        operator_notes_lookup = build_operator_notes_lookup(vault_dir)
        if operator_notes_lookup:
            logging.info(f"Loaded operator notes for {len(operator_notes_lookup)} host key(s)")

    # Build known hosts lookup for content-based association in loot/misc
    known_hosts_lookup = _build_known_hosts_lookup(vault_dir)

    # ------------------------------------------------------------------ #
    # RAG — load or build index                                            #
    # ------------------------------------------------------------------ #
    global _RAG_INDEX, _RAG_BUILDER, _RAG_OLLAMA_URL, _RAG_EMBEDDING_MODEL
    _RAG_OLLAMA_URL = args.ollama_url
    _RAG_EMBEDDING_MODEL = getattr(args, "rag_embedding_model", RAG_DEFAULT_EMBEDDING_MODEL)

    rag_enabled = (
        not getattr(args, "no_rag", False)
        and not args.no_ollama
        and (getattr(args, "rag_docs_dir", None) or getattr(args, "rag_hacktricks_dir", None))
    )

    if rag_enabled:
        existing_meta = _load_rag_index()
        if existing_meta:
            _RAG_INDEX = existing_meta
            logging.info(f"RAG index loaded: {existing_meta['chunk_count']} chunks")

        docs_dir_rag = Path(args.rag_docs_dir).resolve() if args.rag_docs_dir else None
        ht_dir_rag = Path(args.rag_hacktricks_dir).resolve() if args.rag_hacktricks_dir else None

        all_pdfs = list(docs_dir_rag.rglob("*.pdf")) if docs_dir_rag and docs_dir_rag.exists() else []
        all_mds = [m for m in (ht_dir_rag.rglob("*.md") if ht_dir_rag and ht_dir_rag.exists() else [])
                   if not m.name.startswith(".") and "SUMMARY" not in m.name]

        indexed_mtimes = _load_rag_file_mtimes() if existing_meta else {}
        new_pdfs = [p for p in all_pdfs if str(p) not in indexed_mtimes or p.stat().st_mtime != indexed_mtimes[str(p)]]
        new_mds = [m for m in all_mds if str(m) not in indexed_mtimes or m.stat().st_mtime != indexed_mtimes[str(m)]]
        new_count = len(new_pdfs) + len(new_mds)

        if new_count > 0:
            parts = []
            if new_pdfs:
                parts.append(f"{len(new_pdfs)} PDFs")
            if new_mds:
                parts.append(f"{len(new_mds)} markdown files")
            if existing_meta:
                print(f"\n  [*] RAG: {', '.join(parts)} new/changed since last index build.")
            else:
                print(f"\n  [*] RAG: Found {', '.join(parts)} not yet indexed.")
            try:
                answer = input("      Build index now in background? (y/n): ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                answer = "n"
            if answer in ("y", "yes"):
                print("  [*] RAG: Building index in background — you can work while it runs.")
                _RAG_BUILDER = _RagIndexBuilder(
                    docs_dir_rag if all_pdfs else None,
                    ht_dir_rag if all_mds else None,
                    args.ollama_url, _RAG_EMBEDDING_MODEL,
                    RAG_DEFAULT_CHUNK_SIZE, RAG_DEFAULT_CHUNK_OVERLAP,
                    _rag_db_path(),
                )
                _RAG_BUILDER.start()
            else:
                print("  [*] RAG: Skipped. Run /build-index when ready.")
        elif len(all_pdfs) + len(all_mds) == 0 and not existing_meta:
            logging.info("RAG: No source files found in docs or hacktricks directories.")

    # ------------------------------------------------------------------ #
    # Eat Me — process operator drop zone first                            #
    # ------------------------------------------------------------------ #
    eat_me_result = _process_eat_me_page(vault_dir, args=args)
    if eat_me_result:
        scan_host_map[eat_me_result["scan_stem"]] = eat_me_result["host_stems"]
        if eat_me_result["new_hosts"]:
            logging.info(
                f"Eat Me: {len(eat_me_result['new_hosts'])} new host(s) created: "
                f"{', '.join(eat_me_result['new_hosts'])}"
            )

    # ------------------------------------------------------------------ #
    # Nmap                                                                 #
    # ------------------------------------------------------------------ #
    all_nmap_scans: list[tuple[str, str, dict]] = []

    xml_files: list[Path] = []
    if args.xml:
        xml_path = Path(args.xml).resolve()
        if not xml_path.exists():
            logging.error(f"File not found: {xml_path}")
            return
        xml_files = [xml_path]
    else:
        nmap_dir = base / "nmap"
        if not nmap_dir.exists():
            nmap_dir = base / "Nmap"
        if nmap_dir.exists():
            xml_files = sorted(nmap_dir.glob("*.xml"))
        else:
            logging.info("No scans/nmap/ directory found; skipping Nmap processing")

    if xml_files:
        logging.info(f"Processing {len(xml_files)} Nmap file(s)")
    for xml_path in xml_files:
        if not _file_needs_analysis(xml_path, analysis_state):
            logging.info(f"Nmap → {xml_path.name} (unchanged, skipping)")
            continue
        logging.info(f"Nmap → {xml_path.name}")
        scan_data = parse_nmap_xml(xml_path)

        if not scan_data["hosts"]:
            logging.info(f"Skipping {xml_path.name}: no hosts with open ports")
            _mark_analyzed(xml_path, analysis_state)
            continue

        analysis: str | None = None
        nmap_warnings: list[str] = []
        if not args.no_ollama and not skip_llm:
            try:
                raw_analysis = ollama_chat(
                    args.ollama_url, args.model,
                    build_ollama_prompt(scan_data, operator_notes_by_ip=operator_notes_lookup),
                    args.temperature,
                )
                if not args.skip_validation:
                    analysis, nmap_warnings = validate_ai_output(
                        raw_analysis, scan_data, "nmap"
                    )
                    for w in nmap_warnings:
                        logging.warning(f"[Nmap Validation] {w}")
                else:
                    analysis = raw_analysis
            except Exception as exc:
                logging.warning(f"Ollama failed ({xml_path.name}): {exc}")
                analysis = f"_Ollama failed: {exc}_"
        else:
            logging.info("Skipping Ollama (--no-ollama)")

        result = create_obsidian_vault(
            vault_dir,
            xml_path.stem,
            scan_data,
            args.tool_name,
            analysis,
            None if (args.no_ollama or skip_llm) else args.model,
            validation_warnings=nmap_warnings or None,
        )

        scan_host_map[result["scan_stem"]] = result["host_stems"]
        if analysis and not args.no_ollama:
            all_analyses.append((result["scan_display"], analysis))
        all_nmap_scans.append((result["scan_stem"], result["scan_display"], scan_data))
        _mark_analyzed(xml_path, analysis_state)

    # ------------------------------------------------------------------ #
    # Nessus                                                               #
    # ------------------------------------------------------------------ #
    all_nessus_scans: list[tuple[str, str, dict]] = []

    if not args.no_nessus:
        nessus_dir = base / "nessus"
        if not nessus_dir.exists():
            nessus_dir = base / "Nessus"
        nessus_files: list[Path] = sorted(nessus_dir.glob("*.nessus")) if nessus_dir.exists() else []

        if nessus_files:
            logging.info(f"Processing {len(nessus_files)} Nessus file(s)")
        for nessus_path in nessus_files:
            if not _file_needs_analysis(nessus_path, analysis_state):
                logging.info(f"Nessus → {nessus_path.name} (unchanged, skipping)")
                continue
            logging.info(f"Nessus → {nessus_path.name}")
            nessus_data = parse_nessus_xml(nessus_path)

            if not nessus_data["hosts"]:
                logging.info(f"Skipping {nessus_path.name}: no hosts with actionable findings")
                _mark_analyzed(nessus_path, analysis_state)
                continue

            analysis = None
            nessus_warnings: list[str] = []
            if not args.no_ollama and not skip_llm:
                try:
                    # Pass 1: fact extraction only
                    facts: str | None = None
                    try:
                        fact_prompt = _build_nessus_fact_extraction_prompt(nessus_data, operator_notes_by_ip=operator_notes_lookup)
                        facts = ollama_chat(
                            args.ollama_url, args.model, fact_prompt, args.temperature
                        )
                        logging.info(
                            f"Nessus Pass 1 complete ({nessus_path.name}): "
                            f"{len(facts)} chars extracted"
                        )
                    except Exception as exc:
                        logging.warning(
                            f"Nessus Pass 1 failed ({nessus_path.name}): {exc}; "
                            "falling back to single-pass"
                        )

                    # Pass 2: analysis (optionally grounded in Pass 1 facts)
                    p2_prompt = build_nessus_ollama_prompt(nessus_data, operator_notes_by_ip=operator_notes_lookup)
                    if facts:
                        p2_prompt = (
                            "The following facts have been extracted directly from the scan "
                            "data. Base your analysis ONLY on these confirmed facts:\n\n"
                            + facts + "\n\n" + p2_prompt
                        )
                    raw_analysis = ollama_chat(
                        args.ollama_url, args.model, p2_prompt, args.temperature
                    )
                    if not args.skip_validation:
                        analysis, nessus_warnings = validate_ai_output(
                            raw_analysis, nessus_data, "nessus"
                        )
                        for w in nessus_warnings:
                            logging.warning(f"[Nessus Validation] {w}")
                    else:
                        analysis = raw_analysis
                except Exception as exc:
                    logging.warning(f"Ollama failed ({nessus_path.name}): {exc}")
                    analysis = f"_Ollama failed: {exc}_"

            result = create_nessus_vault(
                vault_dir,
                nessus_path.stem,
                nessus_data,
                analysis,
                None if (args.no_ollama or skip_llm) else args.model,
                validation_warnings=nessus_warnings or None,
            )

            scan_host_map[result["scan_stem"]] = result["host_stems"]
            if analysis and not args.no_ollama:
                all_analyses.append((result["scan_display"], analysis))
            all_nessus_scans.append((result["scan_stem"], result["scan_display"], nessus_data))
            _mark_analyzed(nessus_path, analysis_state)
    else:
        logging.info("Skipping Nessus (--no-nessus)")

    # ------------------------------------------------------------------ #
    # Burp Suite                                                           #
    # ------------------------------------------------------------------ #
    all_burp_scans: list[tuple[str, str, dict]] = []

    if not args.no_burp:
        burp_dir = base / "burp"
        if not burp_dir.exists():
            burp_dir = base / "Burp"
        burp_files: list[Path] = sorted(burp_dir.glob("*.xml")) if burp_dir.exists() else []

        if burp_files:
            logging.info(f"Processing {len(burp_files)} Burp file(s)")
        for burp_path in burp_files:
            if not _file_needs_analysis(burp_path, analysis_state):
                logging.info(f"Burp → {burp_path.name} (unchanged, skipping)")
                continue
            logging.info(f"Burp → {burp_path.name}")
            # Distinguish Burp XML from Nmap XML by checking root tag
            try:
                root_tag = ET.parse(burp_path).getroot().tag
            except Exception as exc:
                logging.warning(f"Failed to parse {burp_path.name}: {exc}")
                continue
            if root_tag != "issues":
                logging.debug(f"Skipping non-Burp XML: {burp_path.name} (root tag: {root_tag})")
                continue

            burp_data = parse_burp_xml(burp_path)

            analysis = None
            burp_warnings: list[str] = []
            if not args.no_ollama and not skip_llm:
                try:
                    raw_analysis = ollama_chat(
                        args.ollama_url, args.model,
                        build_burp_ollama_prompt(burp_data, operator_notes_by_ip=operator_notes_lookup),
                        args.temperature,
                    )
                    if not args.skip_validation:
                        analysis, burp_warnings = validate_ai_output(
                            raw_analysis, burp_data, "burp"
                        )
                        for w in burp_warnings:
                            logging.warning(f"[Burp Validation] {w}")
                    else:
                        analysis = raw_analysis
                except Exception as exc:
                    logging.warning(f"Ollama failed ({burp_path.name}): {exc}")
                    analysis = f"_Ollama failed: {exc}_"

            result = create_burp_vault(
                vault_dir,
                burp_path.stem,
                burp_data,
                analysis,
                None if (args.no_ollama or skip_llm) else args.model,
                validation_warnings=burp_warnings or None,
            )

            scan_host_map[result["scan_stem"]] = result["host_stems"]
            if analysis and not args.no_ollama:
                all_analyses.append((result["scan_display"], analysis))
            all_burp_scans.append((result["scan_stem"], result["scan_display"], burp_data))
            _mark_analyzed(burp_path, analysis_state)
    else:
        logging.info("Skipping Burp Suite (--no-burp)")

    # ------------------------------------------------------------------ #
    # AutoRecon                                                            #
    # ------------------------------------------------------------------ #
    all_autorecon_scans: list[tuple[str, str, dict]] = []

    if not args.no_autorecon:
        autorecon_dir: Path | None = None
        if args.autorecon:
            autorecon_dir = Path(args.autorecon).resolve()
            if not autorecon_dir.exists():
                logging.error(f"AutoRecon directory not found: {autorecon_dir}")
                autorecon_dir = None
        else:
            candidate = base / "autorecon"
            if not candidate.exists():
                candidate = base / "AutoRecon"
            if candidate.exists():
                autorecon_dir = candidate

        if autorecon_dir:
            logging.info(f"Processing AutoRecon results: {autorecon_dir}")

            # Layer 1: Ingest nmap XMLs into existing Nmap pipeline
            for target_subdir in sorted(autorecon_dir.iterdir()):
                if not target_subdir.is_dir():
                    continue
                xml_dir = target_subdir / "scans" / "xml"
                if not xml_dir.exists():
                    continue
                ar_xml_files = sorted(xml_dir.glob("*.xml"))
                for xml_file in ar_xml_files:
                    if not _file_needs_analysis(xml_file, analysis_state):
                        logging.info(f"AutoRecon Layer 1 → {target_subdir.name}/{xml_file.name} (unchanged, skipping)")
                        continue
                    logging.info(f"AutoRecon Layer 1 (Nmap) → {target_subdir.name}/{xml_file.name}")
                    try:
                        scan_data = parse_nmap_xml(xml_file)
                    except Exception as exc:
                        logging.warning(f"Failed to parse AutoRecon nmap XML {xml_file}: {exc}")
                        continue

                    if not scan_data["hosts"]:
                        logging.info(f"Skipping AutoRecon nmap {xml_file.name}: no hosts with open ports")
                        _mark_analyzed(xml_file, analysis_state)
                        continue

                    nmap_analysis: str | None = None
                    nmap_warnings_ar: list[str] = []
                    if not args.no_ollama and not skip_llm:
                        try:
                            raw = ollama_chat(
                                args.ollama_url, args.model,
                                build_ollama_prompt(scan_data, operator_notes_by_ip=operator_notes_lookup), args.temperature,
                            )
                            if not args.skip_validation:
                                nmap_analysis, nmap_warnings_ar = validate_ai_output(
                                    raw, scan_data, "nmap"
                                )
                                for w in nmap_warnings_ar:
                                    logging.warning(f"[AutoRecon Nmap Validation] {w}")
                            else:
                                nmap_analysis = raw
                        except Exception as exc:
                            logging.warning(f"Ollama failed (AutoRecon Nmap {xml_file.name}): {exc}")
                            nmap_analysis = f"_Ollama failed: {exc}_"

                    result = create_obsidian_vault(
                        vault_dir,
                        xml_file.stem,
                        scan_data,
                        f"Nmap (AutoRecon {target_subdir.name})",
                        nmap_analysis,
                        None if (args.no_ollama or skip_llm) else args.model,
                        validation_warnings=nmap_warnings_ar or None,
                    )
                    scan_host_map[result["scan_stem"]] = result["host_stems"]
                    if nmap_analysis and not args.no_ollama:
                        all_analyses.append((result["scan_display"], nmap_analysis))
                    all_nmap_scans.append((result["scan_stem"], result["scan_display"], scan_data))
                    _mark_analyzed(xml_file, analysis_state)

            # Layer 2: Parse tool outputs and run AutoRecon-specific analysis
            autorecon_data = parse_autorecon_results(autorecon_dir)

            for target in autorecon_data.get("targets", []):
                target_name = target.get("target", "unknown")
                logging.info(f"AutoRecon Layer 2 → {target_name}")

                ar_analysis: str | None = None
                ar_warnings: list[str] = []
                if not args.no_ollama and not skip_llm:
                    try:
                        # Pass 1: fact extraction
                        facts: str | None = None
                        try:
                            fact_prompt = _build_autorecon_fact_extraction_prompt(
                                target, operator_notes=operator_notes_lookup.get(target.get("ip", ""), "")
                            )
                            facts = ollama_chat(
                                args.ollama_url, args.model, fact_prompt, args.temperature,
                            )
                            logging.info(
                                f"AutoRecon Pass 1 complete ({target_name}): {len(facts)} chars"
                            )
                        except Exception as exc:
                            logging.warning(
                                f"AutoRecon Pass 1 failed ({target_name}): {exc}; "
                                "falling back to single-pass"
                            )

                        # Pass 2: analysis (grounded in Pass 1 facts if available)
                        p2_prompt = build_autorecon_ollama_prompt(
                            target, operator_notes=operator_notes_lookup.get(target.get("ip", ""), "")
                        )
                        if facts:
                            p2_prompt = (
                                "The following facts have been extracted directly from the "
                                "enumeration data. Base your analysis ONLY on these confirmed "
                                "facts:\n\n" + facts + "\n\n" + p2_prompt
                            )
                        raw_analysis = ollama_chat(
                            args.ollama_url, args.model, p2_prompt, args.temperature,
                        )
                        if not args.skip_validation:
                            ar_analysis, ar_warnings = validate_ai_output(
                                raw_analysis, {"hosts": [target]}, "autorecon"
                            )
                            for w in ar_warnings:
                                logging.warning(f"[AutoRecon Validation] {w}")
                        else:
                            ar_analysis = raw_analysis
                    except Exception as exc:
                        logging.warning(f"Ollama failed (AutoRecon {target_name}): {exc}")
                        ar_analysis = f"_Ollama failed: {exc}_"

                ar_scan_data = {
                    "source_file": str(autorecon_dir),
                    "parsed_at": autorecon_data["parsed_at"],
                    "report_name": target_name,
                    "targets": [target],
                    "hosts": [target],
                }

                result = create_autorecon_vault(
                    vault_dir,
                    target_name,
                    ar_scan_data,
                    ar_analysis,
                    None if (args.no_ollama or skip_llm) else args.model,
                    validation_warnings=ar_warnings or None,
                )
                scan_host_map[result["scan_stem"]] = result["host_stems"]
                if ar_analysis and not args.no_ollama:
                    all_analyses.append((result["scan_display"], ar_analysis))
                all_autorecon_scans.append((result["scan_stem"], result["scan_display"], ar_scan_data))
    else:
        logging.info("Skipping AutoRecon (--no-autorecon)")

    # ------------------------------------------------------------------ #
    # Loot                                                                 #
    # ------------------------------------------------------------------ #
    all_loot_data: dict | None = None

    if not args.no_loot:
        loot_dir: Path | None = None
        if args.loot:
            loot_dir = Path(args.loot).resolve()
            if not loot_dir.exists():
                logging.error(f"Loot directory not found: {loot_dir}")
                loot_dir = None
        else:
            candidate = base / "loot"
            if not candidate.exists():
                candidate = base / "Loot"
            if candidate.exists():
                loot_dir = candidate

        if loot_dir:
            loot_data = parse_loot_dir(loot_dir, known_hosts=known_hosts_lookup)
            all_loot_data = loot_data

            if not loot_data["summary"]["total_files"]:
                logging.info("Loot directory empty or no processable files found")
            else:
                analysis_by_host: dict[str, str] = {}
                if not args.no_ollama and not skip_llm:
                    hosts_dir = vault_dir / "Hosts"
                    for host_key, loot_files in loot_data["host_loot"].items():
                        host_context = _build_host_context_for_loot(hosts_dir, host_key)
                        op_notes = operator_notes_lookup.get(host_key, "")
                        try:
                            prompt = build_loot_ollama_prompt(
                                host_key, loot_files,
                                host_context=host_context,
                                operator_notes=op_notes,
                            )
                            raw = ollama_chat(
                                args.ollama_url, args.model, prompt, args.temperature,
                            )
                            if not args.skip_validation:
                                analysis_text, loot_warnings = validate_ai_output(
                                    raw, {"hosts": []}, "loot"
                                )
                                for w in loot_warnings:
                                    logging.warning(f"[Loot Validation] {w}")
                            else:
                                analysis_text = raw
                            analysis_by_host[host_key] = analysis_text
                        except Exception as exc:
                            logging.warning(f"Ollama failed (Loot {host_key}): {exc}")
                            analysis_by_host[host_key] = f"_Ollama failed: {exc}_"

                    if loot_data["campaign_loot"]:
                        try:
                            prompt = build_loot_ollama_prompt(
                                "campaign", loot_data["campaign_loot"],
                            )
                            raw = ollama_chat(
                                args.ollama_url, args.model, prompt, args.temperature,
                            )
                            analysis_by_host["_campaign"] = raw
                        except Exception as exc:
                            logging.warning(f"Ollama failed (campaign loot): {exc}")

                result = create_loot_vault(
                    vault_dir, loot_data,
                    analysis_by_host if not args.no_ollama and not skip_llm else None,
                    None if (args.no_ollama or skip_llm) else args.model,
                )
                scan_host_map[result["scan_stem"]] = result["host_stems"]
                if result.get("scan_path_rel"):
                    scan_path_overrides[result["scan_stem"]] = result["scan_path_rel"]
                for host_key, analysis_text in analysis_by_host.items():
                    if analysis_text and not analysis_text.startswith("_Ollama failed"):
                        all_analyses.append((f"Loot ({host_key})", analysis_text))
    else:
        logging.info("Skipping loot (--no-loot)")

    # ------------------------------------------------------------------ #
    # Misc                                                                 #
    # ------------------------------------------------------------------ #
    all_misc_data: dict | None = None
    if not args.no_misc:
        misc_dir: Path | None = None
        if args.misc:
            misc_dir = Path(args.misc).resolve()
            if not misc_dir.exists():
                logging.error(f"Misc directory not found: {misc_dir}")
                misc_dir = None
        else:
            candidate = base / "misc"
            if not candidate.exists():
                candidate = base / "Misc"
            if candidate.exists():
                misc_dir = candidate

        if misc_dir:
            misc_data = parse_misc_dir(misc_dir, known_hosts=known_hosts_lookup)
            all_misc_data = misc_data

            if not misc_data["summary"]["total_files"]:
                logging.info("Misc directory empty or no processable files found")
            else:
                logging.info(f"Processing {misc_data['summary']['total_files']} misc file(s)")
                tool_summary = misc_data["summary"].get("tool_types", {})
                if tool_summary:
                    logging.info(f"Misc tool detection: {tool_summary}")
                analysis_by_file: dict[str, str] = {}
                if not args.no_ollama and not skip_llm:
                    hosts_dir = vault_dir / "Hosts"
                    for file_info in misc_data["files"]:
                        tool_type = file_info.get("tool_type", "unknown")
                        analysis_level = file_info.get("analysis_level", "standard")
                        logging.info(
                            f"Misc → {file_info['filename']} "
                            f"(tool={tool_type}, level={analysis_level})"
                        )
                        host_context = ""
                        if file_info["host_key"]:
                            host_context = _build_host_context_for_loot(
                                hosts_dir, file_info["host_key"]
                            )
                        op_notes = operator_notes_lookup.get(
                            file_info["host_key"] or "", ""
                        )
                        try:
                            prompt = build_misc_ollama_prompt(
                                file_info["filename"],
                                file_info["content"],
                                host_context=host_context,
                                operator_notes=op_notes,
                                tool_type=tool_type,
                                analysis_level=analysis_level,
                            )
                            raw = ollama_chat(
                                args.ollama_url, args.model,
                                prompt, args.temperature,
                            )
                            if not args.skip_validation:
                                analysis_text, misc_warnings = validate_ai_output(
                                    raw, {"hosts": []}, "misc"
                                )
                                for w in misc_warnings:
                                    logging.warning(f"[Misc Validation] {w}")
                            else:
                                analysis_text = raw
                            analysis_by_file[file_info["filename"]] = analysis_text
                        except Exception as exc:
                            logging.warning(
                                f"Ollama failed (Misc {file_info['filename']}): {exc}"
                            )
                            analysis_by_file[file_info["filename"]] = (
                                f"_Ollama failed: {exc}_"
                            )

                misc_results = create_misc_vault(
                    vault_dir, misc_data,
                    analysis_by_file if not args.no_ollama and not skip_llm else None,
                    None if (args.no_ollama or skip_llm) else args.model,
                )
                for mr in misc_results:
                    scan_host_map[mr["scan_stem"]] = mr["host_stems"]
                if not args.no_ollama and not skip_llm:
                    for file_info in misc_data["files"]:
                        a = analysis_by_file.get(file_info["filename"], "")
                        if a and not a.startswith("_Ollama failed"):
                            stem = Path(file_info["filename"]).stem
                            all_analyses.append((f"{stem} - Misc", a))
    else:
        logging.info("Skipping misc (--no-misc)")

    # ------------------------------------------------------------------ #
    # Deep Dives — process checked investigation checkboxes               #
    # ------------------------------------------------------------------ #
    if not args.no_ollama and not skip_llm:
        dd_count = _process_deep_dives(
            vault_dir,
            args.ollama_url,
            args.model,
            args.temperature,
            args.skip_validation,
            operator_notes_lookup,
        )
        if dd_count:
            logging.info(f"Processed {dd_count} deep dive(s)")

    # ------------------------------------------------------------------ #
    # Canvas                                                               #
    # ------------------------------------------------------------------ #
    if not args.no_canvas:
        # Build Priority Targets text: ask Ollama if we have any analyses,
        # otherwise fall back to a static severity-sorted list.
        priority_targets_text: str | None = None
        if not args.no_ollama and not skip_llm and all_analyses:
            try:
                pt_prompt = build_priority_targets_prompt(vault_dir, all_analyses,
                                                          operator_notes_by_ip=operator_notes_lookup)
                raw_pt    = ollama_chat(
                    args.ollama_url, args.model, pt_prompt, args.temperature
                )
                if not args.skip_validation:
                    # Build a combined source for cross-reference (all Nmap hosts)
                    combined_hosts: list[dict] = []
                    for _, _, sd in all_nmap_scans:
                        combined_hosts.extend(sd.get("hosts", []))
                    raw_pt, pt_warnings = validate_ai_output(
                        raw_pt, {"hosts": combined_hosts}, "nmap"
                    )
                    for w in pt_warnings:
                        logging.warning(f"[Priority Targets Validation] {w}")
                # Prepend the standard header so the canvas node is self-labelled
                priority_targets_text = "## Priority Targets\n\n" + raw_pt.strip()
            except Exception as exc:
                logging.warning(f"Ollama priority targets failed: {exc}")
                priority_targets_text = _build_priority_targets_fallback(vault_dir)
        else:
            priority_targets_text = _build_priority_targets_fallback(vault_dir)

        build_canvas(
            vault_dir,
            args.canvas_name,
            scan_host_map,
            all_analyses,
            canvas_cols=args.canvas_cols,
            max_groups_per_row=args.canvas_groups_per_row,
            priority_targets_text=priority_targets_text,
            scan_path_overrides=scan_path_overrides,
        )
    else:
        logging.info("Skipping canvas (--no-canvas)")

    # ------------------------------------------------------------------ #
    # Users Canvas                                                         #
    # ------------------------------------------------------------------ #
    if not args.no_users_canvas and not args.no_canvas:
        if all_loot_data:
            build_users_canvas(vault_dir, all_loot_data)
        else:
            logging.info("Users Canvas: no loot data available, skipping")
    elif args.no_users_canvas:
        logging.info("Skipping Users Canvas (--no-users-canvas)")

    # ------------------------------------------------------------------ #
    # Campaign Targets Note                                                #
    # ------------------------------------------------------------------ #
    _write_campaign_targets_note(vault_dir, loot_data=all_loot_data, misc_data=all_misc_data)

    # ------------------------------------------------------------------ #
    # Excel export                                                         #
    # ------------------------------------------------------------------ #
    if args.excel:
        out = export_excel(vault_dir, all_nmap_scans, all_nessus_scans, all_burp_scans,
                           all_autorecon_scans or None, all_loot_data)
        if out:
            logging.info(f"Excel export: {out}")
        else:
            logging.error("Excel export failed — is openpyxl installed?")

    _save_analysis_state(vault_dir, analysis_state)
    logging.info("Done.")


if __name__ == "__main__":
    main()
